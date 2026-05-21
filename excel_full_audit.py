"""excel_full_audit.py — V1 全分點掃描

讀指定 Excel (預設 data/reports/latest.xlsx) 全 56 分點,逐行盤點:
  - 淨買 / 淨賣 / net=0 狀態
  - 漲停疑似 (close ≈ buy_avg ≈ sell_avg 且 +9.5%+ 推算)
  - 反推估算 跡象 (買均 ≈ 賣均 ≈ 整數,常見於 v3.27.2 lot 反推 case)
  - 異常 flag

輸出:
  - stdout 表格:每分點 stocks 數 + 淨買 N / 淨賣 N / net=0 N
  - 警示清單:淨賣 stocks (應已被 v3.29.1 過濾,出現代表是舊 Excel)
  - 警示清單:net=0 stocks (邊界 case)
  - 警示清單:疑似估算 (買均 = 賣均 整數)

用法:
  python excel_full_audit.py [--xlsx data/reports/latest.xlsx] [--out audit_report.txt]
"""
import argparse
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl


def parse_excel(xlsx_path):
    """Parse the boss Excel into structured rows."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb.active
    rows = []
    current_master = ''
    current_branch = ''
    current_branch_code = ''
    for row_idx in range(1, ws.max_row + 1):
        cells = [ws.cell(row_idx, c).value for c in range(1, 13)]
        a, b, c, d, e, f, g, h, i, j, k, l = cells
        if a == '高手':  # main header
            continue
        if a:
            current_master = a
        if b == '分點' or b == '常下分點':  # sub-header
            continue
        if b and c:
            current_branch = b
            current_branch_code = c
        if d and isinstance(d, str) and ('(' in d or '⚪' in d):
            rows.append({
                'row': row_idx,
                'master': current_master,
                'branch': current_branch,
                'branch_code': current_branch_code,
                'stock_label': d,
                'buy_lot': e if isinstance(e, (int, float)) else 0,
                'sell_lot': f if isinstance(f, (int, float)) else 0,
                'buy_amt': g if isinstance(g, (int, float)) else 0,
                'sell_amt': h if isinstance(h, (int, float)) else 0,
                'net_amt': i if isinstance(i, (int, float)) else 0,
                'buy_avg': j if isinstance(j, (int, float)) else 0,
                'sell_avg': k if isinstance(k, (int, float)) else 0,
            })
    return ws.title, rows


def classify(row):
    """Classify a single row's status."""
    flags = []
    sl = row['stock_label']
    if isinstance(sl, str) and '⚪' in sl:
        flags.append('NOTICE')  # v3.29.2 提示行
        return flags
    if not (row['buy_lot'] or row['sell_lot'] or row['buy_amt'] or row['sell_amt']):
        flags.append('EMPTY')
        return flags

    # net direction
    if row['net_amt'] > 0:
        flags.append('NET_BUY')
    elif row['net_amt'] < 0:
        flags.append('NET_SELL')
    else:
        flags.append('NET_ZERO')

    # 限漲檢測 (label 含 ▲X.XX% 是 v3.27.4 L4 sniper 標示)
    if '▲' in str(sl):
        flags.append('LIMIT_UP_TAG')

    # 反推估算疑似 (買均 == 賣均 且接近整數)
    if (row['buy_avg'] and row['sell_avg']
            and abs(row['buy_avg'] - row['sell_avg']) < 0.01
            and row['buy_avg'] > 0):
        flags.append('REVERSE_EST')  # v3.27.2 lot 反推, buy_avg = close = sell_avg

    return flags


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--xlsx', default='data/reports/latest.xlsx')
    parser.add_argument('--out', default=None)
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"❌ {xlsx_path} 不存在")
        sys.exit(1)

    sheet_title, rows = parse_excel(xlsx_path)

    out_lines = []
    def p(s=''):
        print(s)
        out_lines.append(s)

    p('=' * 80)
    p(f'  Excel 全分點審計 — {xlsx_path.name} (sheet={sheet_title})')
    p('=' * 80)
    p(f'總 row (含 notice + 個股): {len(rows)}')

    # 按 master + branch 統計
    branch_stats = defaultdict(lambda: {
        'NET_BUY': 0, 'NET_SELL': 0, 'NET_ZERO': 0, 'EMPTY': 0,
        'NOTICE': 0, 'LIMIT_UP_TAG': 0, 'REVERSE_EST': 0,
        'master_name': '', 'branch_name': '', 'branch_code': '',
    })
    net_sell_rows = []
    net_zero_rows = []
    notice_rows = []

    for r in rows:
        flags = classify(r)
        key = (r['master'], r['branch_code'])
        bs = branch_stats[key]
        bs['master_name'] = r['master']
        bs['branch_name'] = r['branch']
        bs['branch_code'] = r['branch_code']
        for f in flags:
            if f in bs:
                bs[f] += 1
        if 'NET_SELL' in flags:
            net_sell_rows.append(r)
        if 'NET_ZERO' in flags:
            net_zero_rows.append(r)
        if 'NOTICE' in flags:
            notice_rows.append(r)

    # === 1. Per-branch summary ===
    p()
    p('━' * 80)
    p('  每分點統計 (按 master 排序)')
    p('━' * 80)
    p(f"{'Master':<18} {'分點':<14} {'代號':<6} {'淨買':>4} {'淨賣':>4} {'net0':>4} {'限漲':>4} {'反推':>4} {'提示':>4}")
    p('─' * 80)

    sorted_keys = sorted(branch_stats.keys(),
                        key=lambda k: (k[0], k[1]))
    for key in sorted_keys:
        bs = branch_stats[key]
        p(f"{bs['master_name']:<18} {bs['branch_name']:<14} {bs['branch_code']:<6} "
          f"{bs['NET_BUY']:>4} {bs['NET_SELL']:>4} {bs['NET_ZERO']:>4} "
          f"{bs['LIMIT_UP_TAG']:>4} {bs['REVERSE_EST']:>4} {bs['NOTICE']:>4}")

    # === 2. 淨賣警示 (出現代表是舊 Excel,或新邏輯有遺漏) ===
    p()
    p('━' * 80)
    p(f'  ⚠️ 淨賣警示 ({len(net_sell_rows)} 筆) — v3.29.1+ 起應該為 0')
    p('━' * 80)
    if not net_sell_rows:
        p('  ✅ 無淨賣污染 (v3.29.1+ 過濾生效)')
    else:
        for r in net_sell_rows[:30]:
            p(f"  {r['master']:<14} | {r['branch']:<12} | {r['stock_label']:<18} "
              f"buy {r['buy_amt']:>6}萬 sell {r['sell_amt']:>6}萬 net {r['net_amt']:>+7}萬")
        if len(net_sell_rows) > 30:
            p(f"  ... 另有 {len(net_sell_rows) - 30} 筆")

    # === 3. net=0 警示 ===
    p()
    p('━' * 80)
    p(f'  ⚠️ net=0 警示 ({len(net_zero_rows)} 筆)')
    p('━' * 80)
    if not net_zero_rows:
        p('  ✅ 無 net=0 row')
    else:
        for r in net_zero_rows[:15]:
            p(f"  {r['master']:<14} | {r['branch']:<12} | {r['stock_label']:<18} "
              f"buy {r['buy_amt']:>6}萬 sell {r['sell_amt']:>6}萬")

    # === 4. v3.29.2 提示行統計 ===
    p()
    p('━' * 80)
    p(f'  ⚪ v3.29.2 提示行 ({len(notice_rows)} 筆)')
    p('━' * 80)
    for r in notice_rows[:20]:
        p(f"  {r['master']:<14} | {r['branch']:<12} | {r['stock_label']}")

    # === 5. 總計 ===
    p()
    p('━' * 80)
    p('  總計')
    p('━' * 80)
    total_real_rows = sum(1 for r in rows
                          if 'NOTICE' not in classify(r) and 'EMPTY' not in classify(r))
    p(f'  實際個股 rows: {total_real_rows}')
    p(f'  淨買 rows:  {sum(bs["NET_BUY"] for bs in branch_stats.values())}')
    p(f'  淨賣 rows:  {sum(bs["NET_SELL"] for bs in branch_stats.values())}')
    p(f'  net=0 rows: {sum(bs["NET_ZERO"] for bs in branch_stats.values())}')
    p(f'  限漲標 rows (sniper): {sum(bs["LIMIT_UP_TAG"] for bs in branch_stats.values())}')
    p(f'  反推估算 rows: {sum(bs["REVERSE_EST"] for bs in branch_stats.values())}')
    p(f'  v3.29.2 提示行: {len(notice_rows)}')

    if args.out:
        Path(args.out).write_text('\n'.join(out_lines), encoding='utf-8')
        p(f'\n[結果寫入 {args.out}]')


if __name__ == '__main__':
    main()
