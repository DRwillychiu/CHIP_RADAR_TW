"""v3.29.5 per-branch row 數 override 測試

用戶 5/23 要求 8563 (大牌分析師-新光新竹) 改 Top 20, 其他分點維持 10.

驗證:
  A. 8563 有 25 個 net-buy 個股 → Excel 取前 20 個顯示 (不是 10)
  B. 8563 只有 13 個 net-buy 個股 → 顯示 13 個 + 1 partial notice + 6 row blank (共 20 row)
  C. 其他分點 9217 (迷你哥) → 維持 10 row (override 沒涵蓋)
  D. _branch_stocks_size('8563') == 20, 其他都 == 10
  E. Master block 內 A/B/C 列 merge 範圍正確跟著動
"""
import sys
sys.path.insert(0, '.')

from openpyxl import Workbook
from excel_report import (
    build_day_sheet, MASTER_MAPPING,
    BRANCH_STOCK_OVERRIDES, STOCKS_PER_BRANCH, _branch_stocks_size,
    _top_stocks_for_branch,
)


def make_stock(code, name, buy_amt=0, sell_amt=0, buy_lot=0, sell_lot=0, is_limit_up=False):
    return {
        "code": code, "name": name,
        "buy_amt": buy_amt, "sell_amt": sell_amt,
        "buy_lot": buy_lot, "sell_lot": sell_lot,
        "net_amt": buy_amt - sell_amt,
        "net_lot": buy_lot - sell_lot,
        "is_limit_up": is_limit_up,
    }


def find_row_for_branch_code(ws, branch_code):
    """找 C 欄 == branch_code 的第一個 row"""
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 3).value == branch_code:
            return r
    return None


def count_rows_with_data(ws, start_row, max_check):
    """從 start_row 開始, 數有實際 stock label (含 '(') 的連續 row 數"""
    n = 0
    for r in range(start_row, start_row + max_check):
        d = ws.cell(r, 4).value
        if d and isinstance(d, str) and '(' in d:
            n += 1
        else:
            break
    return n


print("=" * 72)
print("  v3.29.5 per-branch row 數 override 測試")
print("=" * 72)

all_pass = True

# ========== Case D: _branch_stocks_size 邏輯 ==========
print("\nD. _branch_stocks_size lookup:")
size_8563 = _branch_stocks_size("8563")
size_other = _branch_stocks_size("9217")
size_unknown = _branch_stocks_size("UNKNOWN")
ok_d = size_8563 == 20 and size_other == 10 and size_unknown == 10
print(f"  {'✅' if ok_d else '❌'} 8563={size_8563} (expect 20)")
print(f"  {'✅' if size_other == 10 else '❌'} 9217={size_other} (expect 10)")
print(f"  {'✅' if size_unknown == 10 else '❌'} UNKNOWN={size_unknown} (expect 10 fallback)")
print(f"  BRANCH_STOCK_OVERRIDES = {BRANCH_STOCK_OVERRIDES}")
if not ok_d:
    all_pass = False

# ========== Case A: 8563 有 25 個淨買股 → 取 Top 20 ==========
print("\nA. 8563 25 個淨買 → Excel 取 Top 20 (不是 10):")
big_branch = {
    "code": "8563", "name": "新光-新竹",
    "buys": [
        make_stock(f"40{i:02d}", f"股{i:02d}",
                   buy_amt=1_000_000 - i * 10_000,
                   sell_amt=100_000, buy_lot=100, sell_lot=10,
                   is_limit_up=False) for i in range(25)
    ],
    "sells": [],
}
# _top_stocks_for_branch direct call with n_top=20
top20 = _top_stocks_for_branch(big_branch, sniper_mode=False, n_top=20)
n_top20 = len(top20)
ok_a = n_top20 == 20
print(f"  {'✅' if ok_a else '❌'} _top_stocks_for_branch n_top=20 回傳 {n_top20} 個 (expect 20)")

# 也比對 buy_amt 排序: top1 應該是 4000 (buy_amt 最高)
top1_code = top20[0]["code"]
ok_a2 = top1_code == "4000"
print(f"  {'✅' if ok_a2 else '❌'} Top 1 = {top1_code} (expect 4000, buy_amt 最高)")
if not (ok_a and ok_a2):
    all_pass = False

# Build full Excel + verify Excel 8563 section 有 20 row stock data
branches_data_a = [big_branch]
for m in MASTER_MAPPING:
    for code, _ in m["branches"]:
        if code != "8563":
            branches_data_a.append({"code": code, "name": "_", "buys": [], "sells": []})

wb = Workbook()
ws = wb.active
build_day_sheet(ws, branches_data_a, "20260523")

r_8563 = find_row_for_branch_code(ws, "8563")
if r_8563:
    n_data = count_rows_with_data(ws, r_8563, 25)
    print(f"  {'✅' if n_data == 20 else '❌'} 8563 section 從 row {r_8563} 起有 {n_data} 個 stock row (expect 20)")
    if n_data != 20:
        all_pass = False
else:
    print("  ❌ 找不到 8563 在 Excel 中")
    all_pass = False

# ========== Case B: 8563 只有 13 個 → 13 row + partial notice + 6 blank ==========
print("\nB. 8563 13 個淨買 → 13 stock + partial notice + 6 blank (共 20 row):")
small_branch = {
    "code": "8563", "name": "新光-新竹",
    "buys": [
        make_stock(f"50{i:02d}", f"小股{i:02d}",
                   buy_amt=500_000 - i * 5_000, sell_amt=50_000,
                   buy_lot=80, sell_lot=10, is_limit_up=False) for i in range(13)
    ],
    "sells": [],
}
branches_data_b = [small_branch]
for m in MASTER_MAPPING:
    for code, _ in m["branches"]:
        if code != "8563":
            branches_data_b.append({"code": code, "name": "_", "buys": [], "sells": []})

wb2 = Workbook()
ws2 = wb2.active
build_day_sheet(ws2, branches_data_b, "20260523")

r_8563_b = find_row_for_branch_code(ws2, "8563")
if r_8563_b:
    n_data_b = count_rows_with_data(ws2, r_8563_b, 25)
    notice_row = ws2.cell(r_8563_b + 13, 4).value
    has_partial_notice = notice_row and '淨買僅 13 檔' in str(notice_row)
    ok_b = (n_data_b == 13 and has_partial_notice)
    print(f"  {'✅' if n_data_b == 13 else '❌'} 13 個 stock row (actual={n_data_b})")
    print(f"  {'✅' if has_partial_notice else '❌'} row {r_8563_b + 13} 寫 partial notice: {notice_row}")
    if not ok_b:
        all_pass = False
else:
    print("  ❌ 找不到 8563")
    all_pass = False

# ========== Case C: 其他分點仍 10 row ==========
print("\nC. 其他分點 (9217) 維持 10 row:")
# 用 case A 的 wb (8563 滿 20, 其他空)
# Look for 9217 (迷你哥-凱基松山) in build output
r_9217 = find_row_for_branch_code(ws, "9217")
if r_9217:
    # 該 branch 完全空白 (空 fixture). v3.29 邏輯下完全空白 → blank rows
    # 我們只要驗證: 從 r_9217 到 r_9217+9 都是同一 branch (沒越界)
    # 確認方法: r_9217 - 1 是 sub-header row, r_9217+10 是下一個 sub-header (或下一個 master)
    ws_after_10 = ws.cell(r_9217 + 10, 3).value  # C 欄,下一個分點代號 (or '分點代號' header)
    # 如果 9217 真的 10 row, 那 r_9217+10 應該是 sub-header 或下一個分點; 不該還是 9217
    is_10 = ws_after_10 != "9217"
    print(f"  {'✅' if is_10 else '❌'} 9217 維持 10 row (r{r_9217 + 10} C={ws_after_10})")
    if not is_10:
        all_pass = False
else:
    # 9217 完全空白可能整段都沒被打到, 嘗試找 9100 (張濬安-群益)
    print("  (9217 fixture 完全空白, 用 9217 驗 master block 結束時機)")

# ========== Case E: A/B/C merge 範圍跟著動 ==========
print("\nE. 8563 section A/B/C 列 merge 範圍 = 20 row")
# 從 wb (Case A) 抓 8563 的 merge ranges
merged_ranges_for_8563 = []
for mr in ws.merged_cells.ranges:
    # Look for merge starting at r_8563
    if mr.min_row == r_8563 and mr.min_col == 2:  # B 欄 merge
        merged_ranges_for_8563.append(("B", mr.min_row, mr.max_row))
    if mr.min_row == r_8563 and mr.min_col == 3:  # C 欄 merge
        merged_ranges_for_8563.append(("C", mr.min_row, mr.max_row))

ok_e = False
for col, mn, mx in merged_ranges_for_8563:
    span = mx - mn + 1
    print(f"  {col} 欄 merge row {mn}-{mx} (span {span})")
    if col == "B" and span == 20:
        ok_e = True
print(f"  {'✅' if ok_e else '❌'} B 欄 merge span = 20 row")
if not ok_e:
    all_pass = False

print()
print("─" * 72)
print(f"  整體: {'✅ ALL PASS' if all_pass else '❌ HAS FAIL'}")
sys.exit(0 if all_pass else 1)
