# -*- coding: utf-8 -*-
"""v3.72.4 sniper 漲停股「全市場」買超#1 黃色 highlight 測試

v3.72.4 修訂: 判定範圍改用 histock 全市場分點榜, 不再限於 tracked branches.

驗證:
  1. _fetch_histock_top_buyer 用 cache 避免重複 fetch
  2. _build_top_net_buyer_index 只 fetch sniper_stock_codes 內的
  3. 端對端: histock top = 蔣承翰 bno → 黃色; histock top = 別人 → 不塗黃
  4. sniper 沒買漲停 → 不 fetch (省流量)
  5. histock fetch 失敗 → 不 crash, 不塗黃 (safe fallback)
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.stdout.reconfigure(encoding='utf-8')

from unittest.mock import patch
from openpyxl import Workbook

from src.exports.excel_report import (
    _fetch_histock_top_buyer,
    _build_top_net_buyer_index,
    build_day_sheet,
    SNIPER_MASTER_WHITELIST,
)

YELLOW = "FFFFFF00"

pass_count = 0
fail_count = 0
def check(label, cond):
    global pass_count, fail_count
    if cond:
        print(f"  ✅ {label}")
        pass_count += 1
    else:
        print(f"  ❌ {label}")
        fail_count += 1

# Mock histock 資料
MOCK_HISTOCK = {
    "6577": {  # 蔣承翰 9A9S 是 top #1
        "buys": [
            {"bno": "9A9S", "name": "永豐金-南京", "net": 22000},
            {"bno": "OTHER", "name": "其他-分點", "net": 15000},
        ]
    },
    "2330": {  # 台積電 - top #1 不是蔣承翰
        "buys": [
            {"bno": "1480", "name": "元大", "net": 800000},
            {"bno": "9227", "name": "凱基-城中", "net": 100000},
        ]
    },
    "9999": None,  # fetch 失敗
}

def mock_fetch_histock_branch(stock_code, timeout=8, max_retries=1):
    return MOCK_HISTOCK.get(stock_code)


# ─── 1. cache 避免重複 fetch ───
print("\n1. _fetch_histock_top_buyer cache 生效")
with patch("src.audit.histock_branch_audit.fetch_histock_branch",
           side_effect=mock_fetch_histock_branch) as m:
    cache = {}
    _fetch_histock_top_buyer("6577", cache)
    _fetch_histock_top_buyer("6577", cache)  # 應該直接 hit cache
    check("6577 cache 存在", cache.get("6577") == "9A9S")
    check("僅 fetch 1 次 (cache hit)", m.call_count == 1)

# ─── 2. _build_top_net_buyer_index 只 fetch sniper 買的漲停 ───
print("\n2. _build_top_net_buyer_index 只 fetch sniper_stock_codes")
with patch("src.audit.histock_branch_audit.fetch_histock_branch",
           side_effect=mock_fetch_histock_branch) as m:
    idx = _build_top_net_buyer_index(branches_data=[],
                                      sniper_stock_codes={"6577", "2330"})
    check("6577 → 9A9S", idx.get("6577") == "9A9S")
    check("2330 → 1480", idx.get("2330") == "1480")
    check("fetch 2 次 (每股一次)", m.call_count == 2)

# ─── 3. sniper_stock_codes=None → 不 fetch, 回空 ───
print("\n3. 無 sniper 買漲停 → 不 fetch")
with patch("src.audit.histock_branch_audit.fetch_histock_branch",
           side_effect=mock_fetch_histock_branch) as m:
    idx = _build_top_net_buyer_index(branches_data=[{"code": "9227", "buys": [], "sells": []}],
                                      sniper_stock_codes=None)
    check("空 index", idx == {})
    check("0 次 fetch", m.call_count == 0)

# ─── 4. histock fetch 失敗 → 不塗黃 (safe) ───
print("\n4. histock 失敗 → 不塗黃, 不 crash")
with patch("src.audit.histock_branch_audit.fetch_histock_branch",
           side_effect=mock_fetch_histock_branch):
    cache = {}
    result = _fetch_histock_top_buyer("9999", cache)
    check("9999 (mock None) → 回 None", result is None)
    check("cache 9999 = None (不重試)", cache.get("9999") is None)

# ─── 5. 端對端: 9A9S 買 6577 漲停 + histock 顯示 9A9S #1 → 黃色 ───
print("\n5. 端對端: 9A9S 是 6577 histock top #1 → 黃色 highlight")
sample_stock = {
    "code": "6577", "name": "勁豐",
    "buy_lot": 22, "sell_lot": 0,
    "buy_amt": 200000, "sell_amt": 0,
    "net_amt": 200000, "net_lot": 22,
    "is_limit_up": True,
}
branches_data = [
    {"code": "9227", "name": "凱基-城中", "buys": [], "sells": []},
    {"code": "9B18", "name": "台新-建北", "buys": [], "sells": []},
    {"code": "9A9S", "name": "永豐金-南京", "buys": [sample_stock], "sells": []},
]
wb = Workbook()
ws = wb.active
with patch("src.audit.histock_branch_audit.fetch_histock_branch",
           side_effect=mock_fetch_histock_branch):
    build_day_sheet(ws, branches_data, "20260722")

# 找 6577 row 檢查黃色
found_yellow = False
for row in range(1, ws.max_row + 1):
    val_d = ws.cell(row=row, column=4).value
    if val_d and "6577" in str(val_d):
        buy_lot = ws.cell(row=row, column=5).value
        if buy_lot == 22:  # 9A9S 的
            fill = ws.cell(row=row, column=4).fill
            rgb = getattr(getattr(fill, 'fgColor', None), 'rgb', None)
            found_yellow = (rgb == YELLOW)

check("9A9S 6577 row D 欄 = 黃色 FFFFFF00", found_yellow)

# ─── 6. 若 histock top #1 不是我方 → 不塗黃 ───
print("\n6. histock top 不是蔣承翰 → 不塗黃")
sample_stock_wrong = {
    "code": "2330", "name": "台積電",
    "buy_lot": 5, "sell_lot": 0,
    "buy_amt": 500000, "sell_amt": 0,
    "net_amt": 500000, "net_lot": 5,
    "is_limit_up": True,
}
branches_data2 = [
    {"code": "9227", "name": "凱基-城中", "buys": [sample_stock_wrong], "sells": []},
    {"code": "9B18", "name": "台新-建北", "buys": [], "sells": []},
    {"code": "9A9S", "name": "永豐金-南京", "buys": [], "sells": []},
]
wb2 = Workbook()
ws2 = wb2.active
with patch("src.audit.histock_branch_audit.fetch_histock_branch",
           side_effect=mock_fetch_histock_branch):
    build_day_sheet(ws2, branches_data2, "20260722")

found_not_yellow = False
for row in range(1, ws2.max_row + 1):
    val_d = ws2.cell(row=row, column=4).value
    if val_d and "2330" in str(val_d):
        buy_lot = ws2.cell(row=row, column=5).value
        if buy_lot == 5:  # 9227 的 (蔣承翰買, 但 histock top #1 = 元大 1480)
            fill = ws2.cell(row=row, column=4).fill
            rgb = getattr(getattr(fill, 'fgColor', None), 'rgb', None)
            found_not_yellow = (rgb != YELLOW)

check("9227 買 2330 但 histock top=1480 → 不塗黃", found_not_yellow)

# ─── 7. SNIPER_MASTER_WHITELIST 保持 ───
print("\n7. SNIPER_MASTER_WHITELIST 只含蔣承翰")
check("白名單 = 蔣承翰", SNIPER_MASTER_WHITELIST == {"蔣承翰"})

# ─── 8. v3.72.5 histock 時效 guard ───
print("\n8. v3.72.5 histock date != trade_date → 不塗黃 (防假信號)")
MOCK_HISTOCK_STALE = {
    "8888": {  # 日期是 T-1
        "date": "2026/07/23",  # 昨日
        "buys": [{"bno": "9A9S", "name": "永豐金-南京", "net": 100}]
    }
}
def mock_stale(stock_code, timeout=8, max_retries=1):
    return MOCK_HISTOCK_STALE.get(stock_code)

with patch("src.audit.histock_branch_audit.fetch_histock_branch", side_effect=mock_stale):
    cache = {}
    result = _fetch_histock_top_buyer("8888", cache, trade_date="20260724")  # 今日
    check("stale histock (2026/07/23 vs trade_date 20260724) → None", result is None)

with patch("src.audit.histock_branch_audit.fetch_histock_branch", side_effect=mock_stale):
    cache = {}
    result = _fetch_histock_top_buyer("8888", cache, trade_date="20260723")  # 同日
    check("fresh histock (date match) → 9A9S", result == "9A9S")

with patch("src.audit.histock_branch_audit.fetch_histock_branch", side_effect=mock_stale):
    cache = {}
    result = _fetch_histock_top_buyer("8888", cache, trade_date=None)  # 不 guard
    check("trade_date=None → 不 guard, 正常回 9A9S (b/c compat)", result == "9A9S")

# ─── 9. v3.72.5 drift guard ───
print("\n9. v3.72.5 MASTER_MAPPING drift guard 存在且 non-fatal")
from src.exports.excel_report import _MASTER_MAPPING_WARNINGS
check("_MASTER_MAPPING_WARNINGS list 存在 (可能 empty 也可能有既有 drift)",
      isinstance(_MASTER_MAPPING_WARNINGS, list))
# 蔣承翰 9A9S 這次改動不該產生新 warning
new_warnings_for_new_branch = [w for w in _MASTER_MAPPING_WARNINGS if "9A9S" in w]
check("蔣承翰 9A9S 沒 drift warning (branches.py + MASTER_MAPPING 同步)",
      len(new_warnings_for_new_branch) == 0)

# ─── v3.72.8 helpers 匯入 ───
from src.exports.excel_report import (
    _write_histock_status_notice, _write_histock_timestamp_footer,
    _get_histock_stats, _reset_histock_stats,
)

# ─── 10. v3.72.7 histock stats + net<=0 guard ───
print("\n10. v3.72.7 histock fetch stats 收集")

# Scenario A: success case
_reset_histock_stats()
with patch("src.audit.histock_branch_audit.fetch_histock_branch",
           side_effect=mock_fetch_histock_branch):
    cache = {}
    _fetch_histock_top_buyer("6577", cache)
stats = _get_histock_stats()
check("成功後 stats.success++", stats["success"] == 1 and stats["attempted"] == 1)

# Scenario B: net<=0 guard (histock 買方榜 top net <= 0 → 不算 top #1)
_reset_histock_stats()
NET_ZERO_MOCK = {"XXXX": {"date": "2026/07/24",
                          "buys": [{"bno": "9A9S", "net": 0, "buy_lot": 10, "sell_lot": 10}]}}
def mock_net_zero(code, timeout=15, max_retries=2):
    return NET_ZERO_MOCK.get(code)
with patch("src.audit.histock_branch_audit.fetch_histock_branch", side_effect=mock_net_zero):
    cache = {}
    result = _fetch_histock_top_buyer("XXXX", cache)
stats = _get_histock_stats()
check("net=0 top → return None (bug fix)", result is None)
check("stats.net_zero_or_neg++", stats["net_zero_or_neg"] == 1)

# Scenario C: stale date stats
_reset_histock_stats()
with patch("src.audit.histock_branch_audit.fetch_histock_branch", side_effect=mock_stale):
    cache = {}
    _fetch_histock_top_buyer("8888", cache, trade_date="20260724")
stats = _get_histock_stats()
check("stale_date stats++", stats["stale_date"] == 1)

# Scenario D: v3.72.10 fetch_fail (fetch_histock_branch 回 None) vs empty_buys
_reset_histock_stats()
def mock_none_returned(code, timeout=15, max_retries=2):
    return None  # 模擬 timeout / block
with patch("src.audit.histock_branch_audit.fetch_histock_branch", side_effect=mock_none_returned):
    cache = {}
    _fetch_histock_top_buyer("Z1", cache)
stats = _get_histock_stats()
check("fetch=None → fetch_fail 分類 (v3.72.10)", stats.get("fetch_fail") == 1)

# empty_buys separate
_reset_histock_stats()
def mock_empty_buys(code, timeout=15, max_retries=2):
    return {"date": "2026/07/24", "buys": []}
with patch("src.audit.histock_branch_audit.fetch_histock_branch", side_effect=mock_empty_buys):
    cache = {}
    _fetch_histock_top_buyer("Z2", cache)
stats = _get_histock_stats()
check("empty buys → empty_buys 分類 (v3.72.10)", stats.get("empty_buys") == 1)

# ─── 11. v3.72.8 Bug #4 histock 全 fail → 警示 row ───
print("\n11. v3.72.8 histock 全 fail → 警示 row 寫入")
wb3 = Workbook()
ws3 = wb3.active
# 模擬全 fetch fail (v3.72.10 新 key)
stats_fail = {"attempted": 3, "success": 0, "stale_date": 0, "fetch_fail": 3,
              "empty_buys": 0, "http_error": 0, "net_zero_or_neg": 0}
used = _write_histock_status_notice(ws3, 1, stats_fail)
check("全 fetch_fail → 寫警示 row (used=1)", used == 1)
notice_val = ws3.cell(row=1, column=4).value
check("警示 notice 內容含「無 top-buyer」", "無 top-buyer" in (notice_val or ""))
check("v3.72.10 診斷正確 (「timeout / block」關鍵字)",
      "timeout" in (notice_val or "") or "block" in (notice_val or ""))
# 檢查橘色 fill (FFFFECB3)
fill = ws3.cell(row=1, column=1).fill
rgb = getattr(getattr(fill, 'fgColor', None), 'rgb', None)
check("警示 row 背景 = 橘色 FFFFECB3", rgb == "FFFFECB3")

# 100% success → 不寫警示
wb4 = Workbook()
ws4 = wb4.active
stats_ok = {"attempted": 3, "success": 3, "stale_date": 0, "fetch_fail": 0,
            "empty_buys": 0, "http_error": 0, "net_zero_or_neg": 0}
used = _write_histock_status_notice(ws4, 1, stats_ok)
check("100% success → 不寫警示 (used=0)", used == 0)

# 部分 fail (2/3)
wb5 = Workbook()
ws5 = wb5.active
stats_partial = {"attempted": 3, "success": 1, "stale_date": 1, "fetch_fail": 1,
                 "empty_buys": 0, "http_error": 0, "net_zero_or_neg": 0}
used = _write_histock_status_notice(ws5, 1, stats_partial)
check("部分 fail → 寫警示 (used=1)", used == 1)
notice_val = ws5.cell(row=1, column=4).value
check("警示含「部分 top-buyer」", "部分" in (notice_val or ""))

# ─── 12.5 v3.72.11 build_day_sheet 用 precomputed (skip 二次 fetch) ───
print("\n12.5 v3.72.11 build_day_sheet precomputed path (share single fetch)")
from src.exports.excel_report import build_day_sheet

sample_stock_v11 = {
    "code": "6577", "name": "勁豐",
    "buy_lot": 22, "sell_lot": 0,
    "buy_amt": 200000, "sell_amt": 0,
    "net_amt": 200000, "net_lot": 22,
    "is_limit_up": True,
}
branches_data_v11 = [
    {"code": "9A9S", "name": "永豐金-南京", "buys": [sample_stock_v11], "sells": []},
]
wb_v11 = Workbook()
ws_v11 = wb_v11.active

# Feed precomputed (enricher 已 fetch 過的結果)
precomputed_top = {"6577": "9A9S"}
precomputed_stats = {"attempted": 1, "success": 1, "stale_date": 0,
                     "fetch_fail": 0, "empty_buys": 0,
                     "http_error": 0, "net_zero_or_neg": 0}

fetch_called = [False]
def spy_fetch(code, timeout=15, max_retries=2):
    fetch_called[0] = True
    return None
with patch("src.audit.histock_branch_audit.fetch_histock_branch", side_effect=spy_fetch):
    build_day_sheet(ws_v11, branches_data_v11, "20260731",
                     precomputed_top_buyer=precomputed_top,
                     precomputed_stats=precomputed_stats)

check("precomputed 傳入 → 不再呼叫 histock", fetch_called[0] is False)
# 檢查 9A9S 6577 row 是否黃色
for row in range(1, ws_v11.max_row + 1):
    val_d = ws_v11.cell(row=row, column=4).value
    if val_d and "6577" in str(val_d):
        buy_lot = ws_v11.cell(row=row, column=5).value
        if buy_lot == 22:  # 9A9S 的 row
            fill = ws_v11.cell(row=row, column=4).fill
            rgb = getattr(getattr(fill, 'fgColor', None), 'rgb', None)
            check("precomputed → 6577 row 仍塗黃 (使用 precomputed_top_buyer)", rgb == YELLOW)
            break

# ─── 12.6 v3.72.11 enricher circuit breaker ───
print("\n12.6 v3.72.11 enricher circuit breaker (連續 3 fail → abort)")
from src.analyzers.sniper_top_buyer_enricher import enrich_sniper_top_buyer

# 10 檔股票, 全部 fetch 都失敗 → 應在 3 檔後 circuit break
limit_up_summary_cb = {
    'sniper_ranking': [
        {
            'master': '蔣承翰',
            'branch_code': '9A9S',
            'branch_name': '永豐金-南京',
            'limit_up_details': [
                {'code': f'CB{i:02d}', 'name': f'Stock{i}', 'buy_amt': 100} for i in range(10)
            ]
        }
    ]
}
fetch_count_cb = [0]
def spy_fetch_fail(code, timeout=15, max_retries=2):
    fetch_count_cb[0] += 1
    return None  # 全 fail

with patch("src.audit.histock_branch_audit.fetch_histock_branch", side_effect=spy_fetch_fail):
    result_cb = enrich_sniper_top_buyer(limit_up_summary_cb, {"蔣承翰"}, trade_date="20260731")

# 只有 3 檔被 fetch, 之後 abort
check("Circuit breaker 觸發: 只 fetch 3 檔 (連續 3 fail abort)", fetch_count_cb[0] == 3)
check("Circuit breaker after 記錄 abort 位置", result_cb.get('circuit_break_after') == 3)
check("Total targets = 10", result_cb.get('total_targets') == 10)

# ─── 12. v3.72.8 Bug #8 histock 時間戳 footer + v3.72.10 TW timezone ───
print("\n12. v3.72.8 histock timestamp footer + v3.72.10 TW tz")
wb6 = Workbook()
ws6 = wb6.active
used = _write_histock_timestamp_footer(ws6, 100, stats_ok)
check("attempted>0 → 寫時間戳 (used=1)", used == 1)
val = ws6.cell(row=100, column=1).value
check("時間戳含 'histock top-buyer 資料 fetched'", "histock top-buyer 資料 fetched" in (val or ""))
# v3.72.10: 用 TW timezone
check("v3.72.10 時間戳含 'TW' 標記", "TW" in (val or ""))

# attempted=0 → 不寫
wb7 = Workbook()
ws7 = wb7.active
stats_empty = {"attempted": 0, "success": 0, "stale_date": 0, "fetch_fail": 0,
               "empty_buys": 0, "http_error": 0, "net_zero_or_neg": 0}
used = _write_histock_timestamp_footer(ws7, 100, stats_empty)
check("attempted=0 → 不寫 (used=0)", used == 0)

# ─── 總結 ───
print(f"\n{'─' * 60}")
print(f"整體: {pass_count} pass / {fail_count} fail")
sys.exit(0 if fail_count == 0 else 1)
