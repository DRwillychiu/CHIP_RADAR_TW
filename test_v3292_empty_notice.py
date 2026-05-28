"""v3.29.2 空白分點 by-design 提示行測試

驗證:
  1. sniper master 該分點有 TWSE 資料但 filter 後 0 個漲停淨買 → 第 1 列 D 欄
     寫「⚪ 此分點今日未搶漲停」
  2. swing master 該分點有 TWSE 資料但 filter 後 0 個淨買 → 第 1 列 D 欄
     寫「⚪ 此分點今日無淨買超個股」
  3. 該分點 TWSE 完全沒返回資料 (bdata 空) → 不寫提示, 全空白 (避免誤導)
  4. 該分點有 stocks 入選 → 不寫提示 (跟既有行為一致)
"""
import sys
sys.path.insert(0, '.')

from openpyxl import Workbook
from excel_report import build_day_sheet, MASTER_MAPPING


def make_stock(code, name, buy_amt=0, sell_amt=0, buy_lot=0, sell_lot=0, is_limit_up=False):
    return {
        "code": code, "name": name,
        "buy_amt": buy_amt, "sell_amt": sell_amt,
        "buy_lot": buy_lot, "sell_lot": sell_lot,
        "net_amt": buy_amt - sell_amt,
        "net_lot": buy_lot - sell_lot,
        "is_limit_up": is_limit_up,
    }


def find_master_block(ws, master_name):
    """找指定 master 在 ws 中第一個 branch block 的 first row"""
    for row_idx in range(1, ws.max_row + 1):
        a = ws.cell(row_idx, 1).value
        if a == master_name:
            return row_idx
    return None


print("=" * 70)
print("  v3.29.2 空白分點 by-design 提示行測試")
print("=" * 70)

all_pass = True

# ========== Case A: sniper master 該分點有交易但全濾掉 ==========
# v3.30.5: sniper 白名單僅蔣承翰, 用他第 1 個分點 9227 (凱基-城中)
print("\nA. 蔣承翰 9227: 有買榜+賣榜但無漲停淨買 → 提示「未搶漲停」")
branches_data_a = [
    {
        "code": "9227", "name": "凱基-城中",
        "buys": [
            make_stock("2317", "鴻海", buy_amt=400_000, sell_amt=100_000, is_limit_up=False),
            make_stock("2330", "台積電", buy_amt=600_000, sell_amt=50_000, is_limit_up=False),
        ],
        "sells": [
            make_stock("2454", "聯發科", buy_amt=50_000, sell_amt=400_000, is_limit_up=False),
        ],
    },
]
# 加其他必要分點 0 資料避免錯
for m in MASTER_MAPPING:
    for code, _ in m["branches"]:
        if code != "9227":
            branches_data_a.append({"code": code, "name": "_", "buys": [], "sells": []})

wb = Workbook()
ws = wb.active
build_day_sheet(ws, branches_data_a, "20260519")

# 找 蔣承翰 master block
master_row = find_master_block(ws, "蔣承翰")
if master_row is None:
    print("  ❌ 找不到 蔣承翰 row [FAIL]")
    all_pass = False
else:
    # 該 row 對應的 D 欄應該是「⚪ 此分點今日未搶漲停」(因為 9227 是 master 第一個分點)
    d_value = ws.cell(master_row, 4).value
    if d_value and '未搶漲停' in str(d_value):
        print(f"  ✅ row {master_row} D 欄: {d_value}")
    else:
        # 也可能是另一個分點順序, 看看接下來幾列
        found = False
        for r in range(master_row, master_row + 12):
            d = ws.cell(r, 4).value
            if d and '未搶漲停' in str(d):
                print(f"  ✅ row {r} D 欄: {d}")
                found = True
                break
        if not found:
            print(f"  ❌ 沒找到「未搶漲停」提示 [FAIL]")
            all_pass = False

# ========== Case B: swing master 該分點有交易但 v3.29.1 全濾淨賣 ==========
print("\nB. 民哥 9B25: 有交易但全淨賣 → 提示「無淨買超個股」")
branches_data_b = [
    {
        "code": "9B25", "name": "台新-五權西",
        "buys": [
            # 全部淨賣
            make_stock("2330", "台積電", buy_amt=100_000, sell_amt=300_000, is_limit_up=False),
            make_stock("2317", "鴻海",   buy_amt=80_000,  sell_amt=200_000, is_limit_up=False),
        ],
        "sells": [
            make_stock("2454", "聯發科", buy_amt=50_000, sell_amt=300_000, is_limit_up=False),
        ],
    },
]
for m in MASTER_MAPPING:
    for code, _ in m["branches"]:
        if code != "9B25":
            branches_data_b.append({"code": code, "name": "_", "buys": [], "sells": []})

wb2 = Workbook()
ws2 = wb2.active
build_day_sheet(ws2, branches_data_b, "20260519")

master_row_b = find_master_block(ws2, "民哥")
if master_row_b is None:
    print("  ❌ 找不到 民哥 row [FAIL]")
    all_pass = False
else:
    found_b = False
    for r in range(master_row_b, master_row_b + 12):
        d = ws2.cell(r, 4).value
        if d and '無淨買超個股' in str(d):
            print(f"  ✅ row {r} D 欄: {d}")
            found_b = True
            break
    if not found_b:
        print(f"  ❌ 沒找到「無淨買超個股」提示 [FAIL]")
        all_pass = False

# ========== Case C: 該分點 TWSE 完全沒資料 → 不寫提示 ==========
print("\nC. 完全沒 TWSE 資料的分點 → 全空白 (不寫提示)")
branches_data_c = []
for m in MASTER_MAPPING:
    for code, _ in m["branches"]:
        branches_data_c.append({"code": code, "name": "_", "buys": [], "sells": []})

wb3 = Workbook()
ws3 = wb3.active
build_day_sheet(ws3, branches_data_c, "20260519")

# 掃整張表, 應該沒有任何「⚪」字樣
has_notice = False
for r in range(1, ws3.max_row + 1):
    d = ws3.cell(r, 4).value
    if d and '⚪' in str(d):
        has_notice = True
        break
if not has_notice:
    print(f"  ✅ 全空白, 無 ⚪ 提示")
else:
    print(f"  ❌ 不該有提示但出現 ⚪ [FAIL]")
    all_pass = False

# ========== Case D: 有 stocks 入選 → 不該寫提示 ==========
print("\nD. 有 stocks 入選 → 不寫提示")
branches_data_d = [
    {
        "code": "9B25", "name": "台新-五權西",
        "buys": [
            # 淨買, 應入選
            make_stock("3443", "創意", buy_amt=600_000, sell_amt=50_000, is_limit_up=False),
        ],
        "sells": [],
    },
]
for m in MASTER_MAPPING:
    for code, _ in m["branches"]:
        if code != "9B25":
            branches_data_d.append({"code": code, "name": "_", "buys": [], "sells": []})

wb4 = Workbook()
ws4 = wb4.active
build_day_sheet(ws4, branches_data_d, "20260519")

# 民哥 9B25 第 1 列 D 欄應該是「創意(3443)」不是提示
master_row_d = find_master_block(ws4, "民哥")
if master_row_d:
    d_first = ws4.cell(master_row_d, 4).value
    if d_first and '創意' in str(d_first) and '⚪' not in str(d_first):
        print(f"  ✅ row {master_row_d} D 欄: {d_first} (正常入選, 無提示)")
    else:
        print(f"  ❌ row {master_row_d} D 欄不對: {d_first} [FAIL]")
        all_pass = False

print()
print("─" * 70)
print(f"  整體: {'✅ ALL PASS' if all_pass else '❌ HAS FAIL'}")
sys.exit(0 if all_pass else 1)
