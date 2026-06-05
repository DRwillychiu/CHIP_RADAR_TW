"""
========================================================================
Module: db_excel_import.py  (v3.31.15 Phase 1.2)
功能: 從 Excel 月檔 (chip_radar_YYYY-MM.xlsx) import 進 DB (source='excel')
      + cross_validate raw vs excel 雙源比對

使用者 Q1=C 雙來源 (raw 為主 + Excel 備份):
  - raw source: crawler → upsert_from_raw_dict (已有 v3.31.2)
  - excel source: 本檔 → upsert_from_excel_sheet + cross_validate

Excel 結構 (excel_report.py 產出):
  - 每 sheet = 一個交易日 (sheet name = YYYYMMDD)
  - 12 欄: 高手 / 分點 / 代號 / 標的 / 買進(張) / 賣出(張) /
            買進(萬元) / 賣出(萬元) / 淨買差(萬元) / 買均 / 賣均 / 損益(萬)
  - 合併儲存格 (A=高手, B=分點, C=代號 向下 merge → 需 forward-fill)
  - 分隔標題列 (標的=「標的」) + 備註列 (「⚪ 今日...」) → 需清洗

寫入 DB: 用 daily_chips + daily_records 的 source='excel' 欄位區分.
         UNIQUE(date, trader_id, branch_id, stock_id, source) → 不衝突 raw.

CLI:
  python db_excel_import.py import data/reports/chip_radar_2026-06.xlsx
  python db_excel_import.py import-all data/reports/
  python db_excel_import.py validate --date 20260604
========================================================================
"""
import os
import sys
import re
import json
import sqlite3
import argparse
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ 需要 openpyxl: pip install openpyxl")
    sys.exit(1)

from db_pipeline import init_db, _upsert_dim, _upsert_trader_branch, _now_tw_iso, DEFAULT_DB


# ════════════════════════════════════════════════════════════════════
#  Excel 清洗 + 解析
# ════════════════════════════════════════════════════════════════════

STOCK_PATTERN = re.compile(r'^(.+?)\((\d{4,6}[A-Z]?)\)$')
HEADER_KEYWORDS = {'標的', '高手', '分點', '代號', '買進(張)'}
NOTICE_PREFIX = '⚪'


def parse_excel_sheet(ws, trade_date: str) -> List[Dict[str, Any]]:
    """解析單一 sheet → list of stock dicts (已清洗).

    處理:
      1. forward-fill A/B/C 合併儲存格
      2. 移除分隔標題列 (標的=「標的」)
      3. 移除備註列 (「⚪ 今日...」)
      4. 拆分標的: 「華邦電(2344)」→ name=華邦電, code=2344
      5. 計算 pnl (Excel 公式未快取): sell_lots × (sell_avg - buy_avg)
    """
    rows = []
    last_master = None
    last_branch_name = None
    last_branch_code = None

    for row in ws.iter_rows(min_row=1, max_col=12, values_only=False):
        vals = [cell.value for cell in row]

        # forward-fill A/B/C
        if vals[0] is not None:
            last_master = str(vals[0]).strip()
        if vals[1] is not None:
            last_branch_name = str(vals[1]).strip()
        if vals[2] is not None:
            last_branch_code = str(vals[2]).strip()

        # D 欄 (標的)
        raw_stock = vals[3]
        if raw_stock is None:
            continue
        raw_stock = str(raw_stock).strip()

        # 跳過分隔標題列
        if raw_stock in HEADER_KEYWORDS or any(kw in raw_stock for kw in HEADER_KEYWORDS):
            continue

        # 跳過備註列
        if raw_stock.startswith(NOTICE_PREFIX):
            continue

        # 拆分標的
        m = STOCK_PATTERN.match(raw_stock)
        if not m:
            continue  # 無法解析 → 跳過 (不是真 stock row)
        stock_name = m.group(1).strip()
        stock_code = m.group(2)

        # 數值欄 (E-L, 0-indexed 4-11)
        def _num(v, default=0):
            if v is None:
                return default
            try:
                return float(v)
            except (ValueError, TypeError):
                return default

        buy_lots = int(_num(vals[4]))
        sell_lots = int(_num(vals[5]))
        buy_amt_wan = _num(vals[6])    # 萬元 (Excel 顯示)
        sell_amt_wan = _num(vals[7])
        net_amt_wan = _num(vals[8])
        buy_avg = _num(vals[9])
        sell_avg = _num(vals[10])
        # pnl: 可能是 Excel formula (未快取 → None),自己算
        pnl_raw = vals[11]
        if pnl_raw is not None and not isinstance(pnl_raw, str):
            pnl = float(pnl_raw)
        else:
            # Excel 公式 =F*(K-J) 其中 F=賣出(張), K=賣均(元/股), J=買均(元/股)
            # 結果單位 = 張 × 元/股 = 元, 但 Excel 欄位名「損益(萬)」→ 除 10000
            # 修正: sell_lots × (sell_avg - buy_avg) / 10 = 萬元
            if sell_lots > 0 and buy_avg > 0 and sell_avg > 0:
                pnl = round(sell_lots * (sell_avg - buy_avg) / 10, 2)
            else:
                pnl = 0.0

        # 萬元 → 仟元 (DB 內 buy_amt 單位仟元, Excel 顯示萬元, 1 萬 = 10 仟)
        buy_amt_k = int(round(buy_amt_wan * 10))
        sell_amt_k = int(round(sell_amt_wan * 10))
        net_amt_k = int(round(net_amt_wan * 10))

        rows.append({
            'date': trade_date,
            'trader': last_master,
            'branch_name': last_branch_name,
            'branch_code': last_branch_code,
            'stock_name': stock_name,
            'stock_code': stock_code,
            'buy_lots': buy_lots,
            'sell_lots': sell_lots,
            'buy_amt': buy_amt_k,    # 仟元
            'sell_amt': sell_amt_k,
            'net_amt': net_amt_k,
            'avg_buy_price': buy_avg,
            'avg_sell_price': sell_avg,
            'pnl': pnl,
        })

    return rows


# ════════════════════════════════════════════════════════════════════
#  Excel → DB upsert
# ════════════════════════════════════════════════════════════════════

def upsert_from_excel_sheet(conn: sqlite3.Connection,
                              parsed_rows: List[Dict[str, Any]],
                              trade_date: str,
                              source_file: str = '') -> Dict[str, int]:
    """把 parse_excel_sheet 的結果寫進 DB (source='excel')."""
    cur = conn.cursor()
    stats = {'daily_chips_rows': 0, 'daily_records_rows': 0}

    for r in parsed_rows:
        trader_name = r['trader']
        branch_code = r['branch_code']
        stock_code = r['stock_code']

        if not branch_code or not stock_code:
            continue

        branch_id = _upsert_dim(conn, 'branches', 'code', branch_code,
                                 {'name': r['branch_name']})
        stock_id = _upsert_dim(conn, 'stocks', 'code', stock_code,
                                {'name': r['stock_name']})
        trader_id = _upsert_dim(conn, 'traders', 'name', trader_name) if trader_name else None
        if trader_id:
            _upsert_trader_branch(conn, trader_id, branch_id)

        cur.execute("""
            INSERT OR REPLACE INTO daily_chips
            (date, trader_id, branch_id, stock_id,
             buy_lots, sell_lots, net_lots, buy_amt, sell_amt, net_amt,
             avg_buy_price, avg_sell_price, pnl,
             is_estimated_lot, is_limit_up, trade_style, source)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 0, NULL, 'excel')
        """, (trade_date, trader_id, branch_id, stock_id,
               r['buy_lots'], r['sell_lots'], r['buy_lots'] - r['sell_lots'],
               r['buy_amt'], r['sell_amt'], r['net_amt'],
               r['avg_buy_price'], r['avg_sell_price'], r['pnl']))
        stats['daily_chips_rows'] += 1

        cur.execute("""
            INSERT OR REPLACE INTO daily_records
            (date, trader, branch_name, branch_code, stock_name, stock_code,
             buy_lots, sell_lots, buy_amt, sell_amt, net_amt,
             avg_buy_price, avg_sell_price, pnl,
             is_estimated_lot, is_limit_up, trade_style, source)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 0, NULL, 'excel')
        """, (trade_date, trader_name, r['branch_name'], branch_code,
               r['stock_name'], stock_code,
               r['buy_lots'], r['sell_lots'], r['buy_amt'], r['sell_amt'], r['net_amt'],
               r['avg_buy_price'], r['avg_sell_price'], r['pnl']))
        stats['daily_records_rows'] += 1

    cur.execute("""
        INSERT INTO import_log (date, source, imported_at, row_count, source_file)
        VALUES (?, 'excel', ?, ?, ?)
    """, (trade_date, _now_tw_iso(), stats['daily_chips_rows'], source_file))

    conn.commit()
    return stats


def import_excel_file(db_path: str, excel_path: str) -> Dict[str, int]:
    """Import 一個月檔 (多 sheet) 全部進 DB."""
    wb = load_workbook(excel_path, data_only=True)
    conn = init_db(db_path)
    total = {'sheets': 0, 'daily_chips_rows': 0, 'daily_records_rows': 0}

    for sheet_name in wb.sheetnames:
        # 只處理 YYYYMMDD 名稱的 sheet
        if not (len(sheet_name) == 8 and sheet_name.isdigit()):
            print(f"  ⏭️ 跳過 sheet '{sheet_name}' (非 YYYYMMDD)")
            continue

        ws = wb[sheet_name]
        parsed = parse_excel_sheet(ws, sheet_name)
        if not parsed:
            print(f"  ⏭️ {sheet_name}: 0 筆 (空 sheet 或全被清洗)")
            continue

        stats = upsert_from_excel_sheet(conn, parsed, sheet_name,
                                          source_file=Path(excel_path).name)
        total['sheets'] += 1
        total['daily_chips_rows'] += stats['daily_chips_rows']
        total['daily_records_rows'] += stats['daily_records_rows']
        print(f"  ✓ {sheet_name}: {stats['daily_chips_rows']} rows (excel)")

    conn.close()
    return total


# ════════════════════════════════════════════════════════════════════
#  Cross-validate raw vs excel
# ════════════════════════════════════════════════════════════════════

def cross_validate(db_path: str, trade_date: Optional[str] = None,
                    tolerance_amt_pct: float = 5.0,
                    tolerance_lot_abs: int = 1) -> Dict[str, Any]:
    """比對同天 raw vs excel 數據一致性.

    比對邏輯 (per date × trader × branch × stock):
      buy_lots: 差 > tolerance_lot_abs → mismatch
      buy_amt:  差% > tolerance_amt_pct → mismatch
      (Excel 是 Top N 篩選後, raw 是全量, 所以 Excel 行數 < raw;
       只比「Excel 有的」那些 row 是否跟 raw 一致)

    Returns:
      {verdict, compared, matched, mismatched, excel_only, raw_only, details}
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    date_clause = f"AND r.date = '{trade_date}'" if trade_date else ""

    # 抓 raw 與 excel 都有的 (date, trader, branch, stock) 組合
    sql = f"""
        SELECT
            r.date, r.trader_id, r.branch_id, r.stock_id,
            r.buy_lots AS raw_buy_lots, r.buy_amt AS raw_buy_amt,
            r.sell_lots AS raw_sell_lots, r.sell_amt AS raw_sell_amt,
            e.buy_lots AS excel_buy_lots, e.buy_amt AS excel_buy_amt,
            e.sell_lots AS excel_sell_lots, e.sell_amt AS excel_sell_amt,
            t.name AS trader, b.code AS branch_code, s.code AS stock_code, s.name AS stock_name
        FROM daily_chips r
        JOIN daily_chips e ON r.date = e.date AND r.trader_id IS e.trader_id
                           AND r.branch_id = e.branch_id AND r.stock_id = e.stock_id
        LEFT JOIN traders t ON r.trader_id = t.id
        JOIN branches b ON r.branch_id = b.id
        JOIN stocks s ON r.stock_id = s.id
        WHERE r.source = 'raw' AND e.source = 'excel' {date_clause}
    """
    rows = cur.execute(sql).fetchall()

    mismatches = []
    matched = 0
    for row in rows:
        lot_diff = abs((row['raw_buy_lots'] or 0) - (row['excel_buy_lots'] or 0))
        raw_amt = row['raw_buy_amt'] or 0
        excel_amt = row['excel_buy_amt'] or 0

        # Excel buy_amt 是 萬元→仟元 反轉, 有 rounding (±5 仟元)
        amt_diff = abs(raw_amt - excel_amt)
        amt_pct = (amt_diff / max(raw_amt, 1)) * 100

        if lot_diff > tolerance_lot_abs or amt_pct > tolerance_amt_pct:
            mismatches.append({
                'date': row['date'],
                'trader': row['trader'],
                'branch': row['branch_code'],
                'stock': f"{row['stock_code']} {row['stock_name']}",
                'raw_lot': row['raw_buy_lots'],
                'excel_lot': row['excel_buy_lots'],
                'lot_diff': lot_diff,
                'raw_amt': raw_amt,
                'excel_amt': excel_amt,
                'amt_diff_pct': round(amt_pct, 2),
            })
        else:
            matched += 1

    # 統計 Excel-only 和 Raw-only
    date_clause_plain = f"AND date = '{trade_date}'" if trade_date else ""
    excel_only = cur.execute(f"""
        SELECT COUNT(*) FROM daily_chips dc1 WHERE dc1.source='excel' {date_clause_plain.replace('date', 'dc1.date')}
        AND NOT EXISTS (
            SELECT 1 FROM daily_chips r
            WHERE r.source='raw' AND r.date=dc1.date
            AND r.trader_id IS dc1.trader_id
            AND r.branch_id=dc1.branch_id AND r.stock_id=dc1.stock_id
        )
    """).fetchone()[0]

    raw_only = cur.execute(f"""
        SELECT COUNT(*) FROM daily_chips dc2 WHERE dc2.source='raw' {date_clause_plain.replace('date', 'dc2.date')}
        AND NOT EXISTS (
            SELECT 1 FROM daily_chips e
            WHERE e.source='excel' AND e.date=dc2.date
            AND e.trader_id IS dc2.trader_id
            AND e.branch_id=dc2.branch_id AND e.stock_id=dc2.stock_id
        )
    """).fetchone()[0]

    conn.close()

    verdict = 'PASS' if len(mismatches) == 0 else ('WARN' if len(mismatches) < 5 else 'FAIL')
    return {
        'verdict': verdict,
        'compared': len(rows),
        'matched': matched,
        'mismatched': len(mismatches),
        'excel_only': excel_only,
        'raw_only': raw_only,
        'details': mismatches[:20],
    }


# ════════════════════════════════════════════════════════════════════
#  CLI
# ════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description='Phase 1.2: Excel → DB + cross-validate')
    sub = parser.add_subparsers(dest='cmd', required=True)

    p_imp = sub.add_parser('import', help='Import 一個 xlsx')
    p_imp.add_argument('xlsx', help='月檔路徑 (e.g. data/reports/chip_radar_2026-06.xlsx)')

    p_all = sub.add_parser('import-all', help='Import reports/ 下所有月檔')
    p_all.add_argument('dir', nargs='?', default='data/reports', help='reports 目錄')

    p_val = sub.add_parser('validate', help='Cross-validate raw vs excel')
    p_val.add_argument('--date', default=None, help='YYYYMMDD (省略=全部日期)')

    args = parser.parse_args()

    if args.cmd == 'import':
        print(f"[Excel Import] {args.xlsx}")
        stats = import_excel_file(DEFAULT_DB, args.xlsx)
        print(f"  完成: {stats['sheets']} sheets, {stats['daily_chips_rows']} rows")

    elif args.cmd == 'import-all':
        xlsx_files = sorted(Path(args.dir).glob('chip_radar_*.xlsx'))
        xlsx_files = [f for f in xlsx_files if f.name != 'latest.xlsx']
        print(f"[Excel Import] {len(xlsx_files)} 個月檔")
        for f in xlsx_files:
            print(f"\n  === {f.name} ===")
            stats = import_excel_file(DEFAULT_DB, str(f))
            print(f"  小計: {stats['sheets']} sheets, {stats['daily_chips_rows']} rows")

    elif args.cmd == 'validate':
        print(f"[Cross-Validate] raw vs excel" + (f" date={args.date}" if args.date else " (全部)"))
        result = cross_validate(DEFAULT_DB, args.date)
        print(f"  verdict: {result['verdict']}")
        print(f"  compared: {result['compared']} | matched: {result['matched']} | "
              f"mismatched: {result['mismatched']}")
        print(f"  excel_only: {result['excel_only']} | raw_only: {result['raw_only']}")
        if result['details']:
            print(f"\n  === Mismatch 明細 (前 20) ===")
            for d in result['details']:
                print(f"    {d['date']} {d['trader']} {d['branch']} {d['stock']}: "
                      f"lot {d['raw_lot']}→{d['excel_lot']} (±{d['lot_diff']}) "
                      f"amt {d['raw_amt']}→{d['excel_amt']} (±{d['amt_diff_pct']}%)")


if __name__ == '__main__':
    main()
