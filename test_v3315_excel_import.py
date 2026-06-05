"""v3.31.15 Phase 1.2 Excel import + cross-validate 測試"""
import os
import sys
import tempfile
import sqlite3
sys.path.insert(0, '.')

from openpyxl import Workbook
from db_pipeline import init_db, upsert_from_raw_dict
from db_excel_import import parse_excel_sheet, upsert_from_excel_sheet, cross_validate

all_pass = True
print("=" * 64)
print("  v3.31.15 Phase 1.2 Excel import + cross-validate")
print("=" * 64)


def _build_test_excel_sheet(ws, trade_date):
    """建一個模擬 excel_report 產出的 sheet."""
    # header row
    headers = ['高手', '分點', '代號', '標的', '買進(張)', '賣出(張)',
               '買進(萬元)', '賣出(萬元)', '淨買差(萬元)', '買均', '賣均', '損益(萬)']
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)

    # 民哥 台新-五權西 2 stocks (合併儲存格模擬)
    ws.cell(2, 1, '民哥')       # A2 (merged down to A3)
    ws.cell(2, 2, '台新-五權西')  # B2
    ws.cell(2, 3, '9B25')       # C2
    ws.cell(2, 4, '台積電(2330)')
    ws.cell(2, 5, 100)   # buy_lots
    ws.cell(2, 6, 10)    # sell_lots
    ws.cell(2, 7, 2284)  # buy_amt 萬元
    ws.cell(2, 8, 228)   # sell_amt 萬元
    ws.cell(2, 9, 2056)  # net 萬元
    ws.cell(2, 10, 2284.0)  # buy_avg
    ws.cell(2, 11, 2280.0)  # sell_avg
    ws.cell(2, 12, None)    # pnl formula (未快取)

    # A3/B3/C3 = None (forward-fill from above)
    ws.cell(3, 4, '聯發科(2454)')
    ws.cell(3, 5, 50)
    ws.cell(3, 6, 5)
    ws.cell(3, 7, 6550)
    ws.cell(3, 8, 655)
    ws.cell(3, 9, 5895)
    ws.cell(3, 10, 1310.0)
    ws.cell(3, 11, 1310.0)
    ws.cell(3, 12, 0)

    # 備註列 (應被跳過)
    ws.cell(4, 4, '⚪ 今日淨買僅 2 檔')

    # 分隔標題列 (應被跳過)
    ws.cell(5, 1, '高手')
    ws.cell(5, 4, '標的')


# ── 1. parse_excel_sheet 基本清洗 ──
print("\n1. parse_excel_sheet: forward-fill + 移除備註/分隔 + 拆標的")
wb = Workbook()
ws = wb.active
ws.title = '20260604'
_build_test_excel_sheet(ws, '20260604')
parsed = parse_excel_sheet(ws, '20260604')
ok = (len(parsed) == 2
      and parsed[0]['stock_code'] == '2330'
      and parsed[0]['trader'] == '民哥'
      and parsed[0]['branch_code'] == '9B25'
      and parsed[0]['buy_lots'] == 100
      and parsed[1]['stock_code'] == '2454')
print(f"  {'OK' if ok else 'FAIL'} 解析 {len(parsed)} 筆 (應 2, 備註+分隔被移除)")
if parsed:
    print(f"    r1: {parsed[0]['stock_code']} {parsed[0]['stock_name']} trader={parsed[0]['trader']} "
          f"buy_lots={parsed[0]['buy_lots']} buy_amt={parsed[0]['buy_amt']}(仟元)")
if not ok: all_pass = False

# ── 2. 萬元→仟元 轉換 ──
print("\n2. 萬元→仟元: 2284 萬元 → 22840 仟元")
ok = parsed[0]['buy_amt'] == 22840 and parsed[0]['sell_amt'] == 2280
print(f"  {'OK' if ok else 'FAIL'} buy_amt={parsed[0]['buy_amt']} sell_amt={parsed[0]['sell_amt']}")
if not ok: all_pass = False

# ── 3. pnl 自動計算 (formula 未快取) ──
# sell_lots=10 × (sell_avg=2280 - buy_avg=2284) / 10 = -4.0 萬元
# 推導: 10張 × 1000股/張 × (2280-2284)元/股 = -40,000 元 = -4.0 萬元
# Excel =F*(K-J) 直接得 10×(-4) = -40, 但那等價於 仟元 level (張×元/股=仟元)
# ÷ 10 = 萬元 = -4.0
print("\n3. pnl 自動計算: 10 × (2280-2284) / 10 = -4.0 萬元")
ok = abs(parsed[0]['pnl'] - (-4.0)) < 0.01
print(f"  {'OK' if ok else 'FAIL'} pnl={parsed[0]['pnl']}")
if not ok: all_pass = False

# ── 4. Excel → DB upsert (source='excel') ──
print("\n4. upsert_from_excel_sheet: 2 rows source=excel 寫入 DB")
with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as f:
    db_path = f.name
try:
    conn = init_db(db_path)
    stats = upsert_from_excel_sheet(conn, parsed, '20260604', 'test.xlsx')
    ok = stats['daily_chips_rows'] == 2
    print(f"  {'OK' if ok else 'FAIL'} daily_chips rows={stats['daily_chips_rows']}")
    if not ok: all_pass = False

    # ── 5. source 區分: raw + excel 不衝突 ──
    print("\n5. raw + excel 雙 source 不衝突 (UNIQUE 含 source)")
    raw_output = {
        'trade_date': '20260604',
        'branches': [{
            'code': '9B25', 'name': '台新-五權西', 'master': '民哥', 'co_masters': [],
            'buys': [
                {'code': '2330', 'name': '台積電', 'buy_lot': 100, 'sell_lot': 10,
                 'buy_amt': 22840, 'sell_amt': 2280, 'net_amt': 20560, 'net_lot': 90,
                 'buy_avg': 2284.0, 'sell_avg': 2280.0,
                 'is_limit_up': False, 'trade_style': 'overnight'},
            ], 'sells': [],
        }],
    }
    raw_stats = upsert_from_raw_dict(conn, raw_output, source='raw')
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM daily_chips WHERE date='20260604'")
    total = cur.fetchone()[0]
    # raw 1 筆 (2330 民哥 9B25) + excel 2 筆 (2330+2454 民哥 9B25) = 3
    ok = total == 3
    print(f"  {'OK' if ok else 'FAIL'} total rows={total} (raw 1 + excel 2 = 3)")
    if not ok: all_pass = False

    # ── 6. cross_validate: raw vs excel ──
    print("\n6. cross_validate: 2330 兩源比對 (相同 buy_lots=100, buy_amt=22840)")
    result = cross_validate(db_path, '20260604')
    ok = (result['compared'] >= 1
          and result['matched'] >= 1
          and result['mismatched'] == 0)
    print(f"  {'OK' if ok else 'FAIL'} verdict={result['verdict']} compared={result['compared']} "
          f"matched={result['matched']} mismatched={result['mismatched']}")
    print(f"    excel_only={result['excel_only']} raw_only={result['raw_only']}")
    if not ok: all_pass = False

    # ── 7. cross_validate mismatch 偵測 ──
    print("\n7. cross_validate: 故意改 raw buy_lots → mismatch 偵測")
    cur.execute("UPDATE daily_chips SET buy_lots=999 WHERE source='raw' AND date='20260604'")
    conn.commit()
    result2 = cross_validate(db_path, '20260604')
    ok = result2['mismatched'] >= 1
    print(f"  {'OK' if ok else 'FAIL'} mismatched={result2['mismatched']} (應 ≥1)")
    if result2['details']:
        d = result2['details'][0]
        print(f"    {d['stock']}: raw_lot={d['raw_lot']} excel_lot={d['excel_lot']} diff={d['lot_diff']}")
    if not ok: all_pass = False

    conn.close()
finally:
    for ext in ('', '-wal', '-shm'):
        p = db_path + ext
        if os.path.exists(p):
            os.unlink(p)

print()
print("─" * 64)
print(f"  整體: {'OK ALL PASS' if all_pass else 'FAIL HAS FAIL'}")
sys.exit(0 if all_pass else 1)
