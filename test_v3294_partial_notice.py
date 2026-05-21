"""v3.29.4 partial 空白提示 + 移除 ▲% 標籤 測試

5/22 用戶 review 5/20 Excel 凱基-松山 9217 (sniper 迷你哥) 4 stocks + 6 row 空白看起來像 bug.
v3.29.2 完全空白才提示, partial 沒處理.

V3.29.4 修補:
  1. partial (1-9 stocks): 在第 N+1 列寫「⚪ 今日漲停僅 N 檔」/「⚪ 今日淨買僅 N 檔」
  2. 移除 v3.27.4 L4 sniper master ▲X.XX% 標籤 (user 不希望直接顯示漲幅)

驗證 case:
  A. sniper master 5 stocks + 5 row 空白 → row N+1 寫 partial notice
  B. swing master 3 stocks + 7 row 空白 → row N+1 寫淨買 partial notice
  C. sniper master 10 stocks full → 沒 partial notice (rows full)
  D. sniper master 0 stocks + 有資料 → 寫 v3.29.2 完全空白 notice
  E. 完全沒資料 → 全 blank, 沒提示
  F. 標的欄不再有 ▲X.XX% (sniper 也是純名稱)
"""
import sys
sys.path.insert(0, '.')

from openpyxl import Workbook
from excel_report import build_day_sheet, MASTER_MAPPING


def make_stock(code, name, buy_amt=0, sell_amt=0, buy_lot=0, sell_lot=0, is_limit_up=False, change_pct=None):
    return {
        "code": code, "name": name,
        "buy_amt": buy_amt, "sell_amt": sell_amt,
        "buy_lot": buy_lot, "sell_lot": sell_lot,
        "net_amt": buy_amt - sell_amt,
        "net_lot": buy_lot - sell_lot,
        "is_limit_up": is_limit_up,
        "change_pct": change_pct,
    }


def find_first_row_with(ws, text):
    """找 D 欄含 text 的第一個 row idx, 或 None"""
    for row_idx in range(1, ws.max_row + 1):
        v = ws.cell(row_idx, 4).value
        if v and text in str(v):
            return row_idx
    return None


def get_all_d_values(ws):
    return [ws.cell(r, 4).value for r in range(1, ws.max_row + 1)]


print("=" * 72)
print("  v3.29.4 partial 空白提示 + 移除 ▲% 標籤")
print("=" * 72)

all_pass = True

# ========== Case A: sniper 5 stocks → partial notice ==========
print("\nA. sniper master 5 stocks: row N+1 寫「漲停僅 5 檔」")
# 迷你哥/松山哥 sniper master. 9217 凱基-松山
sniper_branch = {
    "code": "9217", "name": "凱基-松山",
    "buys": [
        make_stock(f"100{i}", f"漲停{i}", buy_amt=500_000, sell_amt=50_000,
                   buy_lot=100, sell_lot=10, is_limit_up=True, change_pct=9.99) for i in range(5)
    ],
    "sells": [],
}
branches_data_a = [sniper_branch]
for m in MASTER_MAPPING:
    for code, _ in m["branches"]:
        if code != "9217":
            branches_data_a.append({"code": code, "name": "_", "buys": [], "sells": []})

wb = Workbook()
ws = wb.active
build_day_sheet(ws, branches_data_a, "20260522")

partial_row = find_first_row_with(ws, "漲停僅 5 檔")
if partial_row:
    print(f"  ✅ row {partial_row} D 欄: {ws.cell(partial_row, 4).value}")
else:
    # Print all D values for debug
    print("  ❌ 沒找到「漲停僅 5 檔」提示")
    all_d = get_all_d_values(ws)
    for r_idx, d_val in enumerate(all_d, 1):
        if d_val and ('漲停' in str(d_val) or '⚪' in str(d_val)):
            print(f"      row {r_idx}: {d_val}")
    all_pass = False

# ========== Case B: swing 3 stocks → partial notice 淨買 ==========
print("\nB. swing master 3 stocks: row N+1 寫「淨買僅 3 檔」")
# 民哥 swing master. 9B25 台新-五權西
swing_branch = {
    "code": "9B25", "name": "台新-五權西",
    "buys": [
        make_stock(f"200{i}", f"淨買股{i}", buy_amt=500_000, sell_amt=50_000,
                   buy_lot=100, sell_lot=10, is_limit_up=False) for i in range(3)
    ],
    "sells": [],
}
branches_data_b = [swing_branch]
for m in MASTER_MAPPING:
    for code, _ in m["branches"]:
        if code != "9B25":
            branches_data_b.append({"code": code, "name": "_", "buys": [], "sells": []})

wb2 = Workbook()
ws2 = wb2.active
build_day_sheet(ws2, branches_data_b, "20260522")

partial_row_b = find_first_row_with(ws2, "淨買僅 3 檔")
if partial_row_b:
    print(f"  ✅ row {partial_row_b} D 欄: {ws2.cell(partial_row_b, 4).value}")
else:
    print("  ❌ 沒找到「淨買僅 3 檔」提示")
    all_pass = False

# ========== Case C: sniper 10 stocks (full) → 沒 partial notice ==========
print("\nC. sniper 10 stocks (full): 沒任何 partial notice")
full_branch = {
    "code": "9227", "name": "凱基-城中",  # 蔣承翰 sniper
    "buys": [
        make_stock(f"300{i}", f"漲停{i}", buy_amt=500_000, sell_amt=50_000,
                   buy_lot=100, sell_lot=10, is_limit_up=True) for i in range(10)
    ],
    "sells": [],
}
branches_data_c = [full_branch]
for m in MASTER_MAPPING:
    for code, _ in m["branches"]:
        if code != "9227":
            branches_data_c.append({"code": code, "name": "_", "buys": [], "sells": []})

wb3 = Workbook()
ws3 = wb3.active
build_day_sheet(ws3, branches_data_c, "20260522")

# 蔣承翰 9227 section 不該有「漲停僅 N 檔」(因為填滿了)
# 但其他空白分點可能有「未搶漲停」(by 5/22 v3.29.4 邏輯都會印)
# 簡單檢查: 沒有「漲停僅 10 檔」
partial_full_row = find_first_row_with(ws3, "漲停僅 10 檔")
if partial_full_row is None:
    print("  ✅ 沒有錯誤的「漲停僅 10 檔」提示 (10 stocks 已填滿不需要 partial 提示)")
else:
    print(f"  ❌ row {partial_full_row} 不該有「漲停僅 10 檔」")
    all_pass = False

# ========== Case D: sniper 0 stocks + 有資料 → 完全空白 notice ==========
print("\nD. sniper 0 stocks + 有 TWSE 交易: 完全空白「未搶漲停」notice (v3.29.2 邏輯保留)")
zero_branch = {
    "code": "9B2n", "name": "台新-西松",  # 巨人傑 sniper
    "buys": [make_stock("2330", "台積電", buy_amt=500_000, sell_amt=100_000, is_limit_up=False)],
    "sells": [],
}
branches_data_d = [zero_branch]
for m in MASTER_MAPPING:
    for code, _ in m["branches"]:
        if code != "9B2n":
            branches_data_d.append({"code": code, "name": "_", "buys": [], "sells": []})

wb4 = Workbook()
ws4 = wb4.active
build_day_sheet(ws4, branches_data_d, "20260522")

empty_row = find_first_row_with(ws4, "未搶漲停")
if empty_row:
    print(f"  ✅ row {empty_row}: {ws4.cell(empty_row, 4).value}")
else:
    print("  ❌ 沒找到「未搶漲停」 (v3.29.2 既有邏輯應保留)")
    all_pass = False

# ========== Case E: 完全沒 TWSE 資料 → 全 blank, 沒提示 ==========
print("\nE. 完全沒資料: 全 blank, 任何 ⚪ 都沒出現")
branches_data_e = []
for m in MASTER_MAPPING:
    for code, _ in m["branches"]:
        branches_data_e.append({"code": code, "name": "_", "buys": [], "sells": []})

wb5 = Workbook()
ws5 = wb5.active
build_day_sheet(ws5, branches_data_e, "20260522")

any_notice = find_first_row_with(ws5, "⚪")
if any_notice is None:
    print("  ✅ 全空白, 0 個 ⚪ 提示")
else:
    print(f"  ❌ row {any_notice} 不該有 ⚪")
    all_pass = False

# ========== Case F: 標的欄不再有 ▲X.XX% ==========
print("\nF. v3.29.4 移除 ▲X.XX% 標籤")
# Case A 的 sniper stock 都應該是「漲停X(1000)」純名稱, 沒 ▲9.99%
has_arrow = False
for r in range(1, ws.max_row + 1):
    d = ws.cell(r, 4).value
    if d and isinstance(d, str) and '▲' in d:
        has_arrow = True
        print(f"  ❌ row {r} D 欄含 ▲: {d}")
        break
if not has_arrow:
    print("  ✅ 整個 sheet 0 個 ▲X.XX% 標籤 (v3.27.4 L4 已 reverse)")
else:
    all_pass = False

print()
print("─" * 72)
print(f"  整體: {'✅ ALL PASS' if all_pass else '❌ HAS FAIL'}")
sys.exit(0 if all_pass else 1)
