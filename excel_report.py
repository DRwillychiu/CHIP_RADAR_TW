"""
excel_report.py v3.26 - Style-aware Excel daily report

v3.26 change: 隔日沖 / 當沖 master 的資料源自動切換為「今天買的漲停股」,
波段 / 長線 master 維持原 Top 10 by 買超。視覺格式與手動版完全一致。

Strict mimicry of user's hand-curated Excel ("分點觀察" boss report).

Layout (per master block):
  Row 1: full header [高手|分點|代號|標的|...|損益(萬)]   ← A="高手" label
  Row 2: first data row of master's first branch          ← A=master_name (merged down)
  Rows 2-11: 10 data rows for branch #1 (padded blank if <10 stocks)
  Row 12: sub-header [(空)|分點|代號|標的|...]            ← A empty, under master merge
  Rows 13-22: 10 data rows for branch #2
  ...
  Next master: full header row again, A=高手 label

Visual style (matches manual file exactly):
  - Font: 新細明體 (PMingLiU), 12pt
  - All cells: center/center alignment
  - NO fills, NO borders
  - Header rows + master/branch/code labels: bold
  - L column: =F*(K-J), format '0.00_ ;[Red]\\-0.00\\ ' (red text on negative)

Data source routing (v3.26):
  - sniper master (next_day_flipper / day_trader) → top 10 漲停股 by buy_amt
  - other master (swing / longterm / unmapped)    → top 10 全部個股 by buy_amt (legacy)
  - sniper master with no limit-up buys today     → 10 blank rows (intentional, not fallback)

Outputs:
  data/reports/chip_radar_YYYY-MM-DD.xlsx  - single-sheet daily snapshot
  data/reports/latest.xlsx                  - multi-sheet, last 30 trading days
  data/reports/README.md                    - history index
"""

from typing import Dict, List, Optional, Tuple
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from branches import MASTER_STYLES
except ImportError:
    MASTER_STYLES = {}

SNIPER_STYLES = {"next_day_flipper", "day_trader"}

# v3.30.5: 使用者要求 Excel 只有「蔣承翰」用漲停股抓取法。
# 原 v3.26 是「所有 next_day_flipper/day_trader style 的 master 都用漲停法」
# (蔣承翰 / 迷你哥-松山哥 / Tradow / 巨人傑),現縮為白名單僅蔣承翰一人,
# 其餘 sniper-style master 全部改回一般買超 Top N (其餘不變)。
# 未來要再加 sniper master,把名字加進這個 set 即可。
SNIPER_MASTER_WHITELIST = {"蔣承翰"}


def _is_sniper_master(master_name: str) -> bool:
    """v3.26: 隔日沖 / 當沖 master → 改用漲停股資料源。
    v3.30.5: 限縮為白名單 (僅蔣承翰);不在名單者即使 style 符合也用一般買超 Top N。"""
    if master_name not in SNIPER_MASTER_WHITELIST:
        return False
    styles = MASTER_STYLES.get(master_name, [])
    return bool(SNIPER_STYLES.intersection(styles))


# ============================================================
#  Master mapping (extracted from user's manual file 5/8 version)
#  - 12 masters / 42 branch slots
#  - Branch codes are TWSE codes (authoritative)
#  - Branch names use canonical names from branches.py where possible
#  - header_label: "分點" or "常下分點" (matches manual section style)
# ============================================================

MASTER_MAPPING: List[Dict] = [
    {
        "name": "民哥",
        "header_label": "分點",
        "branches": [
            ("9B25", "台新-五權西"),
            ("9666", "富邦-南屯"),
            ("779W", "國票-彰化"),
        ],
    },
    {
        "name": "林滄海",
        "header_label": "常下分點",
        "branches": [
            ("9658", "富邦-建國"),
            ("9309", "華南永昌-古亭"),
            ("1260", "宏遠證券"),
            ("9216", "凱基-信義"),
        ],
    },
    {
        "name": "張濬安(航海王)",
        "header_label": "常下分點",
        "branches": [
            ("779Z", "國票-安和"),
            ("9B2E", "台新-城中"),
            ("920F", "凱基-站前"),
            ("6167", "中國信託-松江"),
            ("961M", "富邦-木柵"),
            ("9100", "群益金鼎證券"),
        ],
    },
    {
        # Manual file's first 5-branch block (A144:A198 with master name "陳族元")
        "name": "陳族元",
        "header_label": "常下分點",
        "branches": [
            ("8880", "國泰證券"),
            ("9300", "華南永昌證券"),
            ("9216", "凱基-信義"),
            ("9661", "富邦-新店"),
            ("9A9g", "永豐金-內湖"),
        ],
    },
    {
        # Manual file's unnamed 4-branch block (A200:A242 merge has no master name).
        # Maps to "陳律師" per branches.py canonical naming.
        "name": "陳律師",
        "header_label": "常下分點",
        "branches": [
            ("700c", "兆豐-民生"),
            ("8450", "康和總公司"),
            ("9A9R", "永豐金-信義"),
            ("585c", "統一-仁愛"),
        ],
    },
    {
        "name": "迷你哥/松山哥",
        "header_label": "常下分點",
        "branches": [
            ("9217", "凱基-松山"),
            ("9200", "凱基證券"),
            ("9600", "富邦證券"),
        ],
    },
    {
        "name": "布哥/n_nchang",
        "header_label": "常下分點",
        "branches": [
            ("9A8F", "永豐金-敦南"),
        ],
    },
    {
        "name": "強森",
        "header_label": "分點",
        "branches": [
            ("9B25", "台新-五權西"),
            ("9B2E", "台新-城中"),
            ("9B2r", "台新-城東"),
            ("984K", "元大-館前"),
            ("989N", "元大-內湖"),
            ("9215", "凱基-高美館"),
            ("9B2D", "台新-大昌"),
        ],
    },
    {
        "name": "Tradow",
        "header_label": "分點",
        "branches": [
            ("9B2a", "台新-松德"),
        ],
    },
    {
        "name": "巨人傑",
        "header_label": "分點",
        "branches": [
            ("9B2n", "台新-西松"),
            ("984K", "元大-館前"),
            ("9B2z", "台新-文心"),
        ],
    },
    {
        "name": "蔣承翰",
        "header_label": "分點",
        "branches": [
            ("9227", "凱基-城中"),
            ("9B18", "台新-建北"),
        ],
    },
    {
        "name": "大牌分析師",
        "header_label": "分點",
        "branches": [
            ("8563", "新光-新竹"),
        ],
    },
    {
        "name": "竹科主力分點",
        "header_label": "分點",
        "branches": [
            ("700V", "兆豐-新竹"),
            ("9647", "富邦-新竹"),
        ],
    },
]


# ============================================================
#  Layout constants (matching manual file exactly)
# ============================================================

FONT_NAME = "新細明體"  # PMingLiU - traditional Chinese serif (matches manual)
FONT_SIZE = 12
ROW_HEIGHT = 16.5

COL_WIDTHS = {
    "A": 19.28515625,
    "B": 18.28515625,
    "C": 11.85546875,
    "D": 20.28515625,
    "E": 18.42578125,
    "F": 21.140625,
    "G": 18.42578125,
    "H": 21.140625,
    "I": 13.0,
    "J": 13.140625,
    "K": 11.42578125,
    "L": 15.5703125,
}

STOCKS_PER_BRANCH = 10  # default: 10 rows per branch (pad blank if fewer)

# v3.29.5 (2026-05-23): per-branch override 客製化 row 數
# User 5/23 要求 大牌分析師 新光-新竹 (8563) 改 Top 20 (其他維持 10).
# TWSE 分點頁面 publish Top 15 買榜 + Top 15 賣榜 = 最多 30 unique 個股, Top 20 用既有資料就夠.
# 要再加分點就直接編這個 dict, e.g. {"8563": 20, "9227": 15} (蔣承翰城中改 15).
BRANCH_STOCK_OVERRIDES: Dict[str, int] = {
    "8563": 20,   # 大牌分析師 / 新光-新竹 → Top 20 (5/23 user request)
}


# v3.29.7 (2026-05-24): 排除非個股 — 老闆 Excel 只看個股
# v3.29.6 第一版漏掉:
#   1. market_classifier 回傳 lowercase 'etf' / 'etf_active', 之前用 {"ETF"} 不 match
#   2. heuristic 只擋 4-5 char ('0050', '00878'), 漏 6-char 期信 ETN ('00715L 期街口布蘭特正2', '00738U 期元大道瓊白銀')
# v3.29.7 修法:
#   - EXCLUDED_MARKET_TYPES 改 lowercase + 含 'etf_active'
#   - Heuristic 改成 code.startswith('00') (不限長度)
#     根據 market_classifier.py L72: 「所有 '00' 開頭 code 都歸類為 ETF」
# 未來要排除其他類型 (e.g. 'preferred' 特別股), 加進這個 set 即可
EXCLUDED_MARKET_TYPES: set = {"etf", "etf_active"}


def _branch_stocks_size(branch_code: str) -> int:
    """v3.29.5: 取得單一分點的 row 數 (override 優先, fallback default 10)."""
    return BRANCH_STOCK_OVERRIDES.get(branch_code, STOCKS_PER_BRANCH)


def _is_excluded_by_market_type(stock: Dict) -> bool:
    """v3.29.7: 判斷該 stock 是否為非個股 (ETF / 衍生 / 期信 / 商品) 應排除.

    判斷順序:
      1. stock['market_type'].lower() 在 EXCLUDED_MARKET_TYPES → 排除
         (crawler.py 主流程已注入 market_type, 此為最可靠來源)
      2. code 開頭 '00' → 排除 (heuristic, 不限長度)
         根據 market_classifier.py L72-76:
           「所有 '00' 開頭 code 都歸類為 ETF (含 etf_active)」
         涵蓋: 普通 ETF (0050, 00878, 00646), 6-char ETN/期信 (00715L, 00738U),
              主動型 ETF (006208A)
    """
    mt = (stock.get('market_type') or '').strip().lower()
    if mt in EXCLUDED_MARKET_TYPES:
        return True
    code = (stock.get('code') or '').strip()
    if code.startswith('00'):
        return True
    return False

NUMBER_FMT_INT = "#,##0"
NUMBER_FMT_PRICE = "0.00"
NUMBER_FMT_PNL = '0.00_ ;[Red]\\-0.00\\ '


def _font_bold() -> Font:
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=True)


def _font_normal() -> Font:
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=False)


def _font_pnl_neg() -> Font:
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=False, color="FFFF0000")


def _align_center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _header_row(header_label: str, include_master_label: bool) -> List[str]:
    """
    12-column header.
      header_label: "分點" or "常下分點" (master section style)
      include_master_label: True for master's full header (A="高手"), False for sub-header (A=None)
    """
    code_col_label = "代號" if header_label == "分點" else "分點代號"
    return [
        "高手" if include_master_label else None,
        header_label,
        code_col_label,
        "標的",
        "買進(張)",
        "賣出(張)",
        "買進(萬元)",
        "賣出(萬元)",
        "淨買差(萬元)",
        "買均",
        "賣均",
        "損益(萬)",
    ]


# ============================================================
#  Stock selection: top 10 by buy_amt for a given branch
# ============================================================

def _top_stocks_for_branch(branch_data: Dict, sniper_mode: bool = False,
                            n_top: int = None) -> List[Dict]:
    """
    Combine buys + sells lists, dedupe by code, sort by buy_amt desc, take top 10.
    Returns list of stock dicts with keys: code, name, buy_lot, sell_lot, buy_amt, sell_amt
    (buy_amt/sell_amt in 仟元 = thousand TWD per crawler convention).

    v3.26: when sniper_mode=True, restrict to limit-up stocks only (is_limit_up True).
    v3.28.1: 修補 sniper 過濾遺漏 net-seller 問題。
      使用者 review 發現 5/13 蔣承翰兩個分點都顯示微星,但實際 微星 是淨賣超。
      根因:TWSE 分點頁面 publish 兩榜 (買榜 Top 15 + 賣榜 Top 15),
            一檔股票若 grossly traded 兩邊大量,會同時在兩榜出現。
            v3.26 只 filter is_limit_up,沒檢查 net_amt > 0,
            於是「淨賣 但 gross 進買榜」的漲停股仍被選入 sniper 區段。
      用戶定義的正確語意:「先挑漲停股、觀察分點 *買超* 哪幾檔」 →
            net_amt > 0 (或 net_lot > 0) 才是真正 sniper 動作。
    Returns empty list if sniper master had zero net-bought limit-up stocks today (intentional).
    """
    if not branch_data:
        return []
    seen: Dict[str, Dict] = {}
    for s in (branch_data.get("buys") or []):
        c = s.get("code", "")
        if c and c not in seen:
            seen[c] = s
    for s in (branch_data.get("sells") or []):
        c = s.get("code", "")
        if c and c not in seen:
            seen[c] = s
    candidates = list(seen.values())

    # v3.29.6: 排除 ETF 等非個股 (老闆 Excel 只看個股)
    # 必須在 net_buyer + sniper filter *之前* 跑, 避免 ETF 占用 Top N 名額.
    candidates = [s for s in candidates if not _is_excluded_by_market_type(s)]

    # v3.29.1: 對 *所有* master (包含 swing) 加 net_buyer filter
    # 5/19 用戶 review 發現: 大牌分析師-新光新竹 顯示 6257 但實際淨賣 -1412 萬。
    # 系統性掃描全 Excel: 359 row 有資料,48 row (13%) 是淨賣股被選入 (e.g. 航海王 國票安和 國巨* 淨賣 2.1 億).
    # 根因: v3.28.1 只對 sniper_mode 加 net filter, swing master 仍按 buy_amt desc 排序,
    #       gross buy_amt 高 但 net 為負的個股污染 Excel.
    # 修法: 所有 master 先 filter net > 0, sniper_mode 再額外限定 is_limit_up.
    candidates = [
        s for s in candidates
        if (s.get("net_amt", 0) or 0) > 0 or (s.get("net_lot", 0) or 0) > 0
    ]

    if sniper_mode:
        # sniper master 額外限定: 只看漲停股
        candidates = [s for s in candidates if s.get("is_limit_up")]
    sorted_stocks = sorted(
        candidates,
        key=lambda x: x.get("buy_amt", 0) or 0,
        reverse=True,
    )
    # v3.29.5: n_top 預設 STOCKS_PER_BRANCH, 但允許 per-branch override (e.g. 8563→20)
    limit = n_top if n_top is not None else STOCKS_PER_BRANCH
    return sorted_stocks[:limit]


# ============================================================
#  Cell writing helpers
# ============================================================

def _write_header_row(ws: "Worksheet", row: int, header_label: str, include_master_label: bool):
    """Write a 12-cell header row. All cells bold + centered."""
    labels = _header_row(header_label, include_master_label)
    for ci, val in enumerate(labels, start=1):
        c = ws.cell(row=row, column=ci)
        c.value = val
        c.font = _font_bold()
        c.alignment = _align_center()


def _write_stock_row(ws: "Worksheet", row: int, stock: Dict, sniper_mode: bool = False):
    """Write 9 data columns (D-L) for a single stock. E-I integer, J-K price, L formula.
    v3.27.4 L4: sniper_mode=True 時在標的欄顯示漲幅%, 讓使用者一眼驗證漲停。"""
    code = stock.get("code", "") or ""
    name = stock.get("name", "") or code
    buy_lot = stock.get("buy_lot", 0) or 0
    sell_lot = stock.get("sell_lot", 0) or 0
    buy_amt_k = stock.get("buy_amt", 0) or 0   # in 仟元
    sell_amt_k = stock.get("sell_amt", 0) or 0

    # 仟元 -> 萬元 (divide by 10, round to nearest)
    buy_amt_w = round(buy_amt_k / 10) if buy_amt_k else 0
    sell_amt_w = round(sell_amt_k / 10) if sell_amt_k else 0
    net_w = buy_amt_w - sell_amt_w

    # Average prices: buy_amt_k(仟元) / buy_lot(張) gives elem-wise TWD/張 in thousands;
    # but per manual convention, we want TWD per share. 1張=1000股.
    # Manual shows e.g. 買均=386.99 for 國巨 — that's TWD/share.
    # So: buy_avg = buy_amt_k * 1000 / (buy_lot * 1000) = buy_amt_k / buy_lot (TWD/share)
    buy_avg = round(buy_amt_k / buy_lot, 2) if (buy_lot > 0 and buy_amt_k > 0) else 0
    sell_avg = round(sell_amt_k / sell_lot, 2) if (sell_lot > 0 and sell_amt_k > 0) else 0

    # D: stock label "name(code)"
    # v3.29.4: 移除 v3.27.4 L4 的 ▲X.XX% 標籤 (user 5/22 review 不希望直接看漲幅)
    # 漲停驗證改為間接靠 buy_amt 趨勢 / 額外 audit 工具
    c_d = ws.cell(row=row, column=4)
    c_d.value = f"{name}({code})"
    c_d.font = _font_normal()
    c_d.alignment = _align_center()

    # E-I integer columns
    for ci, val in [(5, buy_lot), (6, sell_lot), (7, buy_amt_w), (8, sell_amt_w), (9, net_w)]:
        c = ws.cell(row=row, column=ci)
        c.value = val
        c.font = _font_normal()
        c.alignment = _align_center()
        c.number_format = NUMBER_FMT_INT

    # J/K price columns
    for ci, val in [(10, buy_avg), (11, sell_avg)]:
        c = ws.cell(row=row, column=ci)
        c.value = val if val else 0
        c.font = _font_normal()
        c.alignment = _align_center()
        c.number_format = NUMBER_FMT_PRICE

    # L: P&L formula, red on negative
    c_l = ws.cell(row=row, column=12)
    c_l.value = f"=F{row}*(K{row}-J{row})"
    c_l.alignment = _align_center()
    c_l.number_format = NUMBER_FMT_PNL
    pnl_value = sell_lot * (sell_avg - buy_avg)
    if pnl_value < 0:
        c_l.font = _font_pnl_neg()
    else:
        c_l.font = _font_normal()


def _write_blank_data_row(ws: "Worksheet", row: int):
    """For padding rows when branch has fewer than 10 stocks. Apply formatting only."""
    for ci in range(4, 13):  # D-L
        c = ws.cell(row=row, column=ci)
        c.font = _font_normal()
        c.alignment = _align_center()
        if ci in (5, 6, 7, 8, 9):
            c.number_format = NUMBER_FMT_INT
        elif ci in (10, 11):
            c.number_format = NUMBER_FMT_PRICE
        elif ci == 12:
            c.number_format = NUMBER_FMT_PNL


def _write_empty_branch_notice_row(ws: "Worksheet", row: int, sniper_mode: bool):
    """v3.29.2: 該分點有 TWSE 資料但 filter 後 *完全* 空白 (0 stocks) → 在 D 欄寫提示行.

    sniper_mode=True:  '⚪ 此分點今日未搶漲停'
    sniper_mode=False: '⚪ 此分點今日無淨買超個股'
    """
    notice = '⚪ 此分點今日未搶漲停' if sniper_mode else '⚪ 此分點今日無淨買超個股'
    _write_notice_row(ws, row, notice)


def _write_partial_branch_notice_row(ws: "Worksheet", row: int, n_stocks: int, sniper_mode: bool):
    """v3.29.4: 該分點 partial fill (1-9 stocks 入選) → 在第 N+1 列寫提示告訴老闆「就這樣」.

    用戶 5/22 反映「凱基-松山 9217 顯示 4 stocks 後 6 row 空白看起來像 bug」.
    這個提示行區分「sniper 沒搶更多漲停」vs「我們漏抓」.

    sniper_mode=True:  '⚪ 今日漲停僅 N 檔'
    sniper_mode=False: '⚪ 今日淨買僅 N 檔'
    """
    if sniper_mode:
        notice = f'⚪ 今日漲停僅 {n_stocks} 檔'
    else:
        notice = f'⚪ 今日淨買僅 {n_stocks} 檔'
    _write_notice_row(ws, row, notice)


def _write_notice_row(ws: "Worksheet", row: int, notice: str):
    """共用 helper: 在 D 欄寫灰色斜體提示, E-L 維持空白格式 (跟 _write_blank_data_row 同)."""
    c_d = ws.cell(row=row, column=4)
    c_d.value = notice
    c_d.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=True, color="FF808080")
    c_d.alignment = _align_center()
    # 其他欄 (E-L) 維持空白格式
    for ci in range(5, 13):
        c = ws.cell(row=row, column=ci)
        c.font = _font_normal()
        c.alignment = _align_center()
        if ci in (5, 6, 7, 8, 9):
            c.number_format = NUMBER_FMT_INT
        elif ci in (10, 11):
            c.number_format = NUMBER_FMT_PRICE
        elif ci == 12:
            c.number_format = NUMBER_FMT_PNL


# ============================================================
#  Build single-day sheet
# ============================================================

def build_day_sheet(ws: "Worksheet", branches_data: List[Dict], trade_date: str):
    """
    Build one sheet matching manual template.

    Args:
      ws: openpyxl Worksheet (will be populated, sheet title set externally)
      branches_data: list of branch dicts from crawler (each has code, name, buys, sells)
      trade_date: YYYYMMDD string (used only for fallback display)

    Returns:
      total_rows: number of rows written (1-indexed)
    """
    # Index branches by code for fast lookup
    by_code: Dict[str, Dict] = {}
    for b in branches_data:
        c = b.get("code")
        if c:
            by_code[c] = b

    # Apply column widths
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    row = 1
    sniper_count = 0
    sniper_with_data = 0
    for master in MASTER_MAPPING:
        master_name = master["name"]
        header_label = master["header_label"]
        branches = master["branches"]
        if not branches:
            continue

        sniper_mode = _is_sniper_master(master_name)
        if sniper_mode:
            sniper_count += 1
            master_has_limit_up = False

        # Full master header row (A="高手" label)
        _write_header_row(ws, row, header_label, include_master_label=True)
        row += 1

        master_data_start = row  # first data row of this master block

        for bi, (branch_code, branch_canonical_name) in enumerate(branches):
            if bi > 0:
                # Sub-header before subsequent branches under same master
                _write_header_row(ws, row, header_label, include_master_label=False)
                row += 1

            # v3.29.5: 該分點是否有 override row 數 (e.g. 8563→20, 其他→10)
            branch_size = _branch_stocks_size(branch_code)

            # Lookup live branch data
            bdata = by_code.get(branch_code, {})
            stocks = _top_stocks_for_branch(bdata, sniper_mode=sniper_mode, n_top=branch_size)
            if sniper_mode and stocks:
                master_has_limit_up = True

            # v3.29.2/v3.29.4: 判斷該分點空白狀態
            has_branch_data = bool(bdata and (bdata.get('buys') or bdata.get('sells')))
            n_stocks = len(stocks)

            branch_first_row = row
            branch_last_row = row + branch_size - 1

            # Write branch_size rows (data + notice + blank padding)
            for ri in range(branch_size):
                r = branch_first_row + ri
                if ri < n_stocks:
                    _write_stock_row(ws, r, stocks[ri], sniper_mode=sniper_mode)
                elif ri == n_stocks and has_branch_data:
                    # 第一個空白 row 加 by-design 提示
                    if n_stocks == 0:
                        _write_empty_branch_notice_row(ws, r, sniper_mode=sniper_mode)
                    else:
                        _write_partial_branch_notice_row(ws, r, n_stocks, sniper_mode=sniper_mode)
                else:
                    _write_blank_data_row(ws, r)

            # B column: branch name (merged across 10 rows)
            cb = ws.cell(row=branch_first_row, column=2)
            cb.value = branch_canonical_name
            cb.font = _font_bold()
            cb.alignment = _align_center()
            ws.merge_cells(
                start_row=branch_first_row, start_column=2,
                end_row=branch_last_row, end_column=2,
            )
            # C column: branch code (merged across 10 rows)
            cc = ws.cell(row=branch_first_row, column=3)
            cc.value = branch_code
            cc.font = _font_bold()
            cc.alignment = _align_center()
            ws.merge_cells(
                start_row=branch_first_row, start_column=3,
                end_row=branch_last_row, end_column=3,
            )

            row = branch_last_row + 1  # advance to next branch

        # A column: master name (merged from first data row to last data row)
        master_data_end = row - 1
        ws.merge_cells(
            start_row=master_data_start, start_column=1,
            end_row=master_data_end, end_column=1,
        )
        ca = ws.cell(row=master_data_start, column=1)
        ca.value = master_name
        ca.font = _font_bold()
        ca.alignment = _align_center()

        if sniper_mode and master_has_limit_up:
            sniper_with_data += 1

    # Apply uniform row height
    for r in range(1, row):
        ws.row_dimensions[r].height = ROW_HEIGHT

    if sniper_count:
        print(f"  [Excel v3.26] sniper-mode masters: {sniper_with_data}/{sniper_count} "
              f"have limit-up buys today (others get blank rows)")

    return row - 1


# ============================================================
#  v3.31.0: 月檔 (一個月一份, chip_radar_YYYY-MM.xlsx)
# ============================================================

def _update_monthly_workbook(monthly_path: Path, branches_data: List[Dict],
                              trade_date: str):
    """v3.31.0: 開啟既有月檔 (若有) 或新建, add/update 該日 sheet, save back.
    sheet 名 = trade_date (YYYYMMDD), 同日重跑會覆寫該 sheet, sheets 按日期 desc 排序."""
    if monthly_path.exists():
        try:
            wb = load_workbook(str(monthly_path))
        except Exception as e:
            print(f"  [Excel] 月檔 unreadable, recreating: {e}")
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
    else:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    # 若該日 sheet 已存在 → 移除 (重跑同日 → 覆寫)
    if trade_date in wb.sheetnames:
        wb.remove(wb[trade_date])

    # 新建該日 sheet (build_day_sheet 在 ws 內 render 老闆版)
    ws = wb.create_sheet(title=trade_date)
    build_day_sheet(ws, branches_data, trade_date)

    # 按日期 desc 排序 (新日期在前)
    sheet_names = sorted(wb.sheetnames, reverse=True)
    wb._sheets = [wb[name] for name in sheet_names]

    wb.save(str(monthly_path))


# ============================================================
#  Multi-sheet latest.xlsx (legacy v3.30.x, retained for backfill/reference)
# ============================================================

def _update_latest_multi_sheet(latest_path: Path, branches_data: List[Dict],
                                trade_date: str, max_sheets: int = 30):
    """
    Update latest.xlsx to include current trade_date as a sheet.
    Keep only the most recent `max_sheets` sheets (by sheet name desc).
    Newest sheet is set as the active one.
    """
    if latest_path.exists():
        try:
            wb = load_workbook(str(latest_path))
        except Exception as e:
            print(f"  [Excel] latest.xlsx unreadable, recreating: {e}")
            wb = Workbook()
            # remove default sheet
            for s in list(wb.sheetnames):
                del wb[s]
    else:
        wb = Workbook()
        for s in list(wb.sheetnames):
            del wb[s]

    # Remove existing sheet with same name (will rebuild)
    if trade_date in wb.sheetnames:
        del wb[trade_date]

    # Create new sheet for trade_date
    ws = wb.create_sheet(title=trade_date)
    build_day_sheet(ws, branches_data, trade_date)

    # Sort sheets by name desc; trim to max_sheets
    sheet_names_sorted = sorted(wb.sheetnames, reverse=True)
    keep = sheet_names_sorted[:max_sheets]
    drop = sheet_names_sorted[max_sheets:]
    for n in drop:
        del wb[n]

    # Reorder so newest first
    wb._sheets = [wb[n] for n in keep]
    wb.active = 0  # show newest sheet on open

    wb.save(str(latest_path))


# ============================================================
#  Main entry point
# ============================================================

def generate_excel_report(branches_data: List[Dict], trade_date: str,
                           output_dir: str = "data/reports") -> Optional[str]:
    """
    Generate Excel daily report (mimics manual 「分點觀察」 layout).

    Args:
      branches_data: list of branch dicts from crawler (with buys/sells)
      trade_date: trading day in YYYYMMDD format (e.g. '20260508')
      output_dir: output directory (default 'data/reports')

    Returns:
      file path of the daily snapshot on success, None on failure.
    """
    if not OPENPYXL_AVAILABLE:
        print("  [Excel] openpyxl not installed, skipping")
        return None

    if not branches_data:
        print("  [Excel] no branch data, skipping")
        return None

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if len(trade_date) == 8 and trade_date.isdigit():
        readable_date = f"{trade_date[:4]}-{trade_date[4:6]}-{trade_date[6:8]}"
    else:
        readable_date = trade_date

    # v3.31.0: 月檔模式 ─ 一個月一份 chip_radar_YYYY-MM.xlsx (取代單日檔)
    #   latest.xlsx 永遠 = 當月月檔的 copy
    #   crawler 主流程每次 daily-full 跑完: 開啟月檔 → add/update 該日 sheet → save → copy latest
    year_month = readable_date[:7]   # "2026-06"
    monthly_path = out_dir / f"chip_radar_{year_month}.xlsx"
    latest_path = out_dir / "latest.xlsx"

    valid_count = sum(1 for b in branches_data if b.get("buys") or b.get("sells"))
    print(f"  [Excel] generating v3.31 monthly ({valid_count} branches, date {readable_date}, "
          f"month {year_month})")

    try:
        # 1. 月檔: 開啟既有 (若有) 或新建, add/update 該日 sheet
        _update_monthly_workbook(monthly_path, branches_data, trade_date)

        # 2. latest.xlsx = 當月月檔的 copy (前端下載按鈕指向不變)
        import shutil
        shutil.copy2(str(monthly_path), str(latest_path))

        # 3. Update README index
        _update_reports_readme(out_dir)

        size_kb = monthly_path.stat().st_size / 1024
        latest_size_kb = latest_path.stat().st_size / 1024
        sheet_count = len(load_workbook(str(monthly_path), read_only=True).sheetnames)
        print(f"  [Excel] OK")
        print(f"     {monthly_path.name} ({size_kb:.1f} KB, {sheet_count} daily sheets)")
        print(f"     latest.xlsx ({latest_size_kb:.1f} KB, = 當月月檔)")
        return str(monthly_path)

    except Exception as e:
        print(f"  [Excel] FAILED: {e}")
        import traceback
        traceback.print_exc()
        return None


def _update_reports_readme(reports_dir: Path):
    """Update reports/README.md with auto-generated file index."""
    try:
        excel_files = sorted(
            reports_dir.glob("chip_radar_*.xlsx"),
            key=lambda p: p.name,
            reverse=True,
        )

        readme = reports_dir / "README.md"
        master_count = len(MASTER_MAPPING)
        branch_count = sum(len(m["branches"]) for m in MASTER_MAPPING)

        lines = [
            "# Chip Radar 老闆版 Excel 日報",
            "",
            "由 `excel_report.py` v3.26 自動生成,模仿手動版「分點觀察」格式。",
            "",
            "## 最新檔案",
            "",
            "- [**latest.xlsx**](./latest.xlsx) — 多 sheet, 最近 30 個交易日",
            "  - 每個 sheet 命名 = `YYYYMMDD`,開啟時顯示最新一日",
            "",
            "## 結構",
            "",
            f"- {master_count} 位高手 / {branch_count} 個分點 slot (含跨高手共用分點)",
            "- 每分點固定 10 列 (不足以空白填補)",
            "- 12 欄: 高手 / 分點 / 代號 / 標的 / 買進(張) / 賣出(張) / "
            "買進(萬元) / 賣出(萬元) / 淨買差(萬元) / 買均 / 賣均 / 損益(萬)",
            "- L 欄公式: `=F*(K-J)` (賣出張數 × (賣均-買均)),負值紅字",
            "",
            "## v3.30.5 風格分流規則 (僅蔣承翰用漲停法)",
            "",
            "Excel 抓取法依 master 切換:",
            "",
            "| Master | Top N 資料源 |",
            "|---|---|",
            "| ⭐ 蔣承翰 (隔日沖) | **今日漲停股 by 買進金額** (漲跌幅 ≥ 9.5%) |",
            "| 其餘所有 master | 全部個股 by 買進金額 (淨買超 Top N) |",
            "",
            "蔣承翰今天若沒搶任何漲停股 → 整列空白 (不 fallback,維持風格純度)。",
            "(v3.26~v3.30.4 原為所有隔日沖/當沖 master 都用漲停法;v3.30.5 依使用者要求縮為僅蔣承翰)",
            "",
            "## 每日歷史",
            "",
        ]

        if excel_files:
            lines.append(f"近 {min(len(excel_files), 30)} 個交易日 (共 {len(excel_files)} 個檔案):")
            lines.append("")
            lines.append("| 日期 | 檔案 | 大小 |")
            lines.append("|------|------|------|")
            for p in excel_files[:30]:
                size_kb = p.stat().st_size / 1024
                # 從檔名抽日期
                stem = p.stem  # chip_radar_2026-05-08
                date_part = stem.replace("chip_radar_", "")
                lines.append(f"| {date_part} | [{p.name}](./{p.name}) | {size_kb:.1f} KB |")
            if len(excel_files) > 30:
                lines.append("")
                lines.append(f"…另有 {len(excel_files) - 30} 個更舊檔案")

        lines.extend([
            "",
            "---",
            "",
            f"*Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}*",
        ])

        readme.write_text("\n".join(lines), encoding="utf-8")
    except Exception as e:
        print(f"  [Excel] README update failed: {e}")
