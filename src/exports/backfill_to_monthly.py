"""
v3.31.0 一次性 backfill: 把所有 data/reports/chip_radar_YYYY-MM-DD.xlsx
單日檔合併成 chip_radar_YYYY-MM.xlsx 月檔, 刪除舊單日檔。

Usage: python backfill_to_monthly.py [--dry-run] [--no-delete]
  --dry-run:    只列計畫不執行
  --no-delete:  合併但不刪除舊單日檔 (保留並存)

跑完後:
  data/reports/chip_radar_2026-04.xlsx (4月所有 daily sheet)
  data/reports/chip_radar_2026-05.xlsx (5月所有 daily sheet)
  ... 舊 chip_radar_2026-MM-DD.xlsx 全刪 (除非 --no-delete)

latest.xlsx 不動 (等 crawler 下次跑會自動 copy 當月月檔覆寫)
"""
import argparse
import re
import sys
from copy import copy
from pathlib import Path
from openpyxl import Workbook, load_workbook

REPORTS_DIR = Path("data/reports")
DAILY_PATTERN = re.compile(r'^chip_radar_(\d{4})-(\d{2})-(\d{2})\.xlsx$')


def copy_sheet(src_ws, dst_ws):
    """跨 workbook 拷 sheet (cells + styles + merges + dims)。"""
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col].width = dim.width
        if dim.hidden:
            dst_ws.column_dimensions[col].hidden = True
    for row_num, dim in src_ws.row_dimensions.items():
        if dim.height:
            dst_ws.row_dimensions[row_num].height = dim.height
    for r in list(src_ws.merged_cells.ranges):
        dst_ws.merge_cells(str(r))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true')
    parser.add_argument('--no-delete', action='store_true')
    args = parser.parse_args()

    if not REPORTS_DIR.exists():
        print(f"❌ {REPORTS_DIR} not found")
        sys.exit(1)

    # 1. group 單日檔 by year-month
    groups = {}
    for f in REPORTS_DIR.glob("chip_radar_*.xlsx"):
        m = DAILY_PATTERN.match(f.name)
        if not m:
            continue   # skip 已合併月檔 / 其他
        ym = f"{m.group(1)}-{m.group(2)}"
        groups.setdefault(ym, []).append(f)

    if not groups:
        print("  沒有需要 backfill 的單日檔 (目錄已乾淨)")
        return

    print(f"  發現 {sum(len(v) for v in groups.values())} 個單日檔, 分 {len(groups)} 個月份:")
    for ym, files in sorted(groups.items()):
        print(f"    {ym}: {len(files)} 檔")

    if args.dry_run:
        print("\n  --dry-run, 不執行. 移除旗標再跑.")
        return

    # 2. 對每月合併
    for ym, files in sorted(groups.items()):
        monthly_path = REPORTS_DIR / f"chip_radar_{ym}.xlsx"
        if monthly_path.exists():
            print(f"  ⚠️ {monthly_path.name} 已存在, skip {ym}")
            continue

        wb_out = Workbook()
        if 'Sheet' in wb_out.sheetnames:
            wb_out.remove(wb_out['Sheet'])

        for f in sorted(files):
            wb_in = load_workbook(str(f))
            for sheet_name in wb_in.sheetnames:
                if sheet_name in wb_out.sheetnames:
                    continue   # 同日 sheet 已加 (跨檔重複, 罕見)
                src = wb_in[sheet_name]
                dst = wb_out.create_sheet(title=sheet_name)
                copy_sheet(src, dst)

        # sheets desc 排序 (新日期在前)
        sheet_names = sorted(wb_out.sheetnames, reverse=True)
        wb_out._sheets = [wb_out[name] for name in sheet_names]

        wb_out.save(str(monthly_path))
        size_kb = monthly_path.stat().st_size / 1024
        print(f"  ✅ {monthly_path.name} ({size_kb:.1f} KB, "
              f"{len(wb_out.sheetnames)} sheets from {len(files)} 單日檔)")

        # 3. 刪舊單日檔
        if not args.no_delete:
            for f in files:
                f.unlink()
            print(f"     🗑️ 已刪 {len(files)} 個舊單日檔")

    print("\n  完成. 接下來:")
    print("    git add data/reports/")
    print("    git commit -m 'feat(v3.31.0): backfill 單日檔 → 月檔'")


if __name__ == '__main__':
    main()
