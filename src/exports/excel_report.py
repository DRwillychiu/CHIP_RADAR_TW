"""
excel_report.py v3.26 - Style-aware Excel daily report

v3.26 change: 隔日沖 / 當沖 master 的資料源自動切換為「今天買的漲停股」,
波段 / 長線 master 維持原 Top 10 by 買超。視覺格式與手動版完全一致。

Strict mimicry of user's hand-curated Excel ("分點觀察" boss report).

Layout (per master block):
  Row 1: full header [高手|分點|代號|標的|...|損益(萬)]   ← A="高手" label
  Row 2: first data row of master's first branch          ← A=master_name (merged down)
  Rows 2-11: 10 data rows for branch #1 (padded blank if <10 stocks)
  Row 12: sub-header [(空)|分點|代號|標的|...]            ← A empty, under master merge
  Rows 13-22: 10 data rows for branch #2
  ...
  Next master: full header row again, A=高手 label

Visual style (matches manual file exactly):
  - Font: 新細明體 (PMingLiU), 12pt
  - All cells: center/center alignment
  - NO fills, NO borders
  - Header rows + master/branch/code labels: bold
  - L column: =F*(K-J), format '0.00_ ;[Red]\\-0.00\\ ' (red text on negative)

Data source routing (v3.26):
  - sniper master (next_day_flipper / day_trader) → top 10 漲停股 by buy_amt
  - other master (swing / longterm / unmapped)    → top 10 全部個股 by buy_amt (legacy)
  - sniper master with no limit-up buys today     → 10 blank rows (intentional, not fallback)

Outputs:
  data/reports/chip_radar_YYYY-MM-DD.xlsx  - single-sheet daily snapshot
  data/reports/latest.xlsx                  - multi-sheet, last 30 trading days
  data/reports/README.md                    - history index
"""

from typing import Dict, List, Optional, Tuple
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule, IconSetRule, CellIsRule, DataBarRule
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# v3.62.0 (Sprint 25 → v3.62.1): 用戶決定把 E1-E4 4 個 section 全合 1 sheet
# 月檔結構: [📋 今日 Dashboard (含全 4 section)] + 日期 sheet desc
DASHBOARD_SHEET_NAME = "📋 今日 Dashboard"
MOBILE_SHEET_NAME = "📱 手機摘要"   # v3.67.1 Phase 2.7
QUAD_TRACK_SHEET_NAME = "📈 Quad 實戰追蹤"   # v3.70.2 Phase 3.2 持續性追蹤
PINNED_TRACK_SHEET_NAME = "📌 Pinned Master 追蹤"   # v3.71.18 L2
QUAD_FAIL_SHEET_NAME = "📉 Quad 失效歸因"   # v3.70.3 Phase 3.2 失效學習
ENRICHMENT_SHEETS = [DASHBOARD_SHEET_NAME, MOBILE_SHEET_NAME,
                     QUAD_TRACK_SHEET_NAME, QUAD_FAIL_SHEET_NAME,
                     PINNED_TRACK_SHEET_NAME]
# 舊 sheet 名 (給 cleanup 移除舊月檔殘留)
LEGACY_ENRICHMENT_NAMES = ["📋 今日摘要", "🚨 異常警報", "📦 連續囤貨", "⚠️ 風險警示"]

try:
    from branches import MASTER_STYLES
except ImportError:
    MASTER_STYLES = {}

SNIPER_STYLES = {"next_day_flipper", "day_trader"}

# v3.30.5: 使用者要求 Excel 只有「蔣承翰」用漲停股抓取法。
# 原 v3.26 是「所有 next_day_flipper/day_trader style 的 master 都用漲停法」
# (蔣承翰 / 迷你哥-松山哥 / Tradow / 巨人傑),現縮為白名單僅蔣承翰一人,
# 其餘 sniper-style master 全部改回一般買超 Top N (其餘不變)。
# 未來要再加 sniper master,把名字加進這個 set 即可。
SNIPER_MASTER_WHITELIST = {"蔣承翰"}


# v3.31.1: 老闆版 Excel 色塊配色 (15 個個人大戶 + 法人類)
# 設計: 風格類別分大色系 + master 之內 lightness 微差
#   暖色 (紅/橘) = sniper (next_day_flipper / day_trader)
#   藍綠系     = swing (波段)
#   灰系       = longterm (長線)
#   外資/官股  = 中性灰 (不重點分析)
# header 較深 (master/branch label 列) + body 較淡 (data 列), 保留字型不刺眼
MASTER_BLOCK_COLORS = {
    # ── Sniper 暖色系 (5 個) ──
    "蔣承翰":                {"header": "FFE57373", "body": "FFFFEBEE"},  # 紅 (主 sniper, 漲停獵手)
    "巨人傑":                {"header": "FFEF9A9A", "body": "FFFFE5E5"},  # 淡紅 (雙風格)
    "Tradow":                {"header": "FFFFAB91", "body": "FFFFEDE7"},  # 橘紅
    "迷你哥/松山哥":         {"header": "FFFFB74D", "body": "FFFFF3E0"},  # 橘 (當沖)
    "Krenz(再多一位數本人)": {"header": "FFFFCC80", "body": "FFFFF8E1"},  # 淡橘
    # ── Swing 波段藍綠系 (8 個) ──
    "民哥":                  {"header": "FF81C784", "body": "FFE8F5E9"},  # 綠
    "林滄海":                {"header": "FFA5D6A7", "body": "FFEEF7EE"},  # 淡綠 (longterm tint)
    "張濬安(航海王)":        {"header": "FF64B5F6", "body": "FFE3F2FD"},  # 海藍 (航運王)
    "陳族元":                {"header": "FF90CAF9", "body": "FFE7F2FB"},  # 淡藍
    "陳律師":                {"header": "FFB39DDB", "body": "FFEDE7F6"},  # 紫
    "布哥/n_nchang":         {"header": "FF80DEEA", "body": "FFE0F7FA"},  # 青
    "強森":                  {"header": "FF80CBC4", "body": "FFE0F2F1"},  # 青綠
    "大牌分析師":            {"header": "FFAED581", "body": "FFF1F8E9"},  # 黃綠
    # ── Longterm 長線灰系 (2 個) ──
    "優式資本":              {"header": "FFBCAAA4", "body": "FFEFEBE9"},  # 灰棕
    "東億資本":              {"header": "FFB0BEC5", "body": "FFECEFF1"},  # 灰藍
}
DEFAULT_MASTER_COLOR = {"header": "FFD7D7D7", "body": "FFF5F5F5"}   # 未在表內的 fallback


def _apply_master_block_color(ws: "Worksheet",
                               header_rows: List[int],
                               data_rows: List[int],
                               master_anchor_row: int,
                               colors: Dict[str, str],
                               cols: int = 12) -> None:
    """v3.31.1: 對 master block 套色塊.
    header_rows: 高手/master label / sub-header 列 → 深色 header_fill
    data_rows:   stock data 列 → 淡色 body_fill
    master_anchor_row: A 欄 master name 的 merge anchor → 深色 (跨整個 block 視覺主色)
    cols: 套色欄數 (預設 A-L = 12)。"""
    if not colors:
        return
    try:
        body_fill = PatternFill("solid", fgColor=colors["body"])
        header_fill = PatternFill("solid", fgColor=colors["header"])
    except Exception:
        return  # color spec 壞掉就不套

    # 先全 block 套 body (淡), L 欄(12)留白給色階 conditional formatting
    # v3.72.3: 若 cell 已有黃色 fill (sniper top buyer highlight) 則保留, 不覆寫
    for r in data_rows:
        for c in range(1, cols + 1):
            if c == 12:
                continue
            cell = ws.cell(row=r, column=c)
            existing = cell.fill
            existing_color = getattr(getattr(existing, 'fgColor', None), 'rgb', None) if existing else None
            if existing_color == "FFFFFF00":
                continue  # preserve top-buyer 黃色
            cell.fill = body_fill

    # 再套 header rows (深)
    for r in header_rows:
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c).fill = header_fill

    # A 欄 master anchor 套 header (整個 block 主色, 覆寫 body)
    ws.cell(row=master_anchor_row, column=1).fill = header_fill


def _is_sniper_master(master_name: str) -> bool:
    """v3.26: 隔日沖 / 當沖 master → 改用漲停股資料源。
    v3.30.5: 限縮為白名單 (僅蔣承翰);不在名單者即使 style 符合也用一般買超 Top N。"""
    if master_name not in SNIPER_MASTER_WHITELIST:
        return False
    styles = MASTER_STYLES.get(master_name, [])
    return bool(SNIPER_STYLES.intersection(styles))


# ============================================================
#  Master mapping (extracted from user's manual file 5/8 version)
#  - 12 masters / 42 branch slots
#  - Branch codes are TWSE codes (authoritative)
#  - Branch names use canonical names from branches.py where possible
#  - header_label: "分點" or "常下分點" (matches manual section style)
# ============================================================

MASTER_MAPPING: List[Dict] = [
    {
        "name": "民哥",
        "header_label": "分點",
        "branches": [
            ("9B25", "台新-五權西"),
            ("9666", "富邦-南屯"),
            ("779W", "國票-彰化"),
        ],
    },
    {
        "name": "林滄海",
        "header_label": "常下分點",
        "branches": [
            ("9658", "富邦-建國"),
            ("9309", "華南永昌-古亭"),
            ("1260", "宏遠證券"),
            ("9216", "凱基-信義"),
        ],
    },
    {
        "name": "張濬安(航海王)",
        "header_label": "常下分點",
        "branches": [
            ("779Z", "國票-安和"),
            ("9B2E", "台新-城中"),
            ("920F", "凱基-站前"),
            ("6167", "中國信託-松江"),
            ("961M", "富邦-木柵"),
            ("9100", "群益金鼎證券"),
        ],
    },
    {
        # Manual file's first 5-branch block (A144:A198 with master name "陳族元")
        "name": "陳族元",
        "header_label": "常下分點",
        "branches": [
            ("8880", "國泰證券"),
            ("9300", "華南永昌證券"),
            ("9216", "凱基-信義"),
            ("9661", "富邦-新店"),
            ("9A9g", "永豐金-內湖"),
        ],
    },
    {
        # Manual file's unnamed 4-branch block (A200:A242 merge has no master name).
        # Maps to "陳律師" per branches.py canonical naming.
        "name": "陳律師",
        "header_label": "常下分點",
        "branches": [
            ("700c", "兆豐-民生"),
            ("8450", "康和總公司"),
            ("9A9R", "永豐金-信義"),
            ("585c", "統一-仁愛"),
        ],
    },
    {
        "name": "迷你哥/松山哥",
        "header_label": "常下分點",
        "branches": [
            ("9217", "凱基-松山"),
            ("9200", "凱基證券"),
            ("9600", "富邦證券"),
        ],
    },
    {
        "name": "布哥/n_nchang",
        "header_label": "常下分點",
        "branches": [
            ("9A8F", "永豐金-敦南"),
        ],
    },
    {
        "name": "強森",
        "header_label": "分點",
        "branches": [
            ("9B25", "台新-五權西"),
            ("9B2E", "台新-城中"),
            ("9B2r", "台新-城東"),
            ("984K", "元大-館前"),
            ("989N", "元大-內湖"),
            ("9215", "凱基-高美館"),
            ("9B2D", "台新-大昌"),
        ],
    },
    {
        "name": "Tradow",
        "header_label": "分點",
        "branches": [
            ("9B2a", "台新-松德"),
        ],
    },
    {
        "name": "巨人傑",
        "header_label": "分點",
        "branches": [
            ("9B2n", "台新-西松"),
            ("984K", "元大-館前"),
            ("9B2z", "台新-文心"),
        ],
    },
    {
        "name": "蔣承翰",
        "header_label": "分點",
        "branches": [
            ("9227", "凱基-城中"),
            ("9B18", "台新-建北"),
            ("9A9S", "永豐金-南京"),
        ],
    },
    {
        "name": "大牌分析師",
        "header_label": "分點",
        "branches": [
            ("8563", "新光-新竹"),
        ],
    },
    {
        "name": "竹科主力分點",
        "header_label": "分點",
        "branches": [
            ("700V", "兆豐-新竹"),
            ("9647", "富邦-新竹"),
        ],
    },
]


# v3.63.2: Dashboard 追蹤範圍 — 嚴格鎖定 MASTER_MAPPING 內的大戶
# 使用者要求: 沒放在 Excel 每日籌碼分點觀察清單 (= MASTER_MAPPING) 中的大戶,
# 絕對禁止出現在 Dashboard.
TRACKED_MASTERS: set = {m["name"] for m in MASTER_MAPPING}


# ══════════════════════════════════════════════════════════════════════
# v3.72.5: MASTER_MAPPING <-> branches.py 一致性 guard
# ══════════════════════════════════════════════════════════════════════
# 問題:MASTER_MAPPING (excel 版面順序 + header_label) 是硬 code 手工維護,
# branches.py 是 crawler 的 source of truth. 兩者若不同步 → excel 顯示的分點
# 跟實際 crawl 到的資料不一致 (crawl 到但沒顯示 / 顯示但抓不到).
#
# 這個 guard 在 module import 時執行, 找出 drift:
#   - MASTER_MAPPING 有 (code, name), 但 branches.py 沒有該 bno
#   - branches.py 有 bno 但 master name 對不上 MASTER_MAPPING
#   - MASTER_MAPPING 的 name 跟 branches.py canonical name 不一致 (warn)
def _validate_master_mapping_vs_branches() -> List[str]:
    """回傳 warning list. 若有 drift 印 stderr, 不 raise (避免 breakage)."""
    warnings = []
    try:
        from src.core.branches import get_branch_by_code, get_branches_by_master
    except Exception:
        return warnings  # 不能 import 就 skip (test env, 部分安裝)

    # 1. MASTER_MAPPING 每筆 code 必須存在於 branches.py
    for m in MASTER_MAPPING:
        master_name = m["name"]
        for code, name in m["branches"]:
            b = get_branch_by_code(code)
            if b is None:
                warnings.append(
                    f"[master_mapping] {master_name} 的 bno={code} 在 branches.py 找不到")
                continue
            # 2. bno 存在, 但 master 對不上
            b_master = b.get("master")
            b_co = b.get("co_masters", []) or []
            if b_master != master_name and master_name not in b_co:
                warnings.append(
                    f"[master_mapping] {master_name}/{code}: "
                    f"branches.py 該 bno 掛在 {b_master!r} (co={b_co}), 不含 {master_name}")
            # 3. 名稱 warn (canonical vs 手工)
            b_name = b.get("name", "")
            if b_name and name != b_name:
                warnings.append(
                    f"[master_mapping-name] {code}: MASTER_MAPPING 用 {name!r} 但 branches.py canonical 是 {b_name!r}")

    # 4. branches.py 有的 sniper master 分點, 卻不在 MASTER_MAPPING → 少算
    for m in MASTER_MAPPING:
        master_name = m["name"]
        mapped = {c for c, _ in m["branches"]}
        try:
            true_branches = get_branches_by_master(master_name, include_disabled=False)
        except Exception:
            continue
        true_codes = {b["code"] for b in true_branches}
        missing = true_codes - mapped
        if missing:
            warnings.append(
                f"[master_mapping-missing] {master_name}: branches.py 有 {missing}, "
                f"MASTER_MAPPING 沒有")
    return warnings


# 執行 (module import time)
_MASTER_MAPPING_WARNINGS = _validate_master_mapping_vs_branches()
if _MASTER_MAPPING_WARNINGS:
    import sys as _sys
    for _w in _MASTER_MAPPING_WARNINGS:
        print(f"  ⚠️ {_w}", file=_sys.stderr)


# v3.71.5 Phase 3.2 Premium Tier: master vol_spike 可靠度 (snapshot 2026-06-26)
# source: scripts/analyze_master_vol_spike_reliability.py
# 過去 33 picks / 9 trigger days backtest:
#   竹科主力分點   9 picks  88.9% hit  +4.65% mean
#   陳族元         6 picks  83.3% hit  +5.22% mean
#   陳律師        18 picks  77.8% hit  +4.90% mean (主力 trigger, 3 days)
#   其他 4 位 ≤75% hit
# 門檻: ≥77% hit AND n ≥ 5 → premium tier (quad alpha 信心高)
# ⚠️ 樣本仍小, 季度 review (next: 2026-09-30 後 60 天累積 → n→80+)
PREMIUM_MASTERS: set = {
    '陳律師',
    '竹科主力分點',
    '陳族元',
}

# v3.71.18 L 系列: PINNED_MASTERS — user 自定「常駐關注」 master
# 跟 PREMIUM 不同:
#   PREMIUM = 高 hit rate 自動篩選 (backtest 結果)
#   PINNED  = 用戶手動設定「我每天都要看」 (個人偏好)
# 觸發: 名稱欄 📌 marker / Mobile 專區 / Enrichment sheet
PINNED_MASTERS: set = {
    '大牌分析師',   # v3.71.18 user 要求, 47/47 active 高頻短打型, quad 不適用
}


# ════════════════════════════════════════════════════════════════════
# v3.67.0 Phase 2.6: Color Tokens + 視覺系統 (語義一致性)
# ════════════════════════════════════════════════════════════════════
# 設計原則: 顏色不是裝飾, 而是語義訊號. 同樣語義 (hot / 信號紅 / 損益) 應該
# 用同樣顏色, 避免散落 random hex 導致 trader 無法快速 pattern match.
COLORS = {
    # 品牌
    'brand_dark':      'FF1F2A48',   # Dashboard 大標題深藍底
    # 訊號 (台股慣例: 紅=漲 綠=跌)
    'signal_red':      'FFC62828',   # 漲停 / 正損益 / 多訊號 (Material 深紅)
    'signal_green':    'FF2E7D32',   # 跌停 / 負損益 / 空訊號 (Material 深綠)
    'tw_red':          'FFDC2626',   # 台股紅 (Section 0 標題 / 偏多 banner)
    'tw_green':        'FF059669',   # 台股綠 (偏空 banner / Q5 hit ✅)
    # Hot / 警示
    'hot_red':         'FFEF5350',   # data bar 紅 (借券 / 量爆)
    'hot_orange':      'FFFB923C',   # 橘紅 (J 集中度 / H ratio extremity)
    'attention_gold':  'FFFFB300',   # 金 (G 注意股 / 0 大戶數)
    'attention_amber': 'FFFFC107',   # 淡金 (Section 0 E 大戶數)
    # 規模 / 累積
    'scale_green':     'FF66BB6A',   # data bar 深綠 (買進規模)
    'scale_light':     'FF81C784',   # data bar 淡綠 (累積買金額)
    # Text / 弱化
    'text_muted':      'FF666666',   # KPI label
    'text_secondary':  'FF4B5563',   # Action card
    'text_neutral':    'FF374151',   # 一般正文
    'text_strong':     'FF000000',   # 強調黑字
    # 背景 (淡色系)
    'bg_tldr':         'FFFEF3C7',   # TL;DR 淡黃
    'bg_action':       'FFF3F4F6',   # Action 淡灰
    'bg_subhead':      'FFF9FAFB',   # 極淡灰 (sub-banner / zebra dark)
    'bg_zebra_light':  'FFFFFFFF',   # 白 (zebra light)
    'bg_zebra_dark':   'FFF9FAFB',   # 極淡灰 (zebra dark)
    'bg_attention':    'FFFEF3C7',   # 淡金 (G 注意股 header)
    'bg_danger':       'FFFEE2E2',   # 淡紅 (Section 0 / E header / 偏多 banner)
    'bg_safe':         'FFD1FAE5',   # 淡綠 (偏空 banner)
    'bg_warning':      'FFFFF7ED',   # 極淡橙 (Section 0 註腳)
    'bg_neutral':      'FFE5E7EB',   # 淡灰 (中性 banner)
    'bg_consensus':    'FFE8F5E9',   # 淡綠 (F 連續囤貨 header)
    'bg_pivot':        'FFE0E7FF',   # 淡藍 (J 標頭)
    'bg_top3_rank':    'FFFEF3C7',   # 金 (Section 0 top 3 rank highlight)
    # v3.70.0 Phase 3.2 落地: quad alpha 視覺 token
    'alpha_gold':      'FFD97706',   # 金/橘 (quad 命中股 — alpha 啟動)
    'bg_alpha_light':  'FFFEF3C7',   # 淡金底 (quad 命中股名稱底)
}


def _zebra_stripes(ws, data_start_row, data_end_row, col_start='B', col_end='N'):
    """v3.67.0 Phase 2.6: 應用斑馬條紋 (zebra stripes) 給資料表格.
    奇數 row (相對) → 淡灰底; 偶數 row → 白底 (空 fill).

    用法: 在 section data 全部 render 完後呼叫一次.
    注意: 跳過已有 master block color / data bar color 的 sections.
    """
    if data_end_row < data_start_row:
        return
    dark = PatternFill('solid', fgColor=COLORS['bg_zebra_dark'])
    from openpyxl.utils import column_index_from_string, get_column_letter
    c_start = column_index_from_string(col_start)
    c_end = column_index_from_string(col_end)
    for i, r in enumerate(range(data_start_row, data_end_row + 1)):
        if i % 2 == 1:   # 第 2, 4, 6... row 套淡灰
            for c_idx in range(c_start, c_end + 1):
                cell_ = ws.cell(r, c_idx)
                # 若已有 fill (master block color), 不覆蓋
                if cell_.fill and cell_.fill.fgColor and \
                   cell_.fill.fgColor.rgb and cell_.fill.fgColor.rgb != '00000000':
                    continue
                cell_.fill = dark


def _is_tracked_master(name: Optional[str]) -> bool:
    """Dashboard filter: 該 master 是否在每日 Excel 追蹤清單內."""
    return bool(name) and name in TRACKED_MASTERS


def _filter_tracked_branches(branches_data: List[Dict]) -> List[Dict]:
    """只保留 master 在追蹤清單內的 branch."""
    return [b for b in branches_data if _is_tracked_master(b.get('master'))]


# ============================================================
#  Layout constants (matching manual file exactly)
# ============================================================

FONT_NAME = "新細明體"  # PMingLiU - traditional Chinese serif (matches manual)
FONT_SIZE = 12
ROW_HEIGHT = 16.5

COL_WIDTHS = {
    "A": 19.28515625,
    "B": 18.28515625,
    "C": 11.85546875,
    "D": 20.28515625,
    "E": 18.42578125,
    "F": 21.140625,
    "G": 18.42578125,
    "H": 21.140625,
    "I": 13.0,
    "J": 13.140625,
    "K": 11.42578125,
    "L": 15.5703125,
}

STOCKS_PER_BRANCH = 10  # default: 10 rows per branch (pad blank if fewer)

# v3.29.5 (2026-05-23): per-branch override 客製化 row 數
# User 5/23 要求 大牌分析師 新光-新竹 (8563) 改 Top 20 (其他維持 10).
# TWSE 分點頁面 publish Top 15 買榜 + Top 15 賣榜 = 最多 30 unique 個股, Top 20 用既有資料就夠.
# 要再加分點就直接編這個 dict, e.g. {"8563": 20, "9227": 15} (蔣承翰城中改 15).
BRANCH_STOCK_OVERRIDES: Dict[str, int] = {
    "8563": 20,   # 大牌分析師 / 新光-新竹 → Top 20 (5/23 user request)
}


# v3.29.7 (2026-05-24): 排除非個股 — 老闆 Excel 只看個股
# v3.29.6 第一版漏掉:
#   1. market_classifier 回傳 lowercase 'etf' / 'etf_active', 之前用 {"ETF"} 不 match
#   2. heuristic 只擋 4-5 char ('0050', '00878'), 漏 6-char 期信 ETN ('00715L 期街口布蘭特正2', '00738U 期元大道瓊白銀')
# v3.29.7 修法:
#   - EXCLUDED_MARKET_TYPES 改 lowercase + 含 'etf_active'
#   - Heuristic 改成 code.startswith('00') (不限長度)
#     根據 market_classifier.py L72: 「所有 '00' 開頭 code 都歸類為 ETF」
# 未來要排除其他類型 (e.g. 'preferred' 特別股), 加進這個 set 即可
EXCLUDED_MARKET_TYPES: set = {"etf", "etf_active"}


def _branch_stocks_size(branch_code: str) -> int:
    """v3.29.5: 取得單一分點的 row 數 (override 優先, fallback default 10)."""
    return BRANCH_STOCK_OVERRIDES.get(branch_code, STOCKS_PER_BRANCH)


def _is_excluded_by_market_type(stock: Dict) -> bool:
    """v3.29.7: 判斷該 stock 是否為非個股 (ETF / 衍生 / 期信 / 商品) 應排除.

    判斷順序:
      1. stock['market_type'].lower() 在 EXCLUDED_MARKET_TYPES → 排除
         (crawler.py 主流程已注入 market_type, 此為最可靠來源)
      2. code 開頭 '00' → 排除 (heuristic, 不限長度)
         根據 market_classifier.py L72-76:
           「所有 '00' 開頭 code 都歸類為 ETF (含 etf_active)」
         涵蓋: 普通 ETF (0050, 00878, 00646), 6-char ETN/期信 (00715L, 00738U),
              主動型 ETF (006208A)
    """
    mt = (stock.get('market_type') or '').strip().lower()
    if mt in EXCLUDED_MARKET_TYPES:
        return True
    code = (stock.get('code') or '').strip()
    if code.startswith('00'):
        return True
    return False

NUMBER_FMT_INT = "#,##0"
NUMBER_FMT_PRICE = "0.00"
# v3.64.0: 純數字 + 千分位 (不靠 format color, 由 font color 處理)
NUMBER_FMT_PNL = '#,##0.00;-#,##0.00'


def _font_bold() -> Font:
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=True)


def _font_normal() -> Font:
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=False)


# v3.64.0: L 欄損益高對比色 (取代 v3.63.1 紅綠色塊 + 白字, 用戶反饋 master block 色底上看不清)
# 台股傳統: 紅 = 賺 (正), 綠 = 虧 (負). 字色設深色, 不靠 cell fill, 跟 master block 不衝突
def _font_pnl_neg() -> Font:
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FF2E7D32")  # 深綠


def _font_pnl_pos() -> Font:
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FFC62828")  # 深紅


def _align_center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _header_row(header_label: str, include_master_label: bool) -> List[str]:
    """
    12-column header.
      header_label: "分點" or "常下分點" (master section style)
      include_master_label: True for master's full header (A="高手"), False for sub-header (A=None)
    """
    code_col_label = "代號" if header_label == "分點" else "分點代號"
    return [
        "高手" if include_master_label else None,
        header_label,
        code_col_label,
        "標的",
        "買進(張)",
        "賣出(張)",
        "買進(萬元)",
        "賣出(萬元)",
        "淨買差(萬元)",
        "買均",
        "賣均",
        "損益(萬)",
    ]


# ============================================================
#  Stock selection: top 10 by buy_amt for a given branch
# ============================================================

def _top_stocks_for_branch(branch_data: Dict, sniper_mode: bool = False,
                            n_top: int = None) -> List[Dict]:
    """
    Combine buys + sells lists, dedupe by code, sort by buy_amt desc, take top 10.
    Returns list of stock dicts with keys: code, name, buy_lot, sell_lot, buy_amt, sell_amt
    (buy_amt/sell_amt in 仟元 = thousand TWD per crawler convention).

    v3.26: when sniper_mode=True, restrict to limit-up stocks only (is_limit_up True).
    v3.28.1: 修補 sniper 過濾遺漏 net-seller 問題。
      使用者 review 發現 5/13 蔣承翰兩個分點都顯示微星,但實際 微星 是淨賣超。
      根因:TWSE 分點頁面 publish 兩榜 (買榜 Top 15 + 賣榜 Top 15),
            一檔股票若 grossly traded 兩邊大量,會同時在兩榜出現。
            v3.26 只 filter is_limit_up,沒檢查 net_amt > 0,
            於是「淨賣 但 gross 進買榜」的漲停股仍被選入 sniper 區段。
      用戶定義的正確語意:「先挑漲停股、觀察分點 *買超* 哪幾檔」 →
            net_amt > 0 (或 net_lot > 0) 才是真正 sniper 動作。
    Returns empty list if sniper master had zero net-bought limit-up stocks today (intentional).
    """
    if not branch_data:
        return []
    seen: Dict[str, Dict] = {}
    for s in (branch_data.get("buys") or []):
        c = s.get("code", "")
        if c and c not in seen:
            seen[c] = s
    for s in (branch_data.get("sells") or []):
        c = s.get("code", "")
        if c and c not in seen:
            seen[c] = s
    candidates = list(seen.values())

    # v3.29.6: 排除 ETF 等非個股 (老闆 Excel 只看個股)
    # 必須在 net_buyer + sniper filter *之前* 跑, 避免 ETF 占用 Top N 名額.
    candidates = [s for s in candidates if not _is_excluded_by_market_type(s)]

    # v3.29.1: 對 *所有* master (包含 swing) 加 net_buyer filter
    # 5/19 用戶 review 發現: 大牌分析師-新光新竹 顯示 6257 但實際淨賣 -1412 萬。
    # 系統性掃描全 Excel: 359 row 有資料,48 row (13%) 是淨賣股被選入 (e.g. 航海王 國票安和 國巨* 淨賣 2.1 億).
    # 根因: v3.28.1 只對 sniper_mode 加 net filter, swing master 仍按 buy_amt desc 排序,
    #       gross buy_amt 高 但 net 為負的個股污染 Excel.
    # 修法: 所有 master 先 filter net > 0, sniper_mode 再額外限定 is_limit_up.
    candidates = [
        s for s in candidates
        if (s.get("net_amt", 0) or 0) > 0 or (s.get("net_lot", 0) or 0) > 0
    ]

    if sniper_mode:
        # sniper master 額外限定: 只看漲停股
        candidates = [s for s in candidates if s.get("is_limit_up")]
    sorted_stocks = sorted(
        candidates,
        key=lambda x: x.get("buy_amt", 0) or 0,
        reverse=True,
    )
    # v3.29.5: n_top 預設 STOCKS_PER_BRANCH, 但允許 per-branch override (e.g. 8563→20)
    limit = n_top if n_top is not None else STOCKS_PER_BRANCH
    return sorted_stocks[:limit]


# ============================================================
#  Cell writing helpers
# ============================================================

def _write_header_row(ws: "Worksheet", row: int, header_label: str, include_master_label: bool):
    """Write a 12-cell header row. All cells bold + centered."""
    labels = _header_row(header_label, include_master_label)
    for ci, val in enumerate(labels, start=1):
        c = ws.cell(row=row, column=ci)
        c.value = val
        c.font = _font_bold()
        c.alignment = _align_center()


# v3.72.7: histock fetch 統計 (P2 #5 — 監控 rate limit / block)
# v3.72.10: 拆分 fetch_fail (fetch_histock_branch 回 None, 通常 HTTP/timeout/block)
#           vs empty_buys (抓到 dict 但 buys 空, 真的沒資料)
# 每次 build_day_sheet 開始清空, 結束 dump 到 stderr + 累加到 module-level 統計
_HISTOCK_STATS = {
    "attempted": 0,       # 呼叫次數
    "success": 0,         # 成功抓到 top #1 且 net>0
    "stale_date": 0,      # v3.72.5 時效不符
    "fetch_fail": 0,      # v3.72.10 fetch_histock_branch 回 None (HTTP/timeout/block)
    "empty_buys": 0,      # v3.72.10 抓到 dict 但 buys 空 (真無資料)
    "http_error": 0,      # Python 例外 (import fail 等)
    "net_zero_or_neg": 0, # buys[0].net <= 0
}


def _reset_histock_stats():
    global _HISTOCK_STATS
    for k in _HISTOCK_STATS:
        _HISTOCK_STATS[k] = 0


def _get_histock_stats() -> Dict[str, int]:
    """公開 stats 給 test / 外部監控."""
    return dict(_HISTOCK_STATS)


def _fetch_histock_top_buyer(stock_code: str, cache: Dict[str, Optional[str]],
                              trade_date: Optional[str] = None) -> Optional[str]:
    """v3.72.4: 透過 histock 個股分點榜找該股當日 top #1 買方 bno.
    v3.72.5: 加時效 guard — 若 histock 頁面日期 != trade_date → 回 None 避免假信號.
    v3.72.7: 加 fetch stats 收集 + net<=0 guard (bug fix).

    Args:
      stock_code: 股票代號 (e.g. '6577')
      cache: dict cache {stock_code: bno or None}, 避免同 run 重複 fetch
      trade_date: YYYYMMDD 我們期待的當日. 若 histock 頁面日期不符 → skip.

    Returns:
      top #1 買方 bno (e.g. '9A9S') / None (fetch 失敗、無資料、日期不符、或 net<=0)
    """
    if stock_code in cache:
        return cache[stock_code]
    _HISTOCK_STATS["attempted"] += 1
    try:
        from src.audit.histock_branch_audit import fetch_histock_branch
        # v3.72.10: 從 timeout=8/retry=1 加大到 timeout=15/retry=2
        # 07-31 real-world 0/40 fail 主因是 GH Actions 冷連線 + histock 遠端 latency 超時
        data = fetch_histock_branch(stock_code, timeout=15, max_retries=2)
        if not data:
            # v3.72.10: fetch_histock_branch 內部 catch 所有 exception 回 None
            # 這裡分開統計 (區分「無法連線」vs「連線 OK 但 buys 空」)
            _HISTOCK_STATS["fetch_fail"] += 1
            cache[stock_code] = None
            return None
        if not data.get('buys'):
            _HISTOCK_STATS["empty_buys"] += 1
            cache[stock_code] = None
            return None
        # v3.72.5: 時效 guard
        if trade_date:
            histock_date = (data.get('date') or '').replace('/', '')
            if histock_date and histock_date != trade_date:
                _HISTOCK_STATS["stale_date"] += 1
                cache[stock_code] = None
                return None
        # v3.72.7: net<=0 guard (histock 排序 desc, 但 buys[0].net 可能仍 <=0 若當日全負)
        top = data['buys'][0]
        top_net = int(top.get('net', 0) or 0)
        if top_net <= 0:
            _HISTOCK_STATS["net_zero_or_neg"] += 1
            cache[stock_code] = None
            return None
        top_bno = top.get('bno')
        _HISTOCK_STATS["success"] += 1
        cache[stock_code] = top_bno
        return top_bno
    except Exception:
        _HISTOCK_STATS["http_error"] += 1
    cache[stock_code] = None
    return None


def _build_top_net_buyer_index(branches_data: List[Dict],
                                sniper_stock_codes: Optional[set] = None,
                                trade_date: Optional[str] = None) -> Dict[str, str]:
    """v3.72.4: 建 {stock_code: bno_of_top_net_buyer} index.
    v3.72.5: 加 trade_date 傳遞供時效 guard.

    ★ 判定範圍: 從 histock 全市場分點榜 top #1 買方 (不再限於 tracked branches).
    ★ 為避免 histock 塞爆, 只 fetch sniper_stock_codes 內的股票 (蔣承翰買的漲停股).
    ★ v3.72.5: histock 頁面日期 != trade_date → 該股不列 index (safe skip).

    Args:
      branches_data: crawler 產出的 tracked branches (v3.72.3 fallback 用)
      sniper_stock_codes: set of stock codes to fetch histock for (通常 <5 檔)
      trade_date: YYYYMMDD, 用於 histock date match check

    Returns:
      {stock_code: top_buyer_bno} - 只包含 histock date 匹配且成功抓到的股票
    """
    top_index: Dict[str, str] = {}
    if sniper_stock_codes:
        # v3.72.4 新路徑: 用 histock 全市場榜
        cache: Dict[str, Optional[str]] = {}
        for scode in sniper_stock_codes:
            top_bno = _fetch_histock_top_buyer(scode, cache, trade_date=trade_date)
            if top_bno:
                top_index[scode] = top_bno
        return top_index

    # Fallback (若沒 sniper 買漲停 or 呼叫者沒指定): 空 index
    return top_index


# v3.72.3: 黃色 highlight (蔣承翰漲停股當日買超#1)
_TOP_BUYER_FILL = None
def _get_top_buyer_fill():
    global _TOP_BUYER_FILL
    if _TOP_BUYER_FILL is None and OPENPYXL_AVAILABLE:
        _TOP_BUYER_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    return _TOP_BUYER_FILL


def _write_stock_row(ws: "Worksheet", row: int, stock: Dict, sniper_mode: bool = False,
                     is_top_net_buyer: bool = False):
    """Write 9 data columns (D-L) for a single stock. E-I integer, J-K price, L formula.
    v3.27.4 L4: sniper_mode=True 時在標的欄顯示漲幅%, 讓使用者一眼驗證漲停。
    v3.72.3: is_top_net_buyer=True → 該 row (D-L) 背景黃色 (該股當日買超#1)。"""
    code = stock.get("code", "") or ""
    name = stock.get("name", "") or code
    buy_lot = stock.get("buy_lot", 0) or 0
    sell_lot = stock.get("sell_lot", 0) or 0
    buy_amt_k = stock.get("buy_amt", 0) or 0   # in 仟元
    sell_amt_k = stock.get("sell_amt", 0) or 0

    # 仟元 -> 萬元 (divide by 10, round to nearest)
    buy_amt_w = round(buy_amt_k / 10) if buy_amt_k else 0
    sell_amt_w = round(sell_amt_k / 10) if sell_amt_k else 0
    net_w = buy_amt_w - sell_amt_w

    # Average prices: buy_amt_k(仟元) / buy_lot(張) gives elem-wise TWD/張 in thousands;
    # but per manual convention, we want TWD per share. 1張=1000股.
    # Manual shows e.g. 買均=386.99 for 國巨 — that's TWD/share.
    # So: buy_avg = buy_amt_k * 1000 / (buy_lot * 1000) = buy_amt_k / buy_lot (TWD/share)
    buy_avg = round(buy_amt_k / buy_lot, 2) if (buy_lot > 0 and buy_amt_k > 0) else 0
    sell_avg = round(sell_amt_k / sell_lot, 2) if (sell_lot > 0 and sell_amt_k > 0) else 0

    # D: stock label "name(code)"
    # v3.29.4: 移除 v3.27.4 L4 的 ▲X.XX% 標籤 (user 5/22 review 不希望直接看漲幅)
    # 漲停驗證改為間接靠 buy_amt 趨勢 / 額外 audit 工具
    c_d = ws.cell(row=row, column=4)
    c_d.value = f"{name}({code})"
    c_d.font = _font_normal()
    c_d.alignment = _align_center()

    # E-I integer columns
    for ci, val in [(5, buy_lot), (6, sell_lot), (7, buy_amt_w), (8, sell_amt_w), (9, net_w)]:
        c = ws.cell(row=row, column=ci)
        c.value = val
        c.font = _font_normal()
        c.alignment = _align_center()
        c.number_format = NUMBER_FMT_INT

    # J/K price columns
    for ci, val in [(10, buy_avg), (11, sell_avg)]:
        c = ws.cell(row=row, column=ci)
        c.value = val if val else 0
        c.font = _font_normal()
        c.alignment = _align_center()
        c.number_format = NUMBER_FMT_PRICE

    # L: P&L formula, red on negative
    c_l = ws.cell(row=row, column=12)
    c_l.value = f"=F{row}*(K{row}-J{row})"
    c_l.alignment = _align_center()
    c_l.number_format = NUMBER_FMT_PNL
    pnl_value = sell_lot * (sell_avg - buy_avg)
    if pnl_value < 0:
        c_l.font = _font_pnl_neg()
    else:
        c_l.font = _font_pnl_pos()

    # v3.72.3: 若是該股當日買超#1 → D-L 背景黃色
    if is_top_net_buyer:
        fill = _get_top_buyer_fill()
        if fill:
            for ci in range(4, 13):  # D-L
                ws.cell(row=row, column=ci).fill = fill


def _write_blank_data_row(ws: "Worksheet", row: int):
    """For padding rows when branch has fewer than 10 stocks. Apply formatting only."""
    for ci in range(4, 13):  # D-L
        c = ws.cell(row=row, column=ci)
        c.font = _font_normal()
        c.alignment = _align_center()
        if ci in (5, 6, 7, 8, 9):
            c.number_format = NUMBER_FMT_INT
        elif ci in (10, 11):
            c.number_format = NUMBER_FMT_PRICE
        elif ci == 12:
            c.number_format = NUMBER_FMT_PNL


def _write_empty_branch_notice_row(ws: "Worksheet", row: int, sniper_mode: bool):
    """v3.29.2: 該分點有 TWSE 資料但 filter 後 *完全* 空白 (0 stocks) → 在 D 欄寫提示行.

    sniper_mode=True:  '⚪ 此分點今日未搶漲停'
    sniper_mode=False: '⚪ 此分點今日無淨買超個股'
    """
    notice = '⚪ 此分點今日未搶漲停' if sniper_mode else '⚪ 此分點今日無淨買超個股'
    _write_notice_row(ws, row, notice)


def _write_partial_branch_notice_row(ws: "Worksheet", row: int, n_stocks: int, sniper_mode: bool):
    """v3.29.4: 該分點 partial fill (1-9 stocks 入選) → 在第 N+1 列寫提示告訴老闆「就這樣」.

    用戶 5/22 反映「凱基-松山 9217 顯示 4 stocks 後 6 row 空白看起來像 bug」.
    這個提示行區分「sniper 沒搶更多漲停」vs「我們漏抓」.

    sniper_mode=True:  '⚪ 今日漲停僅 N 檔'
    sniper_mode=False: '⚪ 今日淨買僅 N 檔'
    """
    if sniper_mode:
        notice = f'⚪ 今日漲停僅 {n_stocks} 檔'
    else:
        notice = f'⚪ 今日淨買僅 {n_stocks} 檔'
    _write_notice_row(ws, row, notice)


def _write_notice_row(ws: "Worksheet", row: int, notice: str):
    """共用 helper: 在 D 欄寫灰色斜體提示, E-L 維持空白格式 (跟 _write_blank_data_row 同)."""
    c_d = ws.cell(row=row, column=4)
    c_d.value = notice
    c_d.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=False, italic=True, color="FF808080")
    c_d.alignment = _align_center()
    # 其他欄 (E-L) 維持空白格式
    for ci in range(5, 13):
        c = ws.cell(row=row, column=ci)
        c.font = _font_normal()
        c.alignment = _align_center()
        if ci in (5, 6, 7, 8, 9):
            c.number_format = NUMBER_FMT_INT
        elif ci in (10, 11):
            c.number_format = NUMBER_FMT_PRICE
        elif ci == 12:
            c.number_format = NUMBER_FMT_PNL


# v3.72.8: histock 失敗警示 + 時間戳 (bug #4 + #8)
def _write_histock_status_notice(ws: "Worksheet", row: int, stats: Dict[str, int],
                                  trade_date: Optional[str] = None) -> int:
    """若 histock 全 fail 或 <50% → 寫警示 row 在 Section 0 頂端 (row 1).

    Returns: 用了幾 rows (0 = 沒寫, 1 = 寫了 1 row 警示)
    """
    attempted = stats.get("attempted", 0)
    success = stats.get("success", 0)
    if attempted == 0:
        return 0  # 沒 fetch 任何 histock (no sniper 買漲停 or 未啟用) → 不寫
    if success == attempted:
        return 0  # 100% 成功 → 不寫警示
    # 有失敗, 分析主因
    stale = stats.get("stale_date", 0)
    no_data = stats.get("no_data", 0)
    http_err = stats.get("http_error", 0)
    net_neg = stats.get("net_zero_or_neg", 0)
    success_rate = int(success / attempted * 100)

    # v3.72.10: 新分類 fetch_fail (HTTP/timeout/block) vs empty_buys (真無資料)
    fetch_fail = stats.get("fetch_fail", 0)
    empty = stats.get("empty_buys", stats.get("no_data", 0))  # backward-compat
    # 主因判定
    if success == 0:
        # 全 fail
        if fetch_fail >= stale and fetch_fail >= empty and fetch_fail > 0:
            reason = f"histock 網站抓取失敗 (timeout / block / server down, {fetch_fail} 次)"
        elif stale >= empty:
            reason = f"histock 資料仍是 T-1 (需等到當日盤後晚間 update, {stale} 次)"
        else:
            reason = f"histock 分點榜真無資料 ({empty} 次, 可能個股冷門)"
        notice = f"⚠️ 本 Excel 無 top-buyer highlight — {reason} | attempted={attempted}, success=0"
    else:
        # 部分 fail
        notice = (f"⚠️ 部分 top-buyer highlight 缺 (histock: {success}/{attempted} = "
                  f"{success_rate}% success | fetch_fail={fetch_fail} stale={stale} empty={empty})")

    # 用 orange fill 讓警示醒目
    _write_notice_row(ws, row, notice)
    orange_fill = PatternFill("solid", fgColor="FFFFECB3")  # 淺橘
    for ci in range(1, 13):
        ws.cell(row=row, column=ci).fill = orange_fill
    return 1


def _write_histock_timestamp_footer(ws: "Worksheet", row: int, stats: Dict[str, int]) -> int:
    """Section 0 尾端加 histock 資料時間戳 (informational).

    v3.72.10: 用 TW timezone (UTC+8), 修正 GH Actions 顯示 UTC 時間的 bug.

    Returns: 用了幾 rows.
    """
    attempted = stats.get("attempted", 0)
    if attempted == 0:
        return 0  # 沒 fetch → 不寫
    success = stats.get("success", 0)
    # v3.72.10: TW timezone (Asia/Taipei = UTC+8)
    from datetime import datetime as _dt, timezone as _tz, timedelta as _td
    tw = _dt.now(_tz(_td(hours=8)))
    now = tw.strftime("%Y-%m-%d %H:%M")
    notice = f"ⓘ histock top-buyer 資料 fetched @ {now} TW | {success}/{attempted} success"
    c_d = ws.cell(row=row, column=1)
    c_d.value = notice
    c_d.font = Font(name=FONT_NAME, size=10, bold=False, italic=True, color="FFAAAAAA")
    c_d.alignment = _align_center()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    return 1


# ============================================================
#  Build single-day sheet
# ============================================================

def build_day_sheet(ws: "Worksheet", branches_data: List[Dict], trade_date: str):
    """
    Build one sheet matching manual template.

    Args:
      ws: openpyxl Worksheet (will be populated, sheet title set externally)
      branches_data: list of branch dicts from crawler (each has code, name, buys, sells)
      trade_date: YYYYMMDD string (used only for fallback display)

    Returns:
      total_rows: number of rows written (1-indexed)
    """
    # Index branches by code for fast lookup
    by_code: Dict[str, Dict] = {}
    for b in branches_data:
        c = b.get("code")
        if c:
            by_code[c] = b

    # v3.72.4: 先掃 sniper master 買的漲停股 → fetch histock 全市場 top #1 買方
    # 這樣才是「該股全市場分點榜 #1」而不是「tracked branches 內 #1」
    # v3.72.7: reset histock fetch stats (per build)
    _reset_histock_stats()
    sniper_stock_codes: set = set()
    for master in MASTER_MAPPING:
        if not _is_sniper_master(master["name"]):
            continue
        for branch_code, _ in master["branches"]:
            bdata = by_code.get(branch_code, {})
            for s in (bdata.get("buys") or []):
                if s.get("is_limit_up") and (s.get("net_amt", 0) > 0 or s.get("net_lot", 0) > 0):
                    scode = s.get("code")
                    if scode and not _is_excluded_by_market_type(s):
                        sniper_stock_codes.add(scode)
    top_net_buyer: Dict[str, str] = _build_top_net_buyer_index(
        branches_data,
        sniper_stock_codes=sniper_stock_codes if sniper_stock_codes else None,
        trade_date=trade_date,  # v3.72.5 時效 guard
    )

    # v3.72.7: dump histock fetch stats to stderr (監控 rate limit / block)
    # v3.72.10: 增 fetch_fail / empty_buys 分類
    stats = _get_histock_stats()
    if stats["attempted"] > 0:
        success_rate = stats["success"] / stats["attempted"] * 100
        import sys as _sys
        print(f"[histock stats] {stats['attempted']} attempted "
              f"→ {stats['success']} success ({success_rate:.0f}%) | "
              f"fetch_fail={stats.get('fetch_fail', 0)} | empty_buys={stats.get('empty_buys', 0)} | "
              f"stale={stats['stale_date']} | http_err={stats['http_error']} | net<=0={stats['net_zero_or_neg']}",
              file=_sys.stderr)
        if success_rate < 50:
            print(f"⚠️ histock 成功率 {success_rate:.0f}% < 50% — 可能被 rate limit / 資料未 update", file=_sys.stderr)

    # Apply column widths
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    row = 1
    # v3.72.8: bug #4 — 若 histock 全 fail 或 <50% → Section 0 頂端加醒目警示 row
    row += _write_histock_status_notice(ws, row, stats, trade_date=trade_date)

    sniper_count = 0
    sniper_with_data = 0
    for master in MASTER_MAPPING:
        master_name = master["name"]
        header_label = master["header_label"]
        branches = master["branches"]
        if not branches:
            continue

        sniper_mode = _is_sniper_master(master_name)
        if sniper_mode:
            sniper_count += 1
            master_has_limit_up = False

        # v3.31.1: 記錄 master block 的 header rows 跟 data rows, 結束後套色
        block_header_rows: List[int] = []
        block_data_rows: List[int] = []

        # Full master header row (A="高手" label)
        _write_header_row(ws, row, header_label, include_master_label=True)
        block_header_rows.append(row)
        row += 1

        master_data_start = row  # first data row of this master block

        for bi, (branch_code, branch_canonical_name) in enumerate(branches):
            if bi > 0:
                # Sub-header before subsequent branches under same master
                _write_header_row(ws, row, header_label, include_master_label=False)
                block_header_rows.append(row)
                row += 1

            # v3.29.5: 該分點是否有 override row 數 (e.g. 8563→20, 其他→10)
            branch_size = _branch_stocks_size(branch_code)

            # Lookup live branch data
            bdata = by_code.get(branch_code, {})
            stocks = _top_stocks_for_branch(bdata, sniper_mode=sniper_mode, n_top=branch_size)
            if sniper_mode and stocks:
                master_has_limit_up = True

            # v3.29.2/v3.29.4: 判斷該分點空白狀態
            has_branch_data = bool(bdata and (bdata.get('buys') or bdata.get('sells')))
            n_stocks = len(stocks)

            branch_first_row = row
            branch_last_row = row + branch_size - 1

            # Write branch_size rows (data + notice + blank padding)
            for ri in range(branch_size):
                r = branch_first_row + ri
                block_data_rows.append(r)   # v3.31.1: 累積 data rows 給套色用
                if ri < n_stocks:
                    # v3.72.3: sniper 且該分點是該股當日買超#1 → 黃色 highlight
                    is_top = (sniper_mode
                              and top_net_buyer.get(stocks[ri].get("code", "")) == branch_code)
                    _write_stock_row(ws, r, stocks[ri], sniper_mode=sniper_mode,
                                     is_top_net_buyer=is_top)
                elif ri == n_stocks and has_branch_data:
                    # 第一個空白 row 加 by-design 提示
                    if n_stocks == 0:
                        _write_empty_branch_notice_row(ws, r, sniper_mode=sniper_mode)
                    else:
                        _write_partial_branch_notice_row(ws, r, n_stocks, sniper_mode=sniper_mode)
                else:
                    _write_blank_data_row(ws, r)

            # B column: branch name (merged across 10 rows)
            cb = ws.cell(row=branch_first_row, column=2)
            cb.value = branch_canonical_name
            cb.font = _font_bold()
            cb.alignment = _align_center()
            ws.merge_cells(
                start_row=branch_first_row, start_column=2,
                end_row=branch_last_row, end_column=2,
            )
            # C column: branch code (merged across 10 rows)
            cc = ws.cell(row=branch_first_row, column=3)
            cc.value = branch_code
            cc.font = _font_bold()
            cc.alignment = _align_center()
            ws.merge_cells(
                start_row=branch_first_row, start_column=3,
                end_row=branch_last_row, end_column=3,
            )

            row = branch_last_row + 1  # advance to next branch

        # A column: master name (merged from first data row to last data row)
        master_data_end = row - 1
        ws.merge_cells(
            start_row=master_data_start, start_column=1,
            end_row=master_data_end, end_column=1,
        )
        ca = ws.cell(row=master_data_start, column=1)
        ca.value = master_name
        ca.font = _font_bold()
        ca.alignment = _align_center()

        if sniper_mode and master_has_limit_up:
            sniper_with_data += 1

        # v3.31.1: 套色 master block (依 master_name 取色, 不在表內用 default)
        block_colors = MASTER_BLOCK_COLORS.get(master_name, DEFAULT_MASTER_COLOR)
        _apply_master_block_color(
            ws,
            header_rows=block_header_rows,
            data_rows=block_data_rows,
            master_anchor_row=master_data_start,
            colors=block_colors,
        )

    # v3.72.8: bug #8 — 尾端 histock 資料時間戳 (informational)
    used = _write_histock_timestamp_footer(ws, row, stats)
    if used:
        row += used

    # Apply uniform row height
    for r in range(1, row):
        ws.row_dimensions[r].height = ROW_HEIGHT

    if sniper_count:
        print(f"  [Excel v3.26] sniper-mode masters: {sniper_with_data}/{sniper_count} "
              f"have limit-up buys today (others get blank rows)")

    # v3.63.0 (E6): freeze A 欄 (master name) + B 欄 (分點 name) → 'C2'
    #   滾右仍能看到 master + 分點對應, header 第 1 row 也固定
    ws.freeze_panes = 'C2'
    return row - 1


# ============================================================
#  v3.31.0: 月檔 (一個月一份, chip_radar_YYYY-MM.xlsx)
# ============================================================

# ════════════════════════════════════════════════════════════════════
# v3.62.0 (Sprint 25 E1-E5): Enrichment sheet builders
# ════════════════════════════════════════════════════════════════════

def _read_json_safely(path: Path) -> Optional[Dict]:
    if not path.exists():
        return None
    try:
        import json
        return json.loads(path.read_text(encoding='utf-8'))
    except Exception:
        return None


def _summary_font_header() -> "Font":
    return Font(name='Noto Sans TC', size=12, bold=True, color='FFFFFFFF')


def _summary_fill(color: str) -> "PatternFill":
    return PatternFill('solid', fgColor=color)


def _wilson_ci(hits: int, n: int, z: float = 1.96) -> tuple:
    """v3.70.2 Wilson binomial confidence interval.

    比 Normal approx 在小樣本 / 極端 p (近 0 或 1) 更穩.
    用 z=1.96 (95% CI), z=2.576 (99% CI).

    Returns: (lo, hi) — both in [0, 1].
    """
    import math
    if n <= 0:
        return (0.0, 0.0)
    p = hits / n
    denom = 1 + z * z / n
    centre = (p + z * z / (2 * n)) / denom
    margin = z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n)) / denom
    return (max(0.0, centre - margin), min(1.0, centre + margin))


def _section_header(ws, row: int, title: str, span_cols: int = 9, color: str = 'FFD4AF37'):
    """共用 section header row."""
    section_font = Font(name='Noto Sans TC', size=12, bold=True, color='FF000000')
    section_fill = _summary_fill(color)
    end_col = chr(ord('B') + span_cols - 1)
    ws.merge_cells(f'B{row}:{end_col}{row}')
    c = ws[f'B{row}']
    c.value = title
    c.font = section_font
    c.fill = section_fill
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 22


def _compute_consensus_count(branches_data):
    """v3.64.3: 共用 helper — 算「強共識股」清單 (≥10 大戶 + ≥2 分點 + 排 ETF + net>0).

    Returns: list of dict {code, name, master_count, branch_count, total_net_amt, masters}
    與 Section 0 使用相同邏輯, 兩處共用避免 drift.
    """
    MIN_MASTER_COUNT = 10
    stock_map = {}
    for b in branches_data:
        m = b.get('master')
        if not m:
            continue
        b_code = b.get('code', '')
        for s in (b.get('buys') or []) + (b.get('sells') or []):
            code = s.get('code')
            if not code or code.startswith('00'):
                continue
            net = (s.get('buy_amt') or 0) - (s.get('sell_amt') or 0)
            if net <= 0:
                continue
            entry = stock_map.setdefault(code, {
                'name': s.get('name', ''),
                'branches': [],
            })
            if not entry['name'] and s.get('name'):
                entry['name'] = s.get('name')
            entry['branches'].append({'master': m, 'branch_code': b_code, 'net_amt': net})
    out = []
    for code, info in stock_map.items():
        if len(info['branches']) < 2:
            continue
        masters = {br['master'] for br in info['branches']}
        if len(masters) < MIN_MASTER_COUNT:
            continue
        out.append({
            'code': code, 'name': info['name'],
            'master_count': len(masters),
            'branch_count': len(info['branches']),
            'masters': masters,
            'total_net_amt': sum(br['net_amt'] for br in info['branches']),
        })
    return out


def _compute_mild_up_picks(consensus_picks, data_dir, trade_date=None):
    """v3.71.0 Phase 3.4: 識別共識 ∩ Q5 偏多 ∩ 近 3 天累積 0-8% (溫和上行) picks.

    ⚠️ v3.71.2 DEPRECATED — alpha overlap audit 揭穿 mild_up_only 是 trap:
       quad_only n=24 hit 79.2% mean +4.19% (強 alpha)
       both     n=13 hit 76.9% mean +4.28% (強 alpha)
       mild_up_only n=12 hit 41.7% mean -0.72% (平均虧錢!)
       → 原 Phase 3.4 backtest 「q5_bull_mild_up 60% n=25」混淆 both 的 quad alpha
       → 純 mild_up_only 沒 vol_spike = end-of-trend trap
       此 helper 不再用於 Section 0 sub-banner / 名稱欄 / Mobile sheet,
       保留供未來研究 (e.g. quad + mild_up overlap 是否比純 quad 更強)

    Phase 3.4 (歷史): hit 60.0% (n=25) vs baseline 44.1% (+15.9pp).

    三條件:
      1. 共識: stock 在 consensus_picks
      2. Q5 偏多: daily_signal.json market_direction.direction == '偏多'
      3. 近 3 天累積 0-8%: today_close / 3d_ago_close - 1 在 [0, 0.08]

    Returns:
      {
        'is_mild_up_day': bool (Q5 偏多 + 有 mild_up 命中),
        'q5_direction': str,
        'mild_up_codes': set,
        'mild_up_picks': list,
      }
    """
    result = {
        'is_mild_up_day': False,
        'q5_direction': None,
        'mild_up_codes': set(),
        'mild_up_picks': [],
    }
    ds = _read_json_safely(data_dir / 'daily_signal.json')
    if ds:
        md = ds.get('market_direction') or {}
        result['q5_direction'] = md.get('direction')

    if result['q5_direction'] != '偏多':
        return result

    sh = _read_json_safely(data_dir / 'stock_history.json')
    if not sh:
        return result
    sh_stocks = sh.get('stocks', {})
    sh_dates = sh.get('dates', [])
    if not sh_dates:
        return result

    # 「今日」基準 — 用 trade_date 或 sh_dates 最後一筆
    today = trade_date if (trade_date and trade_date in sh_dates) else sh_dates[-1]
    try:
        today_idx = sh_dates.index(today)
    except ValueError:
        return result
    if today_idx < 3:
        return result

    for c in consensus_picks:
        code = c.get('code')
        if not code:
            continue
        s_data = sh_stocks.get(code, {}).get('daily', {})
        today_close = (s_data.get(today) or {}).get('close')
        d3_close = (s_data.get(sh_dates[today_idx - 3]) or {}).get('close')
        if today_close is None or d3_close is None or d3_close <= 0:
            continue
        chg_3d = (today_close / d3_close - 1) * 100
        if 0 <= chg_3d <= 8.0:
            result['mild_up_codes'].add(code)
            result['mild_up_picks'].append(c)

    result['is_mild_up_day'] = bool(result['mild_up_picks'])
    return result


def _compute_sector_distribution(picks, data_dir, top_n=3):
    """v3.71.15 N2: 對 picks 統計 industry 分佈, 返 top N 族群.

    Args:
      picks: list of dict (含 'code' 或 直接是 code)
      data_dir: Path
      top_n: 取前 N 大族群

    Returns: list of (industry_name, count, pct), sorted by count desc
    """
    sh = _read_json_safely(data_dir / 'stock_history.json')
    if not sh: return []
    sh_stocks = sh.get('stocks', {})
    counts = {}
    for p in picks:
        code = p.get('code') if isinstance(p, dict) else p
        if not code: continue
        ind = (sh_stocks.get(code, {}) or {}).get('industry') or '其他'
        counts[ind] = counts.get(ind, 0) + 1
    total = sum(counts.values())
    if total == 0: return []
    sorted_ind = sorted(counts.items(), key=lambda kv: -kv[1])
    return [(ind, n, n / total * 100) for ind, n in sorted_ind[:top_n]]


def _get_recent_quad_codes(data_dir, days=7, today=None):
    """v3.71.11 C7: 過去 N 天 trigger 的 quad picks codes set (跨日 dedup 用).

    用途: 若 today picks 包含過去 7 天已 trigger 的 codes,
    user 可能已跟單,Excel 標 🔁 重複提示。

    Returns: set of stock codes
    """
    qhl = _read_json_safely(data_dir / 'quad_hit_log.json')
    if not qhl: return set()
    if today is None:
        from datetime import datetime
        today = datetime.now().strftime('%Y%m%d')
    recent_codes = set()
    for td in qhl.get('trigger_days', []):
        date = td.get('date', '')
        if not date or len(date) != 8: continue
        # date 在過去 N 天內 (不含今日)
        try:
            from datetime import datetime, timedelta
            d = datetime.strptime(date, '%Y%m%d')
            t = datetime.strptime(today, '%Y%m%d')
            delta = (t - d).days
            if 0 < delta <= days:
                for p in (td.get('quad_picks') or []):
                    c = p.get('code')
                    if c: recent_codes.add(c)
        except ValueError:
            continue
    return recent_codes


def _compute_quad_picks(consensus_picks, data_dir):
    """v3.70.0 Phase 3.2 落地: 識別今日符合三訊號疊加的 quad picks.

    三訊號定義 (與 phase32_backtest.py 完全一致):
      1. 共識: stock 在 consensus_picks (>=10 大戶 + >=2 分點)
      2. Q5 偏多: daily_signal.json market_direction.direction == '偏多'
      3. master 量爆: stock 至少 1 位 contributing master 在
                     daily_trading_signals.anomalies (type=volume_spike)

    Returns:
      {
        'is_quad_day': bool (Q5 偏多 + ≥1 vol_spike master 存在),
        'q5_direction': str,
        'vol_spike_masters': set,
        'quad_codes': set of stock codes,
        'quad_picks': list of pick dict (subset of consensus_picks),
      }
    """
    result = {
        'is_quad_day': False,
        'q5_direction': None,
        'vol_spike_masters': set(),
        'premium_vol_spike_masters': set(),   # v3.71.5: subset of vol_spike, premium tier
        'quad_codes': set(),
        'premium_codes': set(),                # v3.71.5: quad picks 有 premium master 配對
        'quad_picks': [],
    }
    # 1. Q5 direction
    ds = _read_json_safely(data_dir / 'daily_signal.json')
    if ds:
        md = ds.get('market_direction') or {}
        result['q5_direction'] = md.get('direction')
    # 2. vol_spike masters (from daily_trading_signals)
    dts = _read_json_safely(data_dir / 'daily_trading_signals.json')
    if dts:
        for a in (dts.get('anomalies') or []):
            if a.get('type') == 'volume_spike' and _is_tracked_master(a.get('master')):
                result['vol_spike_masters'].add(a.get('master'))
    # v3.71.5: 抽出 premium tier
    result['premium_vol_spike_masters'] = result['vol_spike_masters'] & PREMIUM_MASTERS
    # 3. 是否「quad day」 (Q5 偏多 + 有 vol_spike master)
    is_q5_bull = (result['q5_direction'] == '偏多')
    has_vol_spike = bool(result['vol_spike_masters'])
    result['is_quad_day'] = is_q5_bull and has_vol_spike
    # 4. 識別 quad picks (要 quad day + master 交集)
    if result['is_quad_day']:
        for c in consensus_picks:
            masters = c.get('masters') or set()
            if masters & result['vol_spike_masters']:
                result['quad_codes'].add(c['code'])
                # v3.71.5: 若交集含 premium master → 標 premium
                if masters & result['premium_vol_spike_masters']:
                    result['premium_codes'].add(c['code'])
                result['quad_picks'].append(c)
    return result


def _build_section_consensus(ws, branches_data, data_dir, start_row, trade_date=None):
    """v3.63.2 ★ Section 0: 今日追蹤大戶共同買超 (置於 Dashboard 最前).

    定義: ≥2 個追蹤清單內「分點」淨買 > 0 同一檔股票即視為共識.
    (同一大戶不同分點也算 — 例如 民哥 同時用 台新-五權西 + 富邦-南屯 買 2330)

    精準度做法: 不讀預算 signals.consensus (含非追蹤大戶), 而是從 branches_data
              (已過濾為追蹤範圍) 獨立計算 net_amt > 0 by branch by stock,
              100% 對齊使用者實際追蹤的分點.

    排序: 涉及大戶數 desc → 同買分點數 desc → 合計淨買 desc.
    v3.63.6 篩選: 涉及大戶數 ≥ 10 (使用者調強訊號 — 13 位追蹤大戶中 ≥77% 共識才入榜).
    取 Top 30 (soft cap).
    """
    MIN_MASTER_COUNT = 10   # v3.63.6: 強共識門檻 — ≥10 位追蹤大戶同買才入榜
    MAX_PICKS = 30          # 軟上限
    TOP_MASTERS_SHOWN = 5   # 大戶清單只顯示前 5 大金額, 其餘以「+N 位」概括
    hdr_font = Font(name='Noto Sans TC', size=10, bold=True)
    sub_font = Font(name='Noto Sans TC', size=11, bold=True)
    hdr_fill = _summary_fill('FFFEE2E2')   # 淡紅 = high attention
    rank_fill_top = _summary_fill('FFFEF3C7')  # 金 = top 3

    row = start_row
    _section_header(ws, row,
                     f"★ 0. 今日強共識買超 (≥{MIN_MASTER_COUNT} 位追蹤大戶共同淨買, 個股 only — 必看)",
                     color='FFDC2626'); row += 1

    # v3.66.9 Phase 2.5: 強共識股隔日 backtest sub-banner
    bt = _read_json_safely(data_dir / 'consensus_backtest.json')
    if bt and bt.get('summary_30d'):
        s = bt['summary_30d']
        total = s.get('total', 0)
        hits = s.get('hits', 0)
        hit_rate = s.get('hit_rate', 0) * 100
        median = s.get('median_change', 0)
        mean = s.get('mean_change', 0)
        if hit_rate >= 55 and mean > 0.3:
            bt_color = 'FF059669'; bt_icon = '✅'; bt_verdict = '有 alpha'
        elif hit_rate >= 45 and mean > 0:
            bt_color = 'FF666666'; bt_icon = '🟡'; bt_verdict = '中性'
        else:
            bt_color = 'FFDC2626'; bt_icon = '⚠️'; bt_verdict = '無顯著 alpha'
        bt_text = (f"{bt_icon} baseline 30 天 backtest: "
                   f"漲 {hits}/{total} ({hit_rate:.0f}%) | "
                   f"平均 {mean:+.2f}% | 判定: {bt_verdict}")
        bt_cell = ws.cell(row, 2, bt_text)
        ws.merge_cells(f'B{row}:N{row}')
        bt_cell.font = Font(name='Noto Sans TC', size=10, italic=True, color=bt_color)
        bt_cell.fill = _summary_fill('FFF9FAFB')
        bt_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 18
        row += 1

    # v3.69.0 Phase 3.2: 三訊號疊加 alpha sub-banner
    # 共識 ∩ Q5 偏多 ∩ ≥1 master volume_spike = 78.9% hit (vs 44.1% baseline)
    # v3.70.2: +Wilson 95% CI + alpha 失效 alarm
    pb = _read_json_safely(data_dir / 'phase32_backtest.json')
    if pb and pb.get('summary'):
        pb_summary = pb['summary']
        baseline = pb_summary.get('baseline', {})
        triple = pb_summary.get('e_vol_spike_q5_bull', {})
        bl_hr = baseline.get('hit_rate', 0) * 100
        tr_hr = triple.get('hit_rate', 0) * 100
        tr_n = triple.get('n', 0)
        tr_hits = triple.get('hits', 0)
        tr_mean = triple.get('mean_change', 0)
        improvement = tr_hr - bl_hr
        # v3.70.2 P1-G: Wilson 95% CI (binomial 真實精度區間)
        ci_str = ''
        if tr_n >= 5:
            ci_lo, ci_hi = _wilson_ci(tr_hits, tr_n, z=1.96)
            ci_str = f" [{ci_lo*100:.1f}–{ci_hi*100:.1f}% 95% CI]"
        # v3.70.0 落地: 接 quad_hit_log.json 顯示實戰 hit rate
        qhl = _read_json_safely(data_dir / 'quad_hit_log.json')
        live_str = ''
        # v3.70.2 P1-F: alpha 失效 alarm 偵測
        decay_alarm = False
        if qhl and qhl.get('rolling_30d'):
            r30 = qhl['rolling_30d']
            r30_n = r30.get('n', 0)
            r30_hits = r30.get('hits', 0)
            if r30_n > 0:
                r30_hr = r30['hit_rate'] * 100
                live_str = f" | 30d 實戰: {r30_hits}/{r30_n} = {r30_hr:.1f}%"
                # 失效檢測: 30d hit <50% AND n>=20 → 警示
                if r30_n >= 20 and r30_hr < 50:
                    decay_alarm = True
        # verdict (v3.70.2: decay 優先)
        if decay_alarm:
            cc_color = 'FFDC2626'; cc_icon = '⚠️'
            verdict = 'alpha 可能失效 (30d 實戰 <50%) — 建議暫停使用直至改善'
        elif tr_n >= 30 and tr_hr >= 70:
            cc_color = 'FF059669'; cc_icon = '⭐'
            verdict = '強 alpha (p<0.001)'
        elif tr_n >= 30 and tr_hr >= 60:
            cc_color = 'FF059669'; cc_icon = '⭐'
            verdict = '強 alpha 訊號'
        elif tr_n >= 20 and tr_hr >= 60:
            cc_color = 'FF666666'; cc_icon = '🟡'
            verdict = 'alpha 訊號 (樣本待累積)'
        elif tr_n >= 10 and tr_hr >= 50:
            cc_color = 'FF666666'; cc_icon = '🟡'
            verdict = '弱 alpha'
        else:
            cc_color = 'FFDC2626'; cc_icon = '⚠️'
            verdict = '樣本不足'
        cc_text = (f"{cc_icon} Phase 3.2 真 alpha (三訊號): 共識 ∩ Q5 偏多 ∩ master 量爆 "
                   f"hit {tr_hr:.1f}%{ci_str} (n={tr_n}, mean {tr_mean:+.2f}%){live_str} "
                   f"vs baseline {bl_hr:.1f}% = {improvement:+.1f}pp — {verdict}")
        cc_cell = ws.cell(row, 2, cc_text)
        ws.merge_cells(f'B{row}:N{row}')
        cc_cell.font = Font(name='Noto Sans TC', size=10, italic=True, color=cc_color)
        cc_cell.fill = _summary_fill('FFF9FAFB')
        cc_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 18
        row += 1

    # v3.71.7: 跟單實際淨報酬 sub-banner (扣台股交易成本)
    # 揭穿 alpha 是真實淨利還是被成本吃光 — 從 quad_hit_log 算淨報酬
    # 成本: 0.585% (手續費 0.1425% × 2 + 證交稅 0.3%, 無折扣保守估)
    qhl_for_roi = _read_json_safely(data_dir / 'quad_hit_log.json')
    if qhl_for_roi and qhl_for_roi.get('trigger_days'):
        TX_COST_PCT = 0.585    # 保守 (無折扣 + 一般證交稅)
        all_picks_for_roi = []
        for td in qhl_for_roi['trigger_days']:
            for p in td.get('quad_picks', []):
                if p.get('next_change_pct') is not None:
                    all_picks_for_roi.append(p['next_change_pct'])
        n_roi = len(all_picks_for_roi)
        if n_roi >= 5:
            net_chgs = [c - TX_COST_PCT for c in all_picks_for_roi]
            net_hits = sum(1 for x in net_chgs if x > 0)
            net_mean = sum(net_chgs) / n_roi
            sorted_nc = sorted(net_chgs)
            net_median = sorted_nc[n_roi // 2] if n_roi % 2 else (sorted_nc[n_roi//2-1] + sorted_nc[n_roi//2]) / 2
            cum = sum(net_chgs)
            if net_mean >= 1.0:
                roi_color = 'FF059669'; roi_icon = '💰'
                roi_verdict = '淨利 alpha 確認'
            elif net_mean >= 0:
                roi_color = 'FF666666'; roi_icon = '🟡'
                roi_verdict = '勉強損益兩平'
            else:
                roi_color = 'FFDC2626'; roi_icon = '⚠️'
                roi_verdict = 'alpha 被成本吃光'
            roi_text = (f"{roi_icon} 跟單實際淨報酬 (扣 {TX_COST_PCT}% 成本): "
                        f"淨 hit {net_hits}/{n_roi} = {net_hits/n_roi*100:.0f}% | "
                        f"平均 {net_mean:+.2f}% | 中位 {net_median:+.2f}% | "
                        f"累積 {cum:+.0f}% — {roi_verdict}")
            roi_cell = ws.cell(row, 2, roi_text)
            ws.merge_cells(f'B{row}:N{row}')
            roi_cell.font = Font(name='Noto Sans TC', size=10, italic=True, color=roi_color)
            roi_cell.fill = _summary_fill('FFF9FAFB')
            roi_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row].height = 18
            row += 1

    # v3.71.8 Phase 3.5: 多日 alpha (擇高出場 / 持有 3 天) sub-banner
    # Iteration 1 揭穿: peak_5d hit 86.8% +12.78% / premium cum_3d 92.9% +12.16%
    # 樣本 n=38 觀察期, 樣本 ≥60 後 Iteration 2 re-audit
    mb = _read_json_safely(data_dir / 'multiday_backtest.json')
    if mb and mb.get('combos'):
        n_total = mb.get('n_total', 0)
        all_quad = (mb['combos'].get('all_quad') or {}).get('all') or {}
        premium = (mb['combos'].get('premium_only') or {}).get('all') or {}
        peak_q = all_quad.get('peak_5d')
        cum3_q = all_quad.get('cum_3d')
        peak_p = premium.get('peak_5d')
        cum3_p = premium.get('cum_3d')
        if peak_q and cum3_q and n_total >= 20:
            md_color = 'FF7C3AED' if n_total < 60 else 'FF059669'   # 紫 (觀察期) / 綠 (正式)
            md_icon = '🚀' if n_total >= 60 else '🔬'
            md_tag = '正式' if n_total >= 60 else f'觀察期 n={n_total}'
            # v3.71.14 fix: prefix 不入 join, 避免 ':  |  ' 醜空格
            parts = [
                f"5 日內擇高 {peak_q['hit_rate']*100:.1f}% / {peak_q['mean']:+.2f}%",
            ]
            if peak_p and peak_p.get('n', 0) >= 10:
                parts.append(
                    f"premium 擇高 {peak_p['hit_rate']*100:.1f}% / {peak_p['mean']:+.2f}% (n={peak_p['n']})"
                )
            if cum3_p and cum3_p.get('n', 0) >= 10:
                parts.append(
                    f"premium 持有 3 天 {cum3_p['hit_rate']*100:.1f}% / {cum3_p['mean']:+.2f}%"
                )
            md_text = (f"{md_icon} Phase 3.5 多日 alpha ({md_tag}): "
                       + '  |  '.join(parts)
                       + '  — 來源: quad_hit_log × stock_history t+1~t+5')
            md_cell = ws.cell(row, 2, md_text)
            ws.merge_cells(f'B{row}:N{row}')
            md_cell.font = Font(name='Noto Sans TC', size=10, italic=True, color=md_color)
            md_cell.fill = _summary_fill('FFF9FAFB')
            md_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row].height = 18
            row += 1

    # v3.71.2 Phase 3.4 ROLLBACK: alpha overlap audit 揭穿 mild_up_only 是 trap
    # quad_only n=24 hit 79.2% mean +4.19% (強 alpha)
    # both     n=13 hit 76.9% mean +4.28% (強 alpha)
    # mild_up_only n=12 hit 41.7% mean -0.72% (TRAP! 平均虧錢)
    # → 原 v3.71.0 backtest 「q5_bull_mild_up 60% n=25」混淆了 both 的 quad alpha
    # → 純 mild_up_only 沒 vol_spike = end-of-trend trap, 不該推薦
    # 砍掉 sub-banner + ★ 標記 + Mobile section. helper 保留但內部不再使用.

    # v3.70.0 Phase 3.2 落地: 今日 quad 狀態 banner
    pre_consensus = _compute_consensus_count(branches_data)
    quad_info = _compute_quad_picks(pre_consensus, data_dir)
    # v3.71.11 C7: 過去 7 天 trigger codes (跨日 dedup, 提示 user 可能已跟單)
    recent_quad_codes = _get_recent_quad_codes(data_dir, days=7, today=trade_date)
    if quad_info['is_quad_day'] and quad_info['quad_picks']:
        # v3.71.5: premium count + names 優先列前面
        premium_picks = [p for p in quad_info['quad_picks']
                          if p['code'] in quad_info.get('premium_codes', set())]
        std_picks = [p for p in quad_info['quad_picks']
                      if p['code'] not in quad_info.get('premium_codes', set())]
        n_prem = len(premium_picks)
        n_std = len(std_picks)
        # 顯示 premium 在前, 一般在後
        ordered = premium_picks + std_picks
        quad_names = [(p['code'], p['name'], p['code'] in quad_info.get('premium_codes', set()))
                       for p in ordered[:5]]
        names_str = ', '.join(
            [f"{'⭐⭐' if prem else ''}{n}({c})" for c, n, prem in quad_names]
        )
        if len(quad_info['quad_picks']) > 5:
            names_str += f' +{len(quad_info["quad_picks"])-5} 檔'
        tier_str = f"⭐⭐ {n_prem} premium + ⭐ {n_std} 一般" if n_prem else f"⭐ {n_std} 一般 quad"
        q_text = (f"🎯 今日 quad 命中 {len(quad_info['quad_picks'])} 檔 ({tier_str}): "
                  f"{names_str}  |  Q5 偏多 + {len(quad_info['vol_spike_masters'])} 位 master 量爆")
        q_color = 'FF059669'    # 綠
        q_fill = 'FFD1FAE5'      # 淡綠
    elif quad_info['q5_direction'] == '偏多' and not quad_info['vol_spike_masters']:
        q_text = (f"💤 今日 Q5 偏多但無 master 量爆 — quad 三訊號未齊聚 (一般共識可參考但無 alpha 加持)")
        q_color = 'FF666666'; q_fill = 'FFF3F4F6'
    elif quad_info['vol_spike_masters'] and quad_info['q5_direction'] != '偏多':
        q_text = (f"💤 今日有 {len(quad_info['vol_spike_masters'])} 位 master 量爆但 Q5 {quad_info['q5_direction'] or '無'} — quad 未啟動")
        q_color = 'FF666666'; q_fill = 'FFF3F4F6'
    else:
        q_text = f"💤 今日無 quad 訊號 (Q5={quad_info['q5_direction'] or '無'}, 無 master 量爆)"
        q_color = 'FF888888'; q_fill = 'FFF9FAFB'
    q_cell = ws.cell(row, 2, q_text)
    ws.merge_cells(f'B{row}:N{row}')
    q_cell.font = Font(name='Noto Sans TC', size=10, bold=True, color=q_color)
    q_cell.fill = _summary_fill(q_fill)
    q_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 18
    row += 1

    # v3.71.15 N2: sector rotation sub-banner (在註腳前)
    # 統計 today 共識股 industry 分佈, 看主力買哪些族群
    if pre_consensus:
        sector_dist = _compute_sector_distribution(pre_consensus, data_dir, top_n=3)
        if sector_dist:
            parts = [f"{ind} {n} ({pct:.0f}%)" for ind, n, pct in sector_dist]
            sec_text = f"📊 今日共識集中產業 (top {len(parts)}): " + '  |  '.join(parts)
            sec_cell = ws.cell(row, 2, sec_text)
            ws.merge_cells(f'B{row}:N{row}')
            sec_cell.font = Font(name='Noto Sans TC', size=10, italic=True, color='FF6366F1')
            sec_cell.fill = _summary_fill('FFEEF2FF')   # 極淡靛
            sec_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row].height = 18
            row += 1

    # v3.71.18 L1: pinned master stats sub-banner (大牌歷史 hit rate)
    pms = _read_json_safely(data_dir / 'pinned_master_stats.json')
    if pms and pms.get('pinned_masters'):
        for m_name, m_stats in pms['pinned_masters'].items():
            if m_stats.get('status') != 'ok': continue
            ap = m_stats.get('all_picks') or {}
            new_s = m_stats.get('new_stocks') or {}
            acc_s = m_stats.get('accumulation') or {}
            parts = []
            if ap.get('n', 0) >= 5:
                parts.append(f"全 picks n={ap['n']} hit_1d={ap.get('hit_1d',0)*100:.0f}% / hit_3d={ap.get('hit_3d',0)*100:.0f}%")
            if new_s.get('n', 0) >= 3:
                parts.append(f"新標 n={new_s['n']} hit_3d={new_s.get('hit_3d',0)*100:.0f}% mean={new_s.get('mean_3d',0):+.2f}%")
            if acc_s.get('n', 0) >= 2:
                parts.append(f"連加 n={acc_s['n']} hit_5d={acc_s.get('hit_5d',0)*100:.0f}% mean={acc_s.get('mean_5d',0):+.2f}%")
            if parts:
                pm_text = f"📌 {m_name} 歷史 alpha: " + '  |  '.join(parts)
                pm_cell = ws.cell(row, 2, pm_text)
                ws.merge_cells(f'B{row}:N{row}')
                pm_cell.font = Font(name='Noto Sans TC', size=10, italic=True, color='FFB45309')
                pm_cell.fill = _summary_fill('FFFEF3C7')   # 淡金
                pm_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                ws.row_dimensions[row].height = 18
                row += 1

    # v3.71.18 註腳 (加 📌 pinned 標記)
    pinned_str = ' / '.join(sorted(PINNED_MASTERS)) if PINNED_MASTERS else '無'
    note_cell = ws.cell(row, 2,
                         f"ⓘ 排序: 合計淨買金額 ↓  |  ⭐⭐ = premium quad (陳律師/竹科主力/陳族元, ≥77% hit)  |  ⭐ = 一般 quad (78.9%)  |  ⚠️ = 領頭獨佔 ≥50%  |  🔁 = 過去 7 天 quad 重複  |  📌 = 你關注的 master ({pinned_str}) 參與")
    ws.merge_cells(f'B{row}:N{row}')
    note_cell.font = Font(name='Noto Sans TC', size=10, italic=True, color='FF7C2D12')
    note_cell.fill = _summary_fill('FFFFF7ED')   # 極淡橙底
    note_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 18
    row += 1

    # Build code -> {name, branches: [{branch_code, branch_name, master, net_amt}]}
    stock_map: Dict[str, Dict] = {}
    for b in branches_data:
        m = b.get('master')
        if not m:
            continue
        b_code = b.get('code', '')
        b_name = b.get('name', '')
        for s in (b.get('buys') or []) + (b.get('sells') or []):
            code = s.get('code')
            if not code:
                continue
            # v3.63.7: 排除 ETF — code 以 '00' 開頭 (00712 / 00632R / 0050 / 0056 ...)
            # 使用者目標: Section 0 只看「大戶共識買的個股」, 不含 ETF
            # 已驗證 stock_categories.json: 沒有任何 listed/otc/emerging 股票 code 起始 '00'
            if code.startswith('00'):
                continue
            buy_amt = s.get('buy_amt') or 0
            sell_amt = s.get('sell_amt') or 0
            net_amt = buy_amt - sell_amt
            if net_amt <= 0:
                continue   # 必須該分點該股淨買超
            entry = stock_map.setdefault(code, {
                'name': s.get('name', ''),
                'branches': [],   # 每個分點一筆 (允許同 master 多分點)
            })
            if not entry['name'] and s.get('name'):
                entry['name'] = s.get('name')
            entry['branches'].append({
                'branch_code': b_code,
                'branch_name': b_name,
                'master': m,
                'net_amt': net_amt,
            })

    # 過濾: ≥2 個分點 AND ≥MIN_MASTER_COUNT 位大戶 (v3.63.6)
    consensus_list = []
    for code, info in stock_map.items():
        if len(info['branches']) < 2:
            continue
        master_set = {br['master'] for br in info['branches']}
        if len(master_set) < MIN_MASTER_COUNT:
            continue
        total_net = sum(br['net_amt'] for br in info['branches'])
        consensus_list.append({
            'code': code,
            'name': info['name'],
            'branches': info['branches'],
            'branch_count': len(info['branches']),
            'master_count': len(master_set),
            'masters': master_set,
            'total_net_amt': total_net,
        })
    # v3.63.9 (用戶決定): 改主排序為 total_net_amt DESC (資金規模 = 真實共識深度)
    # tie-breaker: master_count DESC → branch_count DESC
    # 已有 master_count >= 10 硬門檻保證廣度, 排序時用金額反映深度.
    # 案例: 群聯 7.4 億 (10 master) 在舊版 rank 9, 新版 rank 2; 彩晶 7591 萬 在舊版 rank 1, 新版 rank 8
    consensus_list.sort(key=lambda x: (-x['total_net_amt'], -x['master_count'],
                                        -x['branch_count']))

    # v3.63.8 Excel-native 13 欄佈局: 一格一值, 金額為 int + 千分位 number_format
    # B=#, C=代號, D=名稱, E=大戶數, F=分點數, G=領頭大戶, H=領頭金額(萬),
    # I=#2 大戶, J=#2 金額, K=#3 大戶, L=#3 金額, M=+更多, N=合計淨買(萬)
    headers = [
        ('B', '#', 5),    ('C', '代號', 40),   ('D', '名稱', 18),
        ('E', '大戶數', 9), ('F', '分點數', 9),
        ('G', '領頭大戶', 14),  ('H', '領頭金額(萬)', 13),
        ('I', '#2 大戶', 14),   ('J', '#2 金額(萬)', 13),
        ('K', '#3 大戶', 14),   ('L', '#3 金額(萬)', 13),
        ('M', '+更多', 8),
        ('N', '合計淨買(萬)', 14),
    ]
    header_row = row
    for col_l, h, w in headers:
        cell = ws[f'{col_l}{row}']
        cell.value = h
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[col_l].width = w
    row += 1

    # ── Data rows ──
    num_fmt = '#,##0'
    muted_font = Font(name='Noto Sans TC', size=10, italic=True, color='FF888888')
    leader_font = Font(name='Noto Sans TC', size=11, bold=True)
    total_font = Font(name='Noto Sans TC', size=11, bold=True)

    start_data = row
    for i, item in enumerate(consensus_list[:MAX_PICKS]):
        c_b = ws.cell(row, 2, i + 1)
        if i < 3:
            c_b.fill = rank_fill_top
            c_b.font = sub_font
        ws.cell(row, 3, item['code'])

        # 排序大戶 by 該大戶在此股總 net_amt desc
        master_total = {}
        for br in item['branches']:
            master_total[br['master']] = master_total.get(br['master'], 0) + br['net_amt']
        masters_sorted = sorted(master_total.items(), key=lambda kv: -kv[1])

        # v3.63.9 用戶決定: 領頭大戶佔合計 ≥50% → 加 ⚠️ 警示 (假共識防呆)
        # 名義 ≥10 大戶共識 但 1 大戶獨大 → 真共識訊號被稀釋
        leader_amt = masters_sorted[0][1] if masters_sorted else 0
        total_net = item['total_net_amt']
        leader_pct = leader_amt / total_net if total_net > 0 else 0
        is_outlier = leader_pct >= 0.5
        # v3.70.0 Phase 3.2 落地: ⭐ 標記 quad 命中股 (在 ⚠️ 之前, 因 alpha > 警示)
        # v3.71.2 Phase 3.4 ROLLBACK: ★ mild_up 標記砍掉 (audit 揭穿 trap)
        # v3.71.5 Phase 3.2 premium tier: ⭐⭐ for premium master (≥77% hit) 配對
        # v3.71.11 C7: 🔁 for 過去 7 天 trigger 重複 (user 可能已跟單)
        # v3.71.18 L4: 📌 for pinned master 參與 (user 自定常駐關注)
        is_quad = item['code'] in quad_info['quad_codes']
        is_premium = item['code'] in quad_info.get('premium_codes', set())
        is_repeat = item['code'] in recent_quad_codes
        # pinned: 該股 buyers 含 PINNED_MASTERS
        master_set = set(b.get('master') for b in item.get('branches', []) if b.get('master'))
        is_pinned = bool(master_set & PINNED_MASTERS)
        display_name = item['name'] or '—'
        if is_premium and is_outlier:
            display_name = f"⭐⭐⚠️ {display_name}"
        elif is_premium:
            display_name = f"⭐⭐ {display_name}"
        elif is_quad and is_outlier:
            display_name = f"⭐⚠️ {display_name}"
        elif is_quad:
            display_name = f"⭐ {display_name}"
        elif is_outlier:
            display_name = f"⚠️ {display_name}"
        if is_repeat:
            display_name = f"🔁 {display_name}"
        if is_pinned:
            display_name = f"📌 {display_name}"
        c_name = ws.cell(row, 4, display_name)
        if is_quad and not is_outlier:
            # quad 但非 outlier → 名稱 cell 淡金底 + 綠字 (alpha 啟動視覺)
            c_name.fill = _summary_fill('FFFEF3C7')   # 淡金
            c_name.font = Font(name='Noto Sans TC', size=11, bold=True, color='FF059669')
        if is_outlier:
            # 領頭佔比高 → 名稱 cell 淡橙底警示 (即使 quad, outlier 警示仍生效)
            c_name.fill = _summary_fill('FFFED7AA')   # 淡橙
            c_name.font = Font(name='Noto Sans TC', size=11, bold=True, color='FF7C2D12')
            # v3.63.9: hover comment 顯示詳細%
            try:
                from openpyxl.comments import Comment
                leader_name = masters_sorted[0][0] if masters_sorted else '?'
                comment_text = (
                    f"⚠️ 假共識警示\n"
                    f"領頭大戶「{leader_name}」獨佔比 {leader_pct*100:.1f}%\n"
                    f"  • 領頭金額 {int(leader_amt/10):,} 萬\n"
                    f"  • 合計淨買 {int(total_net/10):,} 萬\n"
                    f"  • 大戶數 {item['master_count']} 位 (名義共識)\n\n"
                    f"判讀: 名義 {item['master_count']} 位大戶共識,\n"
                    f"實質 1 人下注佔 {leader_pct*100:.0f}%,\n"
                    f"剩餘 {item['master_count']-1} 位為陪襯小單,\n"
                    f"真共識訊號被稀釋.")
                c = Comment(comment_text, 'Chip Radar')
                c.width = 280
                c.height = 180
                c_name.comment = c
            except Exception:
                pass

        c_e = ws.cell(row, 5, item['master_count'])
        c_e.font = sub_font
        c_e.alignment = Alignment(horizontal='center')
        c_f = ws.cell(row, 6, item['branch_count'])
        c_f.alignment = Alignment(horizontal='center')

        # 短名稱去括號內容讓欄位緊湊 (例: "張濬安(航海王)" → "張濬安")
        def _short(m):
            return m.split('(')[0].split('/')[0]

        # G/H = 領頭 (Top 1)
        # I/J = #2
        # K/L = #3
        slot_cols = [('G', 'H'), ('I', 'J'), ('K', 'L')]
        for slot_idx, (name_col, amt_col) in enumerate(slot_cols):
            if slot_idx < len(masters_sorted):
                m_name, m_amt = masters_sorted[slot_idx]
                c_name = ws[f'{name_col}{row}']
                c_name.value = _short(m_name)
                c_amt = ws[f'{amt_col}{row}']
                c_amt.value = int(round(m_amt / 10))   # int → 可 sort/sum
                c_amt.number_format = num_fmt
                c_amt.alignment = Alignment(horizontal='right')
                if slot_idx == 0:
                    c_name.font = leader_font
                    c_amt.font = leader_font
            else:
                ws[f'{name_col}{row}'].value = ''
                ws[f'{amt_col}{row}'].value = None

        # M = +更多 (剩餘大戶數, 用 int 而非字串, 0 顯示為空)
        tail = max(0, len(masters_sorted) - 3)
        c_m = ws.cell(row, 13, tail if tail > 0 else None)
        c_m.font = muted_font
        c_m.alignment = Alignment(horizontal='center')

        # N = 合計淨買(萬), 粗體 + 千分位
        c_n = ws.cell(row, 14, int(round(item['total_net_amt'] / 10)))
        c_n.number_format = num_fmt
        c_n.font = total_font
        c_n.alignment = Alignment(horizontal='right')

        row += 1

    if row == start_data:
        ws.cell(row, 2, f'⚪ 今日無 ≥{MIN_MASTER_COUNT} 位追蹤大戶共同淨買的個股')
        ws.merge_cells(f'B{row}:N{row}')
        row += 1
    else:
        # v3.66.6 Phase 2.2: N 合計淨買加 data bar (深綠 = 越多越強)
        _try_add_data_bar(ws, f'N{start_data}:N{row-1}', 'FF66BB6A')
        # E 大戶數也加 (淡金, 最大 13 看跨度)
        _try_add_data_bar(ws, f'E{start_data}:E{row-1}', 'FFFFC107')
        # v3.63.8: 把資料區包成 Excel Table → 原生 sort/filter 下拉
        try:
            from openpyxl.worksheet.table import Table, TableStyleInfo
            table_range = f'B{header_row}:N{row - 1}'
            tbl = Table(displayName=f"ConsensusTbl_{header_row}", ref=table_range)
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleLight1", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False,
            )
            ws.add_table(tbl)
        except Exception:
            pass   # Table 失敗不影響資料正確性
    row += 1   # 空一行
    return row


def build_mobile_summary_sheet(ws, branches_data, trade_date, data_dir=None):
    """v3.67.1 Phase 2.7: 手機摘要 sheet.

    設計原則 (使用者明確要求):
      1. 精簡 — 每行 1 個資訊, 無複合句
      2. 可讀性高 — 3 級字體階層 (14/12/10pt)
      3. 視覺不雜亂 — 單欄, section 空 1 行, 無格線

    內容 (4 個決策問題):
      📅 明日預測 (Q5 direction)
      🎯 強共識 Top 5 (買什麼)
      🚫 今日避開 (除權息)
      📊 追蹤池方向 (淨買差 + vs 昨/5d)
    """
    data_dir = data_dir or Path('data')
    branches_data = _filter_tracked_branches(branches_data)

    # 單欄佈局: B=2 留白, C=38 主內容, D=2 留白
    ws.column_dimensions['B'].width = 2
    ws.column_dimensions['C'].width = 38
    ws.column_dimensions['D'].width = 2

    title_font = Font(name='Noto Sans TC', size=14, bold=True,
                      color=COLORS['brand_dark'])
    sec_font = Font(name='Noto Sans TC', size=13, bold=True,
                    color=COLORS['text_strong'])
    val_font_big = Font(name='Noto Sans TC', size=16, bold=True)
    val_font = Font(name='Noto Sans TC', size=12)
    sub_font = Font(name='Noto Sans TC', size=10, italic=True,
                    color=COLORS['text_muted'])

    row = 2
    # ── 主標題 ──
    c = ws.cell(row, 3, f"📋 Chip Radar · "
                f"{trade_date[:4]}/{trade_date[4:6]}/{trade_date[6:8]}")
    c.font = title_font
    c.alignment = Alignment(horizontal='left', vertical='center')
    row += 2

    # ── 📅 明日預測 ──
    ws.cell(row, 3, "📅 明日預測").font = sec_font
    row += 1
    daily_signal = _read_json_safely(data_dir / 'daily_signal.json')
    md = (daily_signal or {}).get('market_direction') or {}
    direction = md.get('direction') or '—'
    confidence = md.get('confidence_pct') or 0
    if direction == '偏多':
        arrow, q5_color = '↑', COLORS['tw_red']
    elif direction == '偏空':
        arrow, q5_color = '↓', COLORS['tw_green']
    else:
        arrow, q5_color = '↕', COLORS['text_neutral']
    c_q5 = ws.cell(row, 3, f"{arrow} {direction} {confidence:.1f}%")
    c_q5.font = Font(name='Noto Sans TC', size=16, bold=True, color=q5_color)
    ws.row_dimensions[row].height = 24
    row += 2

    # ── 🎯 強共識買超 Top 5 ──
    consensus = _compute_consensus_count(branches_data)
    consensus.sort(key=lambda x: (-x['total_net_amt'], -x['master_count'],
                                   -x['branch_count']))
    # v3.70.0 Phase 3.2 落地: quad picks 識別
    # v3.71.2 Phase 3.4 ROLLBACK: ★ mild_up section 砍掉 (audit 揭穿 trap)
    mobile_quad = _compute_quad_picks(consensus, data_dir)

    # ── ⭐ Phase 3.2 quad 命中 (alpha 啟動 — 列在共識上方, 最 actionable) ──
    # v3.71.5: premium 在前, 一般在後
    premium_codes = mobile_quad.get('premium_codes', set())
    if mobile_quad['quad_picks']:
        ws.cell(row, 3, "⭐ Quad 命中 (78.9% alpha)").font = Font(
            name='Noto Sans TC', size=13, bold=True, color='FF059669')
        row += 1
        # premium picks 優先列前
        ordered = sorted(mobile_quad['quad_picks'],
                          key=lambda c: 0 if c['code'] in premium_codes else 1)
        for c in ordered[:5]:
            prefix = '⭐⭐' if c['code'] in premium_codes else '🎯'
            cell = ws.cell(row, 3,
                f"{prefix} {c['name']} ({c['code']}) · {c['master_count']} 大戶")
            cell.font = Font(name='Noto Sans TC', size=12, bold=True,
                             color=COLORS.get('alpha_gold', 'FF059669'))
            row += 1
        row += 1

    ws.cell(row, 3, "🎯 強共識買超 Top 5").font = sec_font
    row += 1
    circle = ['①', '②', '③', '④', '⑤']
    for i, c in enumerate(consensus[:5]):
        if c['code'] in premium_codes:
            prefix = '⭐⭐'
        elif c['code'] in mobile_quad['quad_codes']:
            prefix = '⭐'
        else:
            prefix = circle[i]
        ws.cell(row, 3,
                f"{prefix} {c['name']} ({c['code']}) · {c['master_count']} 大戶").font = val_font
        row += 1

    # v3.71.15 N2: 共識集中產業 (Mobile section, 列在共識 Top 5 後)
    if consensus:
        sec_dist = _compute_sector_distribution(consensus, data_dir, top_n=3)
        if sec_dist:
            row += 1
            ws.cell(row, 3, "📊 共識集中產業").font = sec_font
            row += 1
            for ind, n, pct in sec_dist:
                ws.cell(row, 3, f"{ind} · {n} 檔 ({pct:.0f}%)").font = val_font
                row += 1

    # v3.71.18 L3: 📌 pinned master 今日動態 (大牌專區)
    # 列出每個 pinned master 今日 top 3 buys + 共識重疊
    for pinned_master in sorted(PINNED_MASTERS):
        master_buys = []
        for b in branches_data:
            if b.get('master') == pinned_master:
                for s in (b.get('buys') or []):
                    code = s.get('code')
                    if not code or code.startswith('00'): continue
                    master_buys.append({
                        'code': code, 'name': s.get('name', '—'),
                        'amt': s.get('buy_amt') or 0,
                    })
        # 同 master 跨分點同股 dedup + 合計
        agg = {}
        for b in master_buys:
            key = b['code']
            if key in agg:
                agg[key]['amt'] += b['amt']
            else:
                agg[key] = b
        top_buys = sorted(agg.values(), key=lambda x: -x['amt'])[:3]
        if top_buys:
            row += 1
            ws.cell(row, 3, f"📌 {pinned_master} 今日 Top 3").font = sec_font
            row += 1
            consensus_codes = {c['code'] for c in consensus}
            for b in top_buys:
                tag = ' (★共識)' if b['code'] in consensus_codes else ''
                amt_wan = round(b['amt'] / 10)
                ws.cell(row, 3,
                        f"{b['name']} ({b['code']}) · {amt_wan:,} 萬{tag}").font = val_font
                row += 1

    # v3.71.3 用戶要求: Mild_up watch section (反向參考, 非 alpha 推薦)
    # 揭穿: mild_up_only 歷史 hit 41.7% mean -0.72% (n=12) = trap
    # 用戶仍要看 → 顯示為「⚠️ 反向參考」, 不用 ★ (避免誤判為 alpha 訊號)
    mobile_mu = _compute_mild_up_picks(consensus, data_dir, trade_date=trade_date)
    mu_only_codes = mobile_mu['mild_up_codes'] - mobile_quad['quad_codes']
    mu_only_picks = [c for c in mobile_mu['mild_up_picks']
                      if c['code'] in mu_only_codes]
    if mu_only_picks:
        row += 1
        ws.cell(row, 3, "⚠️ Mild_up watch (反向參考)").font = Font(
            name='Noto Sans TC', size=13, bold=True, color='FFB45309')   # 琥珀色
        row += 1
        ws.cell(row, 3, "(歷史 41.7% hit, 平均 -0.72% — 別追)").font = sub_font
        row += 1
        for c in mu_only_picks[:5]:
            ws.cell(row, 3,
                f"⚠️ {c['name']} ({c['code']}) · {c['master_count']} 大戶").font = Font(
                    name='Noto Sans TC', size=12, color='FFB45309')
            row += 1
    if not consensus:
        ws.cell(row, 3, "(今日無強共識)").font = sub_font
        row += 1
    row += 1

    # ── 🚫 今日避開 ──
    ws.cell(row, 3, "🚫 今日避開").font = sec_font
    row += 1
    dividend = _read_json_safely(data_dir / 'dividend_calendar.json')
    today_ex = [i for i in ((dividend or {}).get('upcoming_30d') or [])
                 if i.get('ex_date') == trade_date]
    if today_ex:
        codes_str = ' / '.join(i.get('code', '') for i in today_ex[:3])
        suffix = ' ...' if len(today_ex) > 3 else ''
        ws.cell(row, 3, f"除權息 {len(today_ex)} 檔 ({codes_str}{suffix})").font = val_font
    else:
        ws.cell(row, 3, "今日無除權息").font = sub_font
    row += 1
    # v3.71.7: 處置股 (attstock.tw API)
    disp = _read_json_safely(data_dir / 'disposal_attstock.json')
    if disp:
        n_in = disp.get('count_in_disposal', 0)
        n_pending = disp.get('count_pending_1d', 0)
        if n_in > 0:
            codes = disp.get('codes_in_disposal') or []
            codes_str = ' / '.join(codes[:3]) + (' ...' if len(codes) > 3 else '')
            ws.cell(row, 3, f"處置中 {n_in} 檔 ({codes_str})").font = val_font
            row += 1
        if n_pending > 0:
            codes = disp.get('codes_pending_1d') or []
            codes_str = ' / '.join(codes[:3]) + (' ...' if len(codes) > 3 else '')
            ws.cell(row, 3, f"明日恐處置 {n_pending} 檔 ({codes_str})").font = val_font
            row += 1
        if n_in == 0 and n_pending == 0:
            ws.cell(row, 3, "今日無處置股").font = sub_font
            row += 1
    row += 1

    # ── 📊 追蹤池方向 ──
    ws.cell(row, 3, "📊 追蹤池方向").font = sec_font
    row += 1
    # 計算 Q2 淨買差
    def _bs(blist):
        buy = sum((s.get('buy_amt') or 0) for b in blist for s in (b.get('buys') or []))
        seen = set(); sell = 0
        for b in blist:
            bcode = b.get('code', '')
            for s in (b.get('buys') or []) + (b.get('sells') or []):
                key = (bcode, s.get('code'))
                if key in seen: continue
                seen.add(key)
                sell += (s.get('sell_amt') or 0)
        return buy, sell
    tb, ts_ = _bs(branches_data)
    net_billion = (tb - ts_) / 100000
    # 顏色 + 文字
    net_color = COLORS['tw_red'] if net_billion >= 0 else COLORS['tw_green']
    sign = '+' if net_billion >= 0 else ''
    c_net = ws.cell(row, 3, f"{sign}{net_billion:.0f} 億 淨買")
    c_net.font = Font(name='Noto Sans TC', size=14, bold=True, color=net_color)
    ws.row_dimensions[row].height = 22
    row += 1
    # vs 昨 / 5d (從 timeseries 拿)
    ts_data = _update_load_timeseries(data_dir, trade_date, {
        'q1_active_ratio': 0, 'q2_net_billion': net_billion,
        'q3_consensus_count': 0, 'q3_consensus_net_billion': 0,
        'q4_track_share': 0, 'q4_mkt_net_billion': 0,
    }, update=False)
    y = ts_data['yesterday'].get('q2_net_billion')
    a = ts_data['avg5'].get('q2_net_billion')
    if y is not None and a is not None:
        trend_y = '反彈' if net_billion > y else ('擴空' if net_billion < y else '持平')
        trend_a = '偏弱' if net_billion < a else ('偏強' if net_billion > a else '持平')
        ws.cell(row, 3, f"比昨 {y:+.0f} {trend_y} / 比 5d {a:+.0f} {trend_a}").font = sub_font
    else:
        ws.cell(row, 3, "(歷史資料累積中)").font = sub_font

    # freeze 大標題置頂
    ws.freeze_panes = 'A3'


def build_quad_track_sheet(ws, data_dir):
    """v3.70.2 Phase 3.2 持續性追蹤 — 逐 trigger day inspect 用.

    讀 data/quad_hit_log.json, 列出每個 trigger day:
      日期 | 隔日 | Q5 | vol_spike masters | picks | hits | 命中率 | mean%

    用戶逐筆檢視可建立對 alpha 的 trust + 學習失敗 pattern.
    """
    qhl = _read_json_safely(data_dir / 'quad_hit_log.json')
    title_font = Font(name='Noto Sans TC', size=14, bold=True,
                      color=COLORS['brand_dark'])
    hdr_font = Font(name='Noto Sans TC', size=11, bold=True)
    hdr_fill = _summary_fill('FFFEF3C7')   # 淡金 (alpha)
    sub_font = Font(name='Noto Sans TC', size=10, italic=True,
                    color=COLORS['text_muted'])
    val_font = Font(name='Noto Sans TC', size=11)
    bold_font = Font(name='Noto Sans TC', size=11, bold=True)
    num_fmt = '0.00"%"'

    # 欄寬
    widths = {'B': 12, 'C': 10, 'D': 6, 'E': 18, 'F': 10, 'G': 8,
              'H': 10, 'I': 10, 'J': 40}
    for col_l, w in widths.items():
        ws.column_dimensions[col_l].width = w

    row = 2
    c_title = ws.cell(row, 2, "📈 Quad 實戰追蹤 (Phase 3.2 三訊號)")
    c_title.font = title_font
    ws.row_dimensions[row].height = 22
    row += 1

    if not qhl or not qhl.get('trigger_days'):
        ws.cell(row, 2, "尚無 trigger day 資料 (Q5 偏多 + vol_spike master 尚未齊聚)").font = sub_font
        return

    # ── 累積 + 30d 統計 ──
    ra = qhl.get('rolling_all', {})
    r30 = qhl.get('rolling_30d', {})
    vs_exp = qhl.get('vs_expected', {})

    sum_text = (f"累積: {ra.get('hits',0)}/{ra.get('n',0)} = {ra.get('hit_rate',0)*100:.1f}% "
                f"(mean {ra.get('mean_change',0):+.2f}%) "
                f"  |  30d: {r30.get('hits',0)}/{r30.get('n',0)} = {r30.get('hit_rate',0)*100:.1f}%"
                f"  |  預期 {vs_exp.get('expected_hit_rate',0)*100:.1f}% "
                f"(delta {vs_exp.get('delta_pp',0):+.1f}pp)")
    c_sum = ws.cell(row, 2, sum_text)
    c_sum.font = Font(name='Noto Sans TC', size=10, bold=True,
                      color=COLORS['tw_green'])
    ws.merge_cells(f'B{row}:J{row}')
    c_sum.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 18
    row += 2

    # ── 表頭 ──
    headers = ['日期', '隔日', 'Q5', 'Vol_spike masters', 'picks', 'hits',
               '命中率', 'mean%', '備註 (前 3 picks)']
    for i, h in enumerate(headers):
        c = ws.cell(row, 2 + i, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 22
    row += 1

    # ── 逐 trigger day (倒序 — 最新在前) ──
    trigger_days = sorted(qhl['trigger_days'], key=lambda x: x['date'], reverse=True)
    for td in trigger_days:
        date = td['date']
        d_fmt = f"{date[:4]}/{date[4:6]}/{date[6:8]}"
        nxt = td['next_date']
        nxt_fmt = f"{nxt[4:6]}/{nxt[6:8]}"
        ws.cell(row, 2, d_fmt).font = val_font
        ws.cell(row, 3, nxt_fmt).font = val_font
        # Q5
        c_q5 = ws.cell(row, 4, td['q5_direction'])
        c_q5.font = Font(name='Noto Sans TC', size=11, bold=True,
                         color=COLORS['tw_red'])
        c_q5.alignment = Alignment(horizontal='center')
        # vol_spike masters
        vs_str = ', '.join(td.get('vol_spike_masters') or [])
        ws.cell(row, 5, vs_str[:18]).font = sub_font
        # picks / hits
        n = td['n']; hits = td['hits']
        ws.cell(row, 6, n).font = val_font
        c_hits = ws.cell(row, 7, hits)
        # hit rate color
        hr = td['hit_rate']
        if hr >= 0.7: hr_color = COLORS['tw_green']
        elif hr >= 0.5: hr_color = 'FF666666'
        else: hr_color = COLORS['tw_red']
        c_hr = ws.cell(row, 8, hr * 100)
        c_hr.number_format = num_fmt
        c_hr.font = Font(name='Noto Sans TC', size=11, bold=True, color=hr_color)
        c_hr.alignment = Alignment(horizontal='center')
        c_hits.font = Font(name='Noto Sans TC', size=11, bold=True, color=hr_color)
        # mean
        c_mean = ws.cell(row, 9, td.get('mean_change', 0))
        c_mean.number_format = '+0.00"%";-0.00"%";0"%"'
        mean_color = COLORS['tw_red'] if td.get('mean_change', 0) > 0 else COLORS['tw_green']
        c_mean.font = Font(name='Noto Sans TC', size=11, bold=True, color=mean_color)
        c_mean.alignment = Alignment(horizontal='right')
        # picks preview (front 3 with change%)
        picks = td.get('quad_picks') or []
        preview = ' / '.join(
            f"{p['name']}({p['code']}) {p['next_change_pct']:+.1f}%"
            for p in picks[:3]
        )
        if len(picks) > 3:
            preview += f" +{len(picks)-3}"
        ws.cell(row, 10, preview).font = sub_font
        row += 1

    # ── v3.70.4 P1 研究: per-master vol_spike 可靠度 leaderboard ──
    row += 1
    ws.cell(row, 2, "Per-Master Vol_Spike 可靠度 (排序 by 命中率)").font = Font(
        name='Noto Sans TC', size=12, bold=True, color=COLORS['brand_dark'])
    row += 1
    pm_headers = ['Master', 'trigger days', 'all picks', 'hits', '命中率', 'mean%']
    for i, h in enumerate(pm_headers):
        c = ws.cell(row, 2 + i, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center')
    row += 1

    # 計算 per-master stats
    from collections import defaultdict
    master_picks = defaultdict(list)
    master_triggers = defaultdict(set)
    for td in qhl['trigger_days']:
        for m in (td.get('vol_spike_masters') or []):
            master_triggers[m].add(td['date'])
        for p in td['quad_picks']:
            for m in (p.get('matched_masters') or []):
                master_picks[m].append({'change': p['next_change_pct'],
                                         'hit': p['hit']})

    pm_rows = []
    for m, picks in master_picks.items():
        if not picks: continue
        n = len(picks)
        hits = sum(1 for p in picks if p['hit'])
        hr = hits / n
        mean = sum(p['change'] for p in picks) / n
        pm_rows.append((m, len(master_triggers[m]), n, hits, hr, mean))
    pm_rows.sort(key=lambda x: -x[4])   # by hit rate desc

    for pm in pm_rows:
        master_name, td_n, n_picks, n_hits, hit_rate, mean_chg = pm
        ws.cell(row, 2, master_name).font = val_font
        ws.cell(row, 3, td_n).alignment = Alignment(horizontal='center')
        ws.cell(row, 3, td_n).font = val_font
        ws.cell(row, 4, n_picks).alignment = Alignment(horizontal='center')
        ws.cell(row, 4, n_picks).font = val_font
        ws.cell(row, 5, n_hits).alignment = Alignment(horizontal='center')
        ws.cell(row, 5, n_hits).font = val_font
        c_hr = ws.cell(row, 6, hit_rate * 100)
        c_hr.number_format = '0.0"%"'
        hr_color = (COLORS['tw_green'] if hit_rate >= 0.8
                    else 'FF666666' if hit_rate >= 0.6
                    else COLORS['tw_red'])
        c_hr.font = Font(name='Noto Sans TC', size=11, bold=True, color=hr_color)
        c_hr.alignment = Alignment(horizontal='center')
        c_mn = ws.cell(row, 7, mean_chg)
        c_mn.number_format = '+0.00"%";-0.00"%";0"%"'
        c_mn.font = Font(name='Noto Sans TC', size=11, bold=True,
                         color=COLORS['tw_red'] if mean_chg > 0 else COLORS['tw_green'])
        c_mn.alignment = Alignment(horizontal='right')
        row += 1

    # ── 註腳 ──
    row += 1
    note = (f"註: trigger day = Q5 預測偏多 AND ≥1 master 量爆 (>2σ).\n"
            f"     picks = 該日所有共識股 ∩ ≥1 vol_spike master.\n"
            f"     命中率 = 隔日漲幅 > 0 的比例. 預期 78.9% (Phase 3.2 backtest).\n"
            f"     Per-master 命中率 < 整體 → 該 master 訊號偏弱; > 整體 → 訊號偏強.\n"
            f"     注意樣本小 (trigger days < 5) 時, 命中率 noise 偏大.")
    c_note = ws.cell(row, 2, note)
    c_note.font = sub_font
    c_note.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells(f'B{row}:J{row+3}')
    ws.row_dimensions[row].height = 70

    ws.freeze_panes = 'A6'


def build_pinned_track_sheet(ws, branches_data, data_dir):
    """v3.71.18 L2: pinned master 專屬追蹤 sheet.

    對 PINNED_MASTERS 內每位 master 顯示:
      [Header] master 名 + master_profile narrative + L1/L5/L6 stats
      [Table 1] 今日 top buys (top 10, dedup 跨分點同股 + 合計)
      [Table 2] 過去 30 天 連續加碼 stocks (從 master_profiles.consecutive_accumulation)
    """
    hdr_font = Font(name='Noto Sans TC', size=14, bold=True, color='FFB45309')
    sub_font = Font(name='Noto Sans TC', size=10, color='FF666666')
    val_font = Font(name='Noto Sans TC', size=11)
    th_font = Font(name='Noto Sans TC', size=10, bold=True)
    th_fill = PatternFill('solid', fgColor='FFFEF3C7')

    # column widths
    for col, w in [('A', 3), ('B', 22), ('C', 18), ('D', 14), ('E', 14),
                    ('F', 14), ('G', 14), ('H', 18)]:
        ws.column_dimensions[col].width = w

    pms = _read_json_safely(data_dir / 'pinned_master_stats.json') or {}
    mp = _read_json_safely(data_dir / 'master_profiles.json') or {}
    mp_masters = mp.get('individual_masters') or {}
    if not mp_masters and isinstance(mp.get('masters'), dict):
        mp_masters = mp['masters']

    row = 2
    for m_name in sorted(PINNED_MASTERS):
        # Header
        c = ws.cell(row, 2, f"📌 {m_name}")
        c.font = hdr_font
        row += 1

        # Narrative
        prof = mp_masters.get(m_name, {})
        narr = prof.get('narrative', '')
        if narr:
            c = ws.cell(row, 2, narr[:400])
            c.font = sub_font
            ws.merge_cells(f'B{row}:H{row+1}')
            c.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 28
            ws.row_dimensions[row+1].height = 28
            row += 2
        row += 1

        # L1/L5/L6 stats
        m_stats = (pms.get('pinned_masters') or {}).get(m_name, {})
        if m_stats.get('status') == 'ok':
            for label, key in [('全部 picks', 'all_picks'),
                                ('新標的', 'new_stocks'),
                                ('連續加碼', 'accumulation')]:
                s = m_stats.get(key) or {}
                if not s.get('n'): continue
                txt = (f"{label}: n={s['n']}  hit_1d={s.get('hit_1d',0)*100:.0f}%  "
                       f"mean_1d={s.get('mean_1d',0):+.2f}%  hit_3d={s.get('hit_3d',0)*100:.0f}%  "
                       f"mean_3d={s.get('mean_3d',0):+.2f}%  hit_5d={s.get('hit_5d',0)*100:.0f}%  "
                       f"mean_5d={s.get('mean_5d',0):+.2f}%")
                c = ws.cell(row, 2, txt)
                c.font = Font(name='Noto Sans TC', size=10, color='FFB45309')
                ws.merge_cells(f'B{row}:H{row}')
                row += 1
        else:
            c = ws.cell(row, 2, "(歷史 alpha stats 待 weekly cron 跑 analyze_pinned_master_alpha.py)")
            c.font = sub_font
            row += 1
        row += 1

        # Table 1: 今日 top buys
        ws.cell(row, 2, f"📊 {m_name} 今日 Top 10 買進 (跨分點同股合計)").font = th_font
        row += 1
        for col_i, header in enumerate(['#', '代號', '股名', '買金額(萬)', '買張', '漲跌%']):
            c = ws.cell(row, 2 + col_i, header)
            c.font = th_font; c.fill = th_fill
            c.alignment = Alignment(horizontal='center')
        row += 1

        master_buys = []
        for b in branches_data:
            if b.get('master') == m_name:
                for s in (b.get('buys') or []):
                    code = s.get('code')
                    if not code or code.startswith('00'): continue
                    master_buys.append({
                        'code': code, 'name': s.get('name', '—'),
                        'amt': s.get('buy_amt') or 0,
                        'volume': s.get('volume') or 0,
                        'change_pct': s.get('change_pct'),
                    })
        agg = {}
        for b in master_buys:
            key = b['code']
            if key in agg:
                agg[key]['amt'] += b['amt']
                agg[key]['volume'] += b['volume']
            else:
                agg[key] = b
        top_buys = sorted(agg.values(), key=lambda x: -x['amt'])[:10]
        if top_buys:
            for i, b in enumerate(top_buys, 1):
                ws.cell(row, 2, i)
                ws.cell(row, 3, b['code'])
                ws.cell(row, 4, b['name'])
                ws.cell(row, 5, round(b['amt'] / 10)).number_format = '#,##0'
                ws.cell(row, 6, b['volume']).number_format = '#,##0'
                chg = b.get('change_pct')
                if chg is not None:
                    c_chg = ws.cell(row, 7, chg / 100)
                    c_chg.number_format = '0.00%;[Color10]-0.00%'
                    if chg >= 0.01:
                        c_chg.font = Font(name='Noto Sans TC', size=11, bold=True, color='FFC62828')
                    elif chg <= -0.01:
                        c_chg.font = Font(name='Noto Sans TC', size=11, bold=True, color='FF2E7D32')
                row += 1
        else:
            ws.cell(row, 2, "今日無買進資料").font = sub_font
            row += 1
        row += 2

        # Table 2: 連續加碼 (從 master_profile)
        ws.cell(row, 2, f"📦 {m_name} 連續囤貨 (active)").font = th_font
        row += 1
        for col_i, header in enumerate(['#', '代號', '股名', '連續天數', '累計金額(萬)']):
            c = ws.cell(row, 2 + col_i, header)
            c.font = th_font; c.fill = th_fill
            c.alignment = Alignment(horizontal='center')
        row += 1

        op = prof.get('operation_metrics', {}) if prof else {}
        cons = op.get('consecutive_accumulation') or op.get('consecutive_active') or []
        if isinstance(cons, list) and cons:
            for i, item in enumerate(cons[:10], 1):
                ws.cell(row, 2, i)
                ws.cell(row, 3, item.get('code', '—'))
                ws.cell(row, 4, item.get('name', '—'))
                ws.cell(row, 5, item.get('days') or item.get('streak', '—'))
                amt = item.get('total_amt') or item.get('cumulative_amt')
                if amt:
                    ws.cell(row, 5 if 'days' in item else 6, round(amt / 10)).number_format = '#,##0'
                row += 1
        else:
            ws.cell(row, 2, "無連續囤貨資料 (待 master_profile 更新)").font = sub_font
            row += 1
        row += 3

    ws.freeze_panes = 'A2'


def build_quad_failure_sheet(ws, data_dir):
    """v3.70.3 Phase 3.2 失效歸因 — 從 miss 學習失敗 pattern.

    讀 data/quad_hit_log.json, 列出所有 quad miss (next_change <= 0) + 歸因:
      日期 | 隔日 | 股票 | 漲跌 | TAIEX | 超額 | 領頭% | Q5信心 | 觸發 master | 歸因

    歸因類別:
      1. 資料異常 (next_close 未變動)
      2. TAIEX 整盤跌 (≤ -0.5%)
      3. 假共識 (領頭 ≥50%)
      4. 個股弱勢 (跑輸大盤 >2pp)
      5. Q5 borderline (<55%)
      6. TAIEX 資料缺
      7. alpha noise (無系統性原因)

    用戶從 pattern 學: 若多次同類失效 → 該類訊號要 down-weight.
    """
    qhl = _read_json_safely(data_dir / 'quad_hit_log.json')
    title_font = Font(name='Noto Sans TC', size=14, bold=True,
                      color=COLORS['brand_dark'])
    hdr_font = Font(name='Noto Sans TC', size=11, bold=True)
    hdr_fill = _summary_fill('FFFEE2E2')   # 淡紅 (warning)
    sub_font = Font(name='Noto Sans TC', size=10, italic=True,
                    color=COLORS['text_muted'])
    val_font = Font(name='Noto Sans TC', size=11)
    num_fmt = '+0.00"%";-0.00"%";0"%"'

    widths = {'B': 12, 'C': 10, 'D': 16, 'E': 10, 'F': 10,
              'G': 10, 'H': 8, 'I': 9, 'J': 18, 'K': 32}
    for col_l, w in widths.items():
        ws.column_dimensions[col_l].width = w

    row = 2
    c_title = ws.cell(row, 2, "📉 Quad 失效歸因 (從失敗學習)")
    c_title.font = title_font
    ws.row_dimensions[row].height = 22
    row += 1

    if not qhl or not qhl.get('trigger_days'):
        ws.cell(row, 2, "尚無資料").font = sub_font
        return

    # 收集所有 misses
    misses = []
    for td in qhl['trigger_days']:
        for p in td['quad_picks']:
            if not p.get('hit'):
                misses.append({
                    'date': td['date'], 'next_date': td['next_date'],
                    'q5_conf': td.get('q5_confidence'),
                    'taifex_change': td.get('taifex_change'),
                    **p,
                })

    # 統計 by 歸因類別
    from collections import Counter
    reason_counts = Counter()
    for m in misses:
        for r in (m.get('failure_reasons') or ['未分類']):
            reason_counts[r] += 1

    ra = qhl.get('rolling_all', {})
    total_picks = ra.get('n', 0)
    total_hits = ra.get('hits', 0)
    total_misses = total_picks - total_hits

    # 摘要 banner — 預期值動態從 vs_expected 讀 (v3.70.4)
    vs_exp = qhl.get('vs_expected', {})
    expected_hr = vs_exp.get('expected_hit_rate', 0.857)
    expected_miss_rate = (1 - expected_hr) * 100
    actual_miss_rate = total_misses / max(total_picks, 1) * 100
    sum_text = (f"miss {total_misses}/{total_picks} ({actual_miss_rate:.1f}%) "
                f"| 預期 miss rate {expected_miss_rate:.1f}% "
                f"| 差異 {actual_miss_rate - expected_miss_rate:+.1f}pp")
    c_sum = ws.cell(row, 2, sum_text)
    c_sum.font = Font(name='Noto Sans TC', size=10, bold=True,
                      color=COLORS['tw_red'])
    ws.merge_cells(f'B{row}:K{row}')
    c_sum.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 18
    row += 1

    # 歸因分布 1 行 summary
    if reason_counts:
        sorted_reasons = sorted(reason_counts.items(), key=lambda x: -x[1])
        dist_str = ' | '.join(f"{r}: {c}" for r, c in sorted_reasons[:6])
        c_dist = ws.cell(row, 2, f"歸因分布: {dist_str}")
        c_dist.font = Font(name='Noto Sans TC', size=10, italic=True,
                           color=COLORS['text_secondary'])
        ws.merge_cells(f'B{row}:K{row}')
        c_dist.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 18
        row += 2
    else:
        row += 1

    # 表頭
    headers = ['日期', '隔日', '股票', '漲跌', 'TAIEX', '超額',
               '領頭%', 'Q5%', '觸發 master', '歸因']
    for i, h in enumerate(headers):
        c = ws.cell(row, 2 + i, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 22
    row += 1

    # 逐 miss 列 (倒序 — 最近在前)
    misses_sorted = sorted(misses, key=lambda x: x['date'], reverse=True)
    for m in misses_sorted:
        date = m['date']
        d_fmt = f"{date[:4]}/{date[4:6]}/{date[6:8]}"
        nxt_fmt = f"{m['next_date'][4:6]}/{m['next_date'][6:8]}"
        ws.cell(row, 2, d_fmt).font = val_font
        ws.cell(row, 3, nxt_fmt).font = val_font
        # 股票
        stk = f"{m['name']}({m['code']})"
        ws.cell(row, 4, stk).font = val_font
        # 漲跌
        c_chg = ws.cell(row, 5, m.get('next_change_pct', 0))
        c_chg.number_format = num_fmt
        c_chg.font = Font(name='Noto Sans TC', size=11, bold=True,
                          color=COLORS['tw_green'])
        c_chg.alignment = Alignment(horizontal='right')
        # TAIEX
        taifex = m.get('taifex_change')
        if taifex is not None:
            c_t = ws.cell(row, 6, taifex)
            c_t.number_format = num_fmt
            c_t.alignment = Alignment(horizontal='right')
        else:
            ws.cell(row, 6, 'N/A').font = sub_font
        # 超額
        excess = m.get('excess_return')
        if excess is not None:
            c_e = ws.cell(row, 7, excess)
            c_e.number_format = num_fmt
            c_e.font = Font(name='Noto Sans TC', size=11,
                            color=COLORS['tw_green'] if excess <= 0 else COLORS['tw_red'])
            c_e.alignment = Alignment(horizontal='right')
        else:
            ws.cell(row, 7, 'N/A').font = sub_font
        # 領頭%
        lp = m.get('leader_pct', 0) * 100
        c_lp = ws.cell(row, 8, lp)
        c_lp.number_format = '0"%"'
        c_lp.font = (Font(name='Noto Sans TC', size=11, bold=True,
                          color=COLORS['tw_red'])
                     if lp >= 50 else Font(name='Noto Sans TC', size=11))
        c_lp.alignment = Alignment(horizontal='center')
        # Q5 信心
        qc = m.get('q5_conf', 0)
        c_qc = ws.cell(row, 9, qc)
        c_qc.number_format = '0.0"%"'
        c_qc.font = (Font(name='Noto Sans TC', size=11, bold=True,
                          color=COLORS['hot_orange'])
                     if qc < 55 else Font(name='Noto Sans TC', size=11))
        c_qc.alignment = Alignment(horizontal='center')
        # 觸發 master
        master_str = ', '.join(m.get('matched_masters') or [])[:16]
        ws.cell(row, 10, master_str).font = sub_font
        # 歸因
        reasons = ' | '.join(m.get('failure_reasons') or [])
        ws.cell(row, 11, reasons).font = Font(name='Noto Sans TC', size=10,
                                              color=COLORS['signal_red'])
        row += 1

    if not misses:
        ws.cell(row, 2, "✅ 目前無 quad miss — 全 hit").font = Font(
            name='Noto Sans TC', size=11, bold=True, color=COLORS['tw_green'])

    # 註腳
    row += 2
    note = (
        f"歸因分類: "
        f"flat close = 隔日收盤恰等於今日 (intraday 有波動, TWSE 證實 legit, 非 stale).  "
        f"TAIEX 整盤跌 = 隔日大盤 ≤-0.5%.  "
        f"假共識 = 領頭佔比 ≥50%.  "
        f"個股弱勢 = 跑輸大盤 >2pp.  "
        f"Q5 borderline = 預測信心 <55%.  "
        f"alpha noise = 無系統性原因 (真 alpha 命中率本就 ~79%, 隨機 ~21% miss).")
    c_note = ws.cell(row, 2, note)
    c_note.font = sub_font
    c_note.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells(f'B{row}:K{row+3}')
    ws.row_dimensions[row].height = 60

    ws.freeze_panes = 'A7'


def _compute_q5_hit_rate(data_dir, window_days=30):
    """v3.66.8 Phase 2.4: 計算 Q5 預測歷史命中率.

    Logic:
      - 對 temp_history.json 每筆 entry, 用 signals + infer_market_direction 算 predicted
      - 對比 entry.next_day_change_pct (隔日 TAIEX 漲跌)
      - direction='偏多' AND next_day > 0 → hit
      - direction='偏空' AND next_day < 0 → hit
      - direction='中性' OR next_day_change_pct is None → skip (no bet)

    Returns:
      {'bull': (hits, total), 'bear': (hits, total), 'overall': (hits, total)}
    """
    try:
        import sys as _sys
        _root = data_dir.parent if hasattr(data_dir, 'parent') else None
        if _root and str(_root) not in _sys.path:
            _sys.path.insert(0, str(_root))
        from src.analyzers.signal_engine import infer_market_direction
    except Exception:
        return None

    try:
        with open(data_dir / 'temp_history.json', 'r', encoding='utf-8') as f:
            th = __import__('json').load(f)
    except Exception:
        return None

    history = th.get('history') or []
    history = history[-window_days:] if window_days else history

    # v3.67.3 Phase 2.4 Fix #3: stale guard
    # 揭穿: 6/2-6/8 共 5 個 entry change_pct=0.0 全 false-negative
    # 原因: 兜底排程未抓新 TAIEX, index 跟前一日相同, 但 change_pct=0.0 寫入 history
    # 修補: change_pct=0.0 AND next_day_close=None (證據是 missing) → skip
    bull_hits, bull_total = 0, 0
    bear_hits, bear_total = 0, 0
    skipped_stale = 0
    for e in history:
        nxt = e.get('next_day_change_pct')
        if nxt is None:
            continue
        # v3.67.3: stale = 0.0 + 缺 close 證據 → 視為 missing
        if nxt == 0.0 and e.get('next_day_close') is None:
            skipped_stale += 1
            continue
        signals = e.get('signals') or []
        if not signals:
            continue
        try:
            md = infer_market_direction(signals)
        except Exception:
            continue
        direction = md.get('direction')
        if direction == '偏多':
            bull_total += 1
            if nxt > 0:
                bull_hits += 1
        elif direction == '偏空':
            bear_total += 1
            if nxt < 0:
                bear_hits += 1
        # 中性 → 不算

    overall_total = bull_total + bear_total
    overall_hits = bull_hits + bear_hits
    return {
        'bull': (bull_hits, bull_total),
        'bear': (bear_hits, bear_total),
        'overall': (overall_hits, overall_total),
        'window_days': window_days,
    }


def _update_load_timeseries(data_dir, trade_date, kpis, update=True):
    """v3.66.7 Phase 2.3: 時間維度 cache.

    Schema (data/timeseries.json):
      {dates: ["20260623","20260624"], q1_active_ratio: [...], q2_net_billion: [...],
       q3_consensus_count: [...], q3_consensus_net_billion: [...],
       q4_track_share: [...], q4_mkt_net_billion: [...]}

    Args:
      data_dir: Path
      trade_date: YYYYMMDD
      kpis: dict {q1_active_ratio, q2_net_billion, q3_consensus_count,
                  q3_consensus_net_billion, q4_track_share, q4_mkt_net_billion}
      update: True = 寫入 cache (production); False = 只讀 (test mode)

    Returns:
      {yesterday: {qN: val|None}, avg5: {qN: val|None}, days_history: int}
    """
    import json as _j
    cache_path = data_dir / 'timeseries.json'
    keys = ['q1_active_ratio', 'q2_net_billion', 'q3_consensus_count',
            'q3_consensus_net_billion', 'q4_track_share', 'q4_mkt_net_billion']
    try:
        with open(cache_path, 'r', encoding='utf-8') as f:
            cache = _j.load(f)
    except (FileNotFoundError, _j.JSONDecodeError):
        cache = {'dates': []}
    for k in keys:
        cache.setdefault(k, [])

    if update:
        if trade_date in cache['dates']:
            idx = cache['dates'].index(trade_date)
            for k in keys:
                if idx < len(cache[k]):
                    cache[k][idx] = kpis.get(k, 0)
        else:
            cache['dates'].append(trade_date)
            for k in keys:
                cache[k].append(kpis.get(k, 0))
        # Sort by date + cap 60 days
        sorted_idx = sorted(range(len(cache['dates'])),
                             key=lambda i: cache['dates'][i])
        cache['dates'] = [cache['dates'][i] for i in sorted_idx][-60:]
        for k in keys:
            cache[k] = [cache[k][i] for i in sorted_idx if i < len(cache[k])][-60:]
        try:
            with open(cache_path, 'w', encoding='utf-8') as f:
                _j.dump(cache, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    # 計算 yesterday + 5-day avg (excluding today)
    past_idx = [i for i, d in enumerate(cache['dates']) if d < trade_date]
    yesterday, avg5 = {}, {}
    for k in keys:
        past_vals = [cache[k][i] for i in past_idx if i < len(cache[k])]
        yesterday[k] = past_vals[-1] if past_vals else None
        avg5[k] = (sum(past_vals[-5:]) / min(len(past_vals), 5)) if past_vals else None
    return {'yesterday': yesterday, 'avg5': avg5, 'days_history': len(past_idx)}


def _build_section_summary(ws, branches_data, trade_date, data_dir, start_row,
                            all_branches=None, update_timeseries=True):
    """Section A: 追蹤池摘要 (v3.64.3) — 10 秒判讀今天追蹤大戶在做什麼.

    4 KPI 對應 4 個盤前 decision-making 問題:
      Q1 活躍率      → 是否值得看細節?
      Q2 淨買差      → 偏多/偏空 bias?
      Q3 強共識股    → 是否有 conviction 還是散亂?
      Q4 追蹤佔比    → vs 全市場誰更看多?

    + Top 5 master + Top 5 個股 + 籌碼溫度 (後續 sections 保持不變).
    """
    label_font = Font(name='Noto Sans TC', size=10, color='FF666666')
    val_font = Font(name='Noto Sans TC', size=14, bold=True)
    hdr_font = Font(name='Noto Sans TC', size=10, bold=True)
    hdr_fill = _summary_fill('FFF0F0F0')

    row = start_row
    _section_header(ws, row, "▍ A. 追蹤池摘要 (10 秒判讀今日 13 位大戶在做什麼)")
    row += 1

    # ── 共同計算: sell_amt dedup helper (v3.64.1 Bug 1 fix) ──
    def _compute_buy_sell(blist):
        buy = sum((s.get('buy_amt') or 0) for b in blist for s in (b.get('buys') or []))
        seen = set()
        sell = 0
        for b in blist:
            bcode = b.get('code', '')
            for s in (b.get('buys') or []) + (b.get('sells') or []):
                key = (bcode, s.get('code'))
                if key in seen: continue
                seen.add(key)
                sell += (s.get('sell_amt') or 0)
        return buy, sell

    total_buy, total_sell = _compute_buy_sell(branches_data)
    total_net = total_buy - total_sell
    net_billion = total_net / 100000   # 仟元 → 億元

    # ── Q1: 活躍率 = 今天有 buys 的追蹤大戶 / 全追蹤大戶數 (13) ──
    active_masters = {b.get('master') for b in branches_data
                       if (b.get('buys') or []) and b.get('master')}
    total_masters = len(TRACKED_MASTERS)
    active_count = len(active_masters)
    active_ratio = (active_count / total_masters) if total_masters else 0

    # ── Q3: 強共識股數 + 合計淨買 (Section 0 相同 logic, 集合查詢) ──
    consensus_stocks = _compute_consensus_count(branches_data)
    consensus_count = len(consensus_stocks)
    consensus_net = sum(s['total_net_amt'] for s in consensus_stocks)
    consensus_net_billion = consensus_net / 100000

    # ── Q4: 追蹤佔比 vs 全市場 ──
    if all_branches:
        mkt_buy, mkt_sell = _compute_buy_sell(all_branches)
        mkt_net_billion = (mkt_buy - mkt_sell) / 100000
        track_share = (total_buy / mkt_buy) if mkt_buy else 0
    else:
        mkt_net_billion = 0
        track_share = 0

    # ── v3.66.7 Phase 2.3: 時間維度 cache (今/昨/5日均) ──
    ts = _update_load_timeseries(data_dir, trade_date, {
        'q1_active_ratio': active_ratio,
        'q2_net_billion': net_billion,
        'q3_consensus_count': consensus_count,
        'q3_consensus_net_billion': consensus_net_billion,
        'q4_track_share': track_share,
        'q4_mkt_net_billion': mkt_net_billion,
    }, update=update_timeseries)
    y, a = ts['yesterday'], ts['avg5']

    # 累積中 (歷史不足) 顯示
    has_history = ts['days_history'] >= 1

    # ── v3.66.7 動態 format strings — 緊湊版避免 cell overflow ──
    # 避免 ##### bug: 縮短 sub-text (去空格 / 去單位重複 / 縮 label)
    mkt_sign = '+' if mkt_net_billion >= 0 else ''
    if has_history:
        # Q1: "100% (13/13 ・昨100/5d100)" — 去掉 % 重複, 用 ・ 緊湊分隔
        y_q1 = (y['q1_active_ratio'] or 0) * 100
        a_q1 = (a['q1_active_ratio'] or 0) * 100
        active_fmt = (f'0%" ({active_count}/{total_masters} ・昨{y_q1:.0f}/5d{a_q1:.0f})"')

        # Q2: "-204億 (昨-412/5d-188)" — 去小數點, 去空格
        y_q2 = y['q2_net_billion'] or 0
        a_q2 = a['q2_net_billion'] or 0
        net_fmt = (f'+0" 億 (昨{y_q2:+.0f}/5d{a_q2:+.0f})";'
                   f'-0" 億 (昨{y_q2:+.0f}/5d{a_q2:+.0f})";'
                   f'0" 億"')

        # Q3: "10檔 +185億 (昨14/5d11)" — 去 "淨買" 字
        y_q3 = y['q3_consensus_count'] or 0
        a_q3 = a['q3_consensus_count'] or 0
        consensus_fmt = (f'0" 檔 {consensus_net_billion:+.0f}億 "'
                         f'"(昨{y_q3:.0f}/5d{a_q3:.0f})"')

        # Q4: "21.2% 市-2779億 (昨17/5d19)" — 縮短 "市場" → "市"
        y_q4 = (y['q4_track_share'] or 0) * 100
        a_q4 = (a['q4_track_share'] or 0) * 100
        share_fmt = (f'0.0%" 市{mkt_sign}{mkt_net_billion:.0f}億 "'
                     f'"(昨{y_q4:.0f}/5d{a_q4:.0f})"')
    else:
        active_fmt = f'0%" ({active_count}/{total_masters}, 累積中)"'
        net_fmt = '+0.00" 億 (累積中)";-0.00" 億 (累積中)";0" 億"'
        consensus_fmt = f'0" 檔 (淨買 {consensus_net_billion:+.0f} 億, 累積中)"'
        share_fmt = f'0.0%" (市場 {mkt_sign}{mkt_net_billion:.0f} 億, 累積中)"'

    # ── 4 KPI 2x2 layout (label + value 並排) ──
    # Row 1: B 活躍率 | C-D merged value | F 淨買差 | G-I merged value
    # Row 2: B 強共識股 | C-D merged value | F 追蹤佔比 | G-I merged value
    stats = [
        # (label_col, label_text, value_col, merge_to_col, value, fmt, font_color)
        (row,     'B', 'Q1 活躍率',   'C', 'D', active_ratio,  active_fmt,    None),
        (row,     'F', 'Q2 淨買差',   'G', 'I', net_billion,   net_fmt,
            'FFDC2626' if total_net >= 0 else 'FF059669'),
        (row + 1, 'B', 'Q3 強共識股', 'C', 'D', consensus_count, consensus_fmt, None),
        (row + 1, 'F', 'Q4 追蹤佔比', 'G', 'I', track_share,   share_fmt,
            'FFDC2626' if mkt_net_billion >= 0 else 'FF059669'),
    ]
    # v3.66.7+ font 恢復 14pt — 透過拉寬 C/D 欄解決 overflow (不縮字)
    kpi_font_size = 12 if has_history else 14
    for r, lcol, ltext, vcol_start, vcol_end, val, fmt, color in stats:
        cl = ws[f'{lcol}{r}']
        cl.value = ltext
        cl.font = label_font
        cl.alignment = Alignment(horizontal='right', vertical='center')
        # value cell + merge
        cv = ws[f'{vcol_start}{r}']
        cv.value = val
        cv.number_format = fmt
        # v3.66.7: enable wrap_text 防止 overflow
        cv.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cv.font = (Font(name='Noto Sans TC', size=kpi_font_size, bold=True, color=color)
                   if color else Font(name='Noto Sans TC', size=kpi_font_size, bold=True))
        if vcol_start != vcol_end:
            ws.merge_cells(f'{vcol_start}{r}:{vcol_end}{r}')
    # v3.66.7+ row 高恢復 28 (font 12pt 配 28px 行距正好)
    ws.row_dimensions[row].height = 28
    ws.row_dimensions[row + 1].height = 28
    row += 2

    # ── v3.64.4 Q5: 市場方向 banner (整合 Section D, 全寬, 紅/綠/灰底) ──
    # 資料源 daily_signal.json 在 v3.64+ 已重構, 新 schema:
    #   market_direction: {direction: 偏多/偏空/中性, confidence_pct, contributing[]}
    #   top_focus_stocks: list
    # 顯示: ↑ 偏多 58.7% 信心 — P/C Ratio 主推 — 3 檔焦點 (banner 整列)
    daily_signal = _read_json_safely(data_dir / 'daily_signal.json')
    if daily_signal:
        md = daily_signal.get('market_direction') or {}
        direction = md.get('direction') or '—'
        confidence = md.get('confidence_pct') or 0
        contributing = md.get('contributing') or []
        top_signal = contributing[0].get('name') if contributing else '—'
        focus_n = len(daily_signal.get('top_focus_stocks') or [])

        # icon + bg color by direction (台股慣例: 紅=多 / 綠=空 / 灰=中性)
        if direction == '偏多':
            arrow, bg = '↑', 'FFFEE2E2'   # 淡紅
            color = 'FFDC2626'
        elif direction == '偏空':
            arrow, bg = '↓', 'FFD1FAE5'   # 淡綠
            color = 'FF059669'
        else:
            arrow, bg = '↕', 'FFE5E7EB'   # 淡灰
            color = 'FF374151'

        # v3.64.6 解讀正確性修補: 加「明日預測」前綴
        # 原因: signal_engine.infer_market_direction 註解明確「推 TAIEX 明日方向」
        # 不加前綴用戶可能誤解為「今日市場偏多」, 實際是預測「明日 TAIEX 偏多」
        # 例: 58.7 + fmt '"📅 明日預測 ↑ 偏多 "0.0"% — P/C Ratio 主推 — 3 檔焦點"'
        # → 顯示: "📅 明日預測 ↑ 偏多 58.7% — P/C Ratio 主推 — 3 檔焦點"
        q5_fmt = f'"📅 明日預測 {arrow} {direction} "0.0"% 信心 — {top_signal} 主推 — {focus_n} 檔焦點"'

        ws.merge_cells(f'B{row}:N{row}')
        c_q5 = ws[f'B{row}']
        c_q5.value = float(confidence)
        c_q5.number_format = q5_fmt
        c_q5.alignment = Alignment(horizontal='center', vertical='center')
        c_q5.font = Font(name='Noto Sans TC', size=12, bold=True, color=color)
        c_q5.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        ws.row_dimensions[row].height = 22
        row += 1

        # v3.66.8 Phase 2.4: Q5 hit rate 累積 sub-banner
        hr = _compute_q5_hit_rate(data_dir, window_days=30)
        if hr and hr['overall'][1] > 0:
            bull_h, bull_t = hr['bull']
            bear_h, bear_t = hr['bear']
            ovr_h, ovr_t = hr['overall']
            bull_pct = (bull_h / bull_t * 100) if bull_t else 0
            bear_pct = (bear_h / bear_t * 100) if bear_t else 0
            ovr_pct = (ovr_h / ovr_t * 100)
            # 整體 hit rate 顏色: ≥60% 綠 / 40-60 灰 / <40 紅
            if ovr_pct >= 60:
                hr_color = 'FF059669'   # 綠
                hr_icon = '✅'
            elif ovr_pct >= 40:
                hr_color = 'FF666666'   # 灰
                hr_icon = '🟡'
            else:
                hr_color = 'FFDC2626'   # 紅
                hr_icon = '⚠️'
            hr_text = (f"{hr_icon} 過去 {hr['window_days']} 天 P/C 命中率: "
                       f"偏多 {bull_h}/{bull_t} ({bull_pct:.0f}%) | "
                       f"偏空 {bear_h}/{bear_t} ({bear_pct:.0f}%) | "
                       f"整體 {ovr_h}/{ovr_t} ({ovr_pct:.0f}%)")
            ws.merge_cells(f'B{row}:N{row}')
            c_hr = ws[f'B{row}']
            c_hr.value = hr_text
            c_hr.alignment = Alignment(horizontal='center', vertical='center')
            c_hr.font = Font(name='Noto Sans TC', size=10, italic=True, color=hr_color)
            c_hr.fill = PatternFill(start_color='FFF9FAFB', end_color='FFF9FAFB',
                                     fill_type='solid')   # 極淡灰
            ws.row_dimensions[row].height = 18
            row += 1

    row += 1   # 空一行

    # Top master + Top stocks 並排
    ws.merge_cells(f'B{row}:E{row}')
    s2 = ws[f'B{row}']
    s2.value = "▍ B. Top 5 高手(按今日總買金額)"
    s2.font = Font(name='Noto Sans TC', size=11, bold=True)
    s2.fill = _summary_fill('FFE8F5E9')
    s2.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.merge_cells(f'F{row}:I{row}')
    s3 = ws[f'F{row}']
    s3.value = "▍ C. Top 5 熱門個股(按今日淨買金額)"
    s3.font = Font(name='Noto Sans TC', size=11, bold=True)
    s3.fill = _summary_fill('FFE3F2FD')
    s3.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    row += 1

    master_amt = {}
    for b in branches_data:
        m = b.get('master')
        if not m: continue
        amt = sum((s.get('buy_amt') or 0) for s in (b.get('buys') or []))
        master_amt[m] = master_amt.get(m, 0) + amt
    top_masters = sorted(master_amt.items(), key=lambda x: -x[1])[:5]
    stock_net = {}
    stock_name = {}
    stock_close = {}    # v3.65.0: stock_history 抓今日 close
    stock_prev = {}
    for b in branches_data:
        for s in (b.get('buys') or []):
            c = s.get('code')
            if not c: continue
            # v3.65.0 用戶要求: Dashboard 只顯示個股 (排 ETF 等非個股)
            if _is_excluded_by_market_type(s):
                continue
            stock_net[c] = stock_net.get(c, 0) + (s.get('buy_amt') or 0) - (s.get('sell_amt') or 0)
            stock_name[c] = s.get('name', '')
            if s.get('change_pct') is not None and c not in stock_close:
                stock_close[c] = s.get('close_price')
                stock_prev[c] = s.get('change_pct')
    top_stocks = sorted(stock_net.items(), key=lambda x: -x[1])[:5]

    # v3.65.0: headers — C 改新增「漲跌%」col 給 Top stock 顯示, B 改 8 col → 8 col 保持
    # 維持並排 4-col block: B-E master block, F-I stock block
    for col_letter, txt in [('B', '#'), ('C', 'Master'), ('D', '買進(萬元)'), ('E', '佔比%'),
                              ('F', '#'), ('G', '個股'), ('H', '淨買(萬元)'), ('I', '漲跌%')]:
        ws[f'{col_letter}{row}'] = txt
        ws[f'{col_letter}{row}'].font = hdr_font
        ws[f'{col_letter}{row}'].fill = hdr_fill
    row += 1

    total_all = sum(master_amt.values()) or 1
    bc_data_start = row    # v3.66.6: 記下 B/C 資料開始 row 供 data bar
    for i in range(5):
        m_item = top_masters[i] if i < len(top_masters) else None
        s_item = top_stocks[i] if i < len(top_stocks) else None
        if m_item:
            ws[f'B{row}'] = i + 1
            c_master = ws[f'C{row}']
            c_master.value = m_item[0]
            c_master.font = Font(name='Noto Sans TC', size=11, bold=True)
            # v3.65.0: master 色塊延伸 — 套既有 MASTER_BLOCK_COLORS body 色
            colors = MASTER_BLOCK_COLORS.get(m_item[0]) or DEFAULT_MASTER_COLOR
            c_master.fill = PatternFill('solid', fgColor=colors['body'])
            ws[f'D{row}'] = round(m_item[1] / 10, 0)
            ws[f'D{row}'].number_format = '#,##0'
            ws[f'E{row}'] = f"{m_item[1]/total_all*100:.1f}%"
        if s_item:
            code, net = s_item
            ws[f'F{row}'] = i + 1
            # G 改顯示 「name(code)」格式跟 day sheet 一致
            stock_label = f"{stock_name.get(code, '')}({code})"
            ws[f'G{row}'] = stock_label
            ws[f'G{row}'].font = Font(name='Noto Sans TC', size=11, bold=True)
            ws[f'H{row}'] = round(net / 10, 0)
            ws[f'H{row}'].number_format = '#,##0'
            # v3.65.0: I 欄漲跌% (從 crawler 注入的 change_pct)
            change_pct = stock_prev.get(code)
            if change_pct is not None:
                ws[f'I{row}'] = change_pct / 100
                ws[f'I{row}'].number_format = '0.00%;[Color10]-0.00%'
                # 紅綠字色 (台股傳統)
                if change_pct >= 0.01:
                    ws[f'I{row}'].font = Font(name='Noto Sans TC', size=11, bold=True, color='FFC62828')
                elif change_pct <= -0.01:
                    ws[f'I{row}'].font = Font(name='Noto Sans TC', size=11, bold=True, color='FF2E7D32')
            else:
                ws[f'I{row}'] = '—'
                ws[f'I{row}'].font = Font(name='Noto Sans TC', size=10, color='FF888888')
        row += 1
    # v3.66.6 Phase 2.2: B 買進金額 + C 淨買金額 加 data bar (深綠)
    _try_add_data_bar(ws, f'D{bc_data_start}:D{bc_data_start+4}', 'FF66BB6A')
    _try_add_data_bar(ws, f'H{bc_data_start}:H{bc_data_start+4}', 'FF66BB6A')
    row += 1

    # v3.64.4: 籌碼溫度 / 市場方向 已整合進 Section A Q5 banner (上方).
    # 原 Section D 因 daily_signal.json schema 重構而 silently 失效 (顯示 '—').
    # 新版改讀 market_direction.{direction, confidence_pct, contributing}.
    return row


def _anomaly_severity(a: Dict) -> float:
    """v3.66.0: 統一 anomaly 排序權重.

    volume_spike → 用 |z_score|
    new_stocks   → 用 2.5 + count * 0.2 (5 檔新標 = 3.5, 跟 3σ 量爆同級)
                  原本沒 z_score 預設 0 會被沉底 → 用戶看不到新標進場警報

    用戶 2026-06-23 確認砍 E 重複 sub-section, 修這個 bug 是必修 P0.
    """
    if a.get('type') == 'new_stocks':
        return 2.5 + (a.get('count', 0) or 0) * 0.2
    try:
        return abs(float(a.get('z_score') or 0))
    except (TypeError, ValueError):
        return 0.0


def _build_section_alerts(ws, data_dir, start_row):
    """Section E v3.66.0: 砍掉 🟡 共識 + 🟢 連續加碼 (跟 Section 0/A/F 重複).

    只留 🔴 異常 (volume_spike + new_stocks), top 10.
    用戶 2026-06-23 裁示「Dashboard 要簡潔但有力」.
    """
    hdr_font = Font(name='Noto Sans TC', size=10, bold=True)
    hdr_fill = _summary_fill('FFFEE2E2')

    row = start_row + 1   # 留一行空白
    _section_header(ws, row, "▍ E. 異常行為警報 (z>2σ 量爆 + 新標的進場)",
                     color='FFEF4444'); row += 1
    headers = ['類型', 'Master', '嚴重度', '說明', '金額/檔數']
    for i, h in enumerate(headers):
        cell = ws.cell(row, 2 + i, h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
    row += 1

    signals = _read_json_safely(data_dir / 'daily_trading_signals.json')
    start_data = row
    footer_fill = _summary_fill('FFF3F4F6')
    footer_font = Font(name='Noto Sans TC', size=10, italic=True, color='FF6B7280')
    TOP_N = 10

    if signals:
        # v3.66.0: 砍掉 consensus + accumulations sub-section, 只留 anomalies expand to top 10
        # 排序用 _anomaly_severity 修 new_stocks 沉底 bug
        anomalies = [s for s in (signals.get('anomalies') or [])
                     if _is_tracked_master(s.get('master'))]
        anomalies.sort(key=lambda x: -_anomaly_severity(x))

        for sig in anomalies[:TOP_N]:
            t = sig.get('type')
            if t == 'new_stocks':
                ws.cell(row, 2, '🆕 新標的')
                ws.cell(row, 3, sig.get('master', '—'))
                count = sig.get('count', 0) or 0
                ws.cell(row, 4, 'high' if count >= 5 else 'medium')
                top_new = sig.get('top_new') or []
                codes_preview = ', '.join([n.get('code', '') for n in top_new[:3]])
                ws.cell(row, 5,
                        sig.get('description',
                                f"今日買進 {count} 檔過去從未買過 (top: {codes_preview})"))
                ws.cell(row, 6, f'{count} 檔')
            else:
                ws.cell(row, 2, '🔴 量爆')
                ws.cell(row, 3, sig.get('master', '—'))
                ws.cell(row, 4, _severity_from_z(sig.get('z_score')))
                ws.cell(row, 5, sig.get('description', '—'))
                ws.cell(row, 6, _round_safe(sig.get('today_buy_amt_wan')))
            row += 1

        if len(anomalies) > TOP_N:
            extra = len(anomalies) - TOP_N
            ws.cell(row, 2, f'… 另 {extra} 筆 anomaly (詳見當日 sheet)')
            ws.merge_cells(f'B{row}:F{row}')
            for col in range(2, 7):
                ws.cell(row, col).fill = footer_fill
                ws.cell(row, col).font = footer_font
            ws.cell(row, 2).alignment = Alignment(horizontal='left', indent=1)
            row += 1

    if row == start_data:
        ws.cell(row, 2, '✅ 今日無異常行為 (追蹤範圍內)')
        ws.merge_cells(f'B{row}:F{row}')
        row += 1
    else:
        # v3.67.0 Phase 2.6: E 套 zebra stripes (cols B-F)
        _zebra_stripes(ws, start_data, row - 1, col_start='B', col_end='F')
    return row


def _severity_from_z(z_score) -> str:
    """v3.63.2: z_score -> severity 標籤 (anomaly 用)."""
    try:
        z = abs(float(z_score or 0))
        if z >= 3.0: return 'high'
        if z >= 2.0: return 'medium'
        return 'low'
    except (TypeError, ValueError):
        return 'medium'


def _round_safe(v):
    """v3.63.2: 安全 round int(萬), 失敗回 '—'."""
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return '—'


def _build_section_accumulation(ws, data_dir, start_row):
    """Section F: 跨日連續囤貨 (top 30 by 連續天數). Returns next row.

    v3.63.2: 兩 bug 修復
      (a) 舊版讀 master_profiles.json 的 master_profiles.<m>.op.consecutive_accumulation
          → 結構已重構, 改讀 daily_trading_signals.json 的 accumulations (與 Section E 同源)
      (b) 過濾非追蹤 master
    與 Section E 差異: E 顯示 top 15 (新鮮事), F 顯示 top 30 + 按天數深排序 (深度表).
    """
    hdr_font = Font(name='Noto Sans TC', size=10, bold=True)
    hdr_fill = _summary_fill('FFE8F5E9')

    row = start_row + 1
    _section_header(ws, row, "▍ F. 跨日連續囤貨 Top 30 (按連續天數排序)",
                     color='FF10B981'); row += 1
    # v3.63.2: headers 改為實際存在的欄位 (移除 stock_name/latest_date/is_active —
    # daily_trading_signals.json accumulations 沒這些欄位)
    headers = ['Master', '股票代號', '連續天數', '累計買金額(萬)', '說明']
    for i, h in enumerate(headers):
        cell = ws.cell(row, 2 + i, h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
    row += 1

    signals = _read_json_safely(data_dir / 'daily_trading_signals.json')
    start_data = row
    hot_font_red = Font(name='Noto Sans TC', size=11, bold=True, color='FFC62828')
    HOT_DAYS = 10
    if signals:
        # v3.65.0: 排除 ETF (code 起始 '00')
        # v3.66.0: 連續 ≥10 天標 🔴 hot (master prefix + 紅字粗體連續天數)
        acc_list = [s for s in (signals.get('accumulations') or [])
                    if _is_tracked_master(s.get('master'))
                    and not (s.get('stock_code') or '').startswith('00')]
        acc_list.sort(key=lambda x: -x.get('consecutive_days', 0))
        for s in acc_list[:30]:
            days = s.get('consecutive_days', 0) or 0
            is_hot = days >= HOT_DAYS
            master_label = s.get('master', '—')
            if is_hot:
                master_label = f'🔴 {master_label}'
            ws.cell(row, 2, master_label)
            ws.cell(row, 3, s.get('stock_code', '—'))
            cell_days = ws.cell(row, 4, days)
            if is_hot:
                cell_days.font = hot_font_red
            ws.cell(row, 5, _round_safe(s.get('total_buy_amt_wan')))
            ws.cell(row, 6, s.get('description', '—'))
            row += 1
    if row == start_data:
        ws.cell(row, 2, '尚無連續囤貨紀錄 (追蹤範圍內)')
        ws.merge_cells(f'B{row}:F{row}')
        row += 1
    else:
        # v3.66.6 Phase 2.2: D 連續天數加 data bar (橫條視覺化)
        # 一秒掃出誰囤最久 — 不用讀數字
        _try_add_data_bar(ws, f'D{start_data}:D{row-1}', 'FFEF5350')   # 紅 = hot
        # E 累計買金額也加 data bar (深綠 = 越多越強)
        _try_add_data_bar(ws, f'E{start_data}:E{row-1}', 'FF81C784')
        # v3.67.0 Phase 2.6: F 套 zebra stripes (cols B-F)
        _zebra_stripes(ws, start_data, row - 1, col_start='B', col_end='F')
    return row


def _try_add_data_bar(ws, cell_range, color, show_value=True):
    """v3.66.6: helper — 加 Excel data bar (橫條) 到 cell range. 失敗安全跳過."""
    try:
        from openpyxl.formatting.rule import DataBarRule
        rule = DataBarRule(
            start_type='min', end_type='max',
            color=color, showValue=show_value,
            minLength=5, maxLength=90,
        )
        ws.conditional_formatting.add(cell_range, rule)
    except Exception:
        pass


def _build_section_pivot(ws, branches_data, start_row):
    """E7 (v3.63.0): Section J Master × Top 3 個股 cross-table.

    每個 master 一 row, 把該 master 今日 buys 按 buy_amt sort top 3, 寬展開為 6 cols
    (個股1+金額, 個股2+金額, 個股3+金額) + 總額.
    """
    hdr_font = Font(name='Noto Sans TC', size=10, bold=True)
    hdr_fill = _summary_fill('FFE0E7FF')

    row = start_row + 1
    _section_header(ws, row, "▍ J. Master × Top 3 個股 cross-table (今日)",
                     color='FF6366F1'); row += 1
    # v3.66.2: 加 col J「Top3 合計%」(集中度), Master cell 套 block color
    for h_col, h in [('B', 'Master'), ('C', '今日總買(萬)'),
                      ('D', 'Top1 個股'), ('E', 'Top1 金額'),
                      ('F', 'Top2 個股'), ('G', 'Top2 金額'),
                      ('H', 'Top3 個股'), ('I', 'Top3 金額'),
                      ('J', 'Top3 合計%')]:
        cell = ws[f'{h_col}{row}']
        cell.value = h
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
    row += 1

    # 算每個 master 的 today's stocks
    # v3.65.0: 用戶要求 Dashboard 全只顯示個股, 排除 ETF
    master_stocks = {}
    for b in branches_data:
        m = b.get('master')
        if not m: continue
        for s in (b.get('buys') or []):
            code = s.get('code')
            if not code: continue
            if _is_excluded_by_market_type(s):
                continue
            master_stocks.setdefault(m, {})
            key = (code, s.get('name', ''))
            master_stocks[m][key] = master_stocks[m].get(key, 0) + (s.get('buy_amt') or 0)

    # 每 master sorted top 3
    rows_data = []
    for master, stocks in master_stocks.items():
        sorted_s = sorted(stocks.items(), key=lambda kv: -kv[1])
        total = sum(stocks.values())
        rows_data.append({
            'master': master,
            'total': total,
            'top': sorted_s[:3],
        })
    rows_data.sort(key=lambda x: -x['total'])

    start_data = row
    for d in rows_data[:30]:
        # v3.66.2: Master cell 套 MASTER_BLOCK_COLORS body 色 (跟 B/C 一致)
        colors = MASTER_BLOCK_COLORS.get(d['master']) or DEFAULT_MASTER_COLOR
        top3_amt = sum(amt for _, amt in d['top'])
        top3_pct = (top3_amt / d['total']) if d['total'] > 0 else 0

        master_label = d['master']
        if top3_pct >= 0.80:
            master_label = f'🔥 {master_label}'   # 押大注集中
        c_master = ws.cell(row, 2, master_label)
        c_master.font = Font(name='Noto Sans TC', size=11, bold=True)
        c_master.fill = PatternFill('solid', fgColor=colors['body'])

        ws.cell(row, 3, round(d['total'] / 10, 0)).number_format = '#,##0'
        for i, ((code, name), amt) in enumerate(d['top']):
            col_name = chr(ord('D') + i * 2)
            col_amt = chr(ord('E') + i * 2)
            ws.cell(row, ord(col_name) - 64, f"{name}({code})")
            c_amt = ws.cell(row, ord(col_amt) - 64, round(amt / 10, 0))
            c_amt.number_format = '#,##0'

        c_pct = ws.cell(row, 10, top3_pct)   # col J
        c_pct.number_format = '0.0%'
        if top3_pct >= 0.80:
            c_pct.font = Font(name='Noto Sans TC', size=11, bold=True, color='FFC62828')
        row += 1
    if row == start_data:
        ws.cell(row, 2, '今日無 master 有買進資料')
        ws.merge_cells(f'B{row}:J{row}'); row += 1
    else:
        # v3.66.6 Phase 2.2:
        # C 今日總買 加 data bar (深綠 = 規模)
        _try_add_data_bar(ws, f'C{start_data}:C{row-1}', 'FF66BB6A')
        # J Top3 集中度 加 data bar (橘紅 = 集中越多越警示)
        _try_add_data_bar(ws, f'J{start_data}:J{row-1}', 'FFFB923C')
    return row


def _build_tldr_action_cards(ws, branches_data, all_branches, trade_date, data_dir):
    """v3.66.4 Phase 2.1: TL;DR (Row 3) + Action (Row 4) 首屏摘要.

    用戶設計 spec:
      - 一句話 TL;DR (single-line summary): 共識/Q5/E/F/J/H 6 個 hot 指標
      - Action card 含 進場關注 / 避開 / 訊號強度
    """
    # ── 計算 6 個 hot 指標 ──
    consensus_stocks = _compute_consensus_count(branches_data)
    c_count = len(consensus_stocks)
    # v3.66.5 bug fix: 用 Section 0 相同排序 (-total_net_amt, -master_count, -branch_count)
    # 修前 unsorted dict order → top 3 跟 Section 0 顯示不一致
    top3_consensus = sorted(consensus_stocks,
                             key=lambda x: (-x['total_net_amt'], -x['master_count'],
                                            -x['branch_count']))[:3]

    # Q5
    daily_signal = _read_json_safely(data_dir / 'daily_signal.json')
    md = (daily_signal or {}).get('market_direction') or {}
    direction = md.get('direction') or '?'
    confidence = md.get('confidence_pct') or 0
    arrow = '↑' if direction == '偏多' else ('↓' if direction == '偏空' else '↕')

    # E anomaly (tracked, top 10)
    signals = _read_json_safely(data_dir / 'daily_trading_signals.json')
    e_count = len([a for a in ((signals or {}).get('anomalies') or [])
                    if _is_tracked_master(a.get('master'))][:10])

    # F hot (≥10 days, 非 ETF, tracked)
    f_hot = sum(1 for a in ((signals or {}).get('accumulations') or [])
                 if _is_tracked_master(a.get('master'))
                 and not (a.get('stock_code') or '').startswith('00')
                 and (a.get('consecutive_days', 0) or 0) >= 10)

    # J hot (≥80% concentration)
    j_master_stocks = {}
    for b in branches_data:
        m = b.get('master')
        if not m: continue
        for s in (b.get('buys') or []):
            if _is_excluded_by_market_type(s): continue
            code = s.get('code')
            if not code: continue
            j_master_stocks.setdefault(m, {})
            k = (code, s.get('name', ''))
            j_master_stocks[m][k] = j_master_stocks[m].get(k, 0) + (s.get('buy_amt') or 0)
    j_hot = 0
    for m, stocks in j_master_stocks.items():
        total = sum(stocks.values())
        if total <= 0: continue
        top3 = sorted(stocks.values(), reverse=True)[:3]
        if (sum(top3) / total) >= 0.80:
            j_hot += 1

    # H hot (≥1000x ratio)
    short_lending = _read_json_safely(data_dir / 'short_lending.json')
    h_hot = 0
    for item in ((short_lending or {}).get('top_borrow_sell') or [])[:15]:
        try:
            r = float(item.get('borrow_vs_short_ratio'))
            if r >= 1000:
                h_hot += 1
        except (TypeError, ValueError):
            pass

    # I 今日除權息 (避開用)
    dividend = _read_json_safely(data_dir / 'dividend_calendar.json')
    today_ex_list = [i for i in ((dividend or {}).get('upcoming_30d') or [])
                      if i.get('ex_date') == trade_date]
    today_ex_str = ', '.join(i.get('code', '') for i in today_ex_list[:3])

    # ── TL;DR 一句話 ──
    tldr = (f"🎯 {c_count} 強共識 / "
            f"Q5 {arrow} {direction} {confidence}% / "
            f"E {e_count} 異常 / "
            f"F {f_hot} 長期 / "
            f"J {j_hot} 集中 / "
            f"H {h_hot} 借券壓力")

    # ── v3.70.0 Phase 3.2 落地: Action 進場分級 ──
    # quad 命中股 (Phase 3.2 三訊號齊聚, 預期 78.9% alpha) 優先, 其次一般共識.
    quad_info = _compute_quad_picks(consensus_stocks, data_dir)
    if quad_info['quad_picks']:
        # quad 優先 — 最多列前 3 quad picks
        quad_codes = ' / '.join(p['code'] for p in quad_info['quad_picks'][:3])
        action_buy = f"🎯 quad 進場 (78.9% alpha): {quad_codes}"
        if len(quad_info['quad_picks']) > 3:
            action_buy += f" +{len(quad_info['quad_picks'])-3}"
        # 其他共識 (非 quad) 列為「📌 一般共識」(top 3 by net_amt 內排除 quad)
        non_quad_top = [s for s in top3_consensus
                        if s['code'] not in quad_info['quad_codes']][:3]
        if non_quad_top:
            other_codes = ' / '.join(s['code'] for s in non_quad_top)
            action_buy += f"  |  📌 一般共識: {other_codes}"
    elif c_count > 0 and top3_consensus:
        # 無 quad → 退回原邏輯但加分級警示
        codes = ' / '.join(s['code'] for s in top3_consensus)
        if quad_info['q5_direction'] == '偏多':
            action_buy = f"📌 一般共識 {codes} (Q5 偏多但無 master 量爆 — alpha 未啟動)"
        else:
            action_buy = f"📌 一般共識 {codes} (Q5 {quad_info['q5_direction'] or '無'}, 中性信號)"
    else:
        action_buy = "進場關注: 今日無強共識"

    # v3.66.5 bug fix: 顯示真實檔數避免誤導
    # v3.71.7: 整合處置股 (attstock.tw API) — 兩段提示
    avoid_parts = []
    if today_ex_list:
        n_total = len(today_ex_list)
        suffix = ' ...' if n_total > 3 else ''
        avoid_parts.append(f"除權息 {n_total} 檔: {today_ex_str}{suffix}")
    disp_data = _read_json_safely(data_dir / 'disposal_attstock.json')
    if disp_data:
        n_in = disp_data.get('count_in_disposal', 0)
        n_pending = disp_data.get('count_pending_1d', 0)
        if n_in > 0:
            codes_in = disp_data.get('codes_in_disposal') or []
            codes_str = '/'.join(codes_in[:3]) + ('...' if len(codes_in) > 3 else '')
            avoid_parts.append(f"處置中 {n_in} 檔: {codes_str}")
        if n_pending > 0:
            codes_p = disp_data.get('codes_pending_1d') or []
            codes_str = '/'.join(codes_p[:3]) + ('...' if len(codes_p) > 3 else '')
            avoid_parts.append(f"明日恐處置 {n_pending} 檔: {codes_str}")
    if avoid_parts:
        action_avoid = "避開 — " + " | ".join(avoid_parts)
    else:
        action_avoid = "避開: 今日無除權息+處置股"

    # 訊號強度 (基於 Q5 confidence + E/F/J/H hot signals)
    hot_total = e_count + f_hot + j_hot + h_hot
    if (direction == '偏多' and confidence >= 60) or (direction == '偏空' and confidence <= 40):
        sig_strength = f"訊號: 強{direction} (Q5 {confidence}%, hot={hot_total})"
    elif hot_total >= 5:
        sig_strength = f"訊號: 中性但 hot 多 ({hot_total} 個信號 — 留意)"
    else:
        sig_strength = f"訊號: 中性 (建議觀望)"

    action = f"💡 {action_buy}  |  {action_avoid}  |  {sig_strength}"

    # ── Render TL;DR (Row 3) ──
    ws.merge_cells('B3:N3')
    c_tldr = ws['B3']
    c_tldr.value = tldr
    c_tldr.font = Font(name='Noto Sans TC', size=12, bold=True, color='FF374151')
    c_tldr.fill = PatternFill('solid', fgColor='FFFEF3C7')   # 淡黃
    c_tldr.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 24

    # ── Render Action (Row 4) ──
    ws.merge_cells('B4:N4')
    c_act = ws['B4']
    c_act.value = action
    c_act.font = Font(name='Noto Sans TC', size=11, italic=True, color='FF4B5563')
    c_act.fill = PatternFill('solid', fgColor='FFF3F4F6')    # 淡灰
    c_act.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 22


def _build_section_risk(ws, data_dir, start_row, trade_date: Optional[str] = None):
    """Section G+H+I: 注意股 + 借券 + 除權息. Returns next row.

    v3.66.1 時間正確性修補:
      - I 除權息過濾過期 ex_date (對 6/24 看 Excel 不該還顯示 6/23 已除權的股)
      - G/H/I header 顯示 applicable_date / fetched_at, 用戶知道資料新鮮度
      - trade_date 作為 "今天" 比對基準 (重生歷史 Excel 時亦正確)
    """
    hdr_font = Font(name='Noto Sans TC', size=10, bold=True)
    sub_font = Font(name='Noto Sans TC', size=11, bold=True)

    # 「今天」基準: 用 trade_date (重生 6/18 時 = 20260618; 即時 dispatch = 當日)
    today_yyyymmdd = trade_date or ''

    row = start_row + 1

    # ── G 注意股 ──
    attention = _read_json_safely(data_dir / 'attention_map.json')
    g_apply = (attention or {}).get('applicable_date') or '?'
    g_stale_days = (attention or {}).get('stale_days')
    g_stale_tag = f" / 距 trade_date {g_stale_days} 天" if g_stale_days else ""
    _section_header(ws, row, f"▍ G. 注意股 (資料日 {g_apply}{g_stale_tag})",
                     color='FFFB923C'); row += 1
    for h_col, h in [('B', '代號'), ('C', '名稱'), ('D', '累計次數'), ('E', '收盤價'), ('F', '本益比')]:
        cell = ws[f'{h_col}{row}']; cell.value = h; cell.font = hdr_font
        cell.fill = _summary_fill('FFFEF3C7')
        cell.alignment = Alignment(horizontal='center')
    row += 1
    by_code = (attention or {}).get('by_code') or {}
    g_data_start = row
    if by_code:
        for code, info in list(by_code.items())[:15]:
            ws.cell(row, 2, code); ws.cell(row, 3, info.get('name', '—'))
            ws.cell(row, 4, info.get('cumulative_count', 0))
            ws.cell(row, 5, info.get('close', '—'))
            ws.cell(row, 6, info.get('pe', '—'))
            row += 1
        # v3.66.6 Phase 2.2: 累計次數加 data bar (金色)
        _try_add_data_bar(ws, f'D{g_data_start}:D{row-1}', 'FFFFB300')
        # v3.67.0 Phase 2.6: G 套 zebra stripes (cols B-F)
        _zebra_stripes(ws, g_data_start, row - 1, col_start='B', col_end='F')
    else:
        # v3.66.3: empty state 加 emoji 友善訊息
        ws.cell(row, 2, '✅ 今日無新增注意股 (市場無異常波動標的)')
        ws.merge_cells(f'B{row}:F{row}')
        ws.cell(row, 2).font = Font(name='Noto Sans TC', size=11, italic=True, color='FF10B981')
        row += 1

    row += 1

    # ── H 借券 ──
    short_lending = _read_json_safely(data_dir / 'short_lending.json')
    h_apply = (short_lending or {}).get('applicable_date') or '?'
    _section_header(ws, row,
                     f"▍ H. 借券賣出 Top 15 (機構級反向力量, 資料日 {h_apply})",
                     color='FFDC2626'); row += 1
    for h_col, h in [('B', '代號'), ('C', '名稱'), ('D', '借券張數'), ('E', '融券張數'), ('F', 'ratio')]:
        cell = ws[f'{h_col}{row}']; cell.value = h; cell.font = hdr_font
        cell.fill = _summary_fill('FFFEE2E2')
        cell.alignment = Alignment(horizontal='center')
    row += 1
    top_borrow = ((short_lending or {}).get('top_borrow_sell') or [])
    h_data_start = row
    if top_borrow:
        # v3.66.3: ratio ≥1000x 標 🔴 (極端機構壓力), 紅字粗體
        hot_font_red = Font(name='Noto Sans TC', size=11, bold=True, color='FFC62828')
        RATIO_HOT = 1000.0
        for item in top_borrow[:15]:
            ratio = item.get('borrow_vs_short_ratio')
            try:
                ratio_num = float(ratio) if ratio is not None else None
            except (TypeError, ValueError):
                ratio_num = None
            is_hot = ratio_num is not None and ratio_num >= RATIO_HOT

            code_label = item.get('code', '—')
            if is_hot:
                code_label = f'🔴 {code_label}'
            c_code = ws.cell(row, 2, code_label)
            if is_hot:
                c_code.font = hot_font_red

            ws.cell(row, 3, item.get('name', '—'))
            ws.cell(row, 4, item.get('borrow_sell_lot', 0)).number_format = '#,##0'
            ws.cell(row, 5, item.get('short_sell_lot', 0)).number_format = '#,##0'
            c_ratio = ws.cell(row, 6, ratio_num if ratio_num is not None else '—')
            if ratio_num is not None:
                c_ratio.number_format = '#,##0.0'
            if is_hot:
                c_ratio.font = hot_font_red
            row += 1
        # v3.66.6 Phase 2.2: 借券張數加 data bar (深紅 = 壓力)
        _try_add_data_bar(ws, f'D{h_data_start}:D{row-1}', 'FFEF5350')
        # ratio 也加 (橘紅) — 1000x hot 那筆會超出條 max 區
        _try_add_data_bar(ws, f'F{h_data_start}:F{row-1}', 'FFFB923C')
        # v3.67.0 Phase 2.6: H 套 zebra stripes (cols B-F)
        _zebra_stripes(ws, h_data_start, row - 1, col_start='B', col_end='F')
    else:
        ws.cell(row, 2, '今日無借券資料')
        ws.merge_cells(f'B{row}:F{row}'); row += 1

    row += 1

    # ── I 除權息 ──
    # v3.66.1 修 time bug: 過濾過期 ex_date (≥ trade_date), 避免顯示已除權的股
    dividend = _read_json_safely(data_dir / 'dividend_calendar.json')
    upcoming_raw = (dividend or {}).get('upcoming_30d') or []
    if today_yyyymmdd:
        upcoming = [i for i in upcoming_raw if (i.get('ex_date') or '') >= today_yyyymmdd]
    else:
        upcoming = upcoming_raw
    _section_header(ws, row,
                     f"▍ I. 未來 30 天除權息 (有效 {len(upcoming)} 檔, 已剔除過期)",
                     color='FFFBBF24'); row += 1
    for h_col, h in [('B', '除權息日'), ('C', '代號'), ('D', '名稱'), ('E', '類型'), ('F', '現金股利')]:
        cell = ws[f'{h_col}{row}']; cell.value = h; cell.font = hdr_font
        cell.fill = _summary_fill('FFFEF3C7')
        cell.alignment = Alignment(horizontal='center')
    row += 1
    if upcoming:
        i_data_start = row
        for item in upcoming[:15]:
            ws.cell(row, 2, item.get('ex_date', '—'))
            ws.cell(row, 3, item.get('code', '—'))
            ws.cell(row, 4, item.get('name', '—'))
            ws.cell(row, 5, item.get('type', '—'))
            ws.cell(row, 6, item.get('cash_dividend', '—'))
            row += 1
        # v3.67.0 Phase 2.6: I 套 zebra stripes (cols B-F)
        _zebra_stripes(ws, i_data_start, row - 1, col_start='B', col_end='F')
    else:
        ws.cell(row, 2, '未來 30 天無除權息')
        ws.merge_cells(f'B{row}:F{row}'); row += 1
    return row


def build_dashboard_sheet(ws: "Worksheet", branches_data: List[Dict], trade_date: str,
                            data_dir: Optional[Path] = None,
                            update_timeseries: bool = True):
    """v3.62.1: 把 E1-E4 4 個 section 全部寫到單一 sheet (用戶要求).
    順序: A 規模 → B Top master → C Top stocks → D 籌碼溫度
        → E 異常警報 → F 連續囤貨 → G 注意股 → H 借券 → I 除權息
    """
    data_dir = data_dir or Path('data')
    title_fill = _summary_fill('FF1F2A48')
    title_font = _summary_font_header()

    # v3.63.2: 嚴格只保留追蹤清單內的大戶 (MASTER_MAPPING)
    # v3.64.3: 保留全市場 branches 供 Section A 計算「追蹤佔比 vs 全市場」
    all_branches = branches_data
    branches_data = _filter_tracked_branches(branches_data)

    for col, w in [('A', 4), ('B', 22), ('C', 18), ('D', 22), ('E', 16),
                    ('F', 22), ('G', 18), ('H', 22), ('I', 16)]:
        ws.column_dimensions[col].width = w

    # ── 大標題 ──
    ws.merge_cells('B2:N2')
    c = ws['B2']
    c.value = (f"📋 Chip Radar 今日 Dashboard — "
                f"{trade_date[:4]}/{trade_date[4:6]}/{trade_date[6:]} "
                f"(追蹤 {len(TRACKED_MASTERS)} 位大戶)")
    c.font = title_font
    c.fill = title_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 30

    # ── v3.66.4 Phase 2.1: TL;DR + Action card (首屏 5 秒決策摘要) ──
    _build_tldr_action_cards(ws, branches_data, all_branches, trade_date, data_dir)

    # ── 各 section ──
    row = 6   # v3.66.4: 從 row 4 → row 6 (讓 TL;DR + Action)
    # v3.63.2: ★ Section 0 — 今日共同買超 (置於最前, 使用者最關注)
    row = _build_section_consensus(ws, branches_data, data_dir, row, trade_date=trade_date)
    row = _build_section_summary(ws, branches_data, trade_date, data_dir, row,
                                   all_branches=all_branches,
                                   update_timeseries=update_timeseries)
    row = _build_section_alerts(ws, data_dir, row)
    row = _build_section_accumulation(ws, data_dir, row)
    row = _build_section_pivot(ws, branches_data, row)   # v3.63.0 E7 Pivot
    row = _build_section_risk(ws, data_dir, row, trade_date=trade_date)

    # v3.66.4: freeze pane 延伸到 row 5 (TL;DR + Action 永遠看得到)
    ws.freeze_panes = 'A6'


# v3.62.0 → v3.62.1 backward compat: 舊 builder name 保留但呼叫 dashboard
def build_summary_sheet(ws, branches_data, trade_date, data_dir=None):
    """DEPRECATED v3.62.1: 用戶要求合 dashboard. 此 fn 仍 alias 給 build_dashboard_sheet."""
    return build_dashboard_sheet(ws, branches_data, trade_date, data_dir)


def apply_pnl_color_scale(ws: "Worksheet", first_row: int, last_row: int, col_letter: str = 'L'):
    """E5: 損益欄 L 加紅綠色階 conditional formatting.
    v3.63.1: 端點改深紅/深綠 (C62828 / 2E7D32) 與白字粗體配對, 手機/筆電都高對比."""
    if last_row < first_row:
        return
    rule = ColorScaleRule(
        start_type='num', start_value=-10, start_color='FFC62828',
        mid_type='num', mid_value=0, mid_color='FFFFFFFF',
        end_type='num', end_value=10, end_color='FF2E7D32',
    )
    ws.conditional_formatting.add(f'{col_letter}{first_row}:{col_letter}{last_row}', rule)


def _update_monthly_workbook(monthly_path: Path, branches_data: List[Dict],
                              trade_date: str):
    """v3.31.0: 開啟既有月檔 (若有) 或新建, add/update 該日 sheet, save back.
    sheet 名 = trade_date (YYYYMMDD), 同日重跑會覆寫該 sheet, sheets 按日期 desc 排序."""
    if monthly_path.exists():
        try:
            wb = load_workbook(str(monthly_path))
        except Exception as e:
            print(f"  [Excel] 月檔 unreadable, recreating: {e}")
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
    else:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    # 若該日 sheet 已存在 → 移除 (重跑同日 → 覆寫)
    if trade_date in wb.sheetnames:
        wb.remove(wb[trade_date])

    # 新建該日 sheet (build_day_sheet 在 ws 內 render 老闆版)
    ws = wb.create_sheet(title=trade_date)
    total_rows = build_day_sheet(ws, branches_data, trade_date)
    # v3.64.0: 拿掉 L 欄 ColorScaleRule (用戶反饋: master block 紅/淡綠 fill +
    # 色階重疊變糊, 改用 _font_pnl_pos/neg 字色 (深紅/深綠) 直接清楚)

    # v3.62.1: 清舊版本殘留 (從 v3.62.0 升級時) — 4 個舊 sheet 移除
    for legacy_name in LEGACY_ENRICHMENT_NAMES:
        if legacy_name in wb.sheetnames:
            wb.remove(wb[legacy_name])

    # v3.62.1: rebuild 1 個 dashboard sheet (E1-E4 全 section)
    data_dir = monthly_path.parent.parent if monthly_path.parent.name == 'reports' else Path('data')
    if DASHBOARD_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[DASHBOARD_SHEET_NAME])
    dashboard_ws = wb.create_sheet(title=DASHBOARD_SHEET_NAME)
    try:
        build_dashboard_sheet(dashboard_ws, branches_data, trade_date, data_dir)
    except Exception as _be:
        print(f"  [Excel] dashboard sheet build 失敗: {type(_be).__name__}: {_be}")

    # v3.67.1 Phase 2.7: 手機摘要 sheet (Dashboard 後第 2 個)
    if MOBILE_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[MOBILE_SHEET_NAME])
    mobile_ws = wb.create_sheet(title=MOBILE_SHEET_NAME)
    try:
        build_mobile_summary_sheet(mobile_ws, branches_data, trade_date, data_dir)
    except Exception as _be:
        print(f"  [Excel] mobile summary sheet build 失敗: {type(_be).__name__}: {_be}")

    # v3.70.2 Phase 3.2 持續性追蹤: Quad 實戰追蹤 sheet (Dashboard 後第 3 個)
    if QUAD_TRACK_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[QUAD_TRACK_SHEET_NAME])
    quad_ws = wb.create_sheet(title=QUAD_TRACK_SHEET_NAME)
    try:
        build_quad_track_sheet(quad_ws, data_dir)
    except Exception as _be:
        print(f"  [Excel] quad track sheet build 失敗: {type(_be).__name__}: {_be}")

    # v3.70.3 Phase 3.2 失效歸因: Quad 失效歸因 sheet (Dashboard 後第 4 個)
    if QUAD_FAIL_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[QUAD_FAIL_SHEET_NAME])
    fail_ws = wb.create_sheet(title=QUAD_FAIL_SHEET_NAME)
    try:
        build_quad_failure_sheet(fail_ws, data_dir)
    except Exception as _be:
        print(f"  [Excel] quad failure sheet build 失敗: {type(_be).__name__}: {_be}")

    # v3.71.18 L2: Pinned master 追蹤 sheet (Dashboard 後第 5 個)
    if PINNED_TRACK_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[PINNED_TRACK_SHEET_NAME])
    pinned_ws = wb.create_sheet(title=PINNED_TRACK_SHEET_NAME)
    try:
        build_pinned_track_sheet(pinned_ws, branches_data, data_dir)
    except Exception as _be:
        print(f"  [Excel] pinned track sheet build 失敗: {type(_be).__name__}: {_be}")

    # 排序: dashboard → mobile → quad track → quad fail → pinned → 日期 sheets desc
    enrichment = [DASHBOARD_SHEET_NAME, MOBILE_SHEET_NAME,
                  QUAD_TRACK_SHEET_NAME, QUAD_FAIL_SHEET_NAME,
                  PINNED_TRACK_SHEET_NAME]
    other_sheets = sorted([s for s in wb.sheetnames if s not in enrichment],
                            reverse=True)
    order = [s for s in enrichment if s in wb.sheetnames] + other_sheets
    wb._sheets = [wb[name] for name in order]

    wb.save(str(monthly_path))


# ============================================================
#  Multi-sheet latest.xlsx (legacy v3.30.x, retained for backfill/reference)
# ============================================================

def _update_latest_multi_sheet(latest_path: Path, branches_data: List[Dict],
                                trade_date: str, max_sheets: int = 30):
    """
    Update latest.xlsx to include current trade_date as a sheet.
    Keep only the most recent `max_sheets` sheets (by sheet name desc).
    Newest sheet is set as the active one.
    """
    if latest_path.exists():
        try:
            wb = load_workbook(str(latest_path))
        except Exception as e:
            print(f"  [Excel] latest.xlsx unreadable, recreating: {e}")
            wb = Workbook()
            # remove default sheet
            for s in list(wb.sheetnames):
                del wb[s]
    else:
        wb = Workbook()
        for s in list(wb.sheetnames):
            del wb[s]

    # Remove existing sheet with same name (will rebuild)
    if trade_date in wb.sheetnames:
        del wb[trade_date]

    # Create new sheet for trade_date
    ws = wb.create_sheet(title=trade_date)
    build_day_sheet(ws, branches_data, trade_date)

    # Sort sheets by name desc; trim to max_sheets
    sheet_names_sorted = sorted(wb.sheetnames, reverse=True)
    keep = sheet_names_sorted[:max_sheets]
    drop = sheet_names_sorted[max_sheets:]
    for n in drop:
        del wb[n]

    # Reorder so newest first
    wb._sheets = [wb[n] for n in keep]
    wb.active = 0  # show newest sheet on open

    wb.save(str(latest_path))


# ============================================================
#  Main entry point
# ============================================================

def generate_excel_report(branches_data: List[Dict], trade_date: str,
                           output_dir: str = "data/reports") -> Optional[str]:
    """
    Generate Excel daily report (mimics manual 「分點觀察」 layout).

    Args:
      branches_data: list of branch dicts from crawler (with buys/sells)
      trade_date: trading day in YYYYMMDD format (e.g. '20260508')
      output_dir: output directory (default 'data/reports')

    Returns:
      file path of the daily snapshot on success, None on failure.
    """
    if not OPENPYXL_AVAILABLE:
        print("  [Excel] openpyxl not installed, skipping")
        return None

    if not branches_data:
        print("  [Excel] no branch data, skipping")
        return None

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if len(trade_date) == 8 and trade_date.isdigit():
        readable_date = f"{trade_date[:4]}-{trade_date[4:6]}-{trade_date[6:8]}"
    else:
        readable_date = trade_date

    # v3.31.0: 月檔模式 ─ 一個月一份 chip_radar_YYYY-MM.xlsx (取代單日檔)
    #   latest.xlsx 永遠 = 當月月檔的 copy
    #   crawler 主流程每次 daily-full 跑完: 開啟月檔 → add/update 該日 sheet → save → copy latest
    year_month = readable_date[:7]   # "2026-06"
    monthly_path = out_dir / f"chip_radar_{year_month}.xlsx"
    latest_path = out_dir / "latest.xlsx"

    valid_count = sum(1 for b in branches_data if b.get("buys") or b.get("sells"))
    print(f"  [Excel] generating v3.31 monthly ({valid_count} branches, date {readable_date}, "
          f"month {year_month})")

    try:
        # 1. 月檔: 開啟既有 (若有) 或新建, add/update 該日 sheet
        _update_monthly_workbook(monthly_path, branches_data, trade_date)

        # 2. latest.xlsx = 當月月檔的 copy (前端下載按鈕指向不變)
        import shutil
        shutil.copy2(str(monthly_path), str(latest_path))

        # 3. Update README index
        _update_reports_readme(out_dir)

        size_kb = monthly_path.stat().st_size / 1024
        latest_size_kb = latest_path.stat().st_size / 1024
        sheet_count = len(load_workbook(str(monthly_path), read_only=True).sheetnames)
        print(f"  [Excel] OK")
        print(f"     {monthly_path.name} ({size_kb:.1f} KB, {sheet_count} daily sheets)")
        print(f"     latest.xlsx ({latest_size_kb:.1f} KB, = 當月月檔)")
        return str(monthly_path)

    except Exception as e:
        print(f"  [Excel] FAILED: {e}")
        import traceback
        traceback.print_exc()
        return None


def _update_reports_readme(reports_dir: Path):
    """Update reports/README.md with auto-generated file index."""
    try:
        excel_files = sorted(
            reports_dir.glob("chip_radar_*.xlsx"),
            key=lambda p: p.name,
            reverse=True,
        )

        readme = reports_dir / "README.md"
        master_count = len(MASTER_MAPPING)
        branch_count = sum(len(m["branches"]) for m in MASTER_MAPPING)

        lines = [
            "# Chip Radar 老闆版 Excel 日報",
            "",
            "由 `excel_report.py` v3.26 自動生成,模仿手動版「分點觀察」格式。",
            "",
            "## 最新檔案",
            "",
            "- [**latest.xlsx**](./latest.xlsx) — 多 sheet, 最近 30 個交易日",
            "  - 每個 sheet 命名 = `YYYYMMDD`,開啟時顯示最新一日",
            "",
            "## 結構",
            "",
            f"- {master_count} 位高手 / {branch_count} 個分點 slot (含跨高手共用分點)",
            "- 每分點固定 10 列 (不足以空白填補)",
            "- 12 欄: 高手 / 分點 / 代號 / 標的 / 買進(張) / 賣出(張) / "
            "買進(萬元) / 賣出(萬元) / 淨買差(萬元) / 買均 / 賣均 / 損益(萬)",
            "- L 欄公式: `=F*(K-J)` (賣出張數 × (賣均-買均)),負值紅字",
            "",
            "## v3.30.5 風格分流規則 (僅蔣承翰用漲停法)",
            "",
            "Excel 抓取法依 master 切換:",
            "",
            "| Master | Top N 資料源 |",
            "|---|---|",
            "| ⭐ 蔣承翰 (隔日沖) | **今日漲停股 by 買進金額** (漲跌幅 ≥ 9.5%) |",
            "| 其餘所有 master | 全部個股 by 買進金額 (淨買超 Top N) |",
            "",
            "蔣承翰今天若沒搶任何漲停股 → 整列空白 (不 fallback,維持風格純度)。",
            "(v3.26~v3.30.4 原為所有隔日沖/當沖 master 都用漲停法;v3.30.5 依使用者要求縮為僅蔣承翰)",
            "",
            "## 每日歷史",
            "",
        ]

        if excel_files:
            lines.append(f"近 {min(len(excel_files), 30)} 個交易日 (共 {len(excel_files)} 個檔案):")
            lines.append("")
            lines.append("| 日期 | 檔案 | 大小 |")
            lines.append("|------|------|------|")
            for p in excel_files[:30]:
                size_kb = p.stat().st_size / 1024
                # 從檔名抽日期
                stem = p.stem  # chip_radar_2026-05-08
                date_part = stem.replace("chip_radar_", "")
                lines.append(f"| {date_part} | [{p.name}](./{p.name}) | {size_kb:.1f} KB |")
            if len(excel_files) > 30:
                lines.append("")
                lines.append(f"…另有 {len(excel_files) - 30} 個更舊檔案")

        lines.extend([
            "",
            "---",
            "",
            f"*Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}*",
        ])

        readme.write_text("\n".join(lines), encoding="utf-8")
    except Exception as e:
        print(f"  [Excel] README update failed: {e}")
