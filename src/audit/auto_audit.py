"""auto_audit.py — v3.29.3 W2 自動化 daily audit (V1 邏輯內建)

每次 daily-full crawler 跑完 Excel 後, crawler.py 立刻呼叫 run_audit(),
產 data/daily_audit.json. trigger_chip_radar.ps1 讀此檔 → Windows toast 通知異常.

audit 內容 (V1 全分點掃描的程式化版):
  - 總 row / 淨買 / 淨賣 / net=0 / 限漲標 / 反推估算 / 提示行
  - verdict:
      PASS — 淨賣 = 0 AND net=0 = 0
      WARN — 淨賣 / net=0 / 反推估算 > 預期閾值
      FAIL — 淨賣 row > 0 (v3.29.1 後應永遠為 0)
  - anomalies: detail rows

設計:
  - 純 Python stdlib + openpyxl, 不打外部 API (秒級執行)
  - 失敗不影響主流程 (try/except 包覆)

Output: data/daily_audit.json
"""

import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ════════════════════════════════════════════════════════════════════
#  Excel parsing (與 excel_full_audit.py 相同邏輯, 但 module 化)
# ════════════════════════════════════════════════════════════════════

def _parse_excel(xlsx_path: Path):
    """Returns (sheet_title, list_of_row_dicts)."""
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=False)
    ws = wb.active
    rows = []
    current_master = ''
    current_branch = ''
    current_branch_code = ''
    for row_idx in range(1, ws.max_row + 1):
        cells = [ws.cell(row_idx, c).value for c in range(1, 13)]
        a, b, c, d, e, f, g, h, i, j, k, l = cells
        if a == '高手':
            continue
        if a:
            current_master = a
        if b == '分點' or b == '常下分點':
            continue
        if b and c:
            current_branch = b
            current_branch_code = c
        if d and isinstance(d, str) and ('(' in d or '⚪' in d):
            rows.append({
                'row': row_idx,
                'master': current_master,
                'branch': current_branch,
                'branch_code': current_branch_code,
                'stock_label': d,
                'buy_lot': e if isinstance(e, (int, float)) else 0,
                'sell_lot': f if isinstance(f, (int, float)) else 0,
                'buy_amt': g if isinstance(g, (int, float)) else 0,
                'sell_amt': h if isinstance(h, (int, float)) else 0,
                'net_amt': i if isinstance(i, (int, float)) else 0,
                'buy_avg': j if isinstance(j, (int, float)) else 0,
                'sell_avg': k if isinstance(k, (int, float)) else 0,
            })
    return ws.title, rows


def _classify(row):
    flags = []
    sl = row['stock_label']
    if isinstance(sl, str) and '⚪' in sl:
        flags.append('NOTICE')
        return flags
    if not (row['buy_lot'] or row['sell_lot'] or row['buy_amt'] or row['sell_amt']):
        flags.append('EMPTY')
        return flags
    if row['net_amt'] > 0:
        flags.append('NET_BUY')
    elif row['net_amt'] < 0:
        flags.append('NET_SELL')
    else:
        flags.append('NET_ZERO')
    if '▲' in str(sl):
        flags.append('LIMIT_UP_TAG')
    if (row['buy_avg'] and row['sell_avg']
            and abs(row['buy_avg'] - row['sell_avg']) < 0.01
            and row['buy_avg'] > 0):
        flags.append('REVERSE_EST')
    return flags


# ════════════════════════════════════════════════════════════════════
#  Verdict logic
# ════════════════════════════════════════════════════════════════════

def _judge_verdict(stats, anomalies):
    """Verdict rules (v3.29.3 baseline):
      FAIL if 淨賣 row > 0 (v3.29.1 後應永遠 = 0; 大於 0 代表 filter 失效)
      WARN if net=0 > 5 OR 反推估算 > 20 (邊界 case, 觀察用)
      PASS otherwise
    """
    if stats['net_sell'] > 0:
        return 'FAIL', f"❌ 偵測到 {stats['net_sell']} 筆淨賣污染 (v3.29.1 過濾失效)"
    if stats['net_zero'] > 5:
        return 'WARN', f"⚠️ net=0 row {stats['net_zero']} 筆 (邊界 case 偏多)"
    if stats['reverse_est'] > 20:
        return 'WARN', f"⚠️ 反推估算 {stats['reverse_est']} 筆 (高價股盲點觸發偏多)"
    if stats['total_real'] == 0:
        return 'FAIL', "❌ Excel 完全沒有實際個股 row"
    return 'PASS', (f"✅ {stats['total_real']} row 全淨買, "
                    f"{stats['limit_up_tag']} 限漲標, {stats['reverse_est']} 反推估算")


# ════════════════════════════════════════════════════════════════════
#  Main run_audit
# ════════════════════════════════════════════════════════════════════

def run_audit(data_dir: Path, xlsx_path: Path = None) -> dict:
    """Run V1 audit on the given xlsx (預設 data/reports/latest.xlsx).
    Returns the report dict (also call save_audit_report to persist).
    """
    if xlsx_path is None:
        xlsx_path = data_dir / 'reports' / 'latest.xlsx'

    report = {
        'audit_run_at': datetime.now().isoformat(),
        'xlsx_path': str(xlsx_path),
        'tool': 'auto_audit v3.29.3',
    }

    if not HAS_OPENPYXL:
        report['overall_verdict'] = 'SKIP'
        report['summary'] = '⏭️ openpyxl 未安裝, 跳過 audit'
        return report

    if not xlsx_path.exists():
        report['overall_verdict'] = 'SKIP'
        report['summary'] = f'⏭️ {xlsx_path.name} 不存在'
        return report

    try:
        sheet_title, rows = _parse_excel(xlsx_path)
    except Exception as e:
        report['overall_verdict'] = 'ERROR'
        report['summary'] = f'❌ Excel 讀取失敗: {e}'
        return report

    # 統計
    stats = {
        'total_rows': len(rows),
        'total_real': 0,
        'net_buy': 0,
        'net_sell': 0,
        'net_zero': 0,
        'empty': 0,
        'notice': 0,
        'limit_up_tag': 0,
        'reverse_est': 0,
    }
    anomalies = {
        'net_sell_rows': [],
        'net_zero_rows': [],
    }
    branch_stats = defaultdict(lambda: {'NET_BUY': 0, 'NET_SELL': 0, 'NET_ZERO': 0, 'NOTICE': 0})

    for r in rows:
        flags = _classify(r)
        for f in flags:
            key = f.lower()
            if key in stats:
                stats[key] += 1
        if 'NOTICE' not in flags and 'EMPTY' not in flags:
            stats['total_real'] += 1
        if 'NET_SELL' in flags:
            anomalies['net_sell_rows'].append({
                'master': r['master'], 'branch': r['branch'], 'stock': r['stock_label'],
                'buy_amt': r['buy_amt'], 'sell_amt': r['sell_amt'], 'net_amt': r['net_amt'],
            })
        if 'NET_ZERO' in flags:
            anomalies['net_zero_rows'].append({
                'master': r['master'], 'branch': r['branch'], 'stock': r['stock_label'],
                'buy_amt': r['buy_amt'], 'sell_amt': r['sell_amt'],
            })
        bs_key = (r['master'], r['branch_code'])
        for f in ('NET_BUY', 'NET_SELL', 'NET_ZERO', 'NOTICE'):
            if f in flags:
                branch_stats[bs_key][f] += 1

    verdict, summary = _judge_verdict(stats, anomalies)

    report['sheet'] = sheet_title
    report['stats'] = stats
    report['overall_verdict'] = verdict
    report['summary'] = summary
    report['anomalies'] = {
        'net_sell_count': len(anomalies['net_sell_rows']),
        'net_zero_count': len(anomalies['net_zero_rows']),
        # 限制 examples 數量避免 audit json 爆大
        'net_sell_examples': anomalies['net_sell_rows'][:10],
        'net_zero_examples': anomalies['net_zero_rows'][:5],
    }
    report['branches_audited'] = len(branch_stats)
    return report


def save_audit_report(data_dir: Path, report: dict) -> Path:
    """Write data/daily_audit.json (前端 / trigger script 讀取點).
    v3.33.3 (M5): 同步 append 到 audit_history.json 趨勢化."""
    out_path = data_dir / 'daily_audit.json'
    out_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding='utf-8',
    )
    try:
        append_audit_history(data_dir, report)
    except Exception as e:
        # 趨勢化失敗不影響主 audit 流程
        print(f"[auto_audit] ⚠️ audit_history append 失敗 (不影響主流程): {e}",
              file=sys.stderr)
    return out_path


# ════════════════════════════════════════════════════════════════════
#  v3.33.3 (M5): audit_history.json 趨勢化
# ════════════════════════════════════════════════════════════════════

HISTORY_MAX_ENTRIES = 180   # 保留約 9 個月交易日


def append_audit_history(data_dir: Path, report: dict) -> Path:
    """每日 verdict 累積到 data/audit_history.json.
    - key = sheet 日期 (YYYYMMDD); 同日重跑 → 覆蓋為最新 (兜底排程多跑不會重複)
    - 只存趨勢需要的欄位 (verdict + 核心 stats), 不存 examples 避免膨脹
    - 保留最近 HISTORY_MAX_ENTRIES 筆"""
    hist_path = data_dir / 'audit_history.json'
    history = []
    if hist_path.exists():
        try:
            history = json.loads(hist_path.read_text(encoding='utf-8'))
            if not isinstance(history, list):
                history = []
        except (json.JSONDecodeError, OSError):
            history = []   # 壞檔重建, 不炸主流程

    stats = report.get('stats', {}) or {}
    entry = {
        'date': report.get('sheet') or report.get('audit_run_at', '')[:10],
        'run_at': report.get('audit_run_at', ''),
        'verdict': report.get('overall_verdict', 'SKIP'),
        'summary': report.get('summary', ''),
        'total_real': stats.get('total_real', 0),
        'net_sell': stats.get('net_sell', 0),
        'net_zero': stats.get('net_zero', 0),
        'reverse_est': stats.get('reverse_est', 0),
        'branches_audited': report.get('branches_audited', 0),
    }

    # 同日 dedup: 移除既有同 date 的 entry (兜底排程 22:37/23:47 重跑取最新)
    history = [h for h in history if h.get('date') != entry['date']]
    history.append(entry)
    history.sort(key=lambda h: h.get('date', ''))
    history = history[-HISTORY_MAX_ENTRIES:]

    hist_path.write_text(
        json.dumps(history, ensure_ascii=False, indent=1),
        encoding='utf-8',
    )
    return hist_path


def print_audit_trend(data_dir: Path, last_n: int = 20):
    """CLI: python auto_audit.py --history [N] — 印最近 N 天 verdict 趨勢."""
    hist_path = data_dir / 'audit_history.json'
    if not hist_path.exists():
        print("(audit_history.json 不存在 — 至少跑過一次 audit 後才有趨勢)")
        return
    try:
        history = json.loads(hist_path.read_text(encoding='utf-8'))
    except (json.JSONDecodeError, OSError) as e:
        print(f"❌ audit_history.json 讀取失敗: {e}")
        return
    if not history:
        print("(audit_history.json 是空的)")
        return

    recent = history[-last_n:]
    counts = {'PASS': 0, 'WARN': 0, 'FAIL': 0, 'ERROR': 0, 'SKIP': 0}
    icon = {'PASS': '✅', 'WARN': '⚠️', 'FAIL': '❌', 'ERROR': '💥', 'SKIP': '⏭️'}
    print(f"📈 Audit 趨勢 (最近 {len(recent)} 筆 / 累積 {len(history)} 筆)")
    print("─" * 72)
    for h in recent:
        v = h.get('verdict', '?')
        counts[v] = counts.get(v, 0) + 1
        print(f"  {h.get('date', '?'):>8}  {icon.get(v, '?')} {v:<5} "
              f"rows={h.get('total_real', 0):>4}  "
              f"net_sell={h.get('net_sell', 0)}  net_zero={h.get('net_zero', 0)}  "
              f"反推={h.get('reverse_est', 0)}")
    print("─" * 72)
    total = len(recent)
    pass_rate = counts.get('PASS', 0) / total * 100 if total else 0
    print(f"  PASS {counts.get('PASS', 0)} / WARN {counts.get('WARN', 0)} / "
          f"FAIL {counts.get('FAIL', 0)} / 其他 "
          f"{total - counts.get('PASS', 0) - counts.get('WARN', 0) - counts.get('FAIL', 0)}"
          f"  → PASS 率 {pass_rate:.0f}%")
    # 連續 WARN/FAIL 提示 (趨勢惡化告警)
    streak = 0
    for h in reversed(recent):
        if h.get('verdict') in ('WARN', 'FAIL', 'ERROR'):
            streak += 1
        else:
            break
    if streak >= 2:
        print(f"  🚨 注意: 最近連續 {streak} 天非 PASS — 檢查是否系統性劣化")


# ════════════════════════════════════════════════════════════════════
#  CLI (also used standalone, not just from crawler.py)
# ════════════════════════════════════════════════════════════════════

def _emit_github_actions_marker(report):
    """In GitHub Actions, emit ::notice:: / ::warning:: / ::error:: based on verdict.
    Outside Actions (local PowerShell), this is just stdout text."""
    v = report.get('overall_verdict', 'SKIP')
    summary = report.get('summary', '')
    if v == 'FAIL':
        print(f"::error title=Daily Audit FAIL::{summary}")
    elif v == 'WARN':
        print(f"::warning title=Daily Audit WARN::{summary}")
    elif v == 'PASS':
        print(f"::notice title=Daily Audit PASS::{summary}")


if __name__ == '__main__':
    data_dir = Path('data')
    # v3.33.3 (M5): --history [N] 看趨勢, 不跑 audit
    if len(sys.argv) > 1 and sys.argv[1] == '--history':
        n = int(sys.argv[2]) if len(sys.argv) > 2 else 20
        print_audit_trend(data_dir, n)
        sys.exit(0)
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    rpt = run_audit(data_dir, xlsx)
    out = save_audit_report(data_dir, rpt)
    print(f"[auto_audit] verdict={rpt['overall_verdict']} → {out}")
    print(f"             {rpt.get('summary', '')}")
    _emit_github_actions_marker(rpt)
    if rpt.get('overall_verdict') == 'FAIL':
        sys.exit(1)
    sys.exit(0)
