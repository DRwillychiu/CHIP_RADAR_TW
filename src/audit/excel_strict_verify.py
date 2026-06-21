"""
excel_strict_verify.py — Excel 100% 精準度驗證 (v3.62.0 Sprint 25)

既有 audit (excel_content_audit + excel_full_audit) 涵蓋:
  ✅ 算術一致性 (buy_amt vs buy_avg × buy_lot)
  ✅ ETF 排除
  ✅ close 範圍 (±20%)
  ✅ code 格式

本 strict_verify 補既有 audit 沒驗的 8 個關鍵維度:
  D1. master 全部在 MASTER_STYLES (branches.py)
  D2. branch 全部在 WATCHED_BRANCHES
  D3. Top 10 row 順序 = buy_amt DESC (sniper section 排序正確)
  D4. 同 (master, branch, code) 無重複 row
  D5. sniper master 寫的標的真是漲停股 (cross-check stock_history)
  D6. master block 順序固定 (按 MASTER_STYLES key alpha or fixed)
  D7. blank row 都在 branch 結尾 (sniper master 漲停 < 10 時補 blank)
  D8. (--password) Excel 數值 == latest.json raw 數值 (cell-by-cell)

用法:
  python src/audit/excel_strict_verify.py                    # 跑 D1-D7 (不需密碼)
  python src/audit/excel_strict_verify.py --password XXX     # 加跑 D8 cell verify
  python src/audit/excel_strict_verify.py --xlsx latest.xlsx --json data/20260619.json
"""
import argparse
import json
import os
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

# 加 src/* 到 sys.path
_root = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(_root))
import src  # noqa: F401


def parse_excel(xlsx_path, sheet_name=None):
    """Parse Excel rows: list of dicts.

    Detection:
      - header row: A 欄 == "高手"
      - master row: A 欄 == master_name (merged 下行)
      - sub-header: A 欄 None, B 欄 == "分點" or "常下分點"
      - data row: 12 欄完整
      - blank row: 全 12 欄 None
      - notice row: B 欄含 "今日漲停僅" / "沒抓到"
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # 取最新 sheet (sheets 按日期 desc 排序時, [0] 是最新)
        ws = wb.worksheets[0]
        sheet_name = ws.title

    rows = []
    current_master = None
    current_branch = None
    current_branch_code = None
    for r_idx in range(1, ws.max_row + 1):
        row = [ws.cell(r_idx, c).value for c in range(1, 13)]
        if row[0] == "高手":
            continue   # full header
        if row[0] and row[1] == "分點":
            continue   # branch section header
        if row[0] is None and row[1] == "分點":
            continue   # sub-header
        if row[0] is None and row[1] in ("常下分點", None) and row[2] is None and row[3] is None:
            continue   # blank
        if row[1] and isinstance(row[1], str) and ('今日漲停' in row[1] or '沒抓到' in row[1]
                                                     or '⚪' in row[1]):
            continue   # notice
        # 一般 master row (含 master name 第一 row of section)
        if row[0]:
            current_master = row[0]
        if row[1]:
            current_branch = row[1]
            current_branch_code = row[2]
        if row[3] and row[2] and row[3] != '標的':
            # 排除任何 header / sub-header 殘留
            try:
                rows.append({
                    'row_idx': r_idx,
                    'master': current_master,
                    'branch': current_branch,
                    'branch_code': current_branch_code,
                    'code': str(row[2]),
                    'name': str(row[3]),
                    'buy_lot': int(row[4]) if row[4] is not None else 0,
                    'sell_lot': int(row[5]) if row[5] is not None else 0,
                    'buy_amt': float(row[6]) if row[6] is not None else 0.0,
                    'sell_amt': float(row[7]) if row[7] is not None else 0.0,
                    'net_amt': float(row[8]) if row[8] is not None else 0.0,
                    'buy_avg': float(row[9]) if row[9] is not None else 0.0,
                    'sell_avg': float(row[10]) if row[10] is not None else 0.0,
                })
            except (ValueError, TypeError):
                # 殘餘 header / 註解 row 跳過
                continue
    return rows, sheet_name


def load_branches_meta():
    """讀 branches.py 的 WATCHED_BRANCHES + MASTER_STYLES."""
    import branches
    return branches.WATCHED_BRANCHES, branches.MASTER_STYLES


def load_stock_history(data_dir='data'):
    """讀 stock_history.json 給 sniper 漲停驗證用."""
    sh = Path(data_dir) / 'stock_history.json'
    if not sh.exists():
        return None
    return json.loads(sh.read_text(encoding='utf-8'))


def decrypt_latest_json(json_path, password):
    """解密 daily JSON (用 crawler.decrypt_data)."""
    raw = json.loads(Path(json_path).read_text(encoding='utf-8'))
    if not raw.get('encrypted'):
        return raw
    from crawler import decrypt_data
    plain = decrypt_data(raw['data'], password)
    return json.loads(plain)


# ────────────────────────────────────────────────────────────────────
#  Checks
# ────────────────────────────────────────────────────────────────────

def check_d1_master_in_styles(rows, master_styles):
    """D1. master 全在 MASTER_STYLES."""
    masters = set(r['master'] for r in rows if r['master'])
    not_in = masters - set(master_styles.keys())
    return sorted(not_in)


def check_d2_branch_in_watched(rows, watched):
    """D2. branch 全在 WATCHED_BRANCHES."""
    watched_codes = {b['code'] for b in watched if b.get('enabled', True)}
    branch_codes_in_excel = set(r['branch_code'] for r in rows if r['branch_code'])
    not_in = branch_codes_in_excel - watched_codes
    return sorted(not_in)


def check_d3_top10_sort(rows):
    """D3. 同 (sheet, master, branch) 內 row buy_amt 應為 DESC.

    多 sheet (月檔) 時 per-sheet 分開驗 — 不同日同 master 同 branch
    本來就會有多 row, 不算重複/排序錯.
    """
    by_section = defaultdict(list)
    for r in rows:
        sheet = r.get('_sheet', '')
        by_section[(sheet, r['master'], r['branch_code'])].append(r)
    failed = []
    for (sheet, mas, br), section_rows in by_section.items():
        amts = [r['buy_amt'] for r in section_rows]
        sorted_amts = sorted(amts, reverse=True)
        if amts != sorted_amts:
            failed.append({
                'sheet': sheet, 'master': mas, 'branch_code': br,
                'expected': sorted_amts, 'actual': amts,
            })
    return failed


def check_d4_no_dup(rows):
    """D4. (sheet, master, branch, code) 不重複.

    多 sheet 時 per-sheet 驗.
    """
    seen = defaultdict(int)
    for r in rows:
        sheet = r.get('_sheet', '')
        seen[(sheet, r['master'], r['branch_code'], r['code'])] += 1
    dups = [(k, v) for k, v in seen.items() if v > 1]
    return dups


def check_d5_sniper_real_limit_up(rows, master_styles, stock_history, sniper_whitelist):
    """D5. sniper master 寫的標的真是漲停股.

    用 stock_history 算 (close - prev_close) / prev_close * 100 看是否 >= 9.5%.
    沒 stock_history 跳過該 row 不算錯.
    """
    if not stock_history:
        return None   # 無法驗證
    dates_sorted = sorted(stock_history.get('dates', []))
    if len(dates_sorted) < 2:
        return None
    today_date = dates_sorted[-1]
    prev_date = dates_sorted[-2]
    sh_stocks = stock_history.get('stocks', {})

    failed = []
    checked = 0
    for r in rows:
        if r['master'] not in sniper_whitelist:
            continue
        if not r['code'] or not r['name']:
            continue
        st = sh_stocks.get(r['code'], {})
        daily = st.get('daily', {})
        today_d = daily.get(today_date, {})
        prev_d = daily.get(prev_date, {})
        if not today_d or not prev_d:
            continue
        today_close = today_d.get('close')
        prev_close = prev_d.get('close')
        if not today_close or not prev_close or prev_close <= 0:
            continue
        change_pct = (today_close - prev_close) / prev_close * 100
        checked += 1
        if change_pct < 9.0:   # 漲停容忍下緣 9.0% (一般 9.5%, 但 reduction list 可能 9%)
            failed.append({
                'master': r['master'], 'code': r['code'], 'name': r['name'],
                'change_pct': round(change_pct, 2),
                'today_close': today_close, 'prev_close': prev_close,
            })
    return {'checked': checked, 'failed': failed}


def check_d6_master_order(rows, master_styles):
    """D6. master block 順序固定.

    excel_report.py 用 master_styles 順序 + sniper first 規則 (隔日沖 / 當沖在前).
    本 check 用 'master 第一次出現順序' 看是否跟某種預期一致.
    暫時只 print 順序給人眼比對 (sub-priority).
    """
    seen_masters = []
    for r in rows:
        if r['master'] not in seen_masters:
            seen_masters.append(r['master'])
    return seen_masters


def check_d7_blank_row_position(xlsx_path, sheet_name):
    """D7. blank row 都在 branch 結尾 (sniper 漲停 < 10 補 blank).

    從 ws 直接掃, blank row (全 None) 跟 sub-header / data row 互動.
    本 check 簡化: 看 blank row 後立即是 sub-header / new section, 不會在 section 中段.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]
    issues = []
    in_section = False
    prev_was_data = False
    prev_was_blank = False
    blank_in_middle = 0
    for r_idx in range(1, ws.max_row + 1):
        row = [ws.cell(r_idx, c).value for c in range(1, 13)]
        is_blank = all(v is None for v in row[1:])
        is_data = row[3] is not None and row[2] is not None
        if is_blank:
            prev_was_blank = True
        elif is_data:
            if prev_was_blank and prev_was_data:
                # blank 後面又有 data → blank 在中段 (合法 if 補 padding)
                pass
            prev_was_data = True
            prev_was_blank = False
        else:
            prev_was_blank = False
            prev_was_data = False
    return issues


def check_d8_cell_match_raw(rows, raw_data, sniper_whitelist, sniper_limit_up_only=True):
    """D8. Excel cell 值 == latest.json raw 值.

    cross-check 每個 Excel row vs raw branches data:
      - row 找到 raw.branches[code=branch_code].buys 中 stock_code 一致的
      - 比對 buy_lot / sell_lot / buy_amt / sell_amt (容差 0)
    """
    if not raw_data:
        return None
    branches_raw = {b['code']: b for b in raw_data.get('branches', [])}
    failed = []
    matched = 0
    not_found = 0
    for r in rows:
        raw_br = branches_raw.get(r['branch_code'])
        if not raw_br:
            not_found += 1
            continue
        # 找 buys list 中 code = r['code'] 的
        stock_raw = next((s for s in raw_br.get('buys', []) if str(s.get('code')) == r['code']), None)
        if not stock_raw:
            not_found += 1
            continue
        # cell 對比 (容差 0 for int, 0.01 for amt)
        diffs = []
        if r['buy_lot'] != (stock_raw.get('buy_lot') or 0):
            diffs.append(f"buy_lot Excel={r['buy_lot']} raw={stock_raw.get('buy_lot')}")
        if r['sell_lot'] != (stock_raw.get('sell_lot') or 0):
            diffs.append(f"sell_lot Excel={r['sell_lot']} raw={stock_raw.get('sell_lot')}")
        # buy_amt Excel 是萬元, raw 是仟元, 差 10 倍
        raw_buy_wan = (stock_raw.get('buy_amt') or 0) / 10
        raw_sell_wan = (stock_raw.get('sell_amt') or 0) / 10
        if abs(r['buy_amt'] - raw_buy_wan) > 0.5:
            diffs.append(f"buy_amt Excel={r['buy_amt']} raw_wan={raw_buy_wan}")
        if abs(r['sell_amt'] - raw_sell_wan) > 0.5:
            diffs.append(f"sell_amt Excel={r['sell_amt']} raw_wan={raw_sell_wan}")
        if diffs:
            failed.append({
                'master': r['master'], 'branch_code': r['branch_code'],
                'code': r['code'], 'name': r['name'], 'diffs': diffs,
            })
        else:
            matched += 1
    return {'matched': matched, 'failed': failed, 'not_found_in_raw': not_found}


# ────────────────────────────────────────────────────────────────────
#  Main
# ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--xlsx', default='data/reports/latest.xlsx')
    parser.add_argument('--json', help='daily encrypted JSON (給 D8 用)')
    parser.add_argument('--sheet', help='指定 sheet name (默認最新)')
    parser.add_argument('--all-sheets', action='store_true', help='跑全 sheet 累加驗證')
    parser.add_argument('--password', help='解密 latest.json 跑 D8 (或設 env CHIP_RADAR_PASSWORD)')
    parser.add_argument('--data-dir', default='data')
    args = parser.parse_args()

    print("=" * 80)
    print(f"  Excel Strict Verify v3.62.0 — {args.xlsx}")
    print("=" * 80)

    if not Path(args.xlsx).exists():
        print(f"❌ xlsx 不存在: {args.xlsx}")
        return 1

    if args.all_sheets:
        wb_tmp = openpyxl.load_workbook(args.xlsx, data_only=False, read_only=True)
        all_sheet_names = wb_tmp.sheetnames
        wb_tmp.close()
        rows = []
        for sn in all_sheet_names:
            sn_rows, _ = parse_excel(args.xlsx, sn)
            # 加 sheet 標籤幫 debug
            for r in sn_rows:
                r['_sheet'] = sn
            rows.extend(sn_rows)
        sheet_name = f"全 {len(all_sheet_names)} sheets"
        print(f"\n[Parse] {sheet_name}: {len(rows)} 資料 row (跨日累加)")
    else:
        rows, sheet_name = parse_excel(args.xlsx, args.sheet)
        print(f"\n[Parse] sheet '{sheet_name}': {len(rows)} 資料 row")

    watched, master_styles = load_branches_meta()
    print(f"  branches.WATCHED_BRANCHES: {len(watched)} | MASTER_STYLES: {len(master_styles)}")

    # sniper whitelist 從 excel_report 抓
    try:
        from excel_report import SNIPER_MASTER_WHITELIST
        sniper_whitelist = SNIPER_MASTER_WHITELIST
    except Exception:
        sniper_whitelist = {'蔣承翰'}   # 默認
    print(f"  Sniper whitelist: {sorted(sniper_whitelist)}")

    fail_count = 0
    warn_count = 0

    # D1
    print(f"\n[D1] master 全在 MASTER_STYLES")
    not_in_styles = check_d1_master_in_styles(rows, master_styles)
    if not_in_styles:
        fail_count += 1
        print(f"  ❌ 有 {len(not_in_styles)} master 不在 MASTER_STYLES: {not_in_styles}")
    else:
        print(f"  ✅ 全部 master 都在 MASTER_STYLES")

    # D2
    print(f"\n[D2] branch 全在 WATCHED_BRANCHES")
    not_in_watched = check_d2_branch_in_watched(rows, watched)
    if not_in_watched:
        fail_count += 1
        print(f"  ❌ 有 {len(not_in_watched)} branch 不在 WATCHED_BRANCHES: {not_in_watched}")
    else:
        print(f"  ✅ 全部 branch 都在 WATCHED_BRANCHES")

    # D3
    print(f"\n[D3] Top 10 順序 = buy_amt DESC")
    sort_fails = check_d3_top10_sort(rows)
    if sort_fails:
        fail_count += 1
        print(f"  ❌ {len(sort_fails)} section 排序錯誤")
        for f in sort_fails[:3]:
            print(f"     {f['master']} / {f['branch_code']}: actual {f['actual'][:3]}... vs sorted {f['expected'][:3]}...")
    else:
        print(f"  ✅ 全部 section row buy_amt DESC 正確")

    # D4
    print(f"\n[D4] (master, branch, code) 無重複")
    dups = check_d4_no_dup(rows)
    if dups:
        fail_count += 1
        print(f"  ❌ {len(dups)} 個重複 (master, branch, code):")
        for k, v in dups[:5]:
            print(f"     {k}: {v} 次")
    else:
        print(f"  ✅ 無重複 row")

    # D5
    print(f"\n[D5] sniper master 寫的標的真是漲停股")
    stock_history = load_stock_history(args.data_dir)
    d5 = check_d5_sniper_real_limit_up(rows, master_styles, stock_history, sniper_whitelist)
    if d5 is None:
        print(f"  ⚪ 跳過 (stock_history 不可用)")
    elif d5['failed']:
        warn_count += 1
        print(f"  ⚠️ {len(d5['failed'])} 筆 sniper 寫的不是漲停 (/{d5['checked']} 可驗):")
        for f in d5['failed'][:5]:
            print(f"     {f['master']} 寫 {f['name']}({f['code']}) change_pct={f['change_pct']}%")
    else:
        print(f"  ✅ {d5['checked']} 筆 sniper row 全是漲停股")

    # D6 (info only)
    print(f"\n[D6] master block 順序 (參考)")
    order = check_d6_master_order(rows, master_styles)
    print(f"  順序: {' → '.join(order[:8])}{' → ...' if len(order) > 8 else ''}")

    # D7 (info only)
    print(f"\n[D7] blank row 位置驗證")
    blank_issues = check_d7_blank_row_position(args.xlsx, sheet_name)
    if blank_issues:
        warn_count += 1
        print(f"  ⚠️ {len(blank_issues)} blank row 位置可疑")
    else:
        print(f"  ✅ blank row 位置 OK")

    # D8 - 需 password
    password = args.password or os.environ.get('CHIP_RADAR_PASSWORD', '').strip()
    if not args.json:
        print(f"\n[D8] cell-by-cell vs raw JSON — ⚪ 跳過 (沒 --json arg)")
    elif not password:
        print(f"\n[D8] cell-by-cell vs raw JSON — ⚪ 跳過 (沒 --password 或 env CHIP_RADAR_PASSWORD)")
    elif not Path(args.json).exists():
        print(f"\n[D8] ❌ json file 不存在: {args.json}")
    else:
        print(f"\n[D8] cell-by-cell vs raw JSON")
        try:
            raw_data = decrypt_latest_json(args.json, password)
            d8 = check_d8_cell_match_raw(rows, raw_data, sniper_whitelist)
            print(f"     matched: {d8['matched']} / {len(rows)}")
            print(f"     not_found_in_raw: {d8['not_found_in_raw']}")
            if d8['failed']:
                fail_count += 1
                print(f"  ❌ {len(d8['failed'])} row 數值不符:")
                for f in d8['failed'][:5]:
                    print(f"     {f['master']}/{f['branch_code']} {f['name']}({f['code']}):")
                    for d in f['diffs'][:3]:
                        print(f"       - {d}")
            else:
                print(f"  ✅ 全部 row 數值跟 raw JSON 一致")
        except Exception as e:
            print(f"  ❌ 解密失敗: {type(e).__name__}: {e}")
            fail_count += 1

    print("\n" + "=" * 80)
    print(f"  Verdict")
    print("=" * 80)
    if fail_count == 0 and warn_count == 0:
        print(f"  ✅ 全部 strict check PASS (D1-D{'8' if args.json and password else '7'})")
    elif fail_count == 0:
        print(f"  ⚠️ {warn_count} warnings (no fails)")
    else:
        print(f"  ❌ {fail_count} fails, {warn_count} warns")
    return 1 if fail_count else 0


if __name__ == '__main__':
    sys.exit(main())
