"""6/18 Section A native cell inspection"""
import sys
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')
ROOT = Path(__file__).resolve().parents[1]

from openpyxl import load_workbook
wb = load_workbook(str(ROOT / 'data' / 'reports' / 'chip_radar_2026-06.xlsx'))
print(f"sheets: {wb.sheetnames}")
ws = wb[wb.sheetnames[0]]
print(f"using sheet: {ws.title}")

for r in range(3, ws.max_row + 1):
    v = ws[f'B{r}'].value
    if v and isinstance(v, str) and ('規模統計' in v or 'A.' in v):
        print(f"\nFound section A header at row {r}: {v}")
        r1, r2 = r + 1, r + 2
        print(f"\n--- Row {r1} (Row 1) ---")
        for label_col, val_col in [('B', 'C'), ('E', 'F'), ('H', 'I')]:
            label_v = ws[f'{label_col}{r1}'].value
            val_cell = ws[f'{val_col}{r1}']
            print(f"  {label_v} | value={val_cell.value!r} | type={type(val_cell.value).__name__} | fmt={val_cell.number_format!r}")
        print(f"\n--- Row {r2} (Row 2) ---")
        for label_col, val_col in [('B', 'C'), ('E', 'F'), ('H', 'I')]:
            label_v = ws[f'{label_col}{r2}'].value
            val_cell = ws[f'{val_col}{r2}']
            print(f"  {label_v} | value={val_cell.value!r} | type={type(val_cell.value).__name__} | fmt={val_cell.number_format!r}")
        break
else:
    print("Section A header not found, dumping rows 3-15 col B:")
    for r in range(3, 15):
        print(f"  B{r} = {ws[f'B{r}'].value!r}")
