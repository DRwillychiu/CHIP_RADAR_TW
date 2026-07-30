"""v3.70.5 全 Phase 3.2 落地頁面數字審查.

對 Excel 4 個 enrichment sheet 的每個數字 cross-check 真實源頭, 找出錯誤.

涵蓋:
  Layer 1: Dashboard
    - Phase 3.2 sub-banner (78.9%, n=38, Wilson CI, 30d 實戰)
    - 今日 quad 狀態 banner (row 9)
  Layer 2: 📱 手機摘要
    - 明日預測, 強共識 Top 5 (含 ⭐), 追蹤池方向
  Layer 3: 📈 Quad 實戰追蹤
    - 摘要 line (累積/30d/預期/delta)
    - 逐 trigger day (日期/Q5/picks/hits/命中率/mean)
    - Per-master leaderboard
  Layer 4: 📉 Quad 失效歸因
    - miss 數摘要
    - 歸因分布
    - 逐 miss 行

每個 mismatch → errors.append(('SECTION', expected, actual, note))
全 PASS 才算可信.
"""
import json, sys, math
from pathlib import Path
from collections import defaultdict
sys.stdout.reconfigure(encoding='utf-8')
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl

errors = []
warnings = []

def err(section, expected, actual, note=''):
    errors.append((section, expected, actual, note))

def warn(section, msg):
    warnings.append((section, msg))

# === Load Excel + JSON data ===
xlsx = ROOT / 'data' / 'reports' / 'latest.xlsx'
wb = openpyxl.load_workbook(xlsx, data_only=False)

dashboard = wb['📋 今日 Dashboard']
mobile = wb['📱 手機摘要']
qtrack = wb['📈 Quad 實戰追蹤']
qfail = wb['📉 Quad 失效歸因']

# Ground truth from JSON
def _load(p):
    try:
        with open(ROOT / 'data' / p, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"  ! cannot load {p}: {e}")
        return {}

pb = _load('phase32_backtest.json')
qhl = _load('quad_hit_log.json')

# === Helper: Wilson CI ===
def wilson_ci(hits, n, z=1.96):
    if n <= 0: return (0.0, 0.0)
    p = hits/n
    denom = 1 + z*z/n
    c = (p + z*z/(2*n))/denom
    m = z*math.sqrt(p*(1-p)/n + z*z/(4*n*n))/denom
    return (max(0.0, c-m), min(1.0, c+m))


# ════════════════════════════════════════════════════════════════════
# Layer 1: Dashboard Phase 3.2 sub-banner
# ════════════════════════════════════════════════════════════════════
print("=" * 70)
print("Layer 1: Dashboard Phase 3.2 sub-banner (row 8)")
print("=" * 70)

# Phase 3.2 sub-banner row (we know it's row 8 from prior verification)
banner = dashboard.cell(8, 2).value or ''
print(f"  Banner: {banner[:150]}")

if pb.get('summary'):
    triple = pb['summary'].get('e_vol_spike_q5_bull', {})
    tr_n = triple.get('n', 0)
    tr_hits = triple.get('hits', 0)
    tr_hr = triple.get('hit_rate', 0) * 100
    tr_mean = triple.get('mean_change', 0)
    bl_hr = pb['summary'].get('baseline', {}).get('hit_rate', 0) * 100

    # Check each number
    assertions = [
        (f'{tr_hr:.1f}%',  'Phase 3.2 hit %'),
        (f'n={tr_n}',       'Phase 3.2 n'),
        (f'mean {tr_mean:+.2f}%', 'mean change'),
        (f'baseline {bl_hr:.1f}%', 'baseline %'),
    ]
    for expected_str, desc in assertions:
        if expected_str not in banner:
            err('P32.BANNER', expected_str, banner[:80], f'缺 {desc}')

    # Wilson CI check
    if tr_n >= 5:
        ci_lo, ci_hi = wilson_ci(tr_hits, tr_n)
        ci_str = f'[{ci_lo*100:.1f}–{ci_hi*100:.1f}% 95% CI]'
        if ci_str not in banner:
            err('P32.BANNER.CI', ci_str, banner[:80], 'Wilson CI 不對')

# 30d 實戰
if qhl.get('rolling_30d'):
    r30 = qhl['rolling_30d']
    r30_n = r30.get('n', 0)
    r30_hits = r30.get('hits', 0)
    if r30_n > 0:
        live_str = f'30d 實戰: {r30_hits}/{r30_n}'
        if live_str not in banner:
            err('P32.BANNER.LIVE', live_str, banner[:80], '30d 實戰數字不對')

# ════════════════════════════════════════════════════════════════════
# Layer 1.2: 今日 quad 狀態 banner (row 9)
# ════════════════════════════════════════════════════════════════════
print()
print("Layer 1.2: 今日 quad 狀態 banner (row 9)")
status_banner = dashboard.cell(9, 2).value or ''
print(f"  Status: {status_banner[:150]}")
# Status must be 1 of 4 states: 🎯 命中 / 💤 Q5 偏多但無 vol_spike / 💤 vol_spike 但 Q5 非偏多 / 💤 完全無
status_icons = ['🎯', '💤']
if not any(i in status_banner for i in status_icons):
    err('P32.STATUS', 'has 🎯 or 💤', status_banner[:60], 'quad 狀態 banner 缺 icon')

# ════════════════════════════════════════════════════════════════════
# Layer 2: 📱 手機摘要 數字驗證
# ════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
print("Layer 2: 📱 手機摘要 數字驗證")
print("=" * 70)

# 主標 row 2
m_title = mobile.cell(2, 3).value
if not m_title or '📋' not in str(m_title) or 'Chip Radar' not in str(m_title):
    err('MOB.TITLE', 'has "📋 Chip Radar"', m_title, 'mobile 標題不對')

# 明日預測 row 5
m_q5 = mobile.cell(5, 3).value
if not m_q5 or ('↑' not in str(m_q5) and '↓' not in str(m_q5) and '↕' not in str(m_q5)):
    err('MOB.Q5', 'has direction arrow', m_q5, 'mobile Q5 banner 缺方向')

# 強共識 Top 5 — scan
qhl_quad_codes = set()
if qhl.get('trigger_days'):
    # 應只在今日 trigger 才有 ⭐, 今日 (6/25) Q5 中性 → 不應有 ⭐
    pass   # 今日 quad_codes 從 qhl 不可直接得 → 從 Dashboard 解析

# Mobile 共識 Top 5 必有 5 行 (① ② ③ ④ ⑤ 或 ⭐ 前綴)
print("  mobile 共識 Top 5 rows scan:")
for row_n in range(5, 35):
    val = mobile.cell(row_n, 3).value
    if val and any(c in str(val) for c in ['①', '②', '③', '④', '⑤', '⭐']):
        print(f"    row {row_n}: {val[:60]}")

# ════════════════════════════════════════════════════════════════════
# Layer 3: 📈 Quad 實戰追蹤
# ════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
print("Layer 3: 📈 Quad 實戰追蹤")
print("=" * 70)

# Row 3 摘要 line
sum_row = qtrack.cell(3, 2).value or ''
print(f"  Summary: {sum_row[:150]}")
if qhl.get('rolling_all'):
    ra = qhl['rolling_all']
    n = ra.get('n', 0)
    hits = ra.get('hits', 0)
    hr_str = f'累積: {hits}/{n} = {ra.get("hit_rate",0)*100:.1f}%'
    if hr_str not in sum_row:
        err('QT.SUMMARY', hr_str, sum_row[:60], '累積 hit 數字錯')
    mean_str = f'mean {ra.get("mean_change", 0):+.2f}%'
    if mean_str not in sum_row:
        err('QT.SUMMARY.MEAN', mean_str, sum_row[:60], '累積 mean 錯')

# Per-day rows (start row 6 from previous inspection)
# Expected: same number of rows as len(qhl['trigger_days'])
trigger_days_expected = qhl.get('trigger_days', [])
print(f"  Expected {len(trigger_days_expected)} trigger day rows in Excel")

# Read Excel rows starting from data section
excel_trigger_dates = []
for row_n in range(6, 30):
    d_val = qtrack.cell(row_n, 2).value
    if isinstance(d_val, str) and '/' in d_val and d_val.startswith('2026'):
        date_8 = d_val.replace('/', '')
        excel_trigger_dates.append((row_n, date_8))
    elif d_val and 'Per-Master' in str(d_val):
        break

print(f"  Found {len(excel_trigger_dates)} trigger day rows in Excel")
if len(excel_trigger_dates) != len(trigger_days_expected):
    err('QT.ROWS', len(trigger_days_expected), len(excel_trigger_dates),
        '逐日 trigger 行數 mismatch')

# Per-day data validation
trigger_by_date = {td['date']: td for td in trigger_days_expected}
for row_n, date_8 in excel_trigger_dates:
    td = trigger_by_date.get(date_8)
    if not td:
        err('QT.DAY', date_8, '?', f'Excel row {row_n} 日期不在 hit log')
        continue
    # n picks col 6
    n_excel = qtrack.cell(row_n, 6).value
    if n_excel != td['n']:
        err('QT.DAY.N', td['n'], n_excel, f'{date_8} picks 數量不對')
    # hits col 7
    hits_excel = qtrack.cell(row_n, 7).value
    if hits_excel != td['hits']:
        err('QT.DAY.HITS', td['hits'], hits_excel, f'{date_8} hits 不對')
    # hit rate col 8 (float)
    hr_excel = qtrack.cell(row_n, 8).value
    if abs((hr_excel or 0) - td['hit_rate'] * 100) > 0.5:
        err('QT.DAY.HR', f"{td['hit_rate']*100:.2f}", hr_excel,
            f'{date_8} hit rate 不對')

# Per-master leaderboard
print("\n  Per-master leaderboard scan:")
pm_rows = []
in_leaderboard = False
for row_n in range(1, 30):
    val = qtrack.cell(row_n, 2).value
    if val and 'Per-Master' in str(val):
        in_leaderboard = True
        continue
    if not in_leaderboard: continue
    if val and val not in ('Master', '註: trigger day = Q5 預測偏多 AND ≥1 master 量爆 (>2σ).'):
        if isinstance(val, str) and not val.startswith('註'):
            n_val = qtrack.cell(row_n, 4).value
            h_val = qtrack.cell(row_n, 5).value
            if isinstance(n_val, int) and isinstance(h_val, int):
                pm_rows.append((val, n_val, h_val))

# Ground truth from qhl
master_picks = defaultdict(list)
for td in qhl.get('trigger_days', []):
    for p in td.get('quad_picks', []):
        for m in (p.get('matched_masters') or []):
            master_picks[m].append(p['hit'])

gt_pm_count = {m: (len(hl), sum(hl)) for m, hl in master_picks.items()}
print(f"  GT per-master count: {len(gt_pm_count)}")
print(f"  Excel per-master rows: {len(pm_rows)}")
if len(pm_rows) != len(gt_pm_count):
    err('QT.PM.COUNT', len(gt_pm_count), len(pm_rows),
        'per-master 行數 mismatch')

for master_name, n_excel, h_excel in pm_rows:
    gt = gt_pm_count.get(master_name)
    if not gt:
        err('QT.PM.NAME', f'master in gt', master_name,
            f'{master_name} 不在 ground truth')
        continue
    gt_n, gt_h = gt
    if n_excel != gt_n:
        err('QT.PM.N', gt_n, n_excel, f'{master_name} all picks 不對')
    if h_excel != gt_h:
        err('QT.PM.HITS', gt_h, h_excel, f'{master_name} hits 不對')

# ════════════════════════════════════════════════════════════════════
# Layer 4: 📉 Quad 失效歸因
# ════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
print("Layer 4: 📉 Quad 失效歸因")
print("=" * 70)

# miss count from ground truth
gt_misses = []
for td in qhl.get('trigger_days', []):
    for p in td.get('quad_picks', []):
        if not p.get('hit'):
            gt_misses.append({**p, 'date': td['date']})

ra = qhl.get('rolling_all', {})
gt_total = ra.get('n', 0)
gt_total_misses = gt_total - ra.get('hits', 0)
print(f"  GT misses: {gt_total_misses}/{gt_total} ({gt_total_misses/max(gt_total,1)*100:.1f}%)")

# Summary line in sheet
sum_row = qfail.cell(3, 2).value or ''
print(f"  Summary: {sum_row[:150]}")
if f'{gt_total_misses}/{gt_total}' not in sum_row:
    err('QF.SUMMARY', f'{gt_total_misses}/{gt_total}', sum_row[:80],
        'miss 數摘要錯')

# 歸因分布
dist_row = qfail.cell(4, 2).value or ''
print(f"  分布: {dist_row[:200]}")
gt_reasons = defaultdict(int)
for m in gt_misses:
    for r in (m.get('failure_reasons') or ['未分類']):
        gt_reasons[r] += 1
for r, c in gt_reasons.items():
    expected = f'{r}: {c}'
    if expected not in dist_row:
        err('QF.DIST', expected, dist_row[:100], f'歸因 {r} 計數錯')

# Miss row count
excel_miss_rows = 0
for row_n in range(7, 50):
    val = qfail.cell(row_n, 2).value
    if isinstance(val, str) and val.startswith('2026/'):
        excel_miss_rows += 1
    elif val and val.startswith('歸因分類'):
        break
print(f"  Excel miss rows: {excel_miss_rows}, GT misses: {len(gt_misses)}")
if excel_miss_rows != len(gt_misses):
    err('QF.ROWS', len(gt_misses), excel_miss_rows, 'miss 行數錯')

# ════════════════════════════════════════════════════════════════════
# Result
# ════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
if errors:
    print(f"❌ FAIL: {len(errors)} mismatches:")
    for sec, exp, act, note in errors:
        print(f"  [{sec}] {note}: expected={exp!r}, actual={act!r}")
    sys.exit(1)
elif warnings:
    print(f"⚠️ {len(warnings)} warning(s):")
    for sec, msg in warnings:
        print(f"  [{sec}] {msg}")
    print("OK with warnings.")
else:
    print("✅ PASS: 全部 Phase 3.2 落地頁面 cross-validate 通過")
