"""v3.67.2 Phase 2.7: 萃取手機摘要為純文字 → stdout

用途: GitHub Actions daily-full 跑完後, 抓 latest.xlsx 「📱 手機摘要」sheet
      的內容轉成純文字, 作為 email body.

執行: python scripts/extract_mobile_summary_text.py [path/to/latest.xlsx]
      預設 path = data/reports/latest.xlsx
"""
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')
ROOT = Path(__file__).resolve().parents[1]

xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / 'data' / 'reports' / 'latest.xlsx'

try:
    from openpyxl import load_workbook
    wb = load_workbook(str(xlsx_path), data_only=True)
except Exception as e:
    print(f"(無法讀取 latest.xlsx: {e})", file=sys.stderr)
    sys.exit(1)

MOBILE_SHEET = "📱 手機摘要"
if MOBILE_SHEET not in wb.sheetnames:
    print(f"(找不到 {MOBILE_SHEET} sheet, 可用: {wb.sheetnames[:5]})", file=sys.stderr)
    sys.exit(1)

ws = wb[MOBILE_SHEET]

# 內容都在 C 欄, row 2 起算
lines = []
for r in range(1, ws.max_row + 1):
    v = ws[f'C{r}'].value
    if v is None:
        lines.append('')   # 保留空行 (作為 section 分隔)
    else:
        lines.append(str(v))

# 去除頭尾空行, 中間連續多個空行壓成 1 個
while lines and lines[0] == '':
    lines.pop(0)
while lines and lines[-1] == '':
    lines.pop()

# 連續空行壓縮
out_lines = []
prev_blank = False
for ln in lines:
    if ln == '':
        if not prev_blank:
            out_lines.append('')
        prev_blank = True
    else:
        out_lines.append(ln)
        prev_blank = False

print('\n'.join(out_lines))
