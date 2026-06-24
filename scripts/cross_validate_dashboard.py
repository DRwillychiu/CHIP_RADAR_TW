"""v3.63.2 Dashboard 完整交叉驗證 — 比對每一筆顯示值 vs 原始計算

策略:
  1. 用 6/18 真實 daily_trading_signals.json + master_profiles.json 構造輸入
  2. 用合成 branches_data (從 6/18 一個追蹤 master + 一個非追蹤 master 模擬)
  3. 渲染 Dashboard
  4. 逐 section 比對 Excel cell value vs 獨立計算的 ground truth
  5. 列出所有 mismatch (應為 0)
"""
import sys, json, re
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook, load_workbook
from src.exports.excel_report import (
    build_dashboard_sheet, TRACKED_MASTERS, MASTER_MAPPING,
    _is_tracked_master, _filter_tracked_branches, _round_safe, _severity_from_z,
)

errors = []
def err(section, expected, actual, note=''):
    errors.append((section, expected, actual, note))

def warn(section, msg):
    errors.append((section, '⚠️', msg, 'WARN'))

# === Load data ===
with open(ROOT/'data'/'daily_trading_signals.json', 'r', encoding='utf-8') as f:
    signals = json.load(f)
with open(ROOT/'data'/'master_profiles.json', 'r', encoding='utf-8') as f:
    profiles = json.load(f)

# === Synthetic branches_data — match real Excel display ===
# 用 5 個追蹤 master 各 1 branch + 1 個非追蹤 master 來驗 filter
branches_data = [
    {'code': '9B25', 'name': '台新-五權西', 'master': '民哥', 'buys': [
        {'code': '2330', 'name': '台積電', 'buy_lot': 100, 'sell_lot': 20,
         'buy_amt': 250000, 'sell_amt': 50000, 'net_lot': 80,
         'buy_avg': 2500.0, 'sell_avg': 2500.0, 'close_price': 2500.0,
         'change_pct': 1.0, 'is_limit_up': False},
        {'code': '2454', 'name': '聯發科', 'buy_lot': 30, 'sell_lot': 5,
         'buy_amt': 90000, 'sell_amt': 15000, 'net_lot': 25,
         'buy_avg': 3000.0, 'sell_avg': 3000.0, 'close_price': 3000.0,
         'change_pct': 9.9, 'is_limit_up': True},
    ], 'sells': []},
    # 民哥 第 2 分點也買 2330 — 驗證單一 master 多分點 = ≥2 分點
    {'code': '9666', 'name': '富邦-南屯', 'master': '民哥', 'buys': [
        {'code': '2330', 'name': '台積電', 'buy_lot': 60, 'sell_lot': 10,
         'buy_amt': 150000, 'sell_amt': 25000, 'net_lot': 50,
         'buy_avg': 2500.0, 'sell_avg': 2500.0, 'close_price': 2500.0,
         'change_pct': 1.0, 'is_limit_up': False},
    ], 'sells': []},
    {'code': '9658', 'name': '富邦-建國', 'master': '林滄海', 'buys': [
        {'code': '2330', 'name': '台積電', 'buy_lot': 50, 'sell_lot': 10,
         'buy_amt': 125000, 'sell_amt': 25000, 'net_lot': 40,
         'buy_avg': 2500.0, 'sell_avg': 2500.0, 'close_price': 2500.0,
         'change_pct': 1.0, 'is_limit_up': False},
        {'code': '2317', 'name': '鴻海', 'buy_lot': 200, 'sell_lot': 50,
         'buy_amt': 30000, 'sell_amt': 7500, 'net_lot': 150,
         'buy_avg': 150.0, 'sell_avg': 150.0, 'close_price': 150.0,
         'change_pct': 9.85, 'is_limit_up': True},
    ], 'sells': []},
    {'code': '779Z', 'name': '國票-安和', 'master': '張濬安(航海王)', 'buys': [
        {'code': '00919', 'name': '群益台灣精選高息', 'buy_lot': 10000, 'sell_lot': 0,
         'buy_amt': 200000, 'sell_amt': 0, 'net_lot': 10000,
         'buy_avg': 20.0, 'sell_avg': 0.0, 'close_price': 20.0,
         'change_pct': 0.5, 'is_limit_up': False},
    ], 'sells': []},
    {'code': '9217', 'name': '凱基-松山', 'master': '迷你哥/松山哥', 'buys': [
        {'code': '2330', 'name': '台積電', 'buy_lot': 20, 'sell_lot': 20,
         'buy_amt': 50000, 'sell_amt': 50000, 'net_lot': 0,
         'buy_avg': 2500.0, 'sell_avg': 2500.0, 'close_price': 2500.0,
         'change_pct': 1.0, 'is_limit_up': False},
    ], 'sells': []},
    {'code': '9B2a', 'name': '台新-松德', 'master': 'Tradow', 'buys': [
        {'code': '00919', 'name': '群益台灣精選高息', 'buy_lot': 5000, 'sell_lot': 0,
         'buy_amt': 100000, 'sell_amt': 0, 'net_lot': 5000,
         'buy_avg': 20.0, 'sell_avg': 0.0, 'close_price': 20.0,
         'change_pct': 0.5, 'is_limit_up': False},
    ], 'sells': []},
    # 非追蹤 master — 必須被過濾
    {'code': '888A', 'name': '國泰-館前', 'master': '江士勳', 'buys': [
        {'code': '9999', 'name': '不應出現股', 'buy_lot': 99999, 'sell_lot': 0,
         'buy_amt': 9999999, 'sell_amt': 0, 'net_lot': 99999,
         'buy_avg': 100.0, 'sell_avg': 0.0, 'close_price': 100.0,
         'change_pct': 0.0, 'is_limit_up': False},
    ], 'sells': []},
]

# v3.63.6: 構造 10+ 大戶共識個案 (TESTSTK 9988) 驗證 MIN_MASTER_COUNT=10 篩選
# 從 MASTER_MAPPING 取前 10 位每個各 1 個 branch 都買 9988
from src.exports.excel_report import MASTER_MAPPING
for idx, m in enumerate(MASTER_MAPPING[:10]):
    branch_code, branch_name = m['branches'][0]
    branches_data.append({
        'code': f'TEST{idx:02d}', 'name': branch_name, 'master': m['name'], 'buys': [
            {'code': '9988', 'name': '強共識股', 'buy_lot': 100, 'sell_lot': 10,
             'buy_amt': 100000 + idx * 1000, 'sell_amt': 10000, 'net_lot': 90,
             'buy_avg': 100.0, 'sell_avg': 100.0, 'close_price': 100.0,
             'change_pct': 1.0, 'is_limit_up': False},
            # v3.63.7: 同時加 ETF 00919, 10 大戶都買 -> 應該被 ETF filter 剔除
            {'code': '00919', 'name': '群益台灣精選高息', 'buy_lot': 1000, 'sell_lot': 0,
             'buy_amt': 200000, 'sell_amt': 0, 'net_lot': 1000,
             'buy_avg': 20.0, 'sell_avg': 0.0, 'close_price': 20.0,
             'change_pct': 0.5, 'is_limit_up': False},
        ], 'sells': []
    })

# === Render dashboard ===
wb = Workbook()
ws = wb.active
ws.title = 'Dashboard'
build_dashboard_sheet(ws, branches_data, '20260618', ROOT/'data')

# === Re-load for inspection ===
out_path = ROOT/'data'/'reports'/'_xvalidate_dashboard.xlsx'
wb.save(str(out_path))
wb2 = load_workbook(str(out_path))
ws2 = wb2['Dashboard']

print(f"Total rows rendered: {ws2.max_row}\n")

# Helper: get cell value by addr
def cell(addr):
    v = ws2[addr].value
    return v

# === Ground truth (independent calc, only tracked branches) ===
filtered = [b for b in branches_data if b['master'] in TRACKED_MASTERS]
assert len(filtered) == 16, f"filter should yield 16 (6 original + 10 high-consensus), got {len(filtered)}"

# A. 追蹤池摘要 — v3.64.3 4 KPI 重新設計
# Q1 活躍率 / Q2 淨買差 / Q3 強共識股 / Q4 追蹤佔比
from src.exports.excel_report import _compute_consensus_count, TRACKED_MASTERS

def _compute_buy_sell(blist):
    buy = sum((s.get('buy_amt') or 0) for b in blist for s in (b.get('buys') or []))
    seen = set()
    sell = 0
    for b in blist:
        bcode = b.get('code', '')
        for s in (b.get('buys') or []) + (b.get('sells') or []):
            key = (bcode, s.get('code'))
            if key in seen: continue
            seen.add(key)
            sell += (s.get('sell_amt') or 0)
    return buy, sell

gt_total_buy, gt_total_sell = _compute_buy_sell(filtered)
gt_total_net = gt_total_buy - gt_total_sell
gt_active_masters = {b['master'] for b in filtered if b.get('buys') and b['master']}
gt_active_count = len(gt_active_masters)
gt_total_masters = len(TRACKED_MASTERS)
gt_active_ratio = gt_active_count / gt_total_masters
gt_consensus_list = _compute_consensus_count(filtered)
gt_consensus_count = len(gt_consensus_list)
gt_consensus_net = sum(s['total_net_amt'] for s in gt_consensus_list)

# Q4: vs 全市場 — branches_data 內含全部 (filtered = subset)
gt_all_buy, gt_all_sell = _compute_buy_sell(branches_data)   # all = unfiltered
gt_mkt_net_billion = (gt_all_buy - gt_all_sell) / 100000
gt_track_share = gt_total_buy / gt_all_buy if gt_all_buy else 0

print(f"=== Section A: 追蹤池摘要 (v3.64.3 4 KPI 重新設計) ===")
print(f"  GT Q1 活躍率: {gt_active_count}/{gt_total_masters} = {gt_active_ratio:.4f}")
print(f"  GT Q2 淨買差: {gt_total_net/100000:.2f} 億")
print(f"  GT Q3 強共識股: {gt_consensus_count} 檔, 合計淨買 {gt_consensus_net/100000:.2f} 億")
print(f"  GT Q4 追蹤佔比: {gt_track_share:.4f} (市場 {gt_mkt_net_billion:+.2f} 億)")

# 找 Section A header 列
sec_a_start = None
for r in range(3, ws2.max_row + 1):
    v = cell(f'B{r}')
    if v and isinstance(v, str) and ('追蹤池摘要' in v or 'A.' in v):
        sec_a_start = r + 1
        break

if sec_a_start is None:
    err('A', '找到 Section A', None, 'header missing')
else:
    r1, r2 = sec_a_start, sec_a_start + 1
    # v3.64.3 layout: Q1 at C{r1}, Q2 at G{r1} (merged G-I), Q3 at C{r2}, Q4 at G{r2}
    actual = {
        'Q1 活躍率':  cell(f'C{r1}'),
        'Q2 淨買差':  cell(f'G{r1}'),
        'Q3 強共識股': cell(f'C{r2}'),
        'Q4 追蹤佔比': cell(f'G{r2}'),
    }
    print(f"  Excel row {r1}: Q1={actual['Q1 活躍率']!r}, Q2={actual['Q2 淨買差']!r}")
    print(f"  Excel row {r2}: Q3={actual['Q3 強共識股']!r}, Q4={actual['Q4 追蹤佔比']!r}")

    def _approx(a, b, tol=1e-6):
        if a is None or b is None: return a == b
        return abs(float(a) - float(b)) < tol

    if not _approx(actual['Q1 活躍率'], gt_active_ratio):
        err('A', gt_active_ratio, actual['Q1 活躍率'], 'Q1 活躍率')
    if not _approx(actual['Q2 淨買差'], gt_total_net / 100000):
        err('A', gt_total_net / 100000, actual['Q2 淨買差'], 'Q2 淨買差 億')
    if actual['Q3 強共識股'] != gt_consensus_count:
        err('A', gt_consensus_count, actual['Q3 強共識股'], 'Q3 強共識股 count')
    if not _approx(actual['Q4 追蹤佔比'], gt_track_share):
        err('A', gt_track_share, actual['Q4 追蹤佔比'], 'Q4 追蹤佔比')

    # Excel-native 型別檢查
    for label, val in actual.items():
        if val is not None and isinstance(val, str):
            err('A', 'int/float (Excel-native)', f'str: {val!r}', f'{label} 型別不是 numeric')

    # ════════════════════════════════════════════════════════════════
    # v3.64.5 Q5 banner 三層完整審計 (使用者要求 100% 精準度 + 參考資料源)
    # ════════════════════════════════════════════════════════════════
    # Layer 1: Excel 顯示 = daily_signal.json 內容
    # Layer 2: daily_signal 內部公式 chain (raw → level → weight → conf → direction)
    # Layer 3: temp_history.json raw 信號值支持 daily_signal 的 level 標籤
    print(f"\n=== Q5 banner 三層審計 ===")
    r5 = r2 + 1
    q5_val = cell(f'B{r5}')
    q5_fmt = ws2[f'B{r5}'].number_format
    print(f"  Excel Q5 row {r5}: value={q5_val!r}, fmt={q5_fmt!r}")

    if q5_val is not None:
        from src.exports.excel_report import _read_json_safely
        ds = _read_json_safely(ROOT / 'data' / 'daily_signal.json')
        th = _read_json_safely(ROOT / 'data' / 'temp_history.json')

        if not ds:
            print("  [skip] daily_signal.json 不存在 (test env)")
        else:
            md = ds.get('market_direction') or {}
            gt_direction = md.get('direction')
            gt_confidence = float(md.get('confidence_pct') or 0)
            gt_net = float(md.get('net_weight') or 0)
            gt_contributing = md.get('contributing') or []
            gt_top_signal = gt_contributing[0].get('name') if gt_contributing else '—'
            gt_focus_n = len(ds.get('top_focus_stocks') or [])

            # ── Layer 1: Excel cell value 必須 = daily_signal.confidence_pct ──
            if isinstance(q5_val, str):
                err('Q5.L1', 'numeric', f'str: {q5_val!r}', 'cell type')
            elif not _approx(float(q5_val), gt_confidence, tol=0.01):
                err('Q5.L1', gt_confidence, q5_val, 'banner confidence_pct')
            else:
                print(f"  [L1 PASS] Excel value {q5_val} = daily_signal.confidence_pct {gt_confidence}")

            # ── Layer 1.b: Format string 必須含正確的 direction/arrow/top_signal/focus_n ──
            arrow_map = {'偏多': '↑', '偏空': '↓', '中性': '↕'}
            expected_arrow = arrow_map.get(gt_direction, '↕')
            for expected_substr, label in [
                ('明日預測', '解讀前綴'),   # v3.64.6: 明日預測 vs 今日市場
                (expected_arrow, 'arrow'),
                (gt_direction, 'direction'),
                (gt_top_signal, 'top_signal'),
                (f'{gt_focus_n} 檔焦點', 'focus_n'),
            ]:
                if expected_substr not in (q5_fmt or ''):
                    err('Q5.L1', expected_substr, q5_fmt, f'format string 缺 {label}')
                else:
                    print(f"  [L1 PASS] format 含 '{expected_substr}' ({label})")

            # ── Layer 2: 內部公式 chain ──
            # 2a: confidence = clamp(10, 95, 50 + net*100)
            expected_conf = max(10, min(95, 50 + gt_net * 100))
            if not _approx(expected_conf, gt_confidence, tol=0.1):
                err('Q5.L2', f'{expected_conf:.1f}', f'{gt_confidence:.1f}',
                    f'公式 conf=50+net*100 (net={gt_net})')
            else:
                print(f"  [L2 PASS] 公式 50 + {gt_net}*100 = {expected_conf:.1f} ≈ {gt_confidence}")

            # 2b: direction 由 net thresholds 決定
            if gt_net > 0.05:
                expected_dir = '偏多'
            elif gt_net < -0.05:
                expected_dir = '偏空'
            else:
                expected_dir = '中性'
            if gt_direction != expected_dir:
                err('Q5.L2', expected_dir, gt_direction,
                    f'direction threshold (net={gt_net}, ±0.05)')
            else:
                print(f"  [L2 PASS] direction '{gt_direction}' 對應 net {gt_net} 之 ±0.05 threshold")

            # 2c: net = sum of contributing weights
            recompute_net = sum(c.get('weight', 0) for c in gt_contributing)
            if not _approx(recompute_net, gt_net, tol=0.001):
                err('Q5.L2', recompute_net, gt_net, 'net = Σ contributing.weight')
            else:
                print(f"  [L2 PASS] net {gt_net} = Σ contributing weights {recompute_net}")

            # ── Layer 3: temp_history raw value → level → weight ──
            if th:
                history = th.get('history') or []
                if history:
                    # 找出對應 trade_date 的 entry
                    latest_entry = None
                    for e in history:
                        if e.get('date') == ds.get('date'):
                            latest_entry = e
                            break
                    if latest_entry is None:
                        latest_entry = history[-1]  # fallback
                    th_signals = latest_entry.get('signals') or []
                    print(f"  Layer 3: temp_history entry {latest_entry.get('date')} 有 {len(th_signals)} 信號")

                    # 對每個 contributing signal 驗證 temp_history raw → level → weight
                    try:
                        from src.analyzers.signal_engine import _signal_weight
                        from src.pipelines.crawler_pipeline import TEMP_THRESHOLDS, _temp_signal_score

                        for sig in gt_contributing:
                            name = sig.get('name')
                            ds_level = sig.get('level')
                            ds_weight = sig.get('weight')

                            raw_sig = next((s for s in th_signals if s.get('name') == name), None)
                            if not raw_sig:
                                err('Q5.L3', f'found {name}', 'not in temp_history',
                                    f'{name} 在 temp_history 找不到')
                                continue
                            raw_val = raw_sig.get('value')
                            th_level = raw_sig.get('level')

                            # 3a: temp_history.level == daily_signal.level
                            if th_level != ds_level:
                                err('Q5.L3', th_level, ds_level,
                                    f'{name} level: temp_history vs daily_signal')
                            else:
                                print(f"  [L3 PASS] {name}: temp_history level '{th_level}' == daily_signal level")

                            # 3b: SIGNAL_WEIGHTS lookup
                            sig_data = {'level': th_level, 'value': raw_val}
                            expected_weight = _signal_weight(name, sig_data)
                            if not _approx(expected_weight, ds_weight, tol=0.001):
                                err('Q5.L3', expected_weight, ds_weight,
                                    f'{name} weight: SIGNAL_WEIGHTS[{name}][{th_level}] vs daily_signal.weight')
                            else:
                                print(f"  [L3 PASS] {name} weight {ds_weight} = SIGNAL_WEIGHTS lookup")

                            # 3c: For P/C Ratio, verify raw → level via TEMP_THRESHOLDS
                            if name == 'P/C Ratio' and isinstance(raw_val, (int, float)):
                                thr = TEMP_THRESHOLDS['pc_ratio_oi']
                                expected = _temp_signal_score(raw_val, thr)
                                if expected and expected[1] != th_level:
                                    err('Q5.L3', expected[1], th_level,
                                        f'P/C Ratio threshold: value={raw_val} thresholds={thr}')
                                else:
                                    print(f"  [L3 PASS] P/C Ratio value {raw_val} → level '{th_level}' 符合 thresholds {thr}")
                    except Exception as _e:
                        err('Q5.L3', 'audit ok', f'import error: {_e}', 'cannot run weight verification')

# Check Section B / C
print(f"\n=== Section B: Top 5 高手 ===")
gt_master_amt = {}
for b in filtered:
    m = b['master']
    amt = sum(s['buy_amt'] for s in b['buys'])
    gt_master_amt[m] = gt_master_amt.get(m, 0) + amt
gt_top_masters = sorted(gt_master_amt.items(), key=lambda x: -x[1])
gt_total_all = sum(gt_master_amt.values())
print(f"  GT total_all = {gt_total_all} 千元")
for i, (m, amt) in enumerate(gt_top_masters):
    pct = amt/gt_total_all*100
    print(f"  GT  #{i+1}: {m} = {amt} 千元 (= {round(amt/10,0)} 萬), {pct:.1f}%")

# Find Section B start (look for "Top 5 高手" header)
b_start = None
for r in range(4, 20):
    v = cell(f'B{r}')
    if v and isinstance(v, str) and 'Top 5 高手' in v:
        b_start = r + 2  # +1 hdr +1 col-header
        break
print(f"  Section B data starts at row {b_start}")
if b_start:
    for i in range(5):
        rank = cell(f'B{b_start+i}')
        name = cell(f'C{b_start+i}')
        amt_wan = cell(f'D{b_start+i}')
        pct = cell(f'E{b_start+i}')
        if name is None:
            print(f"  Excel #{i+1}: (empty)")
            continue
        # GT for this rank
        if i < len(gt_top_masters):
            gt_m, gt_a = gt_top_masters[i]
            gt_wan = round(gt_a / 10, 0)
            gt_pct_str = f"{gt_a/gt_total_all*100:.1f}%"
            ok_name = (name == gt_m)
            ok_amt = (amt_wan == gt_wan)
            ok_pct = (pct == gt_pct_str)
            print(f"  Excel #{i+1}: {name}={amt_wan}萬, {pct}")
            if not ok_name: err('B', gt_m, name, f'#{i+1} master')
            if not ok_amt: err('B', gt_wan, amt_wan, f'#{i+1} amt')
            if not ok_pct: err('B', gt_pct_str, pct, f'#{i+1} pct')

# v3.65.0 Section C: Top 5 熱門個股 — 排除 ETF + 新 layout
# G = "name(code)" 合併, H = 淨買(萬), I = 漲跌% (decimal)
print(f"\n=== Section C: Top 5 熱門個股 strict (v3.65.0 ETF excluded) ===")
from src.exports.excel_report import _is_excluded_by_market_type
gt_stock_net = {}
gt_stock_name = {}
gt_stock_change = {}
for b in filtered:
    for s in b.get('buys', []):
        c = s.get('code')
        if not c: continue
        # v3.65.0: 排除 ETF — code 起始 '00' or market_type 在 ETF/債券等
        if _is_excluded_by_market_type(s):
            continue
        gt_stock_net[c] = gt_stock_net.get(c, 0) + (s.get('buy_amt', 0) or 0) - (s.get('sell_amt', 0) or 0)
        gt_stock_name[c] = s.get('name', '')
        if s.get('change_pct') is not None and c not in gt_stock_change:
            gt_stock_change[c] = s.get('change_pct')
gt_top_stocks = sorted(gt_stock_net.items(), key=lambda x: -x[1])[:5]
print(f"  GT top 5 stocks (淨買 desc, ETF 排除):")
for i, (c, net) in enumerate(gt_top_stocks):
    print(f"    #{i+1} {gt_stock_name.get(c, '?')}({c}) = {round(net/10)}萬")
if b_start:
    # v3.65.0: G = "name(code)" 一格, H = 淨買, I = 漲跌% (decimal)
    for i in range(5):
        r = b_start + i
        actual_rank = cell(f'F{r}')
        actual_label = cell(f'G{r}')   # name(code)
        actual_amt = cell(f'H{r}')     # 淨買萬
        actual_chg = cell(f'I{r}')     # 漲跌% decimal
        if i < len(gt_top_stocks):
            gt_c, gt_net = gt_top_stocks[i]
            gt_wan = round(gt_net / 10, 0)
            gt_label = f"{gt_stock_name.get(gt_c, '')}({gt_c})"
            if actual_rank != i + 1:
                err('C', i + 1, actual_rank, f'#{i+1} rank')
            if actual_label != gt_label:
                err('C', gt_label, actual_label, f'#{i+1} name(code)')
            if actual_amt != gt_wan:
                err('C', gt_wan, actual_amt, f'#{i+1} amt')
            # 漲跌% 若 GT 有, decimal 比對 (1e-4 tol)
            gt_chg = gt_stock_change.get(gt_c)
            if gt_chg is not None and actual_chg is not None and not isinstance(actual_chg, str):
                expected_chg_decimal = gt_chg / 100
                if abs(float(actual_chg) - expected_chg_decimal) > 1e-4:
                    err('C', expected_chg_decimal, actual_chg, f'#{i+1} change_pct')

# v3.63.7: ETF (00919) 即使 10 大戶買也應被剔除
print(f"\n=== ETF 排除驗證 ===")
# 直接掃 ws2 整張 sheet 內容找 00919 字串
etf_text = []
for row_cells in ws2.iter_rows(values_only=True):
    for v in row_cells:
        if v is not None:
            etf_text.append(str(v))
sec0_text = []
for r in range(4, ws2.max_row+1):
    v = cell(f'B{r}')
    if v and isinstance(v, str) and '強共識買超' in v:
        # 從這行 +1 (col header) +1 開始, 到下個 section header 為止
        for rr in range(r+2, min(r+50, ws2.max_row+1)):
            for c in 'BCDEFGHIJKLMN':
                vv = cell(f'{c}{rr}')
                if vv is not None:
                    sec0_text.append(str(vv))
            # 偵測下個 section header (▍ 開頭)
            vv = cell(f'B{rr}')
            if vv and isinstance(vv, str) and vv.startswith('▍'):
                break
        break
sec0_str = ' | '.join(sec0_text)
if '00919' in sec0_str:
    err('ETF_FILTER', '不應出現', '00919', 'ETF 出現在 Section 0')
else:
    print(f"  ✓ 00919 群益台灣精選高息 (ETF) 已被 Section 0 排除")

# Check 江士勳 (non-tracked) does NOT appear anywhere
print(f"\n=== 非追蹤 master 過濾驗證 ===")
forbidden = ['江士勳', '何莎', '優式資本', '東億資本', 'Krenz(再多一位數本人)', '志誠資本',
             '林適中', '謝明彧大哥(華南永昌)', '宋福祥', '呂金發', '陳光裕', '謝孟恭(股癌)',
             '丁凌全', '江士勳', '劉子豪', '陳泊澔', '嘉義幫']
all_text = []
for row_cells in ws2.iter_rows(values_only=True):
    for v in row_cells:
        if v is not None:
            all_text.append(str(v))
all_str = ' | '.join(all_text)
for m in forbidden:
    if m in all_str:
        err('FILTER', '不應出現', m, '非追蹤大戶洩漏')
    else:
        print(f"  ✓ {m} 未出現")

# Section E v3.66.0: 砍 consensus + accumulation, 只留 🔴 量爆 + 🆕 新標的 (top 10)
# 排序用 _anomaly_severity (volume_spike=|z|, new_stocks=2.5+count*0.2)
print(f"\n=== Section E v3.66.0: 量爆+新標的 (top 10, severity sort) ===")
from src.exports.excel_report import _anomaly_severity
all_anom = [s for s in (signals.get('anomalies') or [])
            if s.get('master') in TRACKED_MASTERS]
all_anom.sort(key=lambda x: -_anomaly_severity(x))
gt_top10 = all_anom[:10]
print(f"  GT tracked anomalies: {len(all_anom)} 筆, top 10:")
for i, sig in enumerate(gt_top10):
    print(f"    #{i+1} [{sig.get('type')}] {sig.get('master')} sev={_anomaly_severity(sig):.2f}")

# Excel 內 🔴 量爆 或 🆕 新標的 都在 Section E
e_rows = []
for r in range(4, ws2.max_row+1):
    v = cell(f'B{r}')
    if v in ('🔴 量爆', '🆕 新標的'):
        e_rows.append(r)
print(f"  Excel E rows (excluding footer): {len(e_rows)}")
expected_n = len(gt_top10)
if len(e_rows) != expected_n:
    err('E', expected_n, len(e_rows), 'E 筆數 (top 10)')
for i, (r, gt_sig) in enumerate(zip(e_rows, gt_top10)):
    t = gt_sig.get('type')
    if t == 'new_stocks':
        if cell(f'B{r}') != '🆕 新標的':
            err('E', '🆕 新標的', cell(f'B{r}'), f'#{i+1} type label')
        if cell(f'C{r}') != gt_sig.get('master'):
            err('E', gt_sig.get('master'), cell(f'C{r}'), f'#{i+1} master')
        count = gt_sig.get('count', 0) or 0
        expected_sev = 'high' if count >= 5 else 'medium'
        if cell(f'D{r}') != expected_sev:
            err('E', expected_sev, cell(f'D{r}'), f'#{i+1} severity (new_stocks)')
        # F 是 "{count} 檔" 字串
        if cell(f'F{r}') != f'{count} 檔':
            err('E', f'{count} 檔', cell(f'F{r}'), f'#{i+1} count')
    else:
        if cell(f'B{r}') != '🔴 量爆':
            err('E', '🔴 量爆', cell(f'B{r}'), f'#{i+1} type label')
        if cell(f'C{r}') != gt_sig.get('master'):
            err('E', gt_sig.get('master'), cell(f'C{r}'), f'#{i+1} master')
        expected_sev = _severity_from_z(gt_sig.get('z_score'))
        if cell(f'D{r}') != expected_sev:
            err('E', expected_sev, cell(f'D{r}'), f'#{i+1} severity (volume_spike)')
        expected_amt = _round_safe(gt_sig.get('today_buy_amt_wan'))
        if cell(f'F{r}') != expected_amt:
            err('E', expected_amt, cell(f'F{r}'), f'#{i+1} amount')

# Section F v3.66.0 hot 標記: 連續 ≥10 天 master cell 加 "🔴 " prefix
# v3.65.0: ETF (code 起始 '00') 排除
print(f"\n=== Section F: 跨日連續囤貨 全列驗證 (v3.65.0 ETF excluded + v3.66.0 hot 標記) ===")
HOT_DAYS = 10
all_acc_sorted = sorted(
    [s for s in (signals.get('accumulations') or [])
     if s.get('master') in TRACKED_MASTERS
     and not (s.get('stock_code') or '').startswith('00')],
    key=lambda x: -x.get('consecutive_days', 0)
)[:30]
print(f"  GT Section F: {len(all_acc_sorted)} 筆 (top 30, ETF 排除)")
hot_n = sum(1 for s in all_acc_sorted if (s.get('consecutive_days', 0) or 0) >= HOT_DAYS)
print(f"  GT 🔴 hot (≥{HOT_DAYS} 天): {hot_n} 筆")
f_start = None
for r in range(4, ws2.max_row+1):
    v = cell(f'B{r}')
    if v and isinstance(v, str) and '跨日連續囤貨' in v:
        f_start = r + 2  # +1 hdr +1 col-header
        break
if f_start:
    for i, gt_sig in enumerate(all_acc_sorted):
        r = f_start + i
        days = gt_sig.get('consecutive_days', 0) or 0
        # v3.66.0: ≥10 天 master cell prefix '🔴 '
        is_hot = days >= HOT_DAYS
        expected_master = f'🔴 {gt_sig.get("master")}' if is_hot else gt_sig.get('master')
        if cell(f'B{r}') != expected_master:
            err('F', expected_master, cell(f'B{r}'), f'#{i+1} master (hot={is_hot})')
        if cell(f'C{r}') != gt_sig.get('stock_code'):
            err('F', gt_sig.get('stock_code'), cell(f'C{r}'), f'#{i+1} stock_code')
        if cell(f'D{r}') != days:
            err('F', days, cell(f'D{r}'), f'#{i+1} days')
        if cell(f'E{r}') != _round_safe(gt_sig.get('total_buy_amt_wan')):
            err('F', _round_safe(gt_sig.get('total_buy_amt_wan')), cell(f'E{r}'), f'#{i+1} amt')

# Section J: Master × Top 3 (v3.65.0: ETF 排除)
print(f"\n=== Section J: Master × Top 3 cross-table 驗證 (v3.65.0 ETF excluded) ===")
gt_master_stocks = {}
for b in filtered:
    m = b['master']
    for s in b['buys']:
        # v3.65.0: 排 ETF
        if _is_excluded_by_market_type(s):
            continue
        gt_master_stocks.setdefault(m, {})
        k = (s['code'], s['name'])
        gt_master_stocks[m][k] = gt_master_stocks[m].get(k, 0) + s['buy_amt']
gt_rows = []
for m, stocks in gt_master_stocks.items():
    sorted_s = sorted(stocks.items(), key=lambda kv: -kv[1])
    total = sum(stocks.values())
    gt_rows.append({'master': m, 'total': total, 'top': sorted_s[:3]})
gt_rows.sort(key=lambda x: -x['total'])
print(f"  GT top master in J: {gt_rows[0]['master']}, total={round(gt_rows[0]['total']/10,0)}萬")
print(f"    Top stocks: {[(f'{n}({c})', round(a/10,0)) for (c,n),a in gt_rows[0]['top']]}")

# Find J section
j_start = None
for r in range(4, ws2.max_row+1):
    v = cell(f'B{r}')
    if v and isinstance(v, str) and 'cross-table' in v:
        j_start = r + 2
        break
if j_start:
    # v3.64.6: 驗證全部 rows + top 3 stocks 名稱與金額
    for i, gt in enumerate(gt_rows[:30]):
        r = j_start + i
        actual_master = cell(f'B{r}')
        actual_total = cell(f'C{r}')
        expected_total = round(gt['total']/10, 0)
        if actual_master != gt['master']:
            err('J', gt['master'], actual_master, f'#{i+1} master')
        if actual_total != expected_total:
            err('J', expected_total, actual_total, f'#{i+1} total')
        # Verify each top-3 stock + amount (cols D/E for #1, F/G for #2, H/I for #3)
        for slot_idx, ((c, n), amt) in enumerate(gt['top']):
            name_col = chr(ord('D') + slot_idx * 2)  # D F H
            amt_col = chr(ord('E') + slot_idx * 2)   # E G I
            expected_name_code = f"{n}({c})"
            expected_amt = round(amt/10, 0)
            if cell(f'{name_col}{r}') != expected_name_code:
                err('J', expected_name_code, cell(f'{name_col}{r}'),
                    f'#{i+1} top{slot_idx+1} name')
            if cell(f'{amt_col}{r}') != expected_amt:
                err('J', expected_amt, cell(f'{amt_col}{r}'),
                    f'#{i+1} top{slot_idx+1} amt')

# === Section 0: 共同買超 (≥2 分點) verify ===
print(f"\n=== ★ Section 0: 共同買超 (≥2 分點) cross-validate ===")
# Ground truth: 從 filtered branches_data 獨立計算, 以分點為單位
gt_stock_map = {}
for b in filtered:
    m = b['master']
    b_code = b.get('code', '')
    for s in b['buys'] + b.get('sells', []):
        net = (s.get('buy_amt') or 0) - (s.get('sell_amt') or 0)
        if net <= 0: continue
        e = gt_stock_map.setdefault(s['code'], {'name': s.get('name',''), 'branches': []})
        e['branches'].append({'branch_code': b_code, 'master': m, 'net_amt': net})
gt_consensus = []
MIN_MASTER_COUNT = 10  # v3.63.6
for code, info in gt_stock_map.items():
    # v3.63.7: 排除 ETF
    if code.startswith('00'):
        continue
    if len(info['branches']) < 2:
        continue
    master_set = {br['master'] for br in info['branches']}
    if len(master_set) < MIN_MASTER_COUNT:
        continue
    gt_consensus.append({
        'code': code, 'name': info['name'],
        'branch_count': len(info['branches']),
        'master_count': len(master_set),
        'masters': master_set,
        'total': sum(br['net_amt'] for br in info['branches']),
    })
gt_consensus.sort(key=lambda x: (-x['master_count'], -x['branch_count'], -x['total']))

print(f"  GT consensus count: {len(gt_consensus)}")
for i, c in enumerate(gt_consensus[:5]):
    print(f"  GT #{i+1}: {c['code']}({c['name']}), 分點{c['branch_count']}/大戶{c['master_count']}, "
          f"{int(round(c['total']/10))}萬, masters={c['masters']}")

# Find Section 0 in Excel
sec0_start = None
for r in range(4, ws2.max_row+1):
    v = cell(f'B{r}')
    # v3.63.6+: 標題含「強共識買超」或舊「共同買超」
    if v and isinstance(v, str) and ('強共識買超' in v or '共同買超' in v):
        # v3.63.9: 多了註腳 row → +1 hdr, +1 note, +1 col-header = +3
        sec0_start = r + 3
        break
print(f"  Section 0 data starts at row {sec0_start}")
if sec0_start and gt_consensus:
    # v3.63.8 Excel-native 13 欄佈局
    # B=#, C=代號, D=名稱, E=大戶數, F=分點數,
    # G=領頭大戶, H=領頭金額, I=#2 大戶, J=#2 金額, K=#3 大戶, L=#3 金額,
    # M=+更多, N=合計淨買
    def _short(m): return m.split('(')[0].split('/')[0]
    for i in range(min(3, len(gt_consensus))):
        r = sec0_start + i
        actual_rank = cell(f'B{r}')
        actual_code = cell(f'C{r}')
        actual_mc = cell(f'E{r}')
        actual_bc = cell(f'F{r}')
        actual_lead_name = cell(f'G{r}')
        actual_lead_amt = cell(f'H{r}')
        actual_2_name = cell(f'I{r}')
        actual_2_amt = cell(f'J{r}')
        actual_3_name = cell(f'K{r}')
        actual_3_amt = cell(f'L{r}')
        actual_tail = cell(f'M{r}')
        actual_total = cell(f'N{r}')
        gt = gt_consensus[i]
        gt_total = int(round(gt['total']/10))
        print(f"  Excel #{i+1}: code={actual_code}, 大戶={actual_mc}, 分點={actual_bc}, "
              f"領頭={actual_lead_name}({actual_lead_amt}萬), #2={actual_2_name}({actual_2_amt}萬), "
              f"#3={actual_3_name}({actual_3_amt}萬), +{actual_tail}, 合計={actual_total}萬")
        if actual_rank != i+1: err('0', i+1, actual_rank, f'#{i+1} rank')
        if actual_code != gt['code']: err('0', gt['code'], actual_code, f'#{i+1} code')
        if actual_bc != gt['branch_count']: err('0', gt['branch_count'], actual_bc, f'#{i+1} branch_count')
        if actual_mc != gt['master_count']: err('0', gt['master_count'], actual_mc, f'#{i+1} master_count')
        if actual_total != gt_total: err('0', gt_total, actual_total, f'#{i+1} total')
        # 領頭+#2+#3 名字必須都在 gt masters 內
        gt_short = {_short(m) for m in gt['masters']}
        for nm, label in [(actual_lead_name, 'lead'), (actual_2_name, '#2'), (actual_3_name, '#3')]:
            if nm and nm not in gt_short:
                err('0', f'in {gt_short}', nm, f'#{i+1} {label} master not in masters')
        # 領頭+#2+#3+tail 名額總和 = master_count
        visible = sum(1 for x in [actual_lead_name, actual_2_name, actual_3_name] if x)
        tail = actual_tail if isinstance(actual_tail, int) else 0
        if visible + tail != gt['master_count']:
            err('0', gt['master_count'], visible + tail,
                f'#{i+1} visible({visible}) + tail({tail}) != master_count')
        # 領頭金額 ≥ #2 金額 ≥ #3 金額 (排序正確性)
        amts = [a for a in [actual_lead_amt, actual_2_amt, actual_3_amt] if isinstance(a, int)]
        if amts != sorted(amts, reverse=True):
            err('0', 'desc sorted', amts, f'#{i+1} amount sort')

# === Result ===
print(f"\n{'='*60}")
if errors:
    print(f"❌ FAIL: {len(errors)} mismatches:")
    for sec, exp, act, note in errors:
        print(f"  [{sec}] {note}: expected={exp!r}, actual={act!r}")
else:
    print(f"✅ PASS: 全部 cross-validate 通過")
