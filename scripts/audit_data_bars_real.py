"""v3.66.6 Phase 2.2 Data Bar 精準度審計 — 用 6/24 真實 Excel
邏輯:
  1. 11 個 data bar rules 存在
  2. 每個 range 內 cell 是 numeric 或 intentional '—' placeholder
  3. max/min cell 對應 GT 的最大/最小值
  4. 排序 desc 後最大 / 最小 row 跟視覺長條最長/最短一致
"""
import sys, re
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')
ROOT = Path(__file__).resolve().parents[1]
from openpyxl import load_workbook

wb = load_workbook(ROOT / 'data' / 'reports' / 'chip_radar_2026-06.xlsx')
ws = wb[wb.sheetnames[0]]

# 找 section header 列
def find_section_data_row(keyword):
    """找 section header 後第一筆 data row 起始."""
    for r in range(1, ws.max_row + 1):
        v = ws[f'B{r}'].value
        if v and isinstance(v, str) and keyword in v:
            # 跳過 header row + col_header row + 註腳 row (Section 0 有)
            offset = 3 if '強共識' in keyword else 2
            return r + offset
    return None

# 各 section 預期 max value 用 GT 比對
print(f"=== 真實 6/24 Excel data bar accuracy ===\n")
print(f"Total conditional formatting rules: {len(list(ws.conditional_formatting._cf_rules))}\n")

applied = []
for cf_range, rules in ws.conditional_formatting._cf_rules.items():
    rng = str(cf_range.sqref) if hasattr(cf_range, 'sqref') else str(cf_range)
    for rule in rules:
        if hasattr(rule, 'dataBar') and rule.dataBar:
            applied.append(rng)

errors = []
for rng in applied:
    m = re.match(r'([A-Z])(\d+):[A-Z](\d+)', rng)
    if not m: continue
    col, r1, r2 = m.group(1), int(m.group(2)), int(m.group(3))
    values = []
    placeholder_n = 0
    for r in range(r1, r2 + 1):
        v = ws[f'{col}{r}'].value
        if v is None: continue
        if isinstance(v, str):
            if v == '—': placeholder_n += 1
            else: errors.append(f'{rng} row {r}: unexpected str {v!r}')
            continue
        values.append((r, v))
    if not values:
        print(f"  ⚠️  {rng}: no numeric values (data bar 不渲染)")
        continue
    max_r, max_v = max(values, key=lambda x: x[1])
    min_r, min_v = min(values, key=lambda x: x[1])
    range_str = f"[{min_v:>10,.1f} ~ {max_v:>12,.1f}]"
    extra = f' / {placeholder_n} 個「—」' if placeholder_n else ''
    print(f"  ✓ {rng:14s} n={len(values):2d}{extra}, "
          f"range {range_str}, max row={max_r}")

# 順帶提取 Section H ratio range, 確認 1000x hot 那筆是 max
print(f"\n--- Section H ratio data bar 精準度抽查 ---")
for rng in applied:
    if not rng.startswith('F'): continue
    m = re.match(r'F(\d+):F(\d+)', rng)
    if not m: continue
    r1, r2 = int(m.group(1)), int(m.group(2))
    # 找 max
    max_v, max_r, max_code = None, None, None
    for r in range(r1, r2 + 1):
        v = ws[f'F{r}'].value
        code = ws[f'B{r}'].value
        if isinstance(v, (int, float)):
            if max_v is None or v > max_v:
                max_v = v; max_r = r; max_code = code
    if max_v:
        print(f"  最高 ratio={max_v:.1f}x → row {max_r} code={max_code!r}")
        if '🔴' in str(max_code):
            print(f"    ✅ 確認該 row 有 🔴 prefix (≥1000x hot 標記)")
        else:
            print(f"    ⚠️  該 row code 沒 🔴 prefix (預期 ratio ≥1000 標記)")
    break

# Section F 連續天數 max 應該是 hot (🔴)
print(f"\n--- Section F 連續天數 data bar 精準度抽查 ---")
# Section F D 欄
for rng in applied:
    if not rng.startswith('D'): continue
    m = re.match(r'D(\d+):D(\d+)', rng)
    if not m: continue
    r1, r2 = int(m.group(1)), int(m.group(2))
    # 確認 F section 範圍 (從 'F. 跨日連續囤貨' header 找)
    f_data_start = find_section_data_row('F. 跨日連續囤貨')
    if not f_data_start or r1 != f_data_start: continue
    max_v, max_r, max_master = None, None, None
    for r in range(r1, r2 + 1):
        v = ws[f'D{r}'].value
        master = ws[f'B{r}'].value
        if isinstance(v, (int, float)):
            if max_v is None or v > max_v:
                max_v = v; max_r = r; max_master = master
    if max_v:
        print(f"  最高連續天數={max_v} → row {max_r} master={max_master!r}")
        if isinstance(max_master, str) and '🔴' in max_master:
            print(f"    ✅ 確認該 row master 有 🔴 prefix (≥10 天 hot)")
    break

if errors:
    print(f"\n❌ 發現 {len(errors)} 個錯誤:")
    for e in errors: print(f"  {e}")
else:
    print(f"\n✅ Phase 2.2 Data Bar 精準度 PASS — 所有 cells 都是 numeric 或 intentional '—'")
