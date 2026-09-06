"""v3.74.0: 本機排程跑完後,把「📱 手機摘要」+ latest.xlsx 推到 Telegram。
同一個交易日內重跑改為「編輯既有訊息」而非重發。

⚠️ 處置股不在這裡 — 由 scripts/push_disposal_telegram.py 負責 (v3.74.0 起)。
  這支腳本曾自行渲染處置股圖卡,但那是重造輪子: disposal-watch
  (github.com/DRwillychiu/disposal-watch) 早就有完整管線 —
  attstock 撈取 → 反解觸發價 / MoneyDJ 概念標籤 / 快照差集算進出關 /
  命中率回測 → Pillow 產圖 → 寄 Email,排程同為台灣 21:17。
  這裡再畫一份只會得到兩個版本的處置股,故整段移除。
  v3.75.0 起圖卡改由 push_disposal_telegram.py 下載該專案的雲端 artifact
  (四張一組),不在本機重算;收件對象預設跟這支一樣走 TELEGRAM_CHAT_ID。
  → data/disposal_attstock.json 仍由 refresh_attstock_disposal.py 更新,
    因為 excel_report.py 出報表還要讀它 (不是給 Telegram 用的)。

跟雲端 daily-full.yml 的「Send daily summary to Telegram」等價,但多一件事:
  兜底排程重跑時走 editMessageText / editMessageMedia 更新原訊息

為什麼要「更新」而不是「跳過」或「重發」:
  daily-full 有 21:17 / 22:37 / 23:47 三層兜底,每層都完整跑一次爬蟲,
  而後面幾次的資料可能更完整 (實測 21:17 抓到 76 分點、22:37 抓到 77)。
  重發會洗版,單純跳過又拿不到更新後的數字 → 折衷是原地編輯。

多目標 (v3.75.0):
  TELEGRAM_CHAT_ID 可用逗號分隔多個對象,例如「私訊,群組」都要收同一份。
  message_id 是「該 chat 專屬」的,拿 A chat 的 id 去 B chat 編輯會直接失敗,
  所以狀態必須按 chat_id 分開存 (見下方 targets)。

狀態檔 local_data/.tg_last_push (JSON):
  {"trade_date": "20260730",
   "targets": {
     "987654321":  {"summary": {"message_id": 123, "hash": "..."},
                     "document": {"message_id": 125, "hash": "..."}},
     "-1001234567890": {"summary": {...}, "document": {...}}}}
  相容兩種舊格式:
    v3.73.1-3 純文字 (只有 trade_date,無 message_id → 該次改為重發)
    v3.73.4-  summary/document 放在頂層 → 自動歸給第一個 chat_id
  舊狀態檔殘留的 "disposal" 鍵不再讀寫,無害。

用法:
    python scripts/send_daily_telegram.py            # 正常 (新發 or 更新)
    python scripts/send_daily_telegram.py --force    # 略過「內容未變」短路,硬更新一次
    python scripts/send_daily_telegram.py --dry-run  # 只印不推

token 來源 (優先序): 環境變數 → 專案根目錄 .env
兩者皆無 → 印訊息後 exit 0 (不讓排程判定為失敗)
"""
import hashlib
import json
import os
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

FORCE = '--force' in sys.argv
DRY_RUN = '--dry-run' in sys.argv


def load_dotenv_into_env() -> None:
    """把 .env 的值補進 os.environ (不覆蓋既有的) — 對齊 scheduler.ps1:14-26."""
    env_file = ROOT / '.env'
    if not env_file.exists():
        return
    for line in env_file.read_text(encoding='utf-8-sig').splitlines():
        line = line.strip()
        if not line or line.startswith('#') or '=' not in line:
            continue
        k, v = line.split('=', 1)
        k, v = k.strip(), v.strip()
        if v and not os.environ.get(k):
            os.environ[k] = v


# 必須在算 DATA_DIR 之前載入: 排程執行時 scheduler.ps1 已設好
# CHIP_RADAR_DATA_DIR (=local_data),手動執行則只有 .env 有。晚一步載入
# 會讓兩種跑法讀到不同的資料夾與狀態檔,同一天可能因此重發一次。
load_dotenv_into_env()

DATA_DIR = ROOT / os.environ.get('CHIP_RADAR_DATA_DIR', 'data')
XLSX = DATA_DIR / 'reports' / 'latest.xlsx'
LATEST = DATA_DIR / 'latest.json'
STATE = DATA_DIR / '.tg_last_push'

MOBILE_SHEET = '📱 手機摘要'


# ════════════════════════════════════════════════════════════════════
#  基礎工具
# ════════════════════════════════════════════════════════════════════

def _hash(s: str) -> str:
    return hashlib.sha256(s.encode('utf-8')).hexdigest()[:16]


def read_trade_date() -> str:
    """從 latest.json 的明文外層讀 trade_date (不需密碼)."""
    try:
        return str(json.loads(LATEST.read_text(encoding='utf-8')).get('trade_date', ''))
    except Exception as e:
        print(f"  ⚠️ 讀不到 {LATEST}: {e}")
        return ''


def load_state() -> dict:
    """讀狀態檔。相容 v3.73.1-3 的純文字格式 (只有 trade_date)."""
    if not STATE.exists():
        return {}
    raw = ''
    try:
        raw = STATE.read_text(encoding='utf-8').strip()
    except Exception:
        return {}
    if not raw:
        return {}
    try:
        st = json.loads(raw)
    except Exception:
        st = None
    # 注意: 舊格式內容是純日期字串 "20260730",而它本身就是合法 JSON (數字),
    # json.loads 會成功解析成 int → 必須用 isinstance 檢查而非靠例外分支,
    # 否則舊狀態檔會被誤判為「無狀態」,同一天內重發而非更新。
    if isinstance(st, dict):
        return st
    return {'trade_date': raw}


def save_state(st: dict) -> None:
    try:
        STATE.write_text(json.dumps(st, ensure_ascii=False), encoding='utf-8')
    except Exception as e:
        print(f"  [Telegram] ⚠️ 狀態檔寫入失敗 (下次可能重發): {e}")


def chat_ids() -> list:
    """TELEGRAM_CHAT_ID 支援逗號分隔多個對象 (私訊 + 群組都要收同一份)。

    順序有意義: 第一個是「主要對象」,舊狀態檔的頂層 summary/document
    會歸給它 (見 migrate_targets)。
    """
    raw = os.environ.get('TELEGRAM_CHAT_ID', '')
    return [c.strip() for c in raw.split(',') if c.strip()]


def migrate_targets(st: dict, ids: list) -> dict:
    """把舊的頂層 summary/document 搬進 targets[<第一個 chat_id>]。

    不搬的話,加了第二個目標之後主要對象會被當成「全新目標」而重發一次,
    等於洗版 —— 而那正是 v3.73.4 引入 message_id 想避免的事。
    """
    tg = st.get('targets')
    if isinstance(tg, dict):
        return tg
    tg = {}
    legacy = {k: st[k] for k in ('summary', 'document')
              if isinstance(st.get(k), dict)}
    if legacy and ids:
        tg[ids[0]] = legacy
    st['targets'] = tg
    for k in ('summary', 'document'):
        st.pop(k, None)
    return tg


def extract_mobile_summary() -> str:
    """抓 latest.xlsx 的手機摘要 sheet → 純文字 (邏輯同 extract_mobile_summary_text.py)."""
    from openpyxl import load_workbook
    wb = load_workbook(str(XLSX), data_only=True)
    if MOBILE_SHEET not in wb.sheetnames:
        raise RuntimeError(f"找不到「{MOBILE_SHEET}」sheet — 可用: {wb.sheetnames[:5]}")

    ws = wb[MOBILE_SHEET]
    lines = [('' if ws[f'C{r}'].value is None else str(ws[f'C{r}'].value))
             for r in range(1, ws.max_row + 1)]
    while lines and lines[0] == '':
        lines.pop(0)
    while lines and lines[-1] == '':
        lines.pop()

    out, prev_blank = [], False
    for ln in lines:
        if ln == '':
            if not prev_blank:
                out.append('')
            prev_blank = True
        else:
            out.append(ln)
            prev_blank = False
    return '\n'.join(out)


# ════════════════════════════════════════════════════════════════════
#  Telegram 送出 / 編輯
# ════════════════════════════════════════════════════════════════════

def push_text(api: str, chat_id: str, label: str, text: str,
              prev: dict, force_new: bool, force: bool = False) -> dict:
    """送出或編輯一則文字訊息。回傳新的 {message_id, hash}。

    內容沒變 → 完全不打 API (Telegram 也會回 400 message is not modified)。
    有 message_id 且內容有變 → editMessageText。
    編輯失敗 (訊息被刪 / 過舊) → 退回重發,不讓使用者漏掉更新。

    force_new 與 force 是兩件事:
      force_new = 這個 chat 沒有可編輯的舊訊息 (換日 / 新增的目標) → 發新的
      force     = --force,略過「內容未變」的短路,硬打一次編輯。
                  用於手動重送: 你剛修好上游資料,hash 卻可能碰巧沒變。
    """
    import requests
    h = _hash(text)
    mid = (prev or {}).get('message_id')

    if not force and not force_new and mid and (prev or {}).get('hash') == h:
        print(f"  [Telegram] · {label}內容未變,不動")
        return {'message_id': mid, 'hash': h}

    if not force_new and mid:
        r = requests.post(f"{api}/editMessageText", data={
            'chat_id': chat_id, 'message_id': mid,
            'text': text[:4096], 'disable_web_page_preview': 'true',
        }, timeout=20)
        if r.status_code == 200:
            print(f"  [Telegram] ✎ {label}已更新 (msg {mid}, {len(text)} 字元)")
            return {'message_id': mid, 'hash': h}
        desc = ''
        try:
            desc = (r.json() or {}).get('description', '')
        except Exception:
            desc = r.text[:120]
        if 'not modified' in desc:
            print(f"  [Telegram] · {label}內容未變 (API 回報)")
            return {'message_id': mid, 'hash': h}
        print(f"  [Telegram] ⚠️ {label}編輯失敗 ({desc[:80]}),改為重發")

    r = requests.post(f"{api}/sendMessage", data={
        'chat_id': chat_id, 'text': text[:4096],
        'disable_web_page_preview': 'true',
    }, timeout=20)
    if r.status_code != 200:
        print(f"  [Telegram] ✗ {label}推送失敗 HTTP {r.status_code}: {r.text[:200]}")
        return {}
    new_mid = ((r.json() or {}).get('result') or {}).get('message_id')
    print(f"  [Telegram] ✓ {label}已推送 (msg {new_mid}, {len(text)} 字元)")
    return {'message_id': new_mid, 'hash': h}


def push_document(api: str, chat_id: str, caption: str,
                  prev: dict, force_new: bool, data_changed: bool,
                  force: bool = False) -> dict:
    """送出或替換 Excel 附件。

    Excel 每次 regen 內部時間戳都會變,檔案 hash 一定不同,拿來當判斷基準
    會導致每次都重傳 570KB。改以「摘要內容有沒有變」為準 — 摘要沒變表示
    當日數據沒有實質更新,附件也不需要換。
    """
    import requests
    if not XLSX.exists():
        print(f"  [Telegram] ⚠️ 找不到 {XLSX},略過附件")
        return prev or {}

    mid = (prev or {}).get('message_id')
    size_kb = XLSX.stat().st_size / 1024

    if not force and not force_new and mid and not data_changed:
        print(f"  [Telegram] · Excel 資料未變,不動")
        return prev

    if not force_new and mid:
        media = json.dumps({'type': 'document', 'media': 'attach://f',
                            'caption': caption}, ensure_ascii=False)
        with open(XLSX, 'rb') as fh:
            r = requests.post(f"{api}/editMessageMedia",
                              data={'chat_id': chat_id, 'message_id': mid,
                                    'media': media},
                              files={'f': ('latest.xlsx', fh)}, timeout=180)
        if r.status_code == 200:
            print(f"  [Telegram] ✎ Excel 已更新 (msg {mid}, {size_kb:,.0f} KB)")
            return {'message_id': mid}
        desc = ''
        try:
            desc = (r.json() or {}).get('description', '')
        except Exception:
            desc = r.text[:120]
        print(f"  [Telegram] ⚠️ Excel 編輯失敗 ({desc[:80]}),改為重發")

    with open(XLSX, 'rb') as fh:
        r = requests.post(f"{api}/sendDocument",
                          data={'chat_id': chat_id, 'caption': caption},
                          files={'document': ('latest.xlsx', fh)}, timeout=180)
    if r.status_code != 200:
        print(f"  [Telegram] ⚠️ Excel 推送失敗 HTTP {r.status_code}: {r.text[:200]}")
        return prev or {}
    new_mid = ((r.json() or {}).get('result') or {}).get('message_id')
    print(f"  [Telegram] ✓ Excel 已推送 (msg {new_mid}, {size_kb:,.0f} KB)")
    return {'message_id': new_mid}


# ════════════════════════════════════════════════════════════════════
#  主流程
# ════════════════════════════════════════════════════════════════════

def main() -> int:
    token = os.environ.get('TELEGRAM_BOT_TOKEN', '').strip()
    ids = chat_ids()

    if not DRY_RUN and (not token or not ids):
        print("  [Telegram] TELEGRAM_BOT_TOKEN / TELEGRAM_CHAT_ID 未設,跳過推播")
        print("             (見 docs/TELEGRAM_BOT_SETUP.md)")
        return 0

    trade_date = read_trade_date()
    if not trade_date:
        print("  [Telegram] 拿不到 trade_date,跳過推播")
        return 0

    if not XLSX.exists():
        print(f"  [Telegram] 找不到 {XLSX},跳過推播")
        return 0

    try:
        summary = extract_mobile_summary()
    except Exception as e:
        print(f"  [Telegram] 萃取手機摘要失敗: {e}")
        return 0        # 不讓排程判定為失敗 — 資料本身已經抓好了

    d_fmt = (f"{trade_date[:4]}/{trade_date[4:6]}/{trade_date[6:8]}"
             if len(trade_date) == 8 else trade_date)

    if DRY_RUN:
        print(f"  [DRY RUN] trade_date={trade_date}, 摘要 {len(summary)} 字元")
        print(f"  [DRY RUN] 收件對象 {len(ids)}: {', '.join(ids) or '(未設)'}")
        print(f"  [DRY RUN] Excel: {XLSX} ({XLSX.stat().st_size / 1024:,.0f} KB)")
        print('─' * 55)
        print(summary)
        print('─' * 55)
        return 0

    # ── 判斷是「新的一天」還是「同一天重跑」 ──
    st = load_state()
    # 刻意不看 FORCE: --force 的意思是「重新推一次」,不是「重貼一份新的」。
    # 同一交易日仍走原地更新,編輯失敗才由 push_text/push_document 退回重發。
    same_day = st.get('trade_date') == trade_date
    if not same_day:
        st = {'trade_date': trade_date, 'targets': {}}       # 換日 → 全部當新的發
        print(f"  [Telegram] 新交易日 {trade_date} → 發送新訊息")
    else:
        how = '強制更新原訊息' if FORCE else '有異動則更新原訊息'
        print(f"  [Telegram] {trade_date} 今日已推過 → {how}")

    targets = migrate_targets(st, ids)
    if same_day and not targets:
        # 舊版 (v3.73.1-3) 狀態檔只有 trade_date、沒有任何 message_id,無從編輯。
        # 此時「重發」會洗版、「跳過」只是少一次更新 → 選跳過,較不擾人。
        # 隔天換日就會寫入新格式,之後都能正常編輯。
        print(f"  [Telegram] {trade_date} 今日已推過,但狀態檔為舊格式"
              f"(無 message_id 可編輯) → 本次跳過,明日起正常更新")
        return 0

    api = f"https://api.telegram.org/bot{token}"
    caption = f"📋 Chip Radar · {d_fmt} 完整報表"

    for cid in ids:
        tgt = targets.get(cid) or {}
        # 今天才加進來的目標 (例如新增群組) 在本日還沒有自己的 message_id,
        # 必須當新的發 —— 沿用 same_day 會讓它整天都收不到東西。
        force_new = (not same_day) or not tgt

        summary_prev = tgt.get('summary') or {}
        summary_new = push_text(api, cid, f'手機摘要→{cid}', summary,
                                summary_prev, force_new, FORCE)
        if summary_new:
            tgt['summary'] = summary_new
        data_changed = (force_new or FORCE
                        or summary_new.get('hash') != summary_prev.get('hash'))

        res_doc = push_document(api, cid, caption, tgt.get('document') or {},
                                force_new, data_changed, FORCE)
        if res_doc:
            tgt['document'] = res_doc

        targets[cid] = tgt
        # 每個目標推完就存: 中途某個 chat 失敗 (被踢出群組之類) 時,
        # 前面成功的 message_id 不會跟著丟掉而在下次重發。
        save_state(st)

    save_state(st)
    return 0


if __name__ == '__main__':
    sys.exit(main())
